from pathlib import Path
from datetime import datetime
from typing import Dict, List
import hashlib
import re

import pandas as pd

import cadastro_service
import config

from leitor_excel_clientes import ler_excel_cliente
from leitor_pdf_clientes import ler_pdf_cliente
from padronizador_pedidos import padronizar_pedidos
from gerador_erro import gerar_arquivo_erro, gerar_arquivo_validacao_importacao, gerar_excel_validacao_completa
from terminal_logger import get_terminal_logger
import rastreabilidade_layouts
from parsers_excel.coelho_diniz import (
    MISSING_PAIR_MESSAGE as COELHO_DINIZ_MISSING_PAIR_MESSAGE,
    gerar_excel_validacao_coelho,
    processar_lote_coelho_diniz,
)


terminal_log = get_terminal_logger("importacao")


def nome_arquivo_curto(valor: str, limite: int = 42) -> str:
    texto = re.sub(r"[^A-Za-z0-9]+", "_", str(valor or "arquivo")).strip("_")
    if not texto:
        texto = "arquivo"
    if len(texto) <= limite:
        return texto
    digest = hashlib.sha1(texto.encode("utf-8", errors="ignore")).hexdigest()[:8]
    return f"{texto[: max(8, limite - 9)].rstrip('_')}_{digest}"


def selecionar_leitor(layout: Dict[str, str]):
    tipo_arquivo = str(layout.get("tipo_arquivo", "")).strip().upper()

    if tipo_arquivo == "EXCEL":
        return ler_excel_cliente

    if tipo_arquivo == "PDF":
        return ler_pdf_cliente

    return None


def layout_eh_coelho_diniz(layout: Dict[str, str] | None) -> bool:
    if not layout:
        return False
    return "COELHO DINIZ" in str(layout.get("nome_layout", "")).upper()


def layout_eh_grancoffee(layout: Dict[str, str] | None) -> bool:
    if not layout:
        return False
    texto = " ".join(str(layout.get(c, "")) for c in ["layout_id", "cliente_id", "nome_layout", "observacoes"]).upper()
    return "GRANCOFFEE" in texto or "GRAN COFFEE" in texto


def montar_resultado_arquivo(
    nome_arquivo: str,
    nome_layout: str,
    tipo_arquivo: str,
    status: str,
    mensagem: str,
    qtd_linhas_lidas: int = 0,
    qtd_linhas_validas: int = 0,
    qtd_linhas_descartadas: int = 0,
    qtd_linhas_inseridas: int = 0,
    alertas: List[str] | None = None,
    arquivo_validacao: str | None = None,
    arquivo_erro: str | None = None,
    rastreabilidade: Dict[str, object] | None = None,
    resumo_conversao: Dict[str, object] | None = None,
) -> Dict[str, object]:
    return {
        "nome_arquivo": nome_arquivo,
        "nome_layout": nome_layout,
        "tipo_arquivo": tipo_arquivo,
        "status": status,
        "mensagem": mensagem,
        "qtd_linhas_lidas": qtd_linhas_lidas,
        "qtd_linhas_validas": qtd_linhas_validas,
        "qtd_linhas_descartadas": qtd_linhas_descartadas,
        "qtd_linhas_inseridas": qtd_linhas_inseridas,
        "alertas": alertas or [],
        "arquivo_validacao": arquivo_validacao or "",
        "arquivo_erro": arquivo_erro or "",
        "layout_possui_conversao": bool((resumo_conversao or {}).get("layout_possui_conversao", False)),
        "qtd_itens_convertidos": int((resumo_conversao or {}).get("itens_convertidos", 0) or 0),
        "qtd_itens_nao_convertidos": int((resumo_conversao or {}).get("itens_nao_convertidos", 0) or 0),
        "qtd_itens_validar_conversao": int((resumo_conversao or {}).get("itens_validar_conversao", 0) or 0),
        "qtd_itens_sem_conversao": int((resumo_conversao or {}).get("itens_sem_conversao", 0) or 0),
        "rastreabilidade": rastreabilidade or {},
    }



def contar_linhas_modelo_grancoffee(arquivo_validacao: str | None) -> int:
    """Conta itens gravados na aba Modelo Robô KOF para Enviar.

    A rotina dedicada da Grancoffee não insere fila automática antes da validação,
    mas deve informar corretamente quantas linhas foram geradas no Excel para não
    parecer que o pedido ficou vazio na tela.
    """
    if not arquivo_validacao:
        return 0
    try:
        from openpyxl import load_workbook
        wb = load_workbook(arquivo_validacao, read_only=True, data_only=True)
        if "Modelo Robô KOF para Enviar" not in wb.sheetnames:
            return 0
        ws = wb["Modelo Robô KOF para Enviar"]
        total = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            matricula = row[1] if len(row) > 1 else None
            sku = row[2] if len(row) > 2 else None
            qtd = row[3] if len(row) > 3 else None
            pedido = row[4] if len(row) > 4 else None
            if matricula and sku and qtd not in (None, "") and pedido:
                total += 1
        return total
    except Exception as exc:
        terminal_log.warning("[IMPORTACAO][GRANCOFFEE] Não foi possível contar linhas do Excel de validação: %s", exc)
        return 0


def salvar_descartes_em_erro(df_descartadas: pd.DataFrame | None, nome_arquivo: str):
    """Não gera mais Excel separado de erro na importação.

    Padrão operacional Robô KOF: todo descarte/alerta/pendência deve ficar
    dentro do Excel consolidado de validação, nas abas ALERTAS_ERROS,
    Itens Bloqueados Fila, Cadastrar CNPJ, Pendências GLN e Logs do Processamento
    quando aplicável. Arquivo separado de erro só deve existir em falha fatal
    que impeça a criação do Excel principal.
    """
    if df_descartadas is None or getattr(df_descartadas, "empty", True):
        return None
    terminal_log.warning(
        "[IMPORTACAO][EXCEL_UNICO] %s pendência(s)/descarte(s) mantidos no Excel consolidado; arquivo ERR separado não será gerado. arquivo=%s",
        len(df_descartadas),
        nome_arquivo,
    )
    return None


def salvar_validacao_importacao(
    nome_arquivo: str,
    layout: Dict[str, str],
    leitura: Dict[str, object],
    pad: Dict[str, object],
    alertas: List[str] | None = None,
    sufixo: str = "",
    df_rastreabilidade: pd.DataFrame | None = None,
):
    base_nome = nome_arquivo_curto(Path(nome_arquivo).stem)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    sufixo_nome = f"_{sufixo}" if sufixo else ""
    out_path = config.PEDIDOS_A_VALIDAR_DIR / f"VALIDACAO_IMPORTACAO_{base_nome}{sufixo_nome}_{timestamp}.xlsx"
    terminal_log.info(
        "[IMPORTACAO] Gerando Excel de validacao | arquivo=%s | layout=%s | destino=%s",
        nome_arquivo,
        layout.get("nome_layout", ""),
        out_path,
    )
    resumo = {
        "arquivo": nome_arquivo,
        "layout": layout.get("nome_layout", ""),
        "tipo_arquivo": layout.get("tipo_arquivo", ""),
        "linhas_lidas": leitura.get("qtd_linhas_lidas", 0),
        "linhas_validas": pad.get("qtd_linhas_validas", 0),
        "linhas_descartadas": pad.get("qtd_linhas_descartadas", 0),
        "paginas_pdf_total": leitura.get("paginas_pdf_total", ""),
        "paginas_pdf_processadas": leitura.get("paginas_pdf_processadas", ""),
        "paginas_pdf_sem_texto": leitura.get("paginas_pdf_sem_texto", ""),
        "motores_pdf": leitura.get("motores_pdf", ""),
        "itens_extraidos": leitura.get("qtd_itens_extraidos", ""),
        "itens_ignorados": leitura.get("qtd_itens_ignorados", ""),
        "total_arquivos_lote": leitura.get("total_arquivos_lote", ""),
        "arquivos_consolidados": "; ".join(leitura.get("arquivos_consolidados", []) or []),
        "alertas_parser_registrados_excel": (0 if leitura.get("df_itens_ignorados") is None else len(leitura.get("df_itens_ignorados"))),
        "mensagem": pad.get("mensagem", ""),
        "layout_possui_conversao": pad.get("layout_possui_conversao", False),
        "itens_convertidos": pad.get("qtd_itens_convertidos", 0),
        "itens_nao_convertidos": pad.get("qtd_itens_nao_convertidos", 0),
        "itens_validar_conversao": pad.get("qtd_itens_validar_conversao", 0),
        "itens_sem_conversao": pad.get("qtd_itens_sem_conversao", 0),
        "fluxo": "VALIDACAO_PRIMEIRO_SEM_GERAR_TXT_OU_FILA",
    }
    if df_rastreabilidade is not None and not df_rastreabilidade.empty:
        rast_row = df_rastreabilidade.iloc[0].to_dict()
        resumo.update({
            "modo_rastreabilidade": "SIM",
            "layout_referencia_rastreabilidade": rast_row.get("nome_layout_referencia", ""),
            "confianca_rastreabilidade": rast_row.get("confianca", ""),
            "status_rastreabilidade": rast_row.get("status", ""),
            "motivo_rastreabilidade": rast_row.get("motivo", ""),
        })
    gerar_excel_validacao_completa(
        out_path,
        resumo=resumo,
        df_intermediario=leitura.get("df_intermediario"),
        df_final=pad.get("df_final"),
        df_descartadas=pad.get("df_descartadas"),
        alertas=alertas,
        df_validas=pad.get("df_validas"),
        df_auditoria=leitura.get("df_auditoria_paginas"),
        df_alertas_extracao=leitura.get("df_itens_ignorados"),
        df_rastreabilidade=df_rastreabilidade,
    )
    return str(out_path)


def salvar_validacao_erro_leitura(
    nome_arquivo: str,
    layout: Dict[str, str],
    mensagem: str,
    alertas: List[str] | None = None,
    df_auditoria: pd.DataFrame | None = None,
    df_rastreabilidade: pd.DataFrame | None = None,
):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_nome = nome_arquivo_curto(Path(nome_arquivo).stem)
    out_path = config.PEDIDOS_A_VALIDAR_DIR / f"VALIDACAO_ERRO_LEITURA_{base_nome}_{timestamp}.xlsx"
    terminal_log.error(
        "[IMPORTACAO] Gerando Excel de erro de leitura | arquivo=%s | layout=%s | erro=%s",
        nome_arquivo,
        layout.get("nome_layout", ""),
        mensagem,
    )
    resumo = {
        "arquivo": nome_arquivo,
        "layout": layout.get("nome_layout", ""),
        "tipo_arquivo": layout.get("tipo_arquivo", ""),
        "status": "ERRO_LEITURA",
        "mensagem": mensagem,
        "fluxo": "VALIDACAO_PRIMEIRO_SEM_GERAR_TXT_OU_FILA",
    }
    if df_rastreabilidade is not None and not df_rastreabilidade.empty:
        rast_row = df_rastreabilidade.iloc[0].to_dict()
        resumo.update({
            "modo_rastreabilidade": "SIM",
            "layout_referencia_rastreabilidade": rast_row.get("nome_layout_referencia", ""),
            "confianca_rastreabilidade": rast_row.get("confianca", ""),
            "status_rastreabilidade": rast_row.get("status", ""),
            "motivo_rastreabilidade": rast_row.get("motivo", ""),
        })
    gerar_excel_validacao_completa(
        out_path,
        resumo=resumo,
        alertas=(alertas or []) + [mensagem],
        df_auditoria=df_auditoria,
        df_rastreabilidade=df_rastreabilidade,
    )
    return str(out_path)


def aplicar_data_remessa_manual(leitura: Dict[str, object], layout: Dict[str, str], item_importacao: Dict[str, str]):
    data_manual = str(item_importacao.get("data_remessa_manual", "") or "").strip()
    nome_layout = str(layout.get("nome_layout", "")).upper()

    if not data_manual:
        if "SEMPRE VALE" in nome_layout:
            alertas = list(leitura.get("alertas", []) or [])
            alertas.append("Data remessa manual obrigatoria nao informada; itens mantidos no Excel de validacao com DATA_INVALIDA.")
            leitura["alertas"] = sorted({str(a) for a in alertas if str(a).strip()})
            layout["regra_data_entrega"] = "COLUNA"
            return True, ""
        return True, ""

    df_intermediario = leitura.get("df_intermediario")
    if df_intermediario is not None and not df_intermediario.empty:
        df_intermediario = df_intermediario.copy()
        df_intermediario["data_entrega_lida"] = data_manual
        leitura["df_intermediario"] = df_intermediario
    layout["regra_data_entrega"] = "COLUNA"
    return True, ""




def _valor_int_seguro(valor) -> int:
    """Converte contadores vindos dos leitores para inteiro sem quebrar o lote."""
    try:
        if valor is None or valor == "":
            return 0
        return int(float(valor))
    except Exception:
        return 0


def _somar_contador(leituras: List[Dict[str, object]], chave: str) -> int:
    return sum(_valor_int_seguro(leitura.get(chave, 0)) for leitura in leituras)


def _concatenar_dataframes(frames: List[pd.DataFrame | None]) -> pd.DataFrame:
    validos = [df for df in frames if df is not None and not df.empty]
    if not validos:
        return pd.DataFrame()
    return pd.concat(validos, ignore_index=True, sort=False)


def _anexar_origem(df: pd.DataFrame | None, nome_arquivo: str) -> pd.DataFrame | None:
    if df is None:
        return None
    if df.empty:
        return df
    df = df.copy()
    if "arquivo_origem" not in df.columns:
        df.insert(0, "arquivo_origem", nome_arquivo)
    else:
        df["arquivo_origem"] = df["arquivo_origem"].fillna("").astype(str)
        df.loc[df["arquivo_origem"].str.strip() == "", "arquivo_origem"] = nome_arquivo
    return df


def _montar_df_rastreabilidade_lote(dfs: List[pd.DataFrame | None]) -> pd.DataFrame:
    return _concatenar_dataframes(dfs)


def _nome_lote_layout(layout: Dict[str, str], itens_importacao: List[Dict[str, str]]) -> str:
    nome_layout = nome_arquivo_curto(layout.get("nome_layout", "LOTE_LAYOUT"), limite=32)
    tipo = str(layout.get("tipo_arquivo", "PDF") or "PDF").upper()
    return f"LOTE_{nome_layout}_{tipo}_{len(itens_importacao)}_ARQUIVOS"


def _layout_referencia_item(item: Dict[str, str], layout: Dict[str, str] | None = None) -> str:
    """Retorna o layout real aplicado ao item, inclusive quando veio por rastreabilidade."""
    rast = item.get("rastreabilidade") or {}
    if isinstance(rast, dict):
        nome_ref = str(rast.get("nome_layout_referencia", "") or "").strip()
        if nome_ref:
            return nome_ref.upper()
    if layout:
        return str(layout.get("nome_layout", "") or "").strip().upper()
    return str(item.get("layout_nome", "") or "").replace("RASTREABILIDADE ->", "").strip().upper()


def _rastreabilidade_bloqueia_consolidacao(item: Dict[str, str], layout: Dict[str, str]) -> bool:
    """Bloqueia apenas rastreabilidade genérica/incerta.

    Antes, qualquer item com modo_rastreabilidade=SIM era impedido de consolidar.
    Isso fazia lotes Superlar, e outros reconhecidos automaticamente por CNPJ/GLN,
    saírem em um Excel por pedido. Agora consolidamos quando todos apontam para
    o mesmo layout homologado; continuamos bloqueando genérico/homologação para
    evitar mistura de redes.
    """
    modo = str(item.get("modo_rastreabilidade", "") or "").strip().upper() in {"SIM", "S", "TRUE", "1"}
    if not modo:
        return False

    nome_layout = str(layout.get("nome_layout", "") or "").strip().upper()
    rast = item.get("rastreabilidade") or {}
    if not isinstance(rast, dict) or not rast:
        # Compatibilidade: algumas entradas antigas carregam apenas modo_rastreabilidade=SIM
        # e layout_referencia/layout_nome, sem o dicionário completo de rastreabilidade.
        # Se a referência textual aponta para o mesmo layout homologado, pode consolidar.
        ref_textual = str(item.get("layout_referencia", "") or item.get("layout_nome", "") or "").replace("RASTREABILIDADE ->", "").strip().upper()
        status_textual = str(item.get("status_rastreabilidade", "") or item.get("status", "") or "").strip().upper()
        if ref_textual and ref_textual == nome_layout and not any(x in ref_textual for x in ["GENERIC", "GENÉRIC", "HOMOLOG", "RASTREABILIDADE"]):
            return False
        if ref_textual == nome_layout and ("AUTO" in status_textual or "DEPARA" in status_textual):
            return False
        return True

    nome_ref = str(rast.get("nome_layout_referencia", "") or "").strip().upper()
    status = str(rast.get("status", "") or "").strip().upper()
    sucesso = str(rast.get("sucesso", "") or "").strip().upper() in {"1", "TRUE", "SIM", "S"}
    aplicar_auto = str(rast.get("aplicar_automaticamente", "") or "").strip().upper() in {"1", "TRUE", "SIM", "S"}

    if not nome_ref or "GENERIC" in nome_ref or "GENÉRIC" in nome_ref or "HOMOLOG" in nome_ref:
        return True
    if "RASTREABILIDADE GENERICA" in status or "HOMOLOG" in status:
        return True
    if nome_layout and nome_ref != nome_layout:
        return True
    # De/para automático ou rastreabilidade automática segura podem consolidar.
    if sucesso and (aplicar_auto or "DEPARA" in status or "AUTO" in status):
        return False
    return True


def _deve_consolidar_lote_layout(layout: Dict[str, str] | None, itens: List[Dict[str, str]]) -> bool:
    """Define quando varios arquivos devem gerar um unico Excel de validacao.

    Regra operacional solicitada:
    - varios PDFs ou Excels da mesma rede/layout devem sair em um unico Excel consolidado;
    - Coelho Diniz ja possui rotina propria de lote;
    - rastreabilidade generica nao e agrupada automaticamente, porque pode misturar redes desconhecidas;
    - rastreabilidade automática/de-para para o mesmo layout homologado deve consolidar.
    """
    if not layout or len(itens) <= 1:
        return False
    if layout_eh_coelho_diniz(layout):
        return False
    tipo = str(layout.get("tipo_arquivo", "") or "").strip().upper()
    if tipo not in {"PDF", "EXCEL"}:
        return False
    nome_layout = str(layout.get("nome_layout", "") or "").upper()
    if "RASTREABILIDADE" in nome_layout or "HOMOLOGA" in nome_layout:
        return False

    referencias = {_layout_referencia_item(item, layout) for item in itens}
    referencias = {ref for ref in referencias if ref}
    if len(referencias) > 1:
        terminal_log.warning(
            "[IMPORTACAO][LOTE] Consolidacao bloqueada: referencias de layout diferentes no mesmo grupo: %s",
            sorted(referencias),
        )
        return False

    for item in itens:
        if _rastreabilidade_bloqueia_consolidacao(item, layout):
            terminal_log.info(
                "[IMPORTACAO][LOTE] Consolidacao bloqueada para %s: rastreabilidade generica/incerta.",
                item.get("nome_arquivo") or item.get("caminho_arquivo"),
            )
            return False
    return True


def _resultado_lote_para_item(
    item: Dict[str, str],
    layout: Dict[str, str],
    status: str,
    mensagem: str,
    leitura: Dict[str, object],
    pad: Dict[str, object] | None = None,
    alertas: List[str] | None = None,
    arquivo_validacao: str | None = None,
    arquivo_erro: str | None = None,
) -> Dict[str, object]:
    pad = pad or {}
    return montar_resultado_arquivo(
        item.get("nome_arquivo") or Path(item.get("caminho_arquivo", "")).name,
        layout.get("nome_layout", ""),
        layout.get("tipo_arquivo", ""),
        status,
        mensagem,
        int(leitura.get("qtd_linhas_lidas", 0) or 0),
        int(pad.get("qtd_linhas_validas", 0) or 0),
        int(pad.get("qtd_linhas_descartadas", 0) or 0),
        0,
        alertas or [],
        arquivo_validacao,
        arquivo_erro,
        resumo_conversao=pad.get("resumo_conversao", {}),
    )


def processar_lote_grancoffee_importacao(itens_importacao: List[Dict[str, str]]) -> List[Dict[str, object]]:
    """Processa Grancoffee por rotina dedicada.

    O layout precisa cruzar anexo Excel + corpo do e-mail. Por isso ele não passa
    pelo leitor genérico de Excel/PDF. O primeiro output continua sendo o Excel
    de validação, sem gerar TXT/fila KOF automaticamente.
    """
    if not itens_importacao:
        return []
    layout = cadastro_service.buscar_layout(layout_id=itens_importacao[0].get("layout_id")) or {}
    nome_layout = layout.get("nome_layout", "GRANCOFFEE")
    arquivos = [item.get("caminho_arquivo", "") for item in itens_importacao if item.get("caminho_arquivo")]
    try:
        from layouts.rede_grancoffee import processar_grancoffee
        arquivo_validacao = processar_grancoffee(arquivos, str(config.PEDIDOS_A_VALIDAR_DIR))
        qtd_modelo = contar_linhas_modelo_grancoffee(arquivo_validacao)
        status = "VALIDACAO_GERADA" if arquivo_validacao else "ERRO"
        mensagem = (
            f"Grancoffee processado por rotina dedicada. Arquivos: {len(arquivos)} | "
            f"Linhas no Modelo Robô KOF para Enviar: {qtd_modelo} | "
            f"Excel de validação: {arquivo_validacao}. Nenhuma linha foi enviada para a fila automaticamente; validar o Excel antes de gerar TXT/fila KOF."
        )
        alertas = [
            "LAYOUT_DEDICADO_GRANCOFFEE: data oficial da remessa vem do corpo do e-mail por matrícula + pedido.",
            f"VALIDACAO_GERADA: {arquivo_validacao}",
        ]
        return [
            montar_resultado_arquivo(
                item.get("nome_arquivo") or Path(item.get("caminho_arquivo", "")).name,
                nome_layout,
                layout.get("tipo_arquivo", item.get("tipo_arquivo", "")),
                status,
                mensagem,
                qtd_modelo,
                qtd_modelo,
                0,
                0,
                alertas,
                arquivo_validacao,
            )
            for item in itens_importacao
        ]
    except Exception as exc:
        terminal_log.exception("[IMPORTACAO][GRANCOFFEE] Falha no processamento dedicado.")
        return [
            montar_resultado_arquivo(
                item.get("nome_arquivo") or Path(item.get("caminho_arquivo", "")).name,
                nome_layout,
                layout.get("tipo_arquivo", item.get("tipo_arquivo", "")),
                "ERRO",
                f"Falha no processamento dedicado Grancoffee: {exc}",
            )
            for item in itens_importacao
        ]


def processar_lote_layout_importacao(itens_importacao: List[Dict[str, str]]) -> List[Dict[str, object]]:
    """Processa varios arquivos da mesma rede/layout e gera um unico Excel consolidado.

    O objetivo e evitar 1 Excel por arquivo quando o usuario envia varios pedidos da
    mesma rede, preservando rastreabilidade por arquivo de origem e mantendo o
    primeiro output como Excel de validacao.
    """
    if not itens_importacao:
        return []

    layout_base = cadastro_service.buscar_layout(layout_id=itens_importacao[0].get("layout_id"))
    if not layout_base:
        return [
            montar_resultado_arquivo(
                item.get("nome_arquivo") or Path(item.get("caminho_arquivo", "")).name,
                "",
                item.get("tipo_arquivo", ""),
                "ERRO",
                "Layout nao encontrado para lote consolidado",
            )
            for item in itens_importacao
        ]

    layout = layout_base.copy()
    nome_layout = layout.get("nome_layout", "")
    terminal_log.info(
        "[IMPORTACAO][LOTE] Iniciando lote consolidado | layout=%s | arquivos=%s",
        nome_layout,
        len(itens_importacao),
    )

    if layout_eh_grancoffee(layout):
        # Grancoffee precisa processar o lote completo junto, pois a regra oficial
        # cruza corpo do e-mail + anexos por matrícula e número do pedido.
        return processar_lote_grancoffee_importacao(itens_importacao)

    mapeamentos = cadastro_service.buscar_mapeamentos_do_layout(layout["layout_id"])
    leitor = selecionar_leitor(layout)
    if leitor is None:
        mensagem = f"Tipo de arquivo nao suportado para lote: {layout.get('tipo_arquivo', '')}"
        return [
            montar_resultado_arquivo(
                item.get("nome_arquivo") or Path(item.get("caminho_arquivo", "")).name,
                nome_layout,
                layout.get("tipo_arquivo", ""),
                "ERRO",
                mensagem,
            )
            for item in itens_importacao
        ]

    leituras_sucesso: List[Dict[str, object]] = []
    frames_intermediarios: List[pd.DataFrame | None] = []
    frames_auditoria: List[pd.DataFrame | None] = []
    frames_alertas_extracao: List[pd.DataFrame | None] = []
    frames_rastreabilidade: List[pd.DataFrame | None] = []
    alertas_lote: List[str] = []
    arquivos_com_erro: List[str] = []

    for item in itens_importacao:
        caminho = item["caminho_arquivo"]
        nome_arquivo = item.get("nome_arquivo") or Path(caminho).name
        layout_item = layout_base.copy()
        terminal_log.info("[IMPORTACAO][LOTE] Lendo arquivo do lote: %s", nome_arquivo)
        try:
            leitura = leitor(caminho, layout_item, mapeamentos)
            df_rastreabilidade = aplicar_metadados_rastreabilidade(item, layout_item, leitura)
            if df_rastreabilidade is not None and not df_rastreabilidade.empty:
                frames_rastreabilidade.append(_anexar_origem(df_rastreabilidade, nome_arquivo))
        except Exception as exc:
            leitura = {
                "sucesso": False,
                "mensagem": f"Erro ao ler arquivo no lote consolidado: {exc}",
                "alertas": [f"{nome_arquivo}: erro ao ler arquivo no lote consolidado: {exc}"],
            }

        frames_auditoria.append(_anexar_origem(leitura.get("df_auditoria_paginas"), nome_arquivo))
        frames_alertas_extracao.append(_anexar_origem(leitura.get("df_itens_ignorados"), nome_arquivo))
        for alerta in leitura.get("alertas", []) or []:
            alertas_lote.append(f"{nome_arquivo}: {alerta}")

        if not leitura.get("sucesso"):
            msg = leitura.get("mensagem") or "Falha de leitura no lote consolidado"
            arquivos_com_erro.append(f"{nome_arquivo}: {msg}")
            frames_alertas_extracao.append(pd.DataFrame([
                {
                    "arquivo_origem": nome_arquivo,
                    "status_extracao": "ERRO_LEITURA_LOTE",
                    "alerta_extracao": msg,
                }
            ]))
            continue

        ok_data, msg_data = aplicar_data_remessa_manual(leitura, layout_item, item)
        if not ok_data:
            arquivos_com_erro.append(f"{nome_arquivo}: {msg_data}")
            frames_alertas_extracao.append(pd.DataFrame([
                {
                    "arquivo_origem": nome_arquivo,
                    "status_extracao": "ERRO_DATA_REMESSA_LOTE",
                    "alerta_extracao": msg_data,
                }
            ]))
            continue

        df_inter = _anexar_origem(leitura.get("df_intermediario"), nome_arquivo)
        if df_inter is not None and not df_inter.empty:
            frames_intermediarios.append(df_inter)
        leituras_sucesso.append(leitura)

    df_intermediario_lote = _concatenar_dataframes(frames_intermediarios)
    df_auditoria_lote = _concatenar_dataframes(frames_auditoria)
    df_alertas_lote = _concatenar_dataframes(frames_alertas_extracao)
    df_rastreabilidade_lote = _montar_df_rastreabilidade_lote(frames_rastreabilidade)

    leitura_lote = {
        "sucesso": not df_intermediario_lote.empty,
        "mensagem": "Leitura consolidada em lote concluida" if not df_intermediario_lote.empty else "Nenhum item extraido no lote consolidado",
        "df_intermediario": df_intermediario_lote,
        "df_auditoria_paginas": df_auditoria_lote,
        "df_itens_ignorados": df_alertas_lote,
        "df_rastreabilidade": df_rastreabilidade_lote,
        "alertas": sorted({str(a) for a in alertas_lote if str(a).strip()}),
        "qtd_linhas_lidas": int(len(df_intermediario_lote) + len(df_alertas_lote)),
        "qtd_itens_extraidos": int(len(df_intermediario_lote)),
        "qtd_itens_ignorados": int(len(df_alertas_lote)),
        "total_arquivos_lote": len(itens_importacao),
        "arquivos_consolidados": [item.get("nome_arquivo") or Path(item.get("caminho_arquivo", "")).name for item in itens_importacao],
        "paginas_pdf_total": _somar_contador(leituras_sucesso, "paginas_pdf_total"),
        "paginas_pdf_processadas": _somar_contador(leituras_sucesso, "paginas_pdf_processadas"),
        "paginas_pdf_sem_texto": _somar_contador(leituras_sucesso, "paginas_pdf_sem_texto"),
        "motores_pdf": ", ".join(sorted({str(l.get("motores_pdf", "")).strip() for l in leituras_sucesso if str(l.get("motores_pdf", "")).strip()})),
    }

    nome_lote = _nome_lote_layout(layout, itens_importacao)
    if df_intermediario_lote.empty:
        mensagem = "Nenhum item extraido no lote consolidado. " + " | ".join(arquivos_com_erro[:10])
        arquivo_validacao = salvar_validacao_erro_leitura(
            nome_lote,
            layout,
            mensagem,
            leitura_lote.get("alertas", []) + arquivos_com_erro,
            df_auditoria_lote,
            df_rastreabilidade_lote,
        )
        alertas = leitura_lote.get("alertas", []) + arquivos_com_erro + [f"VALIDACAO_GERADA: {arquivo_validacao}"]
        return [
            _resultado_lote_para_item(item, layout, "ERRO", mensagem, leitura_lote, {}, alertas, arquivo_validacao)
            for item in itens_importacao
        ]

    pad = padronizar_pedidos(df_intermediario_lote, layout)
    alertas = pad.get("alertas", []) + leitura_lote.get("alertas", [])
    if arquivos_com_erro:
        alertas.extend([f"ARQUIVO_COM_ERRO_NO_LOTE: {msg}" for msg in arquivos_com_erro])

    arquivo_validacao = salvar_validacao_importacao(
        nome_lote,
        layout,
        leitura_lote,
        pad,
        alertas,
        sufixo="CONSOLIDADO",
        df_rastreabilidade=df_rastreabilidade_lote,
    )
    alertas.append(f"VALIDACAO_GERADA: {arquivo_validacao}")
    alertas.append(
        f"LOTE_CONSOLIDADO: {len(itens_importacao)} arquivos do layout {nome_layout} consolidados em um unico Excel."
    )

    arquivo_erro = None
    if pad.get("qtd_linhas_descartadas", 0) > 0:
        arquivo_erro = salvar_descartes_em_erro(pad.get("df_descartadas"), f"{nome_lote}.xlsx")

    if not pad.get("sucesso") or arquivos_com_erro:
        status = "VALIDACAO_GERADA_COM_ALERTA"
    else:
        status = "VALIDACAO_GERADA"

    mensagem = (
        f"Lote consolidado para validacao. Arquivos: {len(itens_importacao)} | "
        f"Layout: {nome_layout} | Linhas validas: {pad.get('qtd_linhas_validas', 0)} | "
        f"Pendencias: {pad.get('qtd_linhas_descartadas', 0) + len(arquivos_com_erro)} | "
        f"Convertidos: {pad.get('qtd_itens_convertidos', 0)} | Não convertidos: {pad.get('qtd_itens_nao_convertidos', 0)} | "
        f"Excel de validacao: {arquivo_validacao}. Nenhuma linha foi inserida na fila automaticamente."
    )
    terminal_log.info("[IMPORTACAO][LOTE] Excel consolidado gerado: %s", arquivo_validacao)

    return [
        _resultado_lote_para_item(
            item,
            layout,
            status,
            mensagem,
            leitura_lote,
            pad,
            alertas,
            arquivo_validacao,
            arquivo_erro,
        )
        for item in itens_importacao
    ]

def montar_resultado_coelho_para_item(
    item_importacao: Dict[str, str],
    layout: Dict[str, str],
    status: str,
    mensagem: str,
    leitura: Dict[str, object] | None = None,
    pad: Dict[str, object] | None = None,
    linhas_inseridas: int = 0,
    alertas: List[str] | None = None,
    arquivo_validacao: str | None = None,
    arquivo_erro: str | None = None,
):
    leitura = leitura or {}
    pad = pad or {}
    return montar_resultado_arquivo(
        item_importacao.get("nome_arquivo") or Path(item_importacao.get("caminho_arquivo", "")).name,
        layout.get("nome_layout", ""),
        layout.get("tipo_arquivo", ""),
        status,
        mensagem,
        int(leitura.get("qtd_linhas_lidas", 0) or 0),
        int(pad.get("qtd_linhas_validas", 0) or 0),
        int(pad.get("qtd_linhas_descartadas", 0) or 0),
        linhas_inseridas,
        alertas or [],
        arquivo_validacao,
        arquivo_erro,
        resumo_conversao=pad.get("resumo_conversao", {}),
    )


def processar_lote_coelho_diniz_importacao(itens_importacao: List[Dict[str, str]]) -> List[Dict[str, object]]:
    if not itens_importacao:
        return []

    terminal_log.info("[IMPORTACAO][COELHO DINIZ] Iniciando lote com %s arquivo(s).", len(itens_importacao))
    layout = cadastro_service.buscar_layout(layout_id=itens_importacao[0].get("layout_id"))
    if not layout:
        terminal_log.error("[IMPORTACAO][COELHO DINIZ] Layout nao encontrado para o lote.")
        return [
            montar_resultado_arquivo(
                item.get("nome_arquivo") or Path(item.get("caminho_arquivo", "")).name,
                "",
                item.get("tipo_arquivo", ""),
                "ERRO",
                "Layout nao encontrado para Rede Coelho Diniz",
            )
            for item in itens_importacao
        ]

    layout = layout.copy()
    caminhos = [item["caminho_arquivo"] for item in itens_importacao]
    for caminho in caminhos:
        terminal_log.info("[IMPORTACAO][COELHO DINIZ] Arquivo recebido: %s", caminho)
    leitura = processar_lote_coelho_diniz(caminhos, layout)

    if not leitura.get("sucesso"):
        mensagem = leitura.get("mensagem") or COELHO_DINIZ_MISSING_PAIR_MESSAGE
        alertas = leitura.get("alertas", []) or [mensagem]
        terminal_log.error("[IMPORTACAO][COELHO DINIZ] Processamento bloqueado: %s", mensagem)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo_validacao = str(config.PEDIDOS_A_VALIDAR_DIR / f"VALIDACAO_COELHO_DINIZ_ERRO_{timestamp}.xlsx")
        gerar_excel_validacao_completa(
            Path(arquivo_validacao),
            resumo={
                "layout": layout.get("nome_layout", ""),
                "status": "ERRO_LEITURA",
                "mensagem": mensagem,
                "fluxo": "VALIDACAO_PRIMEIRO_SEM_GERAR_TXT_OU_FILA",
            },
            df_intermediario=leitura.get("df_intermediario"),
            alertas=alertas,
        )
        return [
            montar_resultado_coelho_para_item(
                item,
                layout,
                "ERRO",
                mensagem,
                leitura,
                {},
                0,
                alertas,
                arquivo_validacao,
            )
            for item in itens_importacao
        ]

    pad = padronizar_pedidos(leitura["df_intermediario"], layout)
    terminal_log.info(
        "[IMPORTACAO][COELHO DINIZ] Extracao concluida: %s linha(s) lida(s), %s valida(s), %s pendente(s).",
        leitura.get("qtd_linhas_lidas", 0),
        pad.get("qtd_linhas_validas", 0),
        pad.get("qtd_linhas_descartadas", 0),
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    arquivo_validacao = str(config.PEDIDOS_A_VALIDAR_DIR / f"VALIDACAO_COELHO_DINIZ_{timestamp}.xlsx")
    gerar_excel_validacao_coelho(
        Path(arquivo_validacao),
        leitura,
        pad.get("df_final"),
        pad.get("df_descartadas"),
    )
    terminal_log.info("[IMPORTACAO][COELHO DINIZ] Excel de validacao gerado: %s", arquivo_validacao)

    arquivo_erro = None
    if pad.get("qtd_linhas_descartadas", 0) > 0:
        arquivo_erro = salvar_descartes_em_erro(pad.get("df_descartadas"), f"COELHO_DINIZ_LOTE_{timestamp}.xlsx")

    linhas_inseridas = 0

    alertas = pad.get("alertas", []) + leitura.get("alertas", [])
    alertas.append(f"VALIDACAO_GERADA: {arquivo_validacao}")

    if pad.get("sucesso"):
        status = "VALIDACAO_GERADA_COM_ALERTA" if pad.get("qtd_linhas_descartadas", 0) > 0 else "VALIDACAO_GERADA"
        mensagem = (
            f"Rede Coelho Diniz processada para validacao. Nenhuma linha foi inserida na fila automaticamente. "
            f"Arquivo de validacao: {arquivo_validacao}"
        )
    else:
        status = "VALIDACAO_GERADA_COM_ALERTA"
        mensagem = (
            f"{pad.get('mensagem', 'Nenhuma linha valida apos padronizacao')}. "
            f"Nenhuma linha foi inserida na fila; arquivo de validacao gerado: {arquivo_validacao}"
        )

    return [
        montar_resultado_coelho_para_item(
            item,
            layout,
            status,
            mensagem,
            leitura,
            pad,
            linhas_inseridas,
            alertas,
            arquivo_validacao,
            arquivo_erro,
        )
        for item in itens_importacao
    ]



def _rastreabilidade_layout_seguro_sem_alerta(nome_ref: str) -> bool:
    """Layouts dedicados já homologados que podem ter sido reconhecidos por de/para.

    Para essas redes, a rastreabilidade por CNPJ/GLN/matrícula é apenas
    confirmação de identidade do layout. Ela não pode transformar o layout em
    "RASTREABILIDADE -> ..." nem gerar alerta bloqueante, porque isso
    esvazia indevidamente o Modelo Robô KOF.
    """
    nome = str(nome_ref or "").upper()
    homologados = (
        "ALABARCE",
        "MONACO",
        "MÔNACO",
        "PRIMATO",
        "DAHER",
        "SEMPRE VALE",
        "BAKLIZI",
        "BAZKILI",
        "COOPERCICA",
        "REDE VIP",
    )
    return any(x in nome for x in homologados)


def aplicar_metadados_rastreabilidade(
    item_importacao: Dict[str, object],
    layout: Dict[str, str],
    leitura: Dict[str, object],
) -> pd.DataFrame:
    """Marca a leitura como rastreada quando o layout foi aplicado por similaridade.

    A regra de seguranca e manter o parser de referencia, mas deixar claro no
    Excel e nos logs que o arquivo nao esta homologado com aquele nome/layout.
    """
    rast = item_importacao.get("rastreabilidade") or {}
    modo = str(item_importacao.get("modo_rastreabilidade", "")).upper() in {"SIM", "S", "TRUE", "1"}
    if not modo or not rast:
        return pd.DataFrame()

    df_rast = rastreabilidade_layouts.df_rastreabilidade(rast)
    nome_ref = str(rast.get("nome_layout_referencia") or layout.get("nome_layout", ""))
    confianca = str(rast.get("confianca", ""))
    motivo = str(rast.get("motivo", ""))
    rastreabilidade_segura = _rastreabilidade_layout_seguro_sem_alerta(nome_ref)
    nome_rastreado = nome_ref if rastreabilidade_segura else f"RASTREABILIDADE -> {nome_ref}"

    df_inter = leitura.get("df_intermediario")
    if df_inter is not None:
        df_inter = df_inter.copy()
        df_inter["modo_rastreabilidade"] = "SIM"
        df_inter["layout_referencia"] = nome_ref
        df_inter["confianca_rastreabilidade"] = confianca
        df_inter["motivo_rastreabilidade"] = motivo
        df_inter["layout_usado"] = nome_rastreado
        if "alerta_extracao" in df_inter.columns and not rastreabilidade_segura:
            aviso = f"RASTREABILIDADE_LAYOUT: referência={nome_ref}; confiança={confianca}%"
            df_inter["alerta_extracao"] = df_inter["alerta_extracao"].astype(str).apply(
                lambda atual: " | ".join([v for v in [atual.strip(), aviso] if v])
            )
        leitura["df_intermediario"] = df_inter

    alertas = list(leitura.get("alertas", []) or [])
    if rastreabilidade_segura:
        terminal_log.info(
            "[IMPORTACAO][RASTREABILIDADE][LAYOUT_HOMOLOGADO_OK] arquivo=%s | referencia=%s | confianca=%s | motivo=%s | sem_conversao=SIM",
            item_importacao.get("nome_arquivo", ""),
            nome_ref,
            confianca,
            motivo,
        )
    else:
        alertas.append(
            f"RASTREABILIDADE_LAYOUT: arquivo nao homologado processado com layout referencia {nome_ref}; "
            f"confianca={confianca}%; conferir Excel antes de TXT/fila."
        )
        if motivo:
            alertas.append(f"MOTIVO_RASTREABILIDADE: {motivo}")
        terminal_log.warning(
            "[IMPORTACAO][RASTREABILIDADE] arquivo=%s | referencia=%s | confianca=%s | motivo=%s",
            item_importacao.get("nome_arquivo", ""),
            nome_ref,
            confianca,
            motivo,
        )
    leitura["alertas"] = sorted({str(a) for a in alertas if str(a).strip()})
    leitura["df_rastreabilidade"] = df_rast

    # Mantem tipo_cliente_destino e demais regras tecnicas do layout original,
    # mas evita marcar Alabarce homologado como layout genérico de rastreabilidade.
    layout["nome_layout"] = nome_rastreado
    return df_rast


def processar_arquivo(item_importacao: Dict[str, str]) -> Dict[str, object]:
    caminho = item_importacao["caminho_arquivo"]
    nome_arquivo = item_importacao.get("nome_arquivo") or Path(caminho).name
    terminal_log.info("[IMPORTACAO] Iniciando processamento do arquivo: %s", nome_arquivo)
    terminal_log.info("[IMPORTACAO] Caminho do arquivo: %s", caminho)

    layout = cadastro_service.buscar_layout(layout_id=item_importacao["layout_id"])
    if not layout:
        terminal_log.error("[IMPORTACAO] Layout nao encontrado para arquivo: %s", nome_arquivo)
        return montar_resultado_arquivo(
            nome_arquivo,
            "",
            item_importacao.get("tipo_arquivo", ""),
            "ERRO",
            "Layout nÃ£o encontrado",
        )

    layout = layout.copy()
    terminal_log.info(
        "[IMPORTACAO] Layout identificado: %s | tipo=%s",
        layout.get("nome_layout", ""),
        layout.get("tipo_arquivo", ""),
    )
    mapeamentos = cadastro_service.buscar_mapeamentos_do_layout(layout["layout_id"])
    leitor = selecionar_leitor(layout)

    if leitor is None:
        terminal_log.error("[IMPORTACAO] Tipo de arquivo nao suportado: %s", layout.get("tipo_arquivo", ""))
        return montar_resultado_arquivo(
            nome_arquivo,
            layout.get("nome_layout", ""),
            layout.get("tipo_arquivo", ""),
            "ERRO",
            f"Tipo de arquivo nÃ£o suportado: {layout.get('tipo_arquivo', '')}",
        )

    leitura = leitor(caminho, layout, mapeamentos)
    df_rastreabilidade = aplicar_metadados_rastreabilidade(item_importacao, layout, leitura)

    if not leitura["sucesso"]:
        terminal_log.error("[IMPORTACAO] Erro de leitura | arquivo=%s | layout=%s | erro=%s", nome_arquivo, layout.get("nome_layout", ""), leitura["mensagem"])
        arquivo_validacao = salvar_validacao_erro_leitura(
            nome_arquivo,
            layout,
            leitura["mensagem"],
            leitura.get("alertas", []),
            leitura.get("df_auditoria_paginas"),
            df_rastreabilidade,
        )
        return montar_resultado_arquivo(
            nome_arquivo,
            layout.get("nome_layout", ""),
            layout.get("tipo_arquivo", ""),
            "ERRO",
            f"{leitura['mensagem']}. Arquivo de validacao gerado: {arquivo_validacao}",
            0,
            0,
            0,
            0,
            (leitura.get("alertas", []) or []) + [f"VALIDACAO_GERADA: {arquivo_validacao}"],
            arquivo_validacao,
        )

    terminal_log.info(
        "[IMPORTACAO] Leitura concluida | arquivo=%s | linhas_lidas=%s | itens_extraidos=%s | paginas_pdf=%s/%s | ignorados=%s",
        nome_arquivo,
        leitura.get("qtd_linhas_lidas", 0),
        leitura.get("qtd_itens_extraidos", leitura.get("qtd_linhas_lidas", 0)),
        leitura.get("paginas_pdf_processadas", ""),
        leitura.get("paginas_pdf_total", ""),
        leitura.get("qtd_itens_ignorados", 0),
    )

    data_ok, data_msg = aplicar_data_remessa_manual(leitura, layout, item_importacao)
    if not data_ok:
        terminal_log.warning("[IMPORTACAO] Processamento bloqueado por validacao de data | arquivo=%s | detalhe=%s", nome_arquivo, data_msg)
        arquivo_validacao = salvar_validacao_erro_leitura(
            nome_arquivo,
            layout,
            data_msg,
            leitura.get("alertas", []),
            leitura.get("df_auditoria_paginas"),
            df_rastreabilidade,
        )
        return montar_resultado_arquivo(
            nome_arquivo,
            layout.get("nome_layout", ""),
            layout.get("tipo_arquivo", ""),
            "ERRO",
            f"{data_msg}. Arquivo de validacao gerado: {arquivo_validacao}",
            leitura["qtd_linhas_lidas"],
            0,
            leitura["qtd_linhas_lidas"],
            0,
            (leitura.get("alertas", []) or []) + [f"VALIDACAO_GERADA: {arquivo_validacao}"],
            arquivo_validacao,
        )

    pad = padronizar_pedidos(leitura["df_intermediario"], layout)
    terminal_log.info(
        "[IMPORTACAO] Extracao/padronizacao concluida | arquivo=%s | lidas=%s | validas=%s | descartadas=%s",
        nome_arquivo,
        leitura.get("qtd_linhas_lidas", 0),
        pad.get("qtd_linhas_validas", 0),
        pad.get("qtd_linhas_descartadas", 0),
    )
    alertas = pad.get("alertas", []) + leitura.get("alertas", [])
    arquivo_validacao = salvar_validacao_importacao(
        nome_arquivo,
        layout,
        leitura,
        pad,
        alertas,
        df_rastreabilidade=df_rastreabilidade,
    )

    if not pad["sucesso"]:
        terminal_log.warning("[IMPORTACAO] Validacao gerada com alerta | arquivo=%s | motivo=%s", nome_arquivo, pad["mensagem"])
        arquivo_erro = salvar_descartes_em_erro(pad.get("df_descartadas"), nome_arquivo)
        if arquivo_validacao:
            alertas.append(f"VALIDACAO_GERADA: {arquivo_validacao}")
        return montar_resultado_arquivo(
            nome_arquivo,
            layout.get("nome_layout", ""),
            layout.get("tipo_arquivo", ""),
            "VALIDACAO_GERADA_COM_ALERTA" if arquivo_validacao else "ERRO",
            (
                f"{pad['mensagem']}. Nenhuma linha foi inserida na fila; "
                f"arquivo de validacao gerado para conferencia/de-para: {arquivo_validacao}"
                if arquivo_validacao
                else pad["mensagem"]
            ),
            leitura["qtd_linhas_lidas"],
            pad["qtd_linhas_validas"],
            pad["qtd_linhas_descartadas"],
            0,
            alertas,
            arquivo_validacao,
            arquivo_erro,
            resumo_conversao=pad.get("resumo_conversao", {}),
        )

    if pad["qtd_linhas_descartadas"] > 0:
        arquivo_erro = salvar_descartes_em_erro(pad.get("df_descartadas"), nome_arquivo)
    else:
        arquivo_erro = None

    status = "VALIDACAO_GERADA_COM_ALERTA" if pad["qtd_linhas_descartadas"] > 0 else "VALIDACAO_GERADA"
    if arquivo_validacao:
        alertas.append(f"VALIDACAO_GERADA: {arquivo_validacao}")
        terminal_log.info("[IMPORTACAO] Excel de validacao gerado: %s", arquivo_validacao)

    return montar_resultado_arquivo(
        nome_arquivo,
        layout.get("nome_layout", ""),
        layout.get("tipo_arquivo", ""),
        status,
        f"{pad['mensagem']}. Convertidos: {pad.get('qtd_itens_convertidos', 0)} | Não convertidos: {pad.get('qtd_itens_nao_convertidos', 0)} | Validar conversão: {pad.get('qtd_itens_validar_conversao', 0)}. Nenhuma linha foi inserida na fila automaticamente; arquivo de validacao: {arquivo_validacao}",
        leitura["qtd_linhas_lidas"],
        pad["qtd_linhas_validas"],
        pad["qtd_linhas_descartadas"],
        0,
        alertas,
        arquivo_validacao,
        arquivo_erro,
        resumo_conversao=pad.get("resumo_conversao", {}),
    )


def processar_lista_arquivos(itens_importacao: List[Dict[str, str]]) -> List[Dict[str, object]]:
    resultados = []
    itens_coelho_diniz = []
    itens_grancoffee = []
    grupos_layout: Dict[str, List[Dict[str, str]]] = {}
    itens_individuais: List[Dict[str, str]] = []

    for item in itens_importacao:
        layout = cadastro_service.buscar_layout(layout_id=item.get("layout_id"))
        if layout_eh_coelho_diniz(layout):
            itens_coelho_diniz.append(item)
            continue
        if layout_eh_grancoffee(layout):
            itens_grancoffee.append(item)
            continue

        chave_layout = str(item.get("layout_id", "") or "").strip()
        if chave_layout:
            grupos_layout.setdefault(chave_layout, []).append(item)
        else:
            itens_individuais.append(item)

    for item in itens_individuais:
        resultados.append(processar_arquivo(item))

    for _layout_id, itens_layout in grupos_layout.items():
        layout = cadastro_service.buscar_layout(layout_id=itens_layout[0].get("layout_id"))
        if _deve_consolidar_lote_layout(layout, itens_layout):
            terminal_log.info(
                "[IMPORTACAO][LOTE] Consolidacao habilitada | layout=%s | arquivos=%s",
                (layout or {}).get("nome_layout", ""),
                len(itens_layout),
            )
            resultados.extend(processar_lote_layout_importacao(itens_layout))
        else:
            if len(itens_layout) > 1:
                terminal_log.info(
                    "[IMPORTACAO][LOTE] Consolidacao nao aplicada | layout=%s | arquivos=%s | processamento individual por seguranca",
                    (layout or {}).get("nome_layout", ""),
                    len(itens_layout),
                )
            for item in itens_layout:
                resultados.append(processar_arquivo(item))

    if itens_coelho_diniz:
        resultados.extend(processar_lote_coelho_diniz_importacao(itens_coelho_diniz))

    if itens_grancoffee:
        resultados.extend(processar_lote_grancoffee_importacao(itens_grancoffee))

    return resultados

def processar_importacao(itens_importacao):
    terminal_log.info("[IMPORTACAO] Iniciando processamento geral: %s arquivo(s).", len(itens_importacao or []))
    try:
        resultados = processar_lista_arquivos(itens_importacao)
    except Exception:
        terminal_log.exception("[IMPORTACAO] Falha geral ao processar a importacao.")
        raise

    total = len(resultados)
    sucesso = sum(1 for r in resultados if "VALIDACAO_GERADA" in str(r.get("status", "")).upper())
    erro = sum(1 for r in resultados if str(r.get("status", "")).upper() == "ERRO")
    validacoes_unicas = {
        str(r.get("arquivo_validacao", "")).strip()
        for r in resultados
        if str(r.get("arquivo_validacao", "")).strip()
    }
    validacao = len(validacoes_unicas)

    mensagem_geral = (
        f"Processamento concluido para validacao. "
        f"Arquivos: {total} | Excel de validacao gerado: {validacao} | Prontos para conferencia: {sucesso} | Erro: {erro}"
    )
    terminal_log.info("[IMPORTACAO] Finalizacao: %s", mensagem_geral)

    return {
        "mensagem_geral": mensagem_geral,
        "resultados_por_arquivo": resultados,
    }
