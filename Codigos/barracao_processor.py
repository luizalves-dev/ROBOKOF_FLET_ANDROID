
# -*- coding: utf-8 -*-
"""
Parser homologado para Rede Barracão / Barracao - Robô KOF.
Autor: Kauê Melo

Correção aplicada:
- Barracão PDF RP One deixa de cair como RASTREABILIDADE_LAYOUT bloqueante.
- Itens válidos entram no Modelo Robô KOF para Enviar.
- SKU oficial é sempre o campo Cod Forn do PDF.
- Quando Cod Forn vier ausente, o item vai para pendência/validação manual e NÃO desloca EAN/descrição para SKU.
- Quantidade final é Qtde Emb do PDF, sem conversão unidade -> caixaria.
"""

from __future__ import annotations

import csv
import os
import re
from collections import defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

# De/para utilizado no lote validado de 01/06/2026.
# Pode ser sobrescrito/complementado por CSV externo.
BARRACAO_DEPARA_PADRAO: Dict[str, str] = {
    "62375191000123": "7120081433",  # loja 001
    "62375191000476": "7120411341",  # loja 003
    "62375191000557": "7120073093",  # loja 004
    "62375191000638": "7120464414",  # loja 005
    "62375191000719": "7120199642",  # loja 006
    "62375191000980": "7120204659",  # loja 008
    "62375191001014": "7120410372",  # loja 009
}

FORNECEDOR_SPAL_PREFIXOS = ("61186888", "061186888")
LAYOUT_NAME = "BARRACAO PDF"


def somente_digitos(valor: object) -> str:
    return re.sub(r"\D+", "", str(valor or ""))


def normalizar_cnpj(valor: object) -> str:
    dig = somente_digitos(valor)
    if len(dig) > 14:
        dig = dig[-14:]
    return dig.zfill(14) if dig else ""


def br_num_to_int(valor: object, default: Optional[int] = None) -> Optional[int]:
    """Converte número pt-BR de Qtde Emb para inteiro, preservando 0 quando vier 0."""
    if valor is None:
        return default
    s = str(valor).strip()
    if not s:
        return default
    # Ex.: 1.234,000 -> 1234.000 | 20 -> 20
    s = s.replace(".", "").replace(",", ".")
    try:
        return int(float(s))
    except Exception:
        return default


def formatar_data_remessa(valor: object) -> str:
    """Transforma 11/06/26 em 11.06.2026, padrão aceito no Robô KOF."""
    s = str(valor or "").strip()
    m = re.search(r"(\d{2})/(\d{2})/(\d{2,4})", s)
    if not m:
        return s.replace("/", ".")
    dd, mm, yy = m.groups()
    if len(yy) == 2:
        yy = "20" + yy
    return f"{dd}.{mm}.{yy}"


def carregar_depara_barracao(caminho_csv: Optional[str] = None) -> Dict[str, str]:
    """Lê de/para CNPJ -> matrícula, aceitando CSV com ; ou , e cabeçalhos variados."""
    depara = dict(BARRACAO_DEPARA_PADRAO)
    if not caminho_csv or not os.path.exists(caminho_csv):
        return depara

    with open(caminho_csv, "r", encoding="utf-8-sig", newline="") as f:
        amostra = f.read(4096)
        f.seek(0)
        delimitador = ";" if amostra.count(";") >= amostra.count(",") else ","
        reader = csv.DictReader(f, delimiter=delimitador)
        for row in reader:
            normalizado = {somente_digitos(k).lower() if k else k: v for k, v in row.items()}
            # fallback por nomes textuais também
            row_lower = {str(k or "").strip().lower(): v for k, v in row.items()}
            cnpj = (
                normalizar_cnpj(row_lower.get("cnpj"))
                or normalizar_cnpj(row_lower.get("cnpj_lido"))
                or normalizar_cnpj(row_lower.get("cnpj oficial"))
            )
            matricula = somente_digitos(
                row_lower.get("matricula")
                or row_lower.get("matrícula")
                or row_lower.get("cliente")
                or row_lower.get("kunnr")
                or ""
            )
            if cnpj and matricula:
                depara[cnpj] = matricula
    return depara


def _ler_texto_pdf(caminho_pdf: str) -> str:
    partes: List[str] = []
    try:
        import fitz  # PyMuPDF
        with fitz.open(caminho_pdf) as doc:
            for page in doc:
                partes.append(page.get_text("text") or "")
        return "\n".join(partes)
    except Exception:
        pass

    try:
        import pdfplumber
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                partes.append(page.extract_text() or "")
    except Exception:
        pass
    return "\n".join(partes)


def eh_layout_barracao(texto: str) -> bool:
    t = texto.lower()
    return (
        "rp one" in t
        and ("barracao supermercado" in t or "barracão supermercado" in t or "barracao supermercados" in t)
        and "cod forn" in t
        and "cod barras" in t
        and "número do pedido" in t
    )


def extrair_pedido(texto: str) -> str:
    m = re.search(r"N[úu]mero\s+do\s+Pedido:\s*(\d+)", texto, flags=re.I)
    return m.group(1) if m else ""


def extrair_data_entrega(texto: str) -> str:
    m = re.search(r"Dt\.\s*Entrega:\s*(\d{2}/\d{2}/\d{2,4})", texto, flags=re.I)
    return formatar_data_remessa(m.group(1)) if m else ""


def extrair_cnpj_loja(texto: str) -> str:
    cnpjs = re.findall(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}|\d{14}", texto)
    for cnpj in cnpjs:
        dig = normalizar_cnpj(cnpj)
        if dig and not dig.startswith(FORNECEDOR_SPAL_PREFIXOS):
            return dig
    return ""


def extrair_codigo_loja(texto: str) -> str:
    m = re.search(r"Empresa:\s*(\d{3})\s+Barrac[aã]o", texto, flags=re.I)
    return m.group(1) if m else ""


def _palavras_pdf(caminho_pdf: str) -> List[Tuple[int, float, float, str]]:
    """Retorna (pagina, x0, y0, texto). Usa PyMuPDF e cai para pdfplumber."""
    palavras: List[Tuple[int, float, float, str]] = []
    try:
        import fitz  # PyMuPDF
        with fitz.open(caminho_pdf) as doc:
            for pno, page in enumerate(doc, start=1):
                for w in page.get_text("words"):
                    x0, y0, _x1, _y1, txt = w[:5]
                    if txt:
                        palavras.append((pno, float(x0), float(y0), str(txt)))
        if palavras:
            return palavras
    except Exception:
        pass

    try:
        import pdfplumber
        with pdfplumber.open(caminho_pdf) as pdf:
            for pno, page in enumerate(pdf.pages, start=1):
                for w in page.extract_words() or []:
                    txt = w.get("text", "")
                    if txt:
                        palavras.append((pno, float(w.get("x0", 0)), float(w.get("top", 0)), str(txt)))
    except Exception:
        pass
    return palavras


def _agrupar_linhas(palavras: Iterable[Tuple[int, float, float, str]], tolerancia_y: float = 3.0):
    por_pagina: Dict[int, List[Tuple[float, float, str]]] = defaultdict(list)
    for pagina, x0, y0, txt in palavras:
        por_pagina[pagina].append((y0, x0, txt))

    for pagina, itens in sorted(por_pagina.items()):
        itens = sorted(itens, key=lambda it: (it[0], it[1]))
        grupos: List[List[Tuple[float, float, str]]] = []
        for y0, x0, txt in itens:
            if not grupos or abs(grupos[-1][0][0] - y0) > tolerancia_y:
                grupos.append([(y0, x0, txt)])
            else:
                grupos[-1].append((y0, x0, txt))
        for grupo in grupos:
            yield pagina, sorted(grupo, key=lambda it: it[1])


def _tem_numero_ptbr(txt: str) -> bool:
    return bool(re.fullmatch(r"\d{1,3}(?:\.\d{3})*,\d{1,4}|\d+", str(txt)))


def _parse_linha_item(palavras_linha: Sequence[Tuple[float, float, str]]) -> Optional[Dict[str, object]]:
    """Extrai uma linha de item pela posição das colunas do PDF RP One Barracão."""
    if not palavras_linha:
        return None

    # Colunas fixas observadas no PDF enviado:
    # Código x~16 | Cod Barras x~90 | Cod Forn x~224 | Descrição x~358 | Marca x~807
    # Quant x~1030 | Qtde Emb x~1143/1149 | Emb x~1170
    first_word = palavras_linha[0][2]
    first_x = palavras_linha[0][1]
    if first_x > 60 or not somente_digitos(first_word):
        return None

    codigo_original = ""
    cod_barras = ""
    sku = ""
    descricao_words: List[str] = []
    marca_words: List[str] = []
    qtd_emb: Optional[int] = None
    quant_unidade = ""
    embalagem = ""

    for _y, x, txt in palavras_linha:
        dig = somente_digitos(txt)
        if 0 <= x < 80 and dig and not codigo_original:
            codigo_original = dig
        elif 80 <= x < 200 and dig and not cod_barras:
            # Preserva zeros à esquerda do EAN/código de barras.
            cod_barras = dig
        elif 200 <= x < 330 and dig and re.fullmatch(r"\d{4,6}", dig) and not sku:
            sku = dig
        elif 330 <= x < 790:
            descricao_words.append(txt)
        elif 790 <= x < 1020:
            marca_words.append(txt)
        elif 1015 <= x < 1095 and _tem_numero_ptbr(txt) and not quant_unidade:
            quant_unidade = txt
        elif 1095 <= x < 1170 and _tem_numero_ptbr(txt) and qtd_emb is None:
            qtd_emb = br_num_to_int(txt, default=0)
        elif 1165 <= x < 1245 and not embalagem:
            embalagem = txt

    if not codigo_original or not cod_barras or qtd_emb is None:
        return None

    # Segurança: evita tratar totalizadores como item.
    if len(cod_barras) < 7:
        return None

    return {
        "codigo_original": codigo_original,
        "cod_barras": cod_barras,
        "sku": sku,
        "descricao": " ".join(descricao_words).strip(),
        "marca": " ".join(marca_words).strip(),
        "quant_unidade": quant_unidade,
        "qtd": qtd_emb,
        "embalagem": embalagem,
    }


@dataclass
class BarracaoItem:
    arquivo_origem: str
    pagina_pdf: int
    layout_usado: str
    pedido: str
    data_remessa: str
    codigo_loja: str
    cnpj: str
    matricula: str
    sku: str
    qtd: int
    descricao: str
    cod_barras: str
    codigo_original: str
    status_identidade: str
    motivo_identidade: str
    status_conversao: str
    regra_aplicada: str
    origem_regra: str
    pronto_modelo: bool
    bloqueia_fila: bool
    observacao: str


def extrair_barracao_pdf(caminho_pdf: str, depara: Optional[Dict[str, str]] = None) -> Dict[str, object]:
    depara = depara or dict(BARRACAO_DEPARA_PADRAO)
    arquivo = Path(caminho_pdf).name
    texto = _ler_texto_pdf(caminho_pdf)
    logs: List[str] = []

    if not eh_layout_barracao(texto):
        logs.append(f"{arquivo}: layout Barracão não identificado com segurança.")
        return {"arquivo": arquivo, "itens": [], "pendencias": [], "logs": logs, "layout_ok": False}

    pedido = extrair_pedido(texto)
    cnpj = extrair_cnpj_loja(texto)
    data_remessa = extrair_data_entrega(texto)
    codigo_loja = extrair_codigo_loja(texto)
    matricula = depara.get(cnpj, "A CADASTRAR")

    logs.append(f"{arquivo}: pedido={pedido} cnpj_loja={cnpj} matricula={matricula} data_remessa={data_remessa}")

    itens: List[BarracaoItem] = []
    pendencias: List[BarracaoItem] = []
    palavras = _palavras_pdf(caminho_pdf)
    for pagina, linha in _agrupar_linhas(palavras):
        bruto = _parse_linha_item(linha)
        if not bruto:
            continue

        sku = str(bruto["sku"] or "").strip()
        qtd = int(bruto["qtd"] or 0)
        tem_sku = bool(sku)
        tem_matricula = bool(matricula and matricula != "A CADASTRAR")
        pronto = bool(tem_sku and tem_matricula and pedido and cnpj and qtd is not None)

        if tem_sku:
            status_identidade = "OK - BARRACAO HOMOLOGADO"
            motivo_identidade = "BARRACAO_HOMOLOGADO"
            observacao = "Layout BARRACAO PDF homologado; quantidade final usa Qtde Emb do PDF; sem conversão."
        else:
            status_identidade = "PENDÊNCIA - VALIDAR SKU"
            motivo_identidade = "COD_FORN_AUSENTE"
            observacao = "Cod Forn/SKU ausente no PDF; item mantido para validação manual e fora do Modelo."

        item = BarracaoItem(
            arquivo_origem=arquivo,
            pagina_pdf=pagina,
            layout_usado=LAYOUT_NAME,
            pedido=pedido,
            data_remessa=data_remessa,
            codigo_loja=codigo_loja,
            cnpj=cnpj,
            matricula=matricula,
            sku=sku,
            qtd=qtd,
            descricao=str(bruto["descricao"]),
            cod_barras=str(bruto["cod_barras"]),
            codigo_original=str(bruto["codigo_original"]),
            status_identidade=status_identidade,
            motivo_identidade=motivo_identidade,
            status_conversao="OK SEM CONVERSÃO",
            regra_aplicada="BARRACAO_QTDE_EMB_SEM_CONVERSAO",
            origem_regra="PARSER_BARRACAO",
            pronto_modelo=pronto,
            bloqueia_fila=not pronto,
            observacao=observacao,
        )
        if tem_sku:
            itens.append(item)
        else:
            pendencias.append(item)

    logs.append(f"{arquivo}: {len(itens)} item(ns) válidos; {len(pendencias)} pendência(s) sem Cod Forn/SKU.")
    return {"arquivo": arquivo, "itens": itens, "pendencias": pendencias, "logs": logs, "layout_ok": True}


def extrair_lote_barracao(caminhos_pdf: Sequence[str], caminho_depara_csv: Optional[str] = None) -> Dict[str, object]:
    depara = carregar_depara_barracao(caminho_depara_csv)
    todos: List[BarracaoItem] = []
    pendencias: List[BarracaoItem] = []
    logs: List[str] = []

    for i, pdf in enumerate(caminhos_pdf, start=1):
        logs.append(f"[{i}/{len(caminhos_pdf)}] Iniciando Barracão: {Path(pdf).name}")
        resultado = extrair_barracao_pdf(pdf, depara=depara)
        todos.extend(resultado.get("itens", []))
        pendencias.extend(resultado.get("pendencias", []))
        logs.extend(resultado.get("logs", []))

    modelo = gerar_modelo_robo_kof(todos)
    return {
        "layout": LAYOUT_NAME,
        "itens_validos": todos,
        "itens_cod_nao_identificado": pendencias,
        "modelo_robo_kof": modelo,
        "logs": logs,
        "resumo": {
            "arquivos": len(caminhos_pdf),
            "linhas_validas": len(todos),
            "pendencias_sem_sku": len(pendencias),
            "linhas_modelo": len(modelo),
        },
    }


def gerar_modelo_robo_kof(itens: Sequence[BarracaoItem]) -> List[Dict[str, object]]:
    """Linhas que devem popular a aba Modelo Robô KOF para Enviar."""
    linhas: List[Dict[str, object]] = []
    for item in itens:
        if not item.pronto_modelo:
            continue
        linhas.append({
            "Matricula": item.matricula,
            "CNPJ": item.cnpj,
            "Sku": item.sku,
            "Qtd": item.qtd,
            "Nº Pedido": item.pedido,
            "Data remessa": item.data_remessa,
            "Status Conversão": item.status_conversao,
        })
    return linhas


def to_robo_kof_registros(resultado_lote: Dict[str, object]) -> List[Dict[str, object]]:
    """Converte para chaves comuns usadas na esteira de validação do Robô KOF."""
    registros: List[Dict[str, object]] = []
    for item in list(resultado_lote.get("itens_validos", [])) + list(resultado_lote.get("itens_cod_nao_identificado", [])):
        d = asdict(item)
        registros.append({
            "matricula_lida": item.matricula,
            "cnpj_lido": item.cnpj,
            "cnpj_oficial": item.cnpj,
            "sku_lido": item.sku,
            "codigo_sku_lido": item.sku,
            "quantidade_lida": item.qtd,
            "quantidade_final": item.qtd,
            "numero_pedido_lido": item.pedido,
            "data_entrega_lida": item.data_remessa,
            "arquivo_origem": item.arquivo_origem,
            "pagina_pdf": item.pagina_pdf,
            "layout_usado": item.layout_usado,
            "descricao_lida": item.descricao,
            "ean_lido": item.cod_barras,
            "codigo_origem_lido": item.codigo_original,
            "status_identidade": item.status_identidade,
            "motivo_identidade": item.motivo_identidade,
            "status_conversao": item.status_conversao,
            "regra_aplicada": item.regra_aplicada,
            "origem_regra": item.origem_regra,
            "pronto_modelo": item.pronto_modelo,
            "bloqueia_fila": item.bloqueia_fila,
            "observacao": item.observacao,
            "layout_homologado": True,
            "status_extracao": "OK" if item.pronto_modelo else "PENDENCIA_SKU",
            **d,
        })
    return registros


def gerar_excel_barracao_validacao(caminho_saida: str, resultado_lote: Dict[str, object]) -> None:
    """Gerador standalone para teste. No Robô KOF, prefira plugar os registros no gerador de Excel oficial."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception as exc:
        raise RuntimeError("openpyxl não está instalado. Use esta função apenas no ambiente do Robô KOF ou instale openpyxl.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Robô KOF para Enviar"
    headers = ["Matricula", "CNPJ", "Sku", "Qtd", "Nº Pedido", "Data remessa", "Status Conversão"]
    ws.append(headers)
    for row in resultado_lote.get("modelo_robo_kof", []):
        ws.append([row.get(h) for h in headers])

    val = wb.create_sheet("Validação do Pedido")
    val_headers = [
        "Arquivo Origem", "Página PDF", "Status Identidade", "Motivo Identidade", "CNPJ", "Matrícula",
        "Descrição", "SKU", "EAN", "QTD", "Nº do Pedido", "Data remessa", "Status Conversão", "Regra Aplicada", "Observação / Alerta"
    ]
    val.append(val_headers)
    for item in list(resultado_lote.get("itens_validos", [])) + list(resultado_lote.get("itens_cod_nao_identificado", [])):
        val.append([
            item.arquivo_origem, item.pagina_pdf, item.status_identidade, item.motivo_identidade, item.cnpj, item.matricula,
            item.descricao, item.sku, item.cod_barras, item.qtd, item.pedido, item.data_remessa, item.status_conversao,
            item.regra_aplicada, item.observacao,
        ])

    pend = wb.create_sheet("Itens Bloqueados Fila")
    pend_headers = ["Status Bloqueio", "Motivo Bloqueio", "Nº do Pedido", "Matrícula", "CNPJ", "SKU", "EAN", "QTD Final", "Observação / Alerta"]
    pend.append(pend_headers)
    for item in resultado_lote.get("itens_cod_nao_identificado", []):
        pend.append(["BLOQUEADO - VALIDAR ANTES DA FILA/TXT", "SKU vazio", item.pedido, item.matricula, item.cnpj, "", item.cod_barras, item.qtd, item.observacao])

    logs = wb.create_sheet("Logs do Processamento")
    logs.append(["log"])
    for log in resultado_lote.get("logs", []):
        logs.append([log])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for col in sheet.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)

    wb.save(caminho_saida)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Testar parser Barracão PDF - Robô KOF")
    parser.add_argument("pdfs", nargs="+", help="PDFs Barracão/RP One")
    parser.add_argument("--depara", default=None, help="CSV CNPJ x matrícula opcional")
    parser.add_argument("--saida", default="VALIDACAO_BARRACAO_TESTE.xlsx", help="Excel de teste")
    args = parser.parse_args()
    lote = extrair_lote_barracao(args.pdfs, args.depara)
    print(lote["resumo"])
    gerar_excel_barracao_validacao(args.saida, lote)
    print(f"Excel gerado: {args.saida}")
