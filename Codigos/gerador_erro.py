from pathlib import Path
import os
import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from layout_standard import normalize_intermediate_columns
from terminal_logger import get_terminal_logger


terminal_log = get_terminal_logger("excel")

# Estilização completa célula a célula fica pesada em lotes grandes.
# Acima do limite abaixo, o Excel mantém cabeçalho, filtros, cores de alerta
# e larguras amostradas, mas evita borda/alinhamento em todas as células.
MAX_CELULAS_ESTILO_COMPLETO = int(os.getenv("ROBOKOF_MAX_CELULAS_ESTILO_COMPLETO", "25000"))
MAX_LINHAS_AMOSTRA_LARGURA = int(os.getenv("ROBOKOF_MAX_LINHAS_AMOSTRA_LARGURA", "400"))


SHEET_MODELO_ROBOKOF = "Modelo Robô KOF para Enviar"
SHEET_VALIDACAO_PEDIDO = "Validação do Pedido"
SHEET_CADASTRAR_CNPJ = "Cadastrar CNPJ"
SHEET_PAINEL_EXECUTIVO = "Painel Executivo"
SHEET_ITENS_BLOQUEADOS = "Itens Bloqueados Fila"
SHEET_DUPLICIDADES = "Duplicidades Validação"
SHEET_MANIFESTO = "Manifesto Processamento"
SHEET_ALERTAS_ERROS = "Alertas_Erros"
SHEET_PENDENCIAS_GLN = "Pendências GLN"
SHEET_LOGS_PROCESSAMENTO = "Logs do Processamento"
SHEET_RESUMO_PROCESSAMENTO = "Resumo Processamento"

COLUNAS_MODELO_ROBOKOF = [
    "Matricula",
    "CNPJ",
    "Sku",
    "Qtd",
    "Nº Pedido",
    "Data remessa",
    "Status Conversão",
]

COLUNAS_VALIDACAO_PEDIDO = [
    "Arquivo Origem",
    "Página PDF",
    "Linha Origem",
    "Status Identidade",
    "Motivo Identidade",
    "Alerta Identidade",
    "CNPJ",
    "CNPJ Oficial",
    "Tipo da Chave",
    "Chave Lida",
    "Matrícula",
    "Descrição",
    "Código SKU",
    "EAN",
    "SKU",
    "QTD Original",
    "Tipo QTD Original",
    "Fator Conversão",
    "QTD Convertida",
    "QTD Final",
    "QTD",
    "Nº do Pedido",
    "Centro",
    "Centro de Referência",
    "Status Conversão",
    "Regra Aplicada",
    "Origem da Regra",
    "Observação / Alerta",
]


COLUNAS_CADASTRAR_CNPJ = [
    "Rede/Layout",
    "Tipo da Chave",
    "Chave Lida",
    "CNPJ Lido",
    "CNPJ Oficial",
    "Matrícula",
    "Status",
    "Nº do Pedido",
    "Observação",
]


def _validar_xlsx_abre(out_path: Path):
    if out_path.suffix.lower() != ".xlsx" or not out_path.exists():
        return
    try:
        wb = load_workbook(out_path, read_only=True, data_only=True)
        wb.close()
        terminal_log.info("[EXCEL] Arquivo validado para abertura: %s", out_path)
    except Exception:
        terminal_log.exception("[EXCEL] Arquivo gerado nao abriu corretamente: %s", out_path)
        raise


def _texto(valor) -> str:
    if valor is None:
        return ""
    if pd.isna(valor):
        return ""
    return str(valor).strip()


def _primeiro_valor(row: dict, nomes: list[str]) -> str:
    for nome in nomes:
        valor = _texto(row.get(nome, ""))
        if valor:
            return valor
    return ""


def _descricao_item(row: dict) -> str:
    return _primeiro_valor(
        row,
        [
            "descricao_lida",
            "descrição_lida",
            "descricao",
            "Descrição",
            "produto_descricao",
            "descricao_produto",
            "Produto",
        ],
    )


def _codigo_sku_item(row: dict) -> str:
    return _primeiro_valor(
        row,
        [
            "codigo_sku_lido",
            "codigo_produto_lido",
            "codigo_origem_lido",
            "codigo_lido",
            "cod_forn_lido",
            "SKU",
            "Sku",
            "sku_lido",
        ],
    )


def _ean_item(row: dict) -> str:
    return _primeiro_valor(
        row,
        [
            "ean_lido",
            "EAN",
            "codigo_ean_lido",
            "codigo_barras_lido",
            "cod_barras_lido",
            "cod_barras",
            "referencia_lida",
            "barcode_lido",
        ],
    )


def _status_conversao(row: dict, resumo: dict | None = None, alertas: list[str] | None = None) -> str:
    status_direto = _primeiro_valor(row, ["status_conversao", "Status Conversão", "status conversao"]).upper()
    status_direto = status_direto.replace("OK - SEM CONVERSÃO", "OK SEM CONVERSÃO").replace("OK - SEM CONVERSAO", "OK SEM CONVERSÃO")
    if status_direto in {"OK CONVERTIDO", "OK SEM CONVERSÃO", "ALERTA - NÃO CONVERTIDO", "ALERTA - NAO CONVERTIDO", "VALIDAR CONVERSÃO", "VALIDAR CONVERSAO"}:
        if status_direto == "ALERTA - NAO CONVERTIDO":
            return "ALERTA - NÃO CONVERTIDO"
        if status_direto == "VALIDAR CONVERSAO":
            return "VALIDAR CONVERSÃO"
        return status_direto

    partes = [
        " ".join(_texto(v) for v in row.values()),
        " ".join(f"{k}={v}" for k, v in (resumo or {}).items()),
    ]
    if not any(parte.strip() for parte in partes):
        partes.append(" ".join(str(a) for a in (alertas or [])))
    texto = " ".join(partes).upper()
    if "ALERTA - NÃO CONVERTIDO" in texto or "ALERTA - NAO CONVERTIDO" in texto:
        return "ALERTA - NÃO CONVERTIDO"
    if "NAO_CONVERTIDO" in texto or "NÃO_CONVERTIDO" in texto:
        return "ALERTA - NÃO CONVERTIDO"
    if "VALIDAR_CONVERSAO" in texto or "VALIDAR CONVERSAO" in texto or "VALIDAR CONVERSÃO" in texto:
        return "VALIDAR CONVERSÃO"
    if "OK CONVERTIDO" in texto:
        return "OK CONVERTIDO"
    return "OK SEM CONVERSÃO"


def _regra_aplicada_item(row: dict) -> str:
    regra = _primeiro_valor(row, ["regra_aplicada_conversao", "Regra Aplicada"])
    if regra:
        return regra
    tipo = _primeiro_valor(row, ["tipo_regra_conversao", "Tipo Regra Conversão"])
    prioridade = _primeiro_valor(row, ["prioridade_regra_conversao", "Prioridade Regra Conversão"])
    if tipo and prioridade:
        return f"{tipo} | prioridade {prioridade}"
    return tipo


def _origem_regra_item(row: dict) -> str:
    return _primeiro_valor(row, ["origem_regra_conversao", "Origem da Regra"])


def _observacao_validacao(row: dict, status_conversao: str) -> str:
    motivo = _primeiro_valor(row, ["motivo_descarte", "status_validacao", "mensagem", "alerta", "alerta_identidade", "alerta_extracao"])
    matricula = _primeiro_valor(row, ["matricula_final", "Matricula", "Matrícula", "matricula_lida"])
    cnpj = _primeiro_valor(row, ["cnpj_lido", "CNPJ"])
    observacoes = []
    obs_conv = _primeiro_valor(row, ["observacao_conversao", "Observação Conversão", "observacao conversao"])
    if obs_conv:
        observacoes.append(obs_conv)
    origem_conv = _primeiro_valor(row, ["origem_regra_conversao"])
    tipo_conv = _primeiro_valor(row, ["tipo_regra_conversao"])
    if origem_conv or tipo_conv:
        observacoes.append("Regra conversão: " + " / ".join([v for v in [tipo_conv, origem_conv] if v]))
    status_depara = _primeiro_valor(row, ["status_depara_cliente"])
    if cnpj and (not matricula or motivo == "MATRICULA_NAO_ENCONTRADA" or status_depara == "A_CADASTRAR"):
        observacoes.append("A CADASTRAR")
    if status_depara == "DEPARA_CLIENTES":
        observacoes.append("DE/PARA CLIENTE OK")
    elif status_depara == "A_CADASTRAR":
        observacoes.append("CADASTRAR CHAVE CLIENTE")
    obs_depara = _primeiro_valor(row, ["observacao_depara_cliente"])
    if obs_depara:
        observacoes.append(obs_depara)
    if motivo and motivo not in {"MATRICULA_NAO_ENCONTRADA", "PENDENTE_MATRICULA"}:
        observacoes.append(motivo)
    modo_rastreabilidade = _primeiro_valor(row, ["modo_rastreabilidade"])
    if modo_rastreabilidade.upper() in {"SIM", "S", "TRUE", "1"}:
        ref = _primeiro_valor(row, ["layout_referencia"])
        confianca = _primeiro_valor(row, ["confianca_rastreabilidade"])
        obs_rast = "RASTREABILIDADE"
        if ref:
            obs_rast += f" - referência: {ref}"
        if confianca:
            obs_rast += f" ({confianca}% confiança)"
        observacoes.append(obs_rast)
    if status_conversao not in {"OK CONVERTIDO", "OK SEM CONVERSÃO"}:
        observacoes.append(status_conversao)
    return " | ".join(dict.fromkeys(observacoes)) or "OK"


def _normalizar_df(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    return df.copy().fillna("")


def _linha_pendente_pode_aparecer_no_modelo(row: dict, resumo: dict | None = None, alertas: list[str] | None = None) -> bool:
    """Permite espelhar no Modelo linhas com pendência cadastral, sem liberar fila/TXT.

    Regra operacional Robô KOF: quando há CNPJ, SKU, quantidade e pedido, a aba
    "Modelo Robô KOF para Enviar" deve mostrar a formatação do pedido mesmo que a
    matrícula ainda esteja A CADASTRAR/bloqueada. O bloqueio permanece nas abas
    Validação, Cadastrar CNPJ e Itens Bloqueados Fila.
    """
    motivo = _texto(row.get("motivo_descarte", "")).upper()
    status_depara = _texto(row.get("status_depara_cliente", "")).upper()
    obs = _texto(row.get("observacao_depara_cliente", "")).upper() + " " + _texto(row.get("alerta_extracao", "")).upper()
    if not (
        motivo in {"MATRICULA_NAO_ENCONTRADA", "PENDENTE_MATRICULA"}
        or status_depara == "A_CADASTRAR"
        or "A CADASTRAR" in obs
        or "CADASTRAR" in obs
    ):
        return False

    status_conv = _status_conversao(row, resumo, alertas)
    if status_conv not in {"OK CONVERTIDO", "OK SEM CONVERSÃO"}:
        return False

    cnpj = _primeiro_valor(row, ["cnpj_oficial_final", "CNPJ Oficial", "cnpj_lido", "CNPJ"])
    sku = _primeiro_valor(row, ["Sku", "SKU", "sku_lido"])
    qtd = _primeiro_valor(row, ["qtd_final", "qtd_convertida", "Qtd", "QTD", "quantidade_lida"])
    pedido = _primeiro_valor(row, ["Nº Pedido", "NÂº Pedido", "numero_pedido_final", "numero_pedido_lido"])
    if not (cnpj and sku and qtd and pedido):
        return False
    try:
        qtd_num = float(str(qtd).replace(".", "").replace(",", "."))
    except Exception:
        return False
    return qtd_num > 0


def _linha_modelo_from_dict(combinado: dict, resumo: dict | None = None, alertas: list[str] | None = None, matricula_fallback: str = "") -> dict:
    matricula = _primeiro_valor(combinado, ["Matricula", "matricula_final", "matricula_lida"])
    if not matricula and matricula_fallback:
        matricula = matricula_fallback
    return {
        "Matricula": matricula,
        "CNPJ": _primeiro_valor(combinado, ["cnpj_oficial_final", "CNPJ Oficial", "cnpj_lido", "CNPJ"]),
        "Sku": _primeiro_valor(combinado, ["Sku", "SKU", "sku_lido"]),
        "Qtd": _primeiro_valor(combinado, ["qtd_final", "qtd_convertida", "Qtd", "QTD", "quantidade_lida"]),
        "Nº Pedido": _primeiro_valor(combinado, ["Nº Pedido", "NÂº Pedido", "numero_pedido_final", "numero_pedido_lido"]),
        "Data remessa": _primeiro_valor(combinado, ["Data remessa", "data_remessa_final", "data_entrega_lida"]),
        "Status Conversão": _status_conversao(combinado, resumo, alertas),
    }


def _montar_df_modelo(
    df_final: pd.DataFrame | None,
    df_validas: pd.DataFrame | None = None,
    resumo: dict | None = None,
    alertas: list[str] | None = None,
    df_descartadas: pd.DataFrame | None = None,
) -> pd.DataFrame:
    df_final = _normalizar_df(df_final)
    df_validas = _normalizar_df(df_validas)
    df_descartadas = _normalizar_df(df_descartadas)
    linhas = []
    chaves = set()
    if not df_final.empty:
        for idx, row in df_final.reset_index(drop=True).iterrows():
            row_dict = row.to_dict()
            row_valida = df_validas.iloc[idx].to_dict() if idx < len(df_validas) else {}
            combinado = {**row_valida, **row_dict}
            linha_modelo = _linha_modelo_from_dict(combinado, resumo, alertas)
            linhas.append(linha_modelo)
            chaves.add((linha_modelo["CNPJ"], linha_modelo["Sku"], linha_modelo["Qtd"], linha_modelo["Nº Pedido"]))

    # Complemento visual: linhas bloqueadas somente por cadastro/matrícula também
    # aparecem no Modelo com Matricula="A CADASTRAR", mas continuam bloqueadas nas
    # abas de validação. Isso evita o modelo vazio quando a leitura do pedido está correta.
    if not df_descartadas.empty:
        for _, row in df_descartadas.reset_index(drop=True).iterrows():
            row_dict = row.to_dict()
            if not _linha_pendente_pode_aparecer_no_modelo(row_dict, resumo, alertas):
                continue
            linha_modelo = _linha_modelo_from_dict(row_dict, resumo, alertas, matricula_fallback="A CADASTRAR")
            chave = (linha_modelo["CNPJ"], linha_modelo["Sku"], linha_modelo["Qtd"], linha_modelo["Nº Pedido"])
            if chave in chaves:
                continue
            linhas.append(linha_modelo)
            chaves.add(chave)
    return pd.DataFrame(linhas, columns=COLUNAS_MODELO_ROBOKOF)


def _montar_linhas_validacao(
    df_validas: pd.DataFrame | None,
    df_descartadas: pd.DataFrame | None,
    df_final: pd.DataFrame | None,
    resumo: dict | None = None,
    alertas: list[str] | None = None,
) -> pd.DataFrame:
    df_validas = _normalizar_df(df_validas)
    df_descartadas = _normalizar_df(df_descartadas)
    df_final = _normalizar_df(df_final)
    linhas = []

    if not df_validas.empty:
        for idx, row in df_validas.reset_index(drop=True).iterrows():
            row_dict = row.to_dict()
            row_final = df_final.iloc[idx].to_dict() if idx < len(df_final) else {}
            combinado = {**row_dict, **row_final}
            status_conv = _status_conversao(combinado, resumo, alertas)
            linhas.append(
                {
                    "Arquivo Origem": _primeiro_valor(combinado, ["arquivo_origem", "Arquivo Origem"]),
                    "Página PDF": _primeiro_valor(combinado, ["pagina_pdf", "Página PDF"]),
                    "Linha Origem": _primeiro_valor(combinado, ["linha_origem", "Linha Origem"]),
                    "Status Identidade": _primeiro_valor(combinado, ["status_identidade", "Status Identidade"]),
                    "Motivo Identidade": _primeiro_valor(combinado, ["motivo_identidade", "Motivo Identidade"]),
                    "Alerta Identidade": _primeiro_valor(combinado, ["alerta_identidade", "Alerta Identidade"]),
                    "CNPJ": _primeiro_valor(combinado, ["cnpj_lido", "CNPJ"]),
                    "CNPJ Oficial": _primeiro_valor(combinado, ["cnpj_oficial_final", "CNPJ Oficial", "cnpj_lido", "CNPJ"]),
                    "Tipo da Chave": _primeiro_valor(combinado, ["tipo_chave_depara", "Tipo da Chave"]),
                    "Chave Lida": _primeiro_valor(combinado, ["chave_lida_depara", "Chave Lida", "cnpj_lido", "CNPJ"]),
                    "Matrícula": _primeiro_valor(combinado, ["matricula_final", "Matricula", "Matrícula", "matricula_lida"]),
                    "Descrição": _descricao_item(combinado),
                    "Código SKU": _codigo_sku_item(combinado),
                    "EAN": _ean_item(combinado),
                    "SKU": _primeiro_valor(combinado, ["sku_lido", "Sku", "SKU"]),
                    "QTD Original": _primeiro_valor(combinado, ["qtd_original", "quantidade_lida", "Qtd", "QTD"]),
                    "Tipo QTD Original": _primeiro_valor(combinado, ["tipo_qtd_original"]),
                    "Fator Conversão": _primeiro_valor(combinado, ["fator_conversao"]),
                    "QTD Convertida": _primeiro_valor(combinado, ["qtd_convertida"]),
                    "QTD Final": _primeiro_valor(combinado, ["qtd_final", "qtd_convertida", "quantidade_lida", "Qtd", "QTD"]),
                    "QTD": _primeiro_valor(combinado, ["qtd_final", "qtd_convertida", "quantidade_lida", "Qtd", "QTD"]),
                    "Nº do Pedido": _primeiro_valor(combinado, ["numero_pedido_final", "numero_pedido_lido", "Nº Pedido", "NÂº Pedido"]),
                    "Centro": _primeiro_valor(combinado, ["centro_lido", "centro", "Centro"]),
                    "Centro de Referência": _primeiro_valor(combinado, ["centro_referencia_conversao", "Centro de Referência", "centro_referencia"]),
                    "Status Conversão": status_conv,
                    "Regra Aplicada": _regra_aplicada_item(combinado),
                    "Origem da Regra": _origem_regra_item(combinado),
                    "Observação / Alerta": _observacao_validacao(combinado, status_conv),
                }
            )

    if not df_descartadas.empty:
        for _, row in df_descartadas.iterrows():
            row_dict = row.to_dict()
            status_conv = _status_conversao(row_dict, resumo, alertas)
            linhas.append(
                {
                    "Arquivo Origem": _primeiro_valor(row_dict, ["arquivo_origem", "Arquivo Origem"]),
                    "Página PDF": _primeiro_valor(row_dict, ["pagina_pdf", "Página PDF"]),
                    "Linha Origem": _primeiro_valor(row_dict, ["linha_origem", "Linha Origem"]),
                    "Status Identidade": _primeiro_valor(row_dict, ["status_identidade", "Status Identidade"]),
                    "Motivo Identidade": _primeiro_valor(row_dict, ["motivo_identidade", "Motivo Identidade"]),
                    "Alerta Identidade": _primeiro_valor(row_dict, ["alerta_identidade", "Alerta Identidade"]),
                    "CNPJ": _primeiro_valor(row_dict, ["cnpj_lido", "CNPJ"]),
                    "CNPJ Oficial": _primeiro_valor(row_dict, ["cnpj_oficial_final", "CNPJ Oficial", "cnpj_lido", "CNPJ"]),
                    "Tipo da Chave": _primeiro_valor(row_dict, ["tipo_chave_depara", "Tipo da Chave"]),
                    "Chave Lida": _primeiro_valor(row_dict, ["chave_lida_depara", "Chave Lida", "cnpj_lido", "CNPJ"]),
                    "Matrícula": _primeiro_valor(row_dict, ["matricula_final", "Matricula", "Matrícula", "matricula_lida"]),
                    "Descrição": _descricao_item(row_dict),
                    "Código SKU": _codigo_sku_item(row_dict),
                    "EAN": _ean_item(row_dict),
                    "SKU": _primeiro_valor(row_dict, ["sku_lido", "Sku", "SKU"]),
                    "QTD Original": _primeiro_valor(row_dict, ["qtd_original", "quantidade_lida", "Qtd", "QTD"]),
                    "Tipo QTD Original": _primeiro_valor(row_dict, ["tipo_qtd_original"]),
                    "Fator Conversão": _primeiro_valor(row_dict, ["fator_conversao"]),
                    "QTD Convertida": _primeiro_valor(row_dict, ["qtd_convertida"]),
                    "QTD Final": _primeiro_valor(row_dict, ["qtd_final", "qtd_convertida", "quantidade_lida", "Qtd", "QTD"]),
                    "QTD": _primeiro_valor(row_dict, ["qtd_final", "qtd_convertida", "quantidade_lida", "Qtd", "QTD"]),
                    "Nº do Pedido": _primeiro_valor(row_dict, ["numero_pedido_final", "numero_pedido_lido", "Nº Pedido", "NÂº Pedido"]),
                    "Centro": _primeiro_valor(row_dict, ["centro_lido", "centro", "Centro"]),
                    "Centro de Referência": _primeiro_valor(row_dict, ["centro_referencia_conversao", "Centro de Referência", "centro_referencia"]),
                    "Status Conversão": status_conv,
                    "Regra Aplicada": _regra_aplicada_item(row_dict),
                    "Origem da Regra": _origem_regra_item(row_dict),
                    "Observação / Alerta": _observacao_validacao(row_dict, status_conv),
                }
            )

    return pd.DataFrame(linhas, columns=COLUNAS_VALIDACAO_PEDIDO)


def _montar_df_cadastrar_cnpj(df_validacao: pd.DataFrame, layout_padrao: str = "") -> pd.DataFrame:
    if df_validacao is None or df_validacao.empty:
        return pd.DataFrame([{
            "Rede/Layout": layout_padrao,
            "Tipo da Chave": "",
            "Chave Lida": "",
            "CNPJ Lido": "",
            "CNPJ Oficial": "",
            "Matrícula": "",
            "Status": "OK",
            "Nº do Pedido": "",
            "Observação": "Nenhum CNPJ/GLN pendente de cadastro.",
        }], columns=COLUNAS_CADASTRAR_CNPJ)

    df = df_validacao.copy()
    for coluna in ["CNPJ", "CNPJ Oficial", "Tipo da Chave", "Chave Lida", "Matrícula", "Observação / Alerta", "Nº do Pedido"]:
        if coluna not in df.columns:
            df[coluna] = ""

    pendentes = df[
        (df["Chave Lida"].astype(str).str.strip() != "")
        & (
            (df["Matrícula"].astype(str).str.strip() == "")
            | (df["Observação / Alerta"].astype(str).str.contains("A CADASTRAR|CADASTRAR CHAVE", case=False, na=False))
        )
    ].copy()

    if pendentes.empty:
        return pd.DataFrame([{
            "Rede/Layout": layout_padrao,
            "Tipo da Chave": "",
            "Chave Lida": "",
            "CNPJ Lido": "",
            "CNPJ Oficial": "",
            "Matrícula": "",
            "Status": "OK",
            "Nº do Pedido": "",
            "Observação": "Nenhum CNPJ/GLN pendente de cadastro.",
        }], columns=COLUNAS_CADASTRAR_CNPJ)

    linhas = []
    group_cols = ["Tipo da Chave", "Chave Lida", "CNPJ", "CNPJ Oficial"]
    for chaves, grupo in pendentes.groupby(group_cols, dropna=False):
        tipo_chave, chave_lida, cnpj_lido, cnpj_oficial = [str(v).strip() for v in chaves]
        pedidos = sorted({str(p).strip() for p in grupo["Nº do Pedido"].tolist() if str(p).strip()})
        linhas.append(
            {
                "Rede/Layout": layout_padrao,
                "Tipo da Chave": tipo_chave or "CNPJ",
                "Chave Lida": chave_lida or cnpj_lido,
                "CNPJ Lido": cnpj_lido,
                "CNPJ Oficial": cnpj_oficial if cnpj_oficial != cnpj_lido else "",
                "Matrícula": "",
                "Status": "A CADASTRAR",
                "Nº do Pedido": " | ".join(pedidos),
                "Observação": "Preencher CNPJ Oficial e Matrícula; depois usar o botão Atualizar de/para CNPJ/GLN.",
            }
        )
    return pd.DataFrame(linhas, columns=COLUNAS_CADASTRAR_CNPJ)




def _flag_critico_texto(texto: str) -> bool:
    texto_up = _texto(texto).upper()
    return any(
        marcador in texto_up
        for marcador in [
            "ALERTA",
            "ERRO",
            "VALIDAR",
            "A CADASTRAR",
            "CADASTRAR CHAVE",
            "NÃO CONVERTIDO",
            "NAO CONVERTIDO",
            "CONVERSAO_PENDENTE_VALIDACAO",
            "MATRICULA_NAO_ENCONTRADA",
            "SKU_VAZIO",
            "QTD_VAZIA",
            "PEDIDO_VAZIO",
            "DATA_INVALIDA",
            "DUPLICADO",
        ]
    )


def _motivo_bloqueio_validacao(row: dict) -> str:
    motivos = []
    status_conv = _texto(row.get("Status Conversão", ""))
    obs = _texto(row.get("Observação / Alerta", ""))
    if status_conv and status_conv not in {"OK CONVERTIDO", "OK SEM CONVERSÃO"}:
        motivos.append(f"Conversão pendente: {status_conv}")
    if not _texto(row.get("Matrícula", "")):
        motivos.append("Matrícula vazia/A cadastrar")
    if not _texto(row.get("SKU", "")):
        motivos.append("SKU vazio")
    qtd_final_txt = _texto(row.get("QTD Final", row.get("QTD", ""))) or _texto(row.get("QTD", ""))
    if not qtd_final_txt:
        motivos.append("Quantidade final vazia")
    else:
        try:
            qtd_num = float(str(qtd_final_txt).replace(".", "").replace(",", "."))
            if qtd_num <= 0:
                motivos.append("Quantidade zero/negativa")
        except Exception:
            motivos.append("Quantidade final não numérica")
    if not _texto(row.get("Nº do Pedido", "")):
        motivos.append("Pedido vazio")
    if _flag_critico_texto(obs):
        motivos.append(obs)
    return " | ".join(dict.fromkeys([m for m in motivos if m]))


def _montar_df_itens_bloqueados(df_validacao: pd.DataFrame) -> pd.DataFrame:
    colunas = [
        "Status Bloqueio",
        "Motivo Bloqueio",
        "Nº do Pedido",
        "Matrícula",
        "CNPJ",
        "SKU",
        "EAN",
        "QTD Original",
        "QTD Final",
        "Status Conversão",
        "Regra Aplicada",
        "Origem da Regra",
        "Observação / Alerta",
    ]
    if df_validacao is None or df_validacao.empty:
        return pd.DataFrame([{
            "Status Bloqueio": "OK",
            "Motivo Bloqueio": "Nenhum item bloqueado para fila/TXT.",
        }], columns=colunas)

    linhas = []
    for _, row in df_validacao.fillna("").iterrows():
        row_dict = row.to_dict()
        motivo = _motivo_bloqueio_validacao(row_dict)
        if motivo:
            linhas.append({
                "Status Bloqueio": "BLOQUEADO - VALIDAR ANTES DA FILA/TXT",
                "Motivo Bloqueio": motivo,
                "Nº do Pedido": _texto(row_dict.get("Nº do Pedido", "")),
                "Matrícula": _texto(row_dict.get("Matrícula", "")),
                "CNPJ": _texto(row_dict.get("CNPJ", "")),
                "SKU": _texto(row_dict.get("SKU", "")),
                "EAN": _texto(row_dict.get("EAN", "")),
                "QTD Original": _texto(row_dict.get("QTD Original", "")),
                "QTD Final": _texto(row_dict.get("QTD Final", row_dict.get("QTD", ""))),
                "Status Conversão": _texto(row_dict.get("Status Conversão", "")),
                "Regra Aplicada": _texto(row_dict.get("Regra Aplicada", "")),
                "Origem da Regra": _texto(row_dict.get("Origem da Regra", "")),
                "Observação / Alerta": _texto(row_dict.get("Observação / Alerta", "")),
            })
    if not linhas:
        linhas = [{
            "Status Bloqueio": "OK",
            "Motivo Bloqueio": "Nenhum item bloqueado para fila/TXT.",
        }]
    return pd.DataFrame(linhas, columns=colunas)


def _montar_df_duplicidades_validacao(df_validacao: pd.DataFrame) -> pd.DataFrame:
    colunas = [
        "Status",
        "Chave Duplicidade",
        "Ocorrências",
        "Nº do Pedido",
        "Matrícula",
        "SKU",
        "QTD Total Agrupada",
        "Observação",
    ]
    if df_validacao is None or df_validacao.empty:
        return pd.DataFrame([{"Status": "OK", "Observação": "Sem dados para avaliar duplicidades."}], columns=colunas)
    df = df_validacao.fillna("").copy()
    for coluna in ["Nº do Pedido", "Matrícula", "SKU", "QTD Final", "QTD"]:
        if coluna not in df.columns:
            df[coluna] = ""
    # Detecta apenas linhas com pedido, matrícula e SKU preenchidos para evitar falso positivo de pendências vazias.
    base = df[(df["Nº do Pedido"].astype(str).str.strip() != "") & (df["Matrícula"].astype(str).str.strip() != "") & (df["SKU"].astype(str).str.strip() != "")].copy()
    if base.empty:
        return pd.DataFrame([{"Status": "OK", "Observação": "Sem chaves completas para avaliar duplicidades."}], columns=colunas)
    base["_qtd_num"] = pd.to_numeric(base["QTD Final"].where(base["QTD Final"].astype(str).str.strip() != "", base["QTD"]), errors="coerce").fillna(0)
    base["Chave Duplicidade"] = base["Nº do Pedido"].astype(str).str.strip() + " | " + base["Matrícula"].astype(str).str.strip() + " | " + base["SKU"].astype(str).str.strip()
    grouped = base.groupby(["Chave Duplicidade", "Nº do Pedido", "Matrícula", "SKU"], dropna=False).agg(
        Ocorrências=("Chave Duplicidade", "size"),
        QTD_Total_Agrupada=("_qtd_num", "sum"),
    ).reset_index()
    dup = grouped[grouped["Ocorrências"] > 1]
    if dup.empty:
        return pd.DataFrame([{"Status": "OK", "Observação": "Nenhuma duplicidade por Pedido + Matrícula + SKU."}], columns=colunas)
    linhas = []
    for _, row in dup.iterrows():
        linhas.append({
            "Status": "VALIDAR POSSÍVEL DUPLICIDADE",
            "Chave Duplicidade": row["Chave Duplicidade"],
            "Ocorrências": int(row["Ocorrências"]),
            "Nº do Pedido": row["Nº do Pedido"],
            "Matrícula": row["Matrícula"],
            "SKU": row["SKU"],
            "QTD Total Agrupada": row["QTD_Total_Agrupada"],
            "Observação": "Conferir se são linhas legítimas repetidas ou duplicidade operacional antes de gerar fila/TXT.",
        })
    return pd.DataFrame(linhas, columns=colunas)


def _montar_df_manifesto_processamento(resumo: dict | None, alertas: list[str] | None, df_validacao: pd.DataFrame, df_modelo: pd.DataFrame, df_bloqueados: pd.DataFrame) -> pd.DataFrame:
    resumo = resumo or {}
    alertas = alertas or []
    total_validacao = 0 if df_validacao is None else len(df_validacao)
    total_fila = 0 if df_modelo is None else len(df_modelo)
    total_bloqueados = 0
    if df_bloqueados is not None and not df_bloqueados.empty:
        total_bloqueados = int((df_bloqueados.get("Status Bloqueio", pd.Series(dtype=str)).astype(str).str.contains("BLOQUEADO", case=False, na=False)).sum())
    manifest = {
        "data_hora_geracao": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        "arquivo": resumo.get("arquivo", ""),
        "layout": resumo.get("layout", ""),
        "tipo_arquivo": resumo.get("tipo_arquivo", ""),
        "fluxo": resumo.get("fluxo", "VALIDACAO_PRIMEIRO_SEM_GERAR_TXT_OU_FILA"),
        "linhas_validacao": total_validacao,
        "linhas_modelo_fila": total_fila,
        "linhas_bloqueadas": total_bloqueados,
        "total_alertas": len(alertas),
        "go_no_go_fila": "NÃO LIBERADO" if total_bloqueados or alertas else "LIBERADO APÓS CONFERÊNCIA MANUAL",
    }
    rows = [{"campo": key, "valor": value} for key, value in manifest.items()]
    for key, value in resumo.items():
        if key not in manifest:
            rows.append({"campo": f"resumo.{key}", "valor": value})
    return pd.DataFrame(rows)


def _montar_df_painel_executivo(resumo: dict | None, alertas: list[str] | None, df_validacao: pd.DataFrame, df_modelo: pd.DataFrame, df_bloqueados: pd.DataFrame, df_duplicidades: pd.DataFrame, df_cadastrar: pd.DataFrame) -> pd.DataFrame:
    resumo = resumo or {}
    alertas = alertas or []
    total_itens = 0 if df_validacao is None else len(df_validacao)
    total_fila = 0 if df_modelo is None else len(df_modelo)
    total_bloqueados = 0
    if df_bloqueados is not None and not df_bloqueados.empty and "Status Bloqueio" in df_bloqueados.columns:
        total_bloqueados = int(df_bloqueados["Status Bloqueio"].astype(str).str.contains("BLOQUEADO", case=False, na=False).sum())
    total_cnpj_cadastrar = 0
    if df_cadastrar is not None and not df_cadastrar.empty and "Status" in df_cadastrar.columns:
        total_cnpj_cadastrar = int(df_cadastrar["Status"].astype(str).str.upper().eq("A CADASTRAR").sum())
    total_duplicidades = 0
    if df_duplicidades is not None and not df_duplicidades.empty and "Status" in df_duplicidades.columns:
        total_duplicidades = int(df_duplicidades["Status"].astype(str).str.contains("DUPLICIDADE", case=False, na=False).sum())
    status_fila = "BLOQUEADA - VALIDAR PENDÊNCIAS" if total_bloqueados or total_cnpj_cadastrar else "PRONTA PARA CONFERÊNCIA MANUAL"
    rows = [
        {"Indicador": "Status executivo", "Valor": status_fila, "Ação recomendada": "Corrigir pendências antes de gerar TXT/fila" if "BLOQUEADA" in status_fila else "Conferir manualmente e seguir fluxo normal"},
        {"Indicador": "Layout", "Valor": resumo.get("layout", ""), "Ação recomendada": "Confirmar se a rede/layout está correta"},
        {"Indicador": "Arquivo/Lote", "Valor": resumo.get("arquivo", ""), "Ação recomendada": "Validar origem do arquivo"},
        {"Indicador": "Itens na validação", "Valor": total_itens, "Ação recomendada": "Conferir amostragem e totais"},
        {"Indicador": "Itens prontos no modelo/fila", "Valor": total_fila, "Ação recomendada": "Somente gerar TXT após validação manual"},
        {"Indicador": "Itens bloqueados", "Valor": total_bloqueados, "Ação recomendada": "Ver aba Itens Bloqueados Fila"},
        {"Indicador": "CNPJs/Chaves a cadastrar", "Valor": total_cnpj_cadastrar, "Ação recomendada": "Ver aba Cadastrar CNPJ"},
        {"Indicador": "Possíveis duplicidades", "Valor": total_duplicidades, "Ação recomendada": "Ver aba Duplicidades Validação"},
        {"Indicador": "Itens convertidos", "Valor": resumo.get("itens_convertidos", ""), "Ação recomendada": "Conferir fator e regra aplicada"},
        {"Indicador": "Itens não convertidos", "Valor": resumo.get("itens_nao_convertidos", ""), "Ação recomendada": "Não enviar para fila antes de corrigir"},
        {"Indicador": "Itens para validar conversão", "Valor": resumo.get("itens_validar_conversao", ""), "Ação recomendada": "Confirmar regra/fator no mapa/CSV"},
        {"Indicador": "Alertas registrados", "Valor": len(alertas), "Ação recomendada": "Ver LOG_ALERTAS e ALERTAS_ERROS"},
        {"Indicador": "Regra de segurança", "Valor": "TXT/fila não é gerado automaticamente", "Ação recomendada": "Fluxo correto: Excel de validação primeiro"},
    ]
    return pd.DataFrame(rows, columns=["Indicador", "Valor", "Ação recomendada"])

def criar_abas_validacao_padrao(
    writer,
    resumo: dict | None = None,
    df_intermediario: pd.DataFrame | None = None,
    df_final: pd.DataFrame | None = None,
    df_descartadas: pd.DataFrame | None = None,
    alertas: list[str] | None = None,
    df_validas: pd.DataFrame | None = None,
):
    resumo = resumo or {}
    layout_nome = _texto(resumo.get("layout", ""))
    df_validas_padrao = df_validas if df_validas is not None else df_intermediario
    if df_validas_padrao is None:
        df_validas_padrao = pd.DataFrame()
    df_validacao = _montar_linhas_validacao(df_validas_padrao, df_descartadas, df_final, resumo, alertas)
    df_modelo = _montar_df_modelo(df_final, df_validas, resumo, alertas, df_descartadas)
    df_cadastrar = _montar_df_cadastrar_cnpj(df_validacao, layout_nome)
    df_bloqueados = _montar_df_itens_bloqueados(df_validacao)
    df_duplicidades = _montar_df_duplicidades_validacao(df_validacao)
    df_manifesto = _montar_df_manifesto_processamento(resumo, alertas, df_validacao, df_modelo, df_bloqueados)
    df_painel = _montar_df_painel_executivo(resumo, alertas, df_validacao, df_modelo, df_bloqueados, df_duplicidades, df_cadastrar)

    df_painel.to_excel(writer, index=False, sheet_name=SHEET_PAINEL_EXECUTIVO)
    df_modelo.to_excel(writer, index=False, sheet_name=SHEET_MODELO_ROBOKOF)
    df_validacao.to_excel(writer, index=False, sheet_name=SHEET_VALIDACAO_PEDIDO)
    df_bloqueados.to_excel(writer, index=False, sheet_name=SHEET_ITENS_BLOQUEADOS)
    df_duplicidades.to_excel(writer, index=False, sheet_name=SHEET_DUPLICIDADES)
    df_cadastrar.to_excel(writer, index=False, sheet_name=SHEET_CADASTRAR_CNPJ)
    df_manifesto.to_excel(writer, index=False, sheet_name=SHEET_MANIFESTO)

    bloqueados = 0
    if not df_bloqueados.empty and "Status Bloqueio" in df_bloqueados.columns:
        bloqueados = int(df_bloqueados["Status Bloqueio"].astype(str).str.contains("BLOQUEADO", case=False, na=False).sum())
    if bloqueados:
        terminal_log.warning("[EXCEL] Item(ns) bloqueado(s) para fila/TXT antes da validação: %s", bloqueados)

    pendentes = df_cadastrar[df_cadastrar["Status"].astype(str) == "A CADASTRAR"]
    if not pendentes.empty:
        terminal_log.warning("[EXCEL] CNPJ(s) pendente(s) para cadastro: %s", len(pendentes))


def _fill_status_linha(row_text: str, ws_title: str, yellow_fill, green_fill, red_fill, blue_fill):
    if "BLOQUEADO" in row_text or "NÃO LIBERADO" in row_text or "NAO LIBERADO" in row_text:
        return red_fill
    if "DUPLICIDADE" in row_text or "VALIDAR" in row_text:
        return yellow_fill
    if "A CADASTRAR" in row_text:
        return yellow_fill
    if "ALERTA" in row_text or "ERRO" in row_text:
        return red_fill
    if "OK CONVERTIDO" in row_text or "OK SEM CONVERSÃO" in row_text or "OK - SEM CONVERSÃO" in row_text or "PRONTA PARA CONFERÊNCIA" in row_text:
        return green_fill
    if ws_title in {SHEET_PAINEL_EXECUTIVO, SHEET_MANIFESTO}:
        return blue_fill
    return None


def _ajustar_larguras_amostradas(ws, limite_linhas: int = MAX_LINHAS_AMOSTRA_LARGURA, largura_maxima: int = 55):
    max_por_coluna: dict[int, int] = {}
    max_row = min(ws.max_row or 1, int(limite_linhas or 400))
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            max_por_coluna[cell.column] = max(max_por_coluna.get(cell.column, 10), len(_texto(cell.value)))
    for col_idx, max_length in max_por_coluna.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, largura_maxima)


def aplicar_estilo_validacao(out_path: Path):
    wb = load_workbook(out_path)
    header_fill = PatternFill(fill_type="solid", fgColor="FF0000")
    header_font = Font(color="000000", bold=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    green_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    red_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    blue_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")

    for ws in wb.worksheets:
        total_celulas = int((ws.max_row or 0) * (ws.max_column or 0))
        modo_rapido = total_celulas > MAX_CELULAS_ESTILO_COMPLETO
        if modo_rapido:
            terminal_log.info(
                "[EXCEL] Estilo rapido ativado | aba=%s | linhas=%s | colunas=%s | celulas=%s",
                ws.title, ws.max_row, ws.max_column, total_celulas,
            )

        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                if not modo_rapido:
                    cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        if modo_rapido:
            # Mantém cores visuais importantes sem percorrer/formatar todas as
            # células com borda. A amostra cobre o topo da aba, onde ficam as
            # linhas mais usadas na conferência operacional.
            max_row_scan = min(ws.max_row or 1, MAX_LINHAS_AMOSTRA_LARGURA)
            for row in ws.iter_rows(min_row=2, max_row=max_row_scan):
                row_text = " ".join(_texto(cell.value).upper() for cell in row)
                fill = _fill_status_linha(row_text, ws.title, yellow_fill, green_fill, red_fill, blue_fill)
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.number_format = "@"
                    if fill is not None:
                        cell.fill = fill
            _ajustar_larguras_amostradas(ws)
            continue

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.number_format = "@"
            row_text = " ".join(_texto(cell.value).upper() for cell in row)
            fill = _fill_status_linha(row_text, ws.title, yellow_fill, green_fill, red_fill, blue_fill)
            if fill is not None:
                for cell in row:
                    cell.fill = fill

        _ajustar_larguras_amostradas(ws)

    wb.save(out_path)
    wb.close()


def _safe_sheet_name(nome: str, usados: set[str] | None = None) -> str:
    """Normaliza nome de aba para o limite/regras do Excel."""
    usados = usados or set()
    nome_limpo = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(nome or "").strip())[:31] or "Aba"
    base = nome_limpo
    contador = 2
    while nome_limpo in usados:
        sufixo = f"_{contador}"
        nome_limpo = f"{base[:31 - len(sufixo)]}{sufixo}"
        contador += 1
    usados.add(nome_limpo)
    return nome_limpo


def _normalizar_df_excel_unico(data: Any, colunas: list[str] | None = None) -> pd.DataFrame:
    """Converte listas/dicts/DataFrames em DataFrame seguro para abas de diagnóstico."""
    if data is None:
        return pd.DataFrame(columns=colunas or [])
    if isinstance(data, pd.DataFrame):
        df = data.copy()
    elif isinstance(data, list):
        df = pd.DataFrame(data)
    elif isinstance(data, dict):
        df = pd.DataFrame([data])
    else:
        df = pd.DataFrame([{"mensagem": str(data)}])
    if df.empty and colunas:
        return pd.DataFrame(columns=colunas)
    return df.fillna("")


def _escrever_df_em_ws(ws, df: pd.DataFrame):
    if df is None or df.empty:
        df = pd.DataFrame([{"status": "SEM_REGISTROS", "mensagem": "Sem registros para esta aba"}])
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)


def _estilizar_aba_diagnostico(ws):
    header_fill = PatternFill(fill_type="solid", fgColor="FF0000")
    header_font = Font(color="000000", bold=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    red_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    blue_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")

    total_celulas = int((ws.max_row or 0) * (ws.max_column or 0))
    modo_rapido = total_celulas > MAX_CELULAS_ESTILO_COMPLETO

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            if not modo_rapido:
                cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    max_row_scan = min(ws.max_row or 1, MAX_LINHAS_AMOSTRA_LARGURA) if modo_rapido else (ws.max_row or 1)
    for row in ws.iter_rows(min_row=2, max_row=max_row_scan):
        row_text = " ".join(_texto(cell.value).upper() for cell in row)
        fill = None
        if "ERRO" in row_text or "BLOQUEADO" in row_text or "GLN" in row_text or "NÃO" in row_text or "NAO" in row_text:
            fill = red_fill
        elif "ALERTA" in row_text or "VALIDAR" in row_text or "PEND" in row_text:
            fill = yellow_fill
        elif ws.title in {SHEET_RESUMO_PROCESSAMENTO, SHEET_LOGS_PROCESSAMENTO}:
            fill = blue_fill
        for cell in row:
            if not modo_rapido:
                cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.number_format = "@"
            if fill is not None:
                cell.fill = fill

    _ajustar_larguras_amostradas(ws, largura_maxima=60)


def anexar_diagnostico_excel_consolidado(
    excel_path: Path,
    resumo: dict | None = None,
    df_erros: pd.DataFrame | list[dict] | dict | None = None,
    df_pendencias_gln: pd.DataFrame | list[dict] | dict | None = None,
    df_logs: pd.DataFrame | list[dict] | dict | None = None,
    alertas: list[str] | None = None,
) -> Path | None:
    """Registra erros, pendências e logs dentro do próprio Excel consolidado.

    Padrão Robô KOF:
    - erro/alerta normal de layout, validação, GLN, CNPJ, SKU ou conversão não deve
      gerar um segundo arquivo ERR separado;
    - o Excel principal recebe abas próprias de auditoria;
    - só falhas fatais que impeçam abrir/gravar o Excel principal devem cair em arquivo
      separado de erro por fora do fluxo normal.
    """
    excel_path = Path(excel_path)
    if not excel_path.exists() or excel_path.suffix.lower() != ".xlsx":
        terminal_log.warning("[EXCEL_UNICO] Diagnóstico não anexado: caminho inválido ou não-xlsx: %s", excel_path)
        return None

    alertas = [str(a).strip() for a in (alertas or []) if str(a).strip()]
    abas: dict[str, pd.DataFrame] = {}

    resumo_rows = [{"campo": key, "valor": value} for key, value in (resumo or {}).items()]
    resumo_rows.extend(
        {"campo": f"alerta_{idx:03d}", "valor": alerta}
        for idx, alerta in enumerate(alertas, start=1)
    )
    resumo_rows.append({"campo": "gerado_em", "valor": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")})
    abas[SHEET_RESUMO_PROCESSAMENTO] = pd.DataFrame(resumo_rows)

    df_erros_norm = _normalizar_df_excel_unico(df_erros)
    if not df_erros_norm.empty:
        abas[SHEET_ALERTAS_ERROS] = df_erros_norm

    df_gln_norm = _normalizar_df_excel_unico(df_pendencias_gln)
    if not df_gln_norm.empty:
        abas[SHEET_PENDENCIAS_GLN] = df_gln_norm

    logs_rows = []
    if df_logs is not None:
        df_logs_norm = _normalizar_df_excel_unico(df_logs)
        if not df_logs_norm.empty:
            abas[SHEET_LOGS_PROCESSAMENTO] = df_logs_norm
    else:
        for alerta in alertas:
            logs_rows.append({
                "data_hora": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
                "nivel": "ALERTA",
                "mensagem": alerta,
            })
        if logs_rows:
            abas[SHEET_LOGS_PROCESSAMENTO] = pd.DataFrame(logs_rows)

    try:
        wb = load_workbook(excel_path)
        usados = {ws.title for ws in wb.worksheets}

        # Remove apenas as abas de diagnóstico que serão recriadas.
        for nome_aba in list(abas):
            if nome_aba in wb.sheetnames:
                del wb[nome_aba]
                usados.discard(nome_aba)

        for nome_aba, df in abas.items():
            nome_ws = _safe_sheet_name(nome_aba, usados)
            ws = wb.create_sheet(nome_ws)
            _escrever_df_em_ws(ws, df)
            _estilizar_aba_diagnostico(ws)

        wb.save(excel_path)
        wb.close()
        terminal_log.warning("[EXCEL_UNICO] Diagnóstico integrado no Excel consolidado: %s", excel_path)
        return excel_path
    except PermissionError:
        terminal_log.exception("[EXCEL_UNICO] Não foi possível gravar diagnóstico. Feche o Excel e tente novamente: %s", excel_path)
        return None
    except Exception:
        terminal_log.exception("[EXCEL_UNICO] Falha ao anexar diagnóstico no Excel consolidado: %s", excel_path)
        return None


def gerar_arquivo_erro(df_err: pd.DataFrame, out_path: Path):
    if df_err is None or df_err.empty:
        return
    out_path.parent.mkdir(parents=True, exist_ok=True)
    terminal_log.info("[EXCEL] Gerando arquivo de erro: %s", out_path)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_err.to_excel(writer, index=False, sheet_name="ERROS")
    _validar_xlsx_abre(out_path)


def gerar_arquivo_validacao_importacao(
    df_pendencias: pd.DataFrame,
    out_path: Path,
    resumo: dict | None = None,
):
    if df_pendencias is None or df_pendencias.empty:
        return

    out_path.parent.mkdir(parents=True, exist_ok=True)
    terminal_log.info("[EXCEL] Gerando Excel de pendencias para validacao: %s", out_path)

    df_validacao = df_pendencias.copy()
    if "motivo_descarte" in df_validacao.columns:
        df_validacao["status_validacao"] = df_validacao["motivo_descarte"].apply(
            lambda motivo: "PENDENTE_MATRICULA" if str(motivo) == "MATRICULA_NAO_ENCONTRADA" else "PENDENTE_VALIDACAO"
        )
    else:
        df_validacao["status_validacao"] = "PENDENTE_VALIDACAO"

    colunas_prioritarias = [
        "status_validacao",
        "motivo_descarte",
        "arquivo_origem",
        "layout_usado",
        "cnpj_lido",
        "matricula_lida",
        "matricula_final",
        "descricao_lida",
        "codigo_sku_lido",
        "ean_lido",
        "codigo_origem_lido",
        "sku_lido",
        "quantidade_lida",
        "numero_pedido_lido",
        "numero_pedido_final",
        "data_entrega_lida",
        "data_remessa_final",
    ]
    colunas = [col for col in colunas_prioritarias if col in df_validacao.columns]
    colunas += [col for col in df_validacao.columns if col not in colunas]
    df_validacao = df_validacao[colunas]

    df_matriculas = pd.DataFrame()
    if "motivo_descarte" in df_validacao.columns and "cnpj_lido" in df_validacao.columns:
        colunas_matricula = ["cnpj_lido", "matricula_lida", "arquivo_origem", "layout_usado"]
        for coluna in colunas_matricula:
            if coluna not in df_validacao.columns:
                df_validacao[coluna] = ""
        mask = df_validacao["motivo_descarte"].astype(str) == "MATRICULA_NAO_ENCONTRADA"
        df_matriculas = (
            df_validacao.loc[mask, colunas_matricula]
            .drop_duplicates()
            .sort_values(["layout_usado", "cnpj_lido"], kind="stable")
        )

    resumo_rows = [{"indicador": key, "valor": value} for key, value in (resumo or {}).items()]
    resumo_rows.append({"indicador": "linhas_pendentes", "valor": len(df_validacao)})
    if not df_matriculas.empty:
        resumo_rows.append({"indicador": "cnpjs_sem_matricula", "valor": df_matriculas["cnpj_lido"].nunique()})
    df_resumo = pd.DataFrame(resumo_rows)

    resumo_dict = resumo or {}
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        criar_abas_validacao_padrao(
            writer,
            resumo=resumo_dict,
            df_descartadas=df_validacao,
            alertas=["Pendencias para validacao"],
        )
        df_validacao.to_excel(writer, index=False, sheet_name="VALIDACAO")
        if not df_matriculas.empty:
            df_matriculas.to_excel(writer, index=False, sheet_name="CNPJ_SEM_MATRICULA")
        df_resumo.to_excel(writer, index=False, sheet_name="RESUMO")
    aplicar_estilo_validacao(out_path)
    _validar_xlsx_abre(out_path)


def _df_seguro(df: pd.DataFrame | None, colunas: list[str] | None = None) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=colunas or [])
    return df.copy()


def gerar_excel_validacao_completa(
    out_path: Path,
    resumo: dict | None = None,
    df_intermediario: pd.DataFrame | None = None,
    df_final: pd.DataFrame | None = None,
    df_descartadas: pd.DataFrame | None = None,
    alertas: list[str] | None = None,
    df_validas: pd.DataFrame | None = None,
    df_auditoria: pd.DataFrame | None = None,
    df_alertas_extracao: pd.DataFrame | None = None,
    df_rastreabilidade: pd.DataFrame | None = None,
    df_pendencias_gln: pd.DataFrame | None = None,
    df_logs_processamento: pd.DataFrame | None = None,
):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    terminal_log.info("[EXCEL] Gerando Excel de validacao completo: %s", out_path)

    alertas = [str(a) for a in (alertas or []) if str(a).strip()]
    df_resumo = pd.DataFrame([{"indicador": key, "valor": value} for key, value in (resumo or {}).items()])
    if df_resumo.empty:
        df_resumo = pd.DataFrame([{"indicador": "status", "valor": "Validacao gerada"}])

    df_alertas = pd.DataFrame(
        [{"tipo": "ALERTA", "mensagem": alerta} for alerta in alertas]
        or [{"tipo": "INFO", "mensagem": "Sem alertas registrados"}]
    )

    df_fila = _df_seguro(df_final, ["Matricula", "Sku", "Qtd", "Nº Pedido", "Data remessa"])
    df_inter = normalize_intermediate_columns(df_intermediario) if df_intermediario is not None else _df_seguro(df_intermediario)
    df_desc = _df_seguro(df_descartadas)
    df_alertas_ext = _df_seguro(df_alertas_extracao)
    df_rast = _df_seguro(df_rastreabilidade)
    df_pend_gln = _df_seguro(df_pendencias_gln)
    df_logs_proc = _df_seguro(df_logs_processamento)
    if df_logs_proc.empty:
        linhas_log = [
            {
                "data_hora": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
                "nivel": "INFO",
                "etapa": "GERACAO_EXCEL_VALIDACAO",
                "mensagem": f"{key}: {value}",
            }
            for key, value in (resumo or {}).items()
        ]
        linhas_log.extend(
            {
                "data_hora": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
                "nivel": "ALERTA",
                "etapa": "ALERTA_IMPORTACAO",
                "mensagem": alerta,
            }
            for alerta in alertas
        )
        df_logs_proc = pd.DataFrame(linhas_log or [{
            "data_hora": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
            "nivel": "INFO",
            "etapa": "GERACAO_EXCEL_VALIDACAO",
            "mensagem": "Sem logs adicionais para este processamento.",
        }])
    if df_rast.empty:
        df_rast = pd.DataFrame([{"status": "SEM_RASTREABILIDADE", "mensagem": "Arquivo processado por layout homologado/selecionado sem rastreabilidade."}])
    if not df_alertas_ext.empty:
        df_desc = pd.concat([df_desc, df_alertas_ext], ignore_index=True, sort=False) if not df_desc.empty else df_alertas_ext
    df_audit = _df_seguro(df_auditoria)

    if df_desc.empty:
        df_desc = pd.DataFrame([{"status": "SEM_DESCARTES", "mensagem": "Nenhuma linha descartada"}])
    if df_audit.empty:
        df_audit = pd.DataFrame([{"status": "SEM_AUDITORIA", "mensagem": "Sem auditoria detalhada de leitura para este arquivo"}])

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        criar_abas_validacao_padrao(
            writer,
            resumo=resumo,
            df_intermediario=df_intermediario,
            df_final=df_final,
            df_descartadas=df_descartadas,
            alertas=alertas,
            df_validas=df_validas,
        )
        df_resumo.to_excel(writer, index=False, sheet_name="RESUMO")
        df_fila.to_excel(writer, index=False, sheet_name="FILA_KOF_VALIDACAO")
        df_inter.to_excel(writer, index=False, sheet_name="DADOS_EXTRAIDOS")
        df_desc.to_excel(writer, index=False, sheet_name="ALERTAS_ERROS")
        df_audit.to_excel(writer, index=False, sheet_name="AUDITORIA_LEITURA")
        df_rast.to_excel(writer, index=False, sheet_name="RASTREABILIDADE_LAYOUT")
        df_alertas.to_excel(writer, index=False, sheet_name="LOG_ALERTAS")
        df_logs_proc.to_excel(writer, index=False, sheet_name=SHEET_LOGS_PROCESSAMENTO)
        if not df_pend_gln.empty:
            df_pend_gln.to_excel(writer, index=False, sheet_name=SHEET_PENDENCIAS_GLN)
    aplicar_estilo_validacao(out_path)
    _validar_xlsx_abre(out_path)
