# Autor: Kauê Melo
from __future__ import annotations

import csv
import json
import os
import re
import traceback
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
import pandas as pd

from layout_standard import normalize_intermediate_columns
from terminal_logger import get_terminal_logger
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

try:
    from .xls_biff_fallback import read_xls_cells, cells_to_matrix
except Exception:  # execução direta
    from xls_biff_fallback import read_xls_cells, cells_to_matrix



terminal_log = get_terminal_logger("excel_iquegami")

DE_PARA_PADRAO_MATRICULA_CNPJ: dict[str, str] = {
    "7120136174": "01915625000146",
    "7120148030": "01915625000227",
    "7120021397": "01915625000308",
    "7120041167": "01915625000499",
    "7120041165": "01915625000570",
    "7120043621": "01915625000650",
    "7110317767": "01915625000731",
    "7120067250": "01915625000812",
    "7120183043": "01915625000901",
    "7120224787": "01915625001037",
    "7120219208": "01915625001118",
    "7120254231": "01915625001207",
    "7120275165": "01915625001380",
    "7120291224": "01915625001460",
    "7120297546": "01915625001541",
    "7120498825": "01915625001622",
    "7120321089": "01915625001703",
    "7120360096": "01915625001894",
    "7120362796": "01915625001975",
    "7120440837": "01915625002009",
    "7120523572": "01915625002190",
    "7120502631": "01915625002270",
    "7120510870": "01915625002351",
}

@dataclass
class ItemPedido:
    cnpj: str
    matricula_loja: str
    sku: str
    qtd: Any
    numero_pedido: str
    arquivo_origem: str
    aba_origem: str
    linha_origem: int
    status: str = "OK"
    observacao: str = ""


@dataclass
class AlertaProcessamento:
    tipo: str
    arquivo: str
    aba: str
    linha: int | str
    pedido: str
    loja: str
    sku: str
    qtd: Any
    mensagem: str


class IquegamiParser:
    """Parser corporativo para pedidos Iquegami em XLS, XLSX e PDF textual."""

    def __init__(self):
        self.itens: list[ItemPedido] = []
        self.alertas: list[AlertaProcessamento] = []
        self.logs: list[dict[str, Any]] = []
        self.de_para_matricula_cnpj = self.carregar_de_para_matricula_cnpj()
        self.de_para_cnpj_matricula = {cnpj: mat for mat, cnpj in self.de_para_matricula_cnpj.items()}

    def log(self, nivel: str, evento: str, detalhe: str = "", **extra: Any) -> None:
        self.logs.append({
            "data_hora": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "nivel": nivel,
            "evento": evento,
            "detalhe": detalhe,
            **extra,
        })

    @classmethod
    def carregar_de_para_matricula_cnpj(cls) -> dict[str, str]:
        """Carrega o de/para matrícula -> CNPJ.

        O CSV fica em ../data/de_para_matricula_cnpj.csv para manutenção simples.
        Caso o arquivo não exista ou tenha problema, usa o de/para padrão embutido.
        """
        mapa = dict(DE_PARA_PADRAO_MATRICULA_CNPJ)
        csv_path = Path(__file__).resolve().parents[1] / "data" / "de_para_matricula_cnpj.csv"
        if not csv_path.exists():
            return mapa
        try:
            with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f, delimiter=";")
                for row in reader:
                    matricula = cls.somente_digitos(row.get("matricula") or row.get("MATRICULA") or "")
                    cnpj = cls.somente_digitos(row.get("cnpj") or row.get("CNPJ") or "")
                    if matricula and cnpj:
                        mapa[matricula] = cnpj.zfill(14)
        except Exception:
            # Mantém o robô funcionando mesmo se o CSV for editado incorretamente.
            return mapa
        return mapa

    def resolver_matricula_cnpj(self, loja: str) -> tuple[str, str, str]:
        """Resolve a coluna Loja para matrícula e CNPJ.

        Retorna: (matricula, cnpj, observação).
        - Se Loja tiver 14 dígitos, entende como CNPJ e tenta achar a matrícula inversa.
        - Se Loja tiver outro tamanho, entende como matrícula e procura no de/para.
        """
        loja = self.somente_digitos(loja)
        if not loja:
            return "", "", "Loja/Matrícula vazia"
        if len(loja) == 14:
            matricula = self.de_para_cnpj_matricula.get(loja, "")
            return matricula or loja, loja, "" if matricula else "CNPJ informado, matrícula não localizada no de/para"
        cnpj = self.de_para_matricula_cnpj.get(loja, "")
        if cnpj:
            return loja, cnpj, ""
        return loja, "", f"Matrícula {loja} sem CNPJ cadastrado no de/para"

    @staticmethod
    def somente_digitos(valor: Any) -> str:
        if valor is None:
            return ""
        if isinstance(valor, float) and valor.is_integer():
            return str(int(valor))
        if isinstance(valor, int):
            return str(valor)
        txt = str(valor).strip()
        if not txt:
            return ""
        if re.fullmatch(r"\d+\.0", txt):
            txt = txt[:-2]
        return re.sub(r"\D", "", txt)

    @staticmethod
    def normalizar_texto(valor: Any) -> str:
        if valor is None:
            return ""
        txt = str(valor).strip().lower()
        mapa = str.maketrans("áàãâäéèêëíìîïóòõôöúùûüç", "aaaaaeeeeiiiiooooouuuuc")
        return txt.translate(mapa)

    @staticmethod
    def parse_qtd(valor: Any) -> int | float | str | None:
        if valor is None or str(valor).strip() == "":
            return None
        if isinstance(valor, int):
            return valor
        if isinstance(valor, float):
            return int(valor) if valor.is_integer() else valor
        txt = str(valor).strip().replace(" ", "")
        if not txt:
            return None
        txt = txt.replace("−", "-")
        if "," in txt and "." in txt:
            txt = txt.replace(".", "").replace(",", ".")
        elif "," in txt:
            txt = txt.replace(",", ".")
        elif re.fullmatch(r"\d{1,3}(\.\d{3})+", txt):
            txt = txt.replace(".", "")
        try:
            num = float(txt)
            return int(num) if num.is_integer() else num
        except Exception:
            return valor

    @staticmethod
    def detectar_pedido_em_texto(texto: str, nome_arquivo: str = "") -> str:
        """Detecta o número do pedido em cabeçalhos e nomes de arquivo.

        Cobre os layouts Iquegami já observados:
        - Pedido 236369
        - Pedido de Compra Número 313487(...)
        - Pedido de Compra Numero 313487(...)
        - Cabeçalho compacto com o número imediatamente antes do parêntese.

        Retorna somente os dígitos do pedido para manter a saída padronizada.
        """
        texto = str(texto or "")
        padroes = [
            r"Pedido\s+de\s+Compra\s+N[úu]mero\s*[:\-]?\s*([0-9]{4,})",
            r"Pedido\s+de\s+Compra\s+Numero\s*[:\-]?\s*([0-9]{4,})",
            r"Pedido\s*[:\-]?\s*([0-9]{4,})",
            r"N[úu]mero\s*[:\-]?\s*([0-9]{4,})",
            r"\b[A-Z]?([0-9]{5,})\s*\(",
        ]
        for padrao in padroes:
            candidatos = re.findall(padrao, texto, flags=re.I)
            if candidatos:
                return re.sub(r"\D", "", str(candidatos[0]))

        # Fallback por nome do arquivo. Ignora datas simples e usa somente números com 5+ dígitos.
        candidatos = re.findall(r"\b([0-9]{5,})\b", Path(nome_arquivo).stem)
        return candidatos[0] if candidatos else ""

    def detectar_colunas(self, matrix: list[list[Any]]) -> tuple[int | None, dict[str, int]]:
        """Retorna índice da linha cabeçalho e colunas de Loja, SKU e Qtde.

        O Iquegami pode chegar em dois padrões principais:
        1) Loja | Ref. Fornecedor | Qtde. | Pedido
        2) Loja | Descrição do Produto | Ref. Fornec/Fornecedor | Qtde.

        A detecção abaixo aceita abreviações como "Ref. Fornec" e cabeçalhos
        repetidos no fim/meio do arquivo.
        """
        for i, row in enumerate(matrix[:150]):
            norm = [self.normalizar_texto(v) for v in row]
            joined = " | ".join(norm)
            tem_loja = "loja" in joined
            tem_qtd = any(("qtde" in c or "qtd" in c or "quant" in c) for c in norm)
            tem_ref = any((("ref" in c and ("fornec" in c or "fornecedor" in c)) or c in {"sku", "cod", "codigo", "codigo produto"}) for c in norm)
            if tem_loja and tem_qtd and tem_ref:
                cols: dict[str, int] = {}
                for idx, cell in enumerate(norm):
                    if cell == "loja" or "loja" in cell:
                        cols["loja"] = idx
                    if ("ref" in cell and ("fornec" in cell or "fornecedor" in cell)) or cell in {"sku", "cod", "codigo", "codigo produto"}:
                        cols["sku"] = idx
                    if "qtde" in cell or "qtd" in cell or "quant" in cell:
                        cols["qtd"] = idx
                if {"loja", "sku", "qtd"}.issubset(cols):
                    return i, cols
        return None, {}

    def linha_eh_cabecalho(self, row: list[Any]) -> bool:
        norm = [self.normalizar_texto(v) for v in row]
        joined = " | ".join(norm)
        return "loja" in joined and ("qtde" in joined or "qtd" in joined) and ("ref" in joined or "fornec" in joined or "fornecedor" in joined)

    def linha_eh_total(self, row: list[Any]) -> bool:
        norm = [self.normalizar_texto(v) for v in row]
        return any("total" in c or "totais" in c for c in norm)

    @staticmethod
    def linha_vazia(row: list[Any]) -> bool:
        return not any(str(v).strip() for v in row if v is not None)

    def primeira_qtd_numerica(self, row: list[Any]) -> int | float | None:
        for valor in row:
            qtd = self.parse_qtd(valor)
            if isinstance(qtd, (int, float)):
                return qtd
        return None

    def pedido_col_apos_qtd(self, matrix: list[list[Any]], header_idx: int, qtd_col: int) -> int | None:
        col = qtd_col + 1
        hits = 0
        for row in matrix[header_idx + 1: header_idx + 20]:
            if col < len(row) and re.fullmatch(r"\d{4,}", self.somente_digitos(row[col])):
                hits += 1
        return col if hits >= 2 else None

    def processar_matrix(self, matrix: list[list[Any]], arquivo: str, aba: str = "Planilha") -> None:
        texto_geral = "\n".join(" ".join(str(x or "") for x in row) for row in matrix[:40])
        pedido_header = self.detectar_pedido_em_texto(texto_geral, arquivo)
        pedido_atual = pedido_header
        header_idx, cols = self.detectar_colunas(matrix)
        if header_idx is None:
            self.alertas.append(AlertaProcessamento(
                tipo="LAYOUT_NAO_IDENTIFICADO", arquivo=arquivo, aba=aba, linha="-", pedido=pedido_header,
                loja="", sku="", qtd="", mensagem="Não encontrei o cabeçalho esperado: Loja / Ref. Fornecedor / Qtde."
            ))
            self.log("ERRO", "layout_nao_identificado", arquivo=arquivo, aba=aba)
            return

        pedido_col = self.pedido_col_apos_qtd(matrix, header_idx, cols["qtd"])
        self.log("INFO", "layout_identificado", arquivo=arquivo, aba=aba, linha_cabecalho=header_idx + 1,
                 col_loja=cols["loja"] + 1, col_sku=cols["sku"] + 1, col_qtd=cols["qtd"] + 1,
                 col_pedido=(pedido_col + 1 if pedido_col is not None else "cabeçalho/linha"), pedido=pedido_header)

        linhas_lidas = 0
        linhas_totais_ignoradas = 0
        cabecalhos_repetidos_ignorados = 0
        ultima_loja = ""

        for r_idx in range(header_idx + 1, len(matrix)):
            row = matrix[r_idx]

            if self.linha_vazia(row):
                continue

            row_text = " ".join(str(x or "") for x in row)
            pedido_linha = self.detectar_pedido_em_texto(row_text, "")
            if pedido_linha and pedido_linha != pedido_atual:
                pedido_atual = pedido_linha
                self.log("INFO", "pedido_identificado_na_linha", arquivo=arquivo, aba=aba, linha=r_idx + 1, pedido=pedido_atual)

            if self.linha_eh_cabecalho(row):
                cabecalhos_repetidos_ignorados += 1
                self.log("INFO", "cabecalho_repetido_ignorado", arquivo=arquivo, aba=aba, linha=r_idx + 1)
                continue

            if self.linha_eh_total(row):
                linhas_totais_ignoradas += 1
                self.log("INFO", "linha_total_ignorada", arquivo=arquivo, aba=aba, linha=r_idx + 1, qtd_total=self.primeira_qtd_numerica(row))
                continue

            def get(col: int):
                return row[col] if col < len(row) else None

            loja_original = self.somente_digitos(get(cols["loja"]))
            sku = self.somente_digitos(get(cols["sku"]))
            qtd = self.parse_qtd(get(cols["qtd"]))
            pedido = self.somente_digitos(get(pedido_col)) if pedido_col is not None else ""
            pedido = pedido or pedido_linha or pedido_atual or pedido_header

            # Regra operacional Rede Iquegami:
            # alguns arquivos chegam sem número de pedido na origem.
            # Para essa rede, o marcador técnico oficial é ponto (.),
            # permitindo que a linha siga para o Modelo/Fila quando matrícula, SKU,
            # quantidade e data estiverem válidos.
            if not str(pedido or "").strip():
                pedido = "."

            # Alguns arquivos vêm com linhas quebradas: a matrícula aparece na primeira
            # linha do bloco e as linhas seguintes podem carregar somente SKU/Qtd.
            loja = loja_original
            herdou_loja = False
            if loja:
                ultima_loja = loja
            elif ultima_loja and (sku or isinstance(qtd, (int, float))):
                loja = ultima_loja
                herdou_loja = True

            # Linhas sem conteúdo de item real são ignoradas. Isso evita somar títulos,
            # espaçadores e resíduos do Excel como se fossem SKUs.
            if not loja and not sku and (qtd is None or qtd == ""):
                continue
            if not loja and not sku and isinstance(qtd, (int, float)):
                # Número isolado sem Loja/SKU não é item. Total já foi tratado acima;
                # este fallback evita falso positivo quando o layout quebra visualmente.
                self.log("WARN", "linha_numerica_sem_loja_sku_ignorada", arquivo=arquivo, aba=aba, linha=r_idx + 1, qtd=qtd)
                continue

            matricula_loja, cnpj, obs_de_para = self.resolver_matricula_cnpj(loja)
            linhas_lidas += 1

            problemas = []
            if not loja:
                problemas.append("Loja/Matrícula vazia")
            if herdou_loja:
                problemas.append(f"Loja/Matrícula herdada da linha anterior: {loja}")
            if loja and not cnpj:
                problemas.append(obs_de_para or "CNPJ não localizado para a matrícula informada")
            if not sku:
                problemas.append("SKU/Ref. Fornecedor vazio ou não identificado")
            if qtd is None or qtd == "":
                problemas.append("Quantidade vazia")
            elif not isinstance(qtd, (int, float)):
                problemas.append(f"Quantidade inválida/não numérica: {qtd}")
            # Pedido ausente já foi normalizado para "." na Rede Iquegami.
            # Portanto, não gera pendência/bloqueio somente por falta do número.

            status = "OK" if not problemas else "ALERTA"
            obs = "; ".join(problemas)
            item = ItemPedido(
                cnpj=cnpj,
                matricula_loja=matricula_loja,
                sku=sku,
                qtd=qtd,
                numero_pedido=pedido,
                arquivo_origem=arquivo,
                aba_origem=aba,
                linha_origem=r_idx + 1,
                status=status,
                observacao=obs,
            )
            self.itens.append(item)

            if problemas:
                self.alertas.append(AlertaProcessamento(
                    tipo="DADO_PENDENTE", arquivo=arquivo, aba=aba, linha=r_idx + 1, pedido=pedido,
                    loja=matricula_loja, sku=sku, qtd=qtd, mensagem=obs
                ))
        self.log("INFO", "arquivo_processado", arquivo=arquivo, aba=aba, linhas_lidas=linhas_lidas,
                 linhas_totais_ignoradas=linhas_totais_ignoradas, cabecalhos_repetidos_ignorados=cabecalhos_repetidos_ignorados)

    def ler_xlsx(self, path: str) -> Iterable[tuple[str, list[list[Any]]]]:
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            matrix = [list(row) for row in ws.iter_rows(values_only=True)]
            yield ws.title, matrix

    def ler_xls(self, path: str) -> Iterable[tuple[str, list[list[Any]]]]:
        # Preferência: xlrd. Fallback: leitor BIFF simples embutido.
        try:
            import xlrd  # type: ignore
            book = xlrd.open_workbook(path)
            for sheet in book.sheets():
                matrix = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
                yield sheet.name, matrix
            return
        except Exception as e:
            self.log("WARN", "xlrd_indisponivel_ou_falhou", detalhe=str(e), arquivo=os.path.basename(path))
        cells = read_xls_cells(path)
        yield "Planilha", cells_to_matrix(cells)

    def ler_pdf_textual(self, path: str) -> Iterable[tuple[str, list[list[Any]]]]:
        try:
            import pdfplumber  # type: ignore
        except Exception:
            self.alertas.append(AlertaProcessamento(
                tipo="PDF_NAO_PROCESSADO", arquivo=os.path.basename(path), aba="PDF", linha="-", pedido="",
                loja="", sku="", qtd="", mensagem="pdfplumber não instalado. Instale requirements.txt para habilitar leitura de PDF textual."
            ))
            return
        texto_total = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                texto_total.append(page.extract_text() or "")
        texto = "\n".join(texto_total)
        pedido = self.detectar_pedido_em_texto(texto, path)
        matrix = [[f"Pedido {pedido}"], ["Loja", "Ref. Fornecedor", "Qtde."]]
        for line in texto.splitlines():
            # Ex.: 7120291224 92522 25 236369
            m = re.search(r"\b(\d{8,14})\b\s+\b(\d{4,8})\b\s+(\d+(?:[\.,]\d+)?)\s*(\d{4,})?", line)
            if m:
                loja, sku, qtd, ped_linha = m.groups()
                matrix.append([loja, sku, qtd, ped_linha or pedido])
        yield "PDF", matrix

    def processar_arquivo(self, path: str) -> None:
        nome = os.path.basename(path)
        ext = Path(path).suffix.lower()
        self.log("INFO", "iniciando_arquivo", arquivo=nome)
        try:
            if ext == ".xlsx":
                for aba, matrix in self.ler_xlsx(path):
                    self.processar_matrix(matrix, nome, aba)
            elif ext == ".xls":
                for aba, matrix in self.ler_xls(path):
                    self.processar_matrix(matrix, nome, aba)
            elif ext == ".pdf":
                for aba, matrix in self.ler_pdf_textual(path):
                    self.processar_matrix(matrix, nome, aba)
            else:
                self.alertas.append(AlertaProcessamento(
                    tipo="EXTENSAO_NAO_SUPORTADA", arquivo=nome, aba="-", linha="-", pedido="",
                    loja="", sku="", qtd="", mensagem=f"Extensão não suportada: {ext}. Use .xls, .xlsx ou PDF textual."
                ))
                self.log("ERRO", "extensao_nao_suportada", arquivo=nome, extensao=ext)
        except Exception as e:
            self.alertas.append(AlertaProcessamento(
                tipo="ERRO_PROCESSAMENTO", arquivo=nome, aba="-", linha="-", pedido="",
                loja="", sku="", qtd="", mensagem=str(e)
            ))
            self.log("ERRO", "falha_arquivo", arquivo=nome, detalhe=str(e), traceback=traceback.format_exc())

    def gerar_excel(self, output_path: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "PEDIDOS_CONSOLIDADO"
        headers = ["CNPJ", "MATRICULA_LOJA", "SKU", "QTD", "NUMERO_PEDIDO", "ARQUIVO_ORIGEM", "ABA_ORIGEM", "LINHA_ORIGEM", "STATUS", "OBSERVACAO"]
        ws.append(headers)
        for item in self.itens:
            ws.append([item.cnpj, item.matricula_loja, item.sku, item.qtd, item.numero_pedido,
                       item.arquivo_origem, item.aba_origem, item.linha_origem, item.status, item.observacao])
        self._formatar_planilha(ws, cor="#7A0000")
        for col in ["A", "B", "C", "E"]:
            for cell in ws[col][1:]:
                cell.number_format = "@"

        wa = wb.create_sheet("ALERTAS")
        wa.append(["TIPO", "ARQUIVO", "ABA", "LINHA", "PEDIDO", "LOJA", "SKU", "QTD", "MENSAGEM"])
        for a in self.alertas:
            wa.append([a.tipo, a.arquivo, a.aba, a.linha, a.pedido, a.loja, a.sku, a.qtd, a.mensagem])
        self._formatar_planilha(wa, cor="#111111")

        wl = wb.create_sheet("LOG_PROCESSAMENTO")
        log_headers = ["DATA_HORA", "NIVEL", "EVENTO", "DETALHE", "EXTRAS_JSON"]
        wl.append(log_headers)
        for l in self.logs:
            base = [l.get("data_hora", ""), l.get("nivel", ""), l.get("evento", ""), l.get("detalhe", "")]
            extras = {k: v for k, v in l.items() if k not in {"data_hora", "nivel", "evento", "detalhe"}}
            wl.append(base + [json.dumps(extras, ensure_ascii=False)])
        self._formatar_planilha(wl, cor="#7A0000")

        wr = wb.create_sheet("RESUMO")
        total_ok = sum(1 for i in self.itens if i.status == "OK")
        pedidos = sorted({i.numero_pedido for i in self.itens if i.numero_pedido})
        arquivos = sorted({i.arquivo_origem for i in self.itens})
        wr.append(["INDICADOR", "VALOR"])
        wr.append(["Arquivos processados", len(arquivos)])
        wr.append(["Pedidos identificados", len(pedidos)])
        wr.append(["Itens totais", len(self.itens)])
        wr.append(["Itens OK", total_ok])
        wr.append(["Alertas", len(self.alertas)])
        wr.append(["Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
        wr.append([])
        wr.append(["PEDIDOS", ", ".join(pedidos)])
        self._formatar_planilha(wr, cor="#111111")

        wd = wb.create_sheet("DE_PARA_MATRICULA_CNPJ")
        wd.append(["MATRICULA", "CNPJ"])
        for matricula, cnpj in sorted(self.de_para_matricula_cnpj.items()):
            wd.append([matricula, cnpj])
        self._formatar_planilha(wd, cor="#7A0000")
        for col in ["A", "B"]:
            for cell in wd[col][1:]:
                cell.number_format = "@"

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    @staticmethod
    def _formatar_planilha(ws, cor: str) -> None:
        header_fill = PatternFill("solid", fgColor=cor.replace("#", ""))
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.freeze_panes = "A2"
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            max_len = max(len(str(c.value or "")) for c in column_cells[:200])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)
        ws.auto_filter.ref = ws.dimensions




def ler_excel_iquegami(caminho_arquivo: str, layout_config: dict, mapeamentos_df=None) -> dict:
    """Integra o parser Iquegami antigo ao fluxo padrão do Robô KOF.

    Entrada: .xls/.xlsx do Iquegami. O parser antigo também possui leitura textual de PDF,
    mas no Robô KOF esta automação foi cadastrada como EXCEL para não misturar redes/layouts.
    """
    nome_layout = str(layout_config.get("nome_layout", "REDE IQUEGAMI Excel Pedido"))
    parser = IquegamiParser()
    try:
        parser.processar_arquivo(caminho_arquivo)
        rows = []
        for item in parser.itens:
            alerta = item.observacao if item.status != "OK" else ""
            rows.append({
                "matricula_lida": item.matricula_loja,
                "cnpj_lido": item.cnpj,
                "sku_lido": str(item.sku or ""),
                "codigo_sku_lido": str(item.sku or ""),
                "ean_lido": "",
                "descricao_lida": "",
                "quantidade_lida": str(item.qtd or ""),
                "numero_pedido_lido": item.numero_pedido,
                "data_entrega_lida": "",
                "codigo_loja_lido": item.matricula_loja,
                "linha_origem": str(item.linha_origem),
                "origem_extracao": "EXCEL_IQUEGAMI_LEGADO_INTEGRADO",
                "status_extracao": "OK" if item.status == "OK" else "VALIDAR",
                "alerta_extracao": alerta,
                "qtd_original": str(item.qtd or ""),
                "tipo_qtd_original": "CAIXARIA",
                "fator_conversao": "1",
                "qtd_convertida": str(item.qtd or ""),
                "qtd_final": str(item.qtd or ""),
                "status_conversao": "OK SEM CONVERSÃO",
                "regra_aplicada_conversao": "IQUEGAMI_QTD_JA_CAIXARIA",
                "origem_regra_conversao": "robo_iquegami_excel/core/parser_iquegami.py",
                "aba_origem": item.aba_origem,
            })
        df = normalize_intermediate_columns(pd.DataFrame(rows), arquivo_origem=os.path.basename(caminho_arquivo), layout_usado=nome_layout)
        alertas = [a.mensagem for a in parser.alertas if str(a.mensagem).strip()]
        if df.empty:
            alertas.append("IQUEGAMI: nenhum item foi extraído. Verifique cabeçalho Loja / Ref. Fornecedor / Qtde.")
        auditoria = pd.DataFrame(parser.logs)
        terminal_log.info(
            "[IQUEGAMI] leitura concluída | arquivo=%s | itens=%s | alertas=%s",
            os.path.basename(caminho_arquivo), len(df), len(alertas),
        )
        return {
            "sucesso": not df.empty,
            "mensagem": f"Leitura Iquegami concluída com {len(df)} item(ns)" if not df.empty else "Nenhum item Iquegami extraído",
            "df_intermediario": df,
            "qtd_linhas_lidas": len(df),
            "qtd_itens_extraidos": len(df),
            "qtd_linhas_planilha": len(df),
            "alertas": sorted({str(a) for a in alertas if str(a).strip()}),
            "df_auditoria_paginas": auditoria,
        }
    except Exception as exc:
        terminal_log.exception("[IQUEGAMI] Falha ao ler arquivo: %s", caminho_arquivo)
        return {
            "sucesso": False,
            "mensagem": str(exc),
            "df_intermediario": pd.DataFrame(),
            "qtd_linhas_lidas": 0,
            "qtd_itens_extraidos": 0,
            "alertas": [f"IQUEGAMI_ERRO_LEITURA: {exc}"],
            "df_auditoria_paginas": pd.DataFrame(parser.logs),
        }


def processar_lote(arquivos: list[str], output_dir: str) -> dict[str, Any]:
    parser = IquegamiParser()
    for arquivo in arquivos:
        parser.processar_arquivo(arquivo)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_dir, f"consolidado_iquegami_{stamp}.xlsx")
    parser.gerar_excel(excel_path)
    log_json = os.path.join(output_dir, f"log_iquegami_{stamp}.json")
    with open(log_json, "w", encoding="utf-8") as f:
        json.dump({"itens": [asdict(i) for i in parser.itens], "alertas": [asdict(a) for a in parser.alertas], "logs": parser.logs}, f, ensure_ascii=False, indent=2)
    log_txt = os.path.join(output_dir, f"log_iquegami_{stamp}.txt")
    with open(log_txt, "w", encoding="utf-8") as f:
        for l in parser.logs:
            f.write(f"[{l.get('data_hora')}] {l.get('nivel')} - {l.get('evento')} - {l.get('detalhe')} - {l}\n")
    return {
        "excel": excel_path,
        "log_json": log_json,
        "log_txt": log_txt,
        "total_itens": len(parser.itens),
        "total_alertas": len(parser.alertas),
        "pedidos": sorted({i.numero_pedido for i in parser.itens if i.numero_pedido}),
        "arquivos": [os.path.basename(a) for a in arquivos],
    }
