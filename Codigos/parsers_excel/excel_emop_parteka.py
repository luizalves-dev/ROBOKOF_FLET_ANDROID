# Autor: Kauê Melo
from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook

from layout_standard import normalize_intermediate_columns
from terminal_logger import get_terminal_logger

terminal_log = get_terminal_logger("excel_emop_parteka")

# Regras herdadas do projeto antigo emop_atletico_bot.
SKU_OVERRIDES: Dict[str, str] = {
    "7898770420042": "119605",
    "7894900664003": "119181",
    "7894900664010": "119182",
    "7896388010556": "139765",
    "7791540127106": "139738",
    "7804330006717": "139693",
    "7894900530032": "92521",
    "78934115": "139367",
    "0000078934115": "139367",
    "7894900027013": "56600",
    "7896388010303": "139766",
    "8412598009358": "139770",
}

HARDCODED_CNPJ_MATRICULA: Dict[str, str] = {
    "12364137000708": "7120259997",
    "12364137000961": "7120298231",
    "12364137000538": "7120077223",
    "12364137000619": "7120541420",
    "12364137000880": "7120491774",
    "12364137000457": "7120077222",
    "12364137000295": "7120043355",
    "12364137000376": "7120043356",
}


@dataclass
class StoreInfo:
    codigo_loja: str
    cnpj: str
    matricula: str = ""
    nome_loja: str = ""


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\xa0", " ").strip()


def _digits(value: Any) -> str:
    if value is None:
        return ""
    text = _numeric_text(value)
    return re.sub(r"\D+", "", text)


def _numeric_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    text = str(value).strip()
    if re.fullmatch(r"[+-]?(?:\d+(?:[\.,]\d+)?|[\.,]\d+)[eE][+-]?\d+", text):
        try:
            return format(Decimal(text.replace(",", ".")).quantize(Decimal("1")), "f")
        except (InvalidOperation, ValueError):
            return text
    return text


def _normalize_cnpj(value: Any) -> str:
    digits = _digits(value)
    if not digits:
        return ""
    if len(digits) < 14:
        return digits.zfill(14)
    if len(digits) > 14:
        return digits[-14:]
    return digits


def _parse_qty(value: Any) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value) if value > 0 else ""
    if isinstance(value, float):
        if value <= 0:
            return ""
        return str(int(value)) if value.is_integer() else str(value).replace(".", ",")
    text = str(value).strip().replace("\xa0", "").replace(" ", "")
    if not text or text in {"-", "0", "0,0", "0.0"}:
        return ""
    match = re.search(r"[-+]?\d[\d\.,]*", text)
    if not match:
        return ""
    number = match.group(0)
    if "." in number and "," in number:
        number = number.replace(".", "").replace(",", ".")
    elif "," in number:
        number = number.replace(",", ".")
    try:
        val = float(number)
    except ValueError:
        return ""
    if val <= 0:
        return ""
    return str(int(round(val))) if float(val).is_integer() or abs(val - round(val)) < 1e-9 else str(val).replace(".", ",")


def _clean_sku(cod_forn: Any) -> Tuple[str, str]:
    text = _text(cod_forn)
    if not text:
        return "", "missing"
    numbers = re.findall(r"\d+", text)
    if not numbers:
        return "", "missing"
    if len(numbers) > 1:
        return numbers[0], "cod_forn_multiplo"
    return numbers[0], "cod_forn"


def _resolve_sku(cod_forn: Any, barcode: Any) -> Tuple[str, str]:
    sku, origem = _clean_sku(cod_forn)
    if sku:
        return sku, origem
    barcode_digits = _digits(barcode)
    if barcode_digits in SKU_OVERRIDES:
        return SKU_OVERRIDES[barcode_digits], "override_barcode"
    sem_zero = barcode_digits.lstrip("0")
    if sem_zero in SKU_OVERRIDES:
        return SKU_OVERRIDES[sem_zero], "override_barcode"
    return "", "missing"


def _norm_header(value: Any) -> str:
    text = _text(value).lower()
    text = text.replace("á", "a").replace("ã", "a").replace("â", "a").replace("à", "a")
    text = text.replace("é", "e").replace("ê", "e").replace("í", "i")
    text = text.replace("ó", "o").replace("ô", "o").replace("õ", "o").replace("ú", "u").replace("ç", "c")
    return re.sub(r"\s+", " ", text).strip()


def _find_header_row(ws) -> int:
    for row_idx in range(1, ws.max_row + 1):
        vals = [_norm_header(ws.cell(row_idx, c).value) for c in range(1, min(ws.max_column, 30) + 1)]
        joined = " | ".join(v for v in vals if v)
        if "codigo" in vals and "descricao" in vals and any(v in {"emb", "embal", "embalagem"} for v in vals):
            return row_idx
        if "codigo" in joined and "descricao" in joined and "cod barras" in joined and "cod forn" in joined:
            return row_idx
    raise ValueError("Cabeçalho Parteka/EMOP não encontrado: esperado Código / Descrição / Cod Barras / Cod Forn / Emb / lojas.")


def _detect_column_indexes(ws, header_row: int) -> Dict[str, int]:
    aliases = {
        "codigo": {"codigo", "código"},
        "descricao": {"descricao", "descrição"},
        "cod_barras": {"cod barras", "codigo barras", "código barras", "ean"},
        "preco": {"preco", "preço"},
        "cod_forn": {"cod forn", "cod. forn", "cod fornecedor"},
        "emb": {"emb", "embal", "embalagem"},
    }
    found: Dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        value = _norm_header(ws.cell(header_row, col_idx).value)
        for key, options in aliases.items():
            if value in options:
                found[key] = col_idx
    missing = {"codigo", "descricao", "cod_barras", "cod_forn", "emb"} - set(found)
    if missing:
        raise ValueError(f"Cabeçalhos obrigatórios Parteka/EMOP ausentes: {sorted(missing)}")
    return found


def _find_store_section(ws, header_row: int) -> Dict[str, StoreInfo]:
    stores: Dict[str, StoreInfo] = {}
    for row_idx in range(1, header_row):
        values = [_text(ws.cell(row_idx, c).value) for c in range(1, min(ws.max_column, 30) + 1)]
        line = " ".join(v for v in values if v)
        m = re.match(r"\s*(\d{3})\b", line)
        if not m:
            continue
        loja = m.group(1)
        tokens = re.findall(r"\d+", line)
        cnpj = ""
        for token in reversed(tokens):
            if len(token) == 14:
                cnpj = _normalize_cnpj(token)
                break
        if not cnpj:
            continue
        stores[loja] = StoreInfo(
            codigo_loja=loja,
            cnpj=cnpj,
            matricula=HARDCODED_CNPJ_MATRICULA.get(cnpj, ""),
            nome_loja=re.sub(r"^\s*\d{3}\s*", "", line).strip(),
        )
    return stores


def _detect_store_columns(ws, header_row: int) -> Dict[int, str]:
    store_columns: Dict[int, str] = {}
    for col_idx in range(1, ws.max_column + 1):
        value = _text(ws.cell(header_row, col_idx).value)
        if re.fullmatch(r"\d{3}", value):
            store_columns[col_idx] = value
    if not store_columns:
        raise ValueError("Nenhuma coluna de loja Parteka/EMOP foi encontrada no cabeçalho (001, 002, 006...).")
    return store_columns


def _extract_order_number(ws, caminho_arquivo: str) -> str:
    patterns = [
        r"pedido[s]?\s*(?:n[ºo°]?\s*)?[:\-]?\s*(\d{5,12})",
        r"n[ºo°]?\s*pedido\s*[:\-]?\s*(\d{5,12})",
    ]
    for row_idx in range(1, min(ws.max_row, 40) + 1):
        line = " | ".join(_text(ws.cell(row_idx, c).value) for c in range(1, min(ws.max_column, 25) + 1) if _text(ws.cell(row_idx, c).value))
        for pat in patterns:
            m = re.search(pat, line, flags=re.I)
            if m:
                return _digits(m.group(1))
    # O layout antigo muitas vezes não traz número de pedido. Mantemos vazio; o padronizador Excel usa ponto como compatibilidade.
    return ""


def ler_excel_emop_parteka(caminho_arquivo: str, layout_config: dict, mapeamentos_df=None) -> dict:
    nome_layout = str(layout_config.get("nome_layout", "GRUPO E.M.O.P. / PARTEKA Excel Flex"))
    alertas: List[str] = []
    auditoria: List[dict] = []
    rows: List[dict] = []
    try:
        wb = load_workbook(caminho_arquivo, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row = _find_header_row(ws)
        cols = _detect_column_indexes(ws, header_row)
        stores = _find_store_section(ws, header_row)
        store_cols = _detect_store_columns(ws, header_row)
        pedido = _extract_order_number(ws, caminho_arquivo)
        terminal_log.info(
            "[PARTEKA] layout identificado | arquivo=%s | aba=%s | header=%s | lojas=%s | pedido=%s",
            Path(caminho_arquivo).name, ws.title, header_row, len(store_cols), pedido or "vazio",
        )

        if not stores:
            alertas.append("PARTEKA: bloco Loja/Razão Social/CNPJ não localizado; CNPJs podem ficar pendentes.")

        for row_idx in range(header_row + 1, ws.max_row + 1):
            codigo_interno = _digits(ws.cell(row_idx, cols["codigo"]).value)
            descricao = _text(ws.cell(row_idx, cols["descricao"]).value)
            barcode = _digits(ws.cell(row_idx, cols["cod_barras"]).value)
            cod_forn = _text(ws.cell(row_idx, cols["cod_forn"]).value)
            emb = _text(ws.cell(row_idx, cols["emb"]).value)
            preco = _text(ws.cell(row_idx, cols.get("preco", 0)).value) if cols.get("preco") else ""
            if not codigo_interno and not descricao and not barcode and not cod_forn:
                continue
            sku, sku_source = _resolve_sku(cod_forn, barcode)
            for col_idx, loja_codigo in store_cols.items():
                qtd = _parse_qty(ws.cell(row_idx, col_idx).value)
                if not qtd:
                    continue
                store = stores.get(loja_codigo, StoreInfo(codigo_loja=loja_codigo, cnpj="", matricula=""))
                problemas = []
                if not store.cnpj:
                    problemas.append(f"CNPJ não localizado para loja {loja_codigo}")
                if not sku:
                    problemas.append(f"SKU não localizado para Cod Forn/Barcode | cod_forn={cod_forn} | ean={barcode}")
                if sku_source == "override_barcode":
                    problemas.append("SKU resolvido por exceção histórica EAN->SKU")
                alerta = "; ".join(problemas)
                rows.append({
                    "matricula_lida": store.matricula,
                    "cnpj_lido": store.cnpj,
                    "sku_lido": sku,
                    "codigo_sku_lido": sku,
                    "ean_lido": barcode,
                    "descricao_lida": descricao,
                    "quantidade_lida": qtd,
                    "numero_pedido_lido": pedido or ".",
                    "data_entrega_lida": "",
                    "codigo_loja_lido": loja_codigo,
                    "loja_lida": store.nome_loja,
                    "codigo_origem_lido": codigo_interno,
                    "linha_origem": str(row_idx),
                    "origem_extracao": "EXCEL_EMOP_PARTEKA_FLEX",
                    "status_extracao": "OK" if not alerta else "VALIDAR",
                    "alerta_extracao": alerta,
                    "qtd_original": qtd,
                    "tipo_qtd_original": "CAIXARIA",
                    "fator_conversao": "1",
                    "qtd_convertida": qtd,
                    "qtd_final": qtd,
                    "status_conversao": "OK SEM CONVERSÃO",
                    "regra_aplicada_conversao": "PARTEKA_QTD_JA_CAIXARIA",
                    "origem_regra_conversao": "Projeto antigo emop_atletico_bot",
                    "embalagem_lida": emb,
                    "preco_lido": preco,
                    "sku_origem": sku_source,
                })
                if alerta:
                    alertas.append(f"Linha {row_idx} loja {loja_codigo}: {alerta}")
        auditoria.append({
            "arquivo": Path(caminho_arquivo).name,
            "aba": ws.title,
            "linhas_planilha": ws.max_row,
            "colunas_planilha": ws.max_column,
            "linha_cabecalho": header_row,
            "lojas_detectadas": len(store_cols),
            "itens_extraidos": len(rows),
            "pedido_identificado": pedido,
            "status": "OK" if rows else "SEM_ITENS",
        })
        df = normalize_intermediate_columns(pd.DataFrame(rows), arquivo_origem=Path(caminho_arquivo).name, layout_usado=nome_layout)
        if df.empty:
            alertas.append("PARTEKA: nenhum item com quantidade foi encontrado no layout Flex/EMOP.")
        return {
            "sucesso": not df.empty,
            "mensagem": f"Leitura Parteka/EMOP concluída com {len(df)} item(ns)" if not df.empty else "Nenhum item Parteka/EMOP extraído",
            "df_intermediario": df,
            "qtd_linhas_lidas": len(df),
            "qtd_itens_extraidos": len(df),
            "qtd_linhas_planilha": ws.max_row,
            "alertas": sorted({a for a in alertas if a}),
            "df_auditoria_paginas": pd.DataFrame(auditoria),
        }
    except Exception as exc:
        terminal_log.exception("[PARTEKA] Falha ao ler arquivo: %s", caminho_arquivo)
        return {
            "sucesso": False,
            "mensagem": str(exc),
            "df_intermediario": pd.DataFrame(),
            "qtd_linhas_lidas": 0,
            "qtd_itens_extraidos": 0,
            "alertas": [f"PARTEKA_ERRO_LEITURA: {exc}"],
            "df_auditoria_paginas": pd.DataFrame(auditoria),
        }
