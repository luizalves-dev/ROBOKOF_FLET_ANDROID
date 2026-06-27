from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
import json
import re
import unicodedata
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

from gerador_erro import aplicar_estilo_validacao, criar_abas_validacao_padrao
from layout_standard import STANDARD_INTERMEDIATE_COLUMNS, normalize_intermediate_columns


MISSING_PAIR_MESSAGE = (
    "Para processar a Rede Coelho Diniz, é necessário enviar o Excel do pedido "
    "e o Excel com a relação de matrículas e datas."
)

INTERMEDIATE_COLUMNS = STANDARD_INTERMEDIATE_COLUMNS

DE_PARA_MATRICULA_CNPJ: Dict[str, str] = {
    "1700019409": "41930199000134",
    "1700019410": "41930199000215",
    "1700273623": "41930199000304",
    "1700139600": "41930199000487",
    "1700019971": "41930199000568",
    "1700020162": "41930199000649",
    "1700024240": "41930199000720",
    "1700273644": "41930199000991",
    "1700021739": "41930199001025",
    "1700433548": "41930199001106",
    "1700433521": "41930199001297",
    "1700433526": "41930199001378",
    "1700458507": "41930199001459",
    "1700532177": "41930199001530",
    "1700507248": "41930199001610",
    "1700541696": "41930199001700",
    "1700538416": "41930199001882",
    "1700553635": "41930199001963",
    "1700556611": "41930199002005",
    "1700562203": "41930199002269",
    "1700635736": "41930199002340",
    "1700600583": "41930199002420",
}

STOP_HEADERS = {
    "total",
    "total geral",
    "pr.final",
    "pr final",
    "preco final",
    "vr.total",
    "vr total",
    "valor total",
    "verba",
    "total verba",
}


@dataclass
class CoelhoRow:
    cnpj: str
    matricula: str
    sku: str
    qtd: str
    numero_pedido: str
    data_digitacao: str
    atendimento: str
    arquivo_origem: str
    aba: str
    linha_origem: str
    loja: str
    codigo_loja_original: str
    ean: str
    descricao: str
    divisor: str


@dataclass
class CoelhoLog:
    data_hora: str
    nivel: str
    arquivo: str
    contexto: str
    mensagem: str
    detalhe: str = ""


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def only_digits(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\D", "", str(value))


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\xa0", " ").strip()


def normalize_text(value: Any) -> str:
    text = cell_text(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text).lower().strip()
    return text


def normalize_date_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted.strftime("%d/%m/%Y")
        except Exception:
            pass

    text = cell_text(value)
    match = re.search(r"(\d{1,2})[\./-](\d{1,2})[\./-](\d{2,4})", text)
    if not match:
        return text.strip()
    day, month, year = match.groups()
    if len(year) == 2:
        year = "20" + year
    try:
        return datetime(int(year), int(month), int(day)).strftime("%d/%m/%Y")
    except ValueError:
        return text.strip()


def parse_quantity(value: Any) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value) if value > 0 else ""
    if isinstance(value, float):
        if value <= 0:
            return ""
        return str(int(value)) if value.is_integer() else str(value).replace(".", ",")

    text = str(value).replace("\xa0", " ").strip()
    match = re.search(r"[-+]?\d[\d\.,]*", text)
    if not match:
        return ""
    number = match.group(0)
    if "." in number and "," in number:
        number = number.replace(".", "").replace(",", ".")
    elif "." in number:
        parts = number.split(".")
        if len(parts[-1]) == 3 and all(part.isdigit() for part in parts):
            number = "".join(parts)
    elif "," in number:
        parts = number.split(",")
        if len(parts[-1]) == 3 and all(part.isdigit() for part in parts):
            number = "".join(parts)
        else:
            number = number.replace(",", ".")
    try:
        val = float(number)
    except ValueError:
        return ""
    if val <= 0:
        return ""
    return str(int(val)) if val.is_integer() else str(val).replace(".", ",")


def normalizar_pedido_coelho_diniz(value: Any) -> str:
    """Normaliza o número de pedido Coelho Diniz preservando todos os dígitos úteis.

    Regra de negócio validada:
    - arquivo ``Pedido_053641_SPALINDUSTRIA.xlsx`` deve gerar ``53641``;
    - arquivo ``Pedido 053574 SPALINDUSTRIA.xlsx`` deve gerar ``53574``;
    - somente zeros técnicos à esquerda são removidos;
    - nenhum dígito do final do pedido pode ser cortado.
    """
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value)).lstrip("0") or "0"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value).lstrip("0") or "0"

    text = cell_text(value)
    if not text:
        return ""

    # Casos vindos do Excel/CSV como 053641, 53641.0 ou 53641,00.
    decimal_match = re.fullmatch(r"\s*0*([0-9]{3,12})(?:[\.,]0+)?\s*", text)
    if decimal_match:
        return decimal_match.group(1).lstrip("0") or "0"

    grupos = re.findall(r"\d+", text)
    if not grupos:
        return ""

    candidatos: List[str] = []
    for grupo in grupos:
        normalizado = grupo.lstrip("0") or "0"
        # Pedido Coelho Diniz normalmente fica com 5 dígitos, mas deixamos
        # margem para crescimento sem truncar códigos maiores.
        if 3 <= len(normalizado) <= 12:
            candidatos.append(normalizado)

    if not candidatos:
        return ""

    for candidato in candidatos:
        if 5 <= len(candidato) <= 8:
            return candidato
    return candidatos[0]


def extract_pedido_from_text(text: str, allow_loose: bool = False) -> str:
    texto = str(text or "")
    patterns = [
        # Pedido_053641_SPALINDUSTRIA / Pedido 053574 SPALINDUSTRIA
        r"\bpedido\s*[_\s\-]*0*([0-9]{5,8})(?=[^0-9]|$)",
        r"\bped\s*[_\s\-]*0*([0-9]{5,8})(?=[^0-9]|$)",
        # Pedido nº, Pedido no, pedido:, número do pedido etc.
        r"\bpedido\s*(?:n[oº\.]?\s*)?[:\-_/ ]*0*([0-9]{5,8})(?=[^0-9]|$)",
        r"\bnum(?:ero)?\s*(?:do\s*)?pedido\s*[:\-_/ ]*0*([0-9]{5,8})(?=[^0-9]|$)",
        r"\bn[uú]mero\s*(?:do\s*)?pedido\s*[:\-_/ ]*0*([0-9]{5,8})(?=[^0-9]|$)",
        r"\bp\s*[:\-_/ ]*0*([0-9]{5,8})(?=[^0-9]|$)",
    ]
    for pattern in patterns:
        match = re.search(pattern, texto, flags=re.IGNORECASE)
        if match:
            return normalizar_pedido_coelho_diniz(match.group(1))

    if allow_loose:
        # Uso restrito a nomes de arquivo/metadados de anexo. Não usar em corpo de
        # e-mail solto para não confundir datas como 25052026 com pedido.
        for match in re.finditer(r"(?<!\d)0*([0-9]{5,8})(?!\d)", texto):
            return normalizar_pedido_coelho_diniz(match.group(1))
    return ""


def extract_pedido_from_filename(path: Path) -> str:
    stem = path.stem or ""
    return extract_pedido_from_text(stem, allow_loose=True)


def _metadata_outlook_candidates(path: Path) -> List[Path]:
    candidatos = [
        Path(str(path) + ".outlook.json"),
        path.with_suffix(path.suffix + ".outlook.json"),
    ]

    pasta_compacta = path.parent / "_metadata_outlook"
    if pasta_compacta.exists():
        try:
            # Pasta pequena por lote; limitado para evitar varredura pesada.
            candidatos.extend(sorted(pasta_compacta.glob("*.json"))[:250])
        except Exception:
            pass

    vistos = set()
    unicos: List[Path] = []
    for candidato in candidatos:
        chave = str(candidato.resolve()) if candidato.exists() else str(candidato)
        if chave not in vistos:
            vistos.add(chave)
            unicos.append(candidato)
    return unicos


def extract_pedido_from_outlook_metadata(path: Path) -> str:
    """Recupera pedido pelo metadata salvo na importação Outlook.

    Isso protege o fluxo quando o Outlook salva o anexo com nome curto por caminho
    longo, por exemplo ``email001_anexo001_abcd.xlsx``. Nesses casos, o número do
    pedido não está mais no nome salvo, mas permanece em ``nome_original`` no JSON.
    """
    alvo_resolvido = str(path.resolve()) if path.exists() else str(path)
    alvo_nome = path.name.lower()

    for meta_path in _metadata_outlook_candidates(path):
        if not meta_path.exists():
            continue
        try:
            payload = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            continue

        arquivo_salvo = str(payload.get("arquivo_salvo", "") or "")
        nome_salvo = Path(arquivo_salvo).name.lower() if arquivo_salvo else ""
        # Para a pasta compacta, garante que o metadata pertence ao anexo atual.
        if meta_path.parent.name == "_metadata_outlook":
            if arquivo_salvo and str(Path(arquivo_salvo).resolve()) != alvo_resolvido and nome_salvo != alvo_nome:
                continue

        campos_nome = [
            payload.get("nome_original", ""),
            payload.get("arquivo_original", ""),
            payload.get("attachment_name", ""),
            payload.get("arquivo_salvo", ""),
        ]
        for valor in campos_nome:
            if not valor:
                continue
            pedido = extract_pedido_from_text(Path(str(valor)).stem, allow_loose=True)
            if pedido:
                return pedido

        campos_texto = [
            payload.get("assunto", ""),
            payload.get("email_assunto", ""),
            payload.get("contexto", ""),
            payload.get("corpo_email", ""),
            payload.get("body", ""),
        ]
        for valor in campos_texto:
            pedido = extract_pedido_from_text(str(valor), allow_loose=False)
            if pedido:
                return pedido
    return ""


def extract_pedido_from_sources(path: Path) -> str:
    """Ordem de confiança para pedido Coelho Diniz: nome salvo > metadata Outlook."""
    pedido = extract_pedido_from_filename(path)
    if pedido:
        return pedido
    return extract_pedido_from_outlook_metadata(path)

def find_column_by_alias(values: Sequence[Any], aliases: Sequence[str]) -> Optional[int]:
    aliases_norm = [normalize_text(alias) for alias in aliases]
    for idx, value in enumerate(values):
        norm = normalize_text(value)
        if not norm:
            continue
        for alias in aliases_norm:
            if alias in norm:
                return idx
    return None


def detect_header_row_excel(ws) -> Optional[int]:
    for row_idx in range(1, min(ws.max_row, 40) + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        joined = " | ".join(normalize_text(v) for v in values)
        has_desc = "descricao" in joined
        has_sku = "cod.fab" in joined or "cod fab" in joined or "codfab" in joined or "sku" in joined
        has_divisor = "divisor" in joined
        if has_desc and has_sku and has_divisor:
            return row_idx
    return None


def detect_date_layout_header_excel(ws) -> Optional[Tuple[int, int, int, Optional[int]]]:
    for row_idx in range(1, min(ws.max_row, 40) + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        matricula_col = None
        data_col = None
        atendimento_col = None
        for idx, value in enumerate(values):
            norm = normalize_text(value)
            compact = norm.replace(" ", "")
            if matricula_col is None and "matricula" in compact:
                matricula_col = idx
            if data_col is None and "data" in norm and ("entrega" in norm or "digit" in norm or norm == "data"):
                data_col = idx
            if atendimento_col is None and "atendimento" in norm:
                atendimento_col = idx
        if matricula_col is not None and data_col is not None:
            return row_idx, matricula_col, data_col, atendimento_col
    return None


def is_stop_after_store_header(value: Any) -> bool:
    norm = normalize_text(value).replace(" ", "")
    stops = {normalize_text(item).replace(" ", "") for item in STOP_HEADERS}
    return norm in stops or norm.startswith("total")


def is_store_identifier(value: Any) -> bool:
    digits = only_digits(value)
    norm = normalize_text(value)
    if not digits or norm in {"total", "verba", "valor", "preco"}:
        return False
    return 8 <= len(digits) <= 14


def detect_store_columns(header_values: Sequence[Any], store_name_values: Sequence[Any], start_idx: int) -> Tuple[List[int], List[str]]:
    store_cols: List[int] = []
    ignored_headers: List[str] = []
    for idx in range(start_idx, len(header_values)):
        header_value = header_values[idx]
        name_value = store_name_values[idx] if idx < len(store_name_values) else ""
        if is_stop_after_store_header(header_value) or is_stop_after_store_header(name_value):
            ignored_headers.append(cell_text(header_value) or cell_text(name_value))
            break
        if is_store_identifier(header_value) or is_store_identifier(name_value):
            store_cols.append(idx)
            continue
        if cell_text(header_value) or cell_text(name_value):
            ignored_headers.append(cell_text(header_value) or cell_text(name_value))
    return store_cols, ignored_headers


def load_de_para() -> Tuple[Dict[str, str], Dict[str, str], List[CoelhoLog]]:
    logs: List[CoelhoLog] = []
    cnpj_by_matricula: Dict[str, str] = {}
    matricula_by_cnpj: Dict[str, str] = {}
    for matricula_raw, cnpj_raw in DE_PARA_MATRICULA_CNPJ.items():
        matricula = only_digits(matricula_raw)
        cnpj = only_digits(cnpj_raw)
        if matricula and cnpj:
            cnpj_by_matricula[matricula] = cnpj
            matricula_by_cnpj[cnpj] = matricula
    logs.append(CoelhoLog(now_str(), "OK", "DE_PARA_INTERNO", "Base", f"De/para interno carregado com {len(cnpj_by_matricula)} vinculos."))
    return cnpj_by_matricula, matricula_by_cnpj, logs


def resolve_store(raw_code: Any, cnpj_by_matricula: Dict[str, str], matricula_by_cnpj: Dict[str, str]) -> Tuple[str, str]:
    digits = only_digits(raw_code)
    if not digits:
        return "", ""
    if len(digits) in {13, 14}:
        cnpj = digits.zfill(14)
        return cnpj, matricula_by_cnpj.get(cnpj, "")
    matricula = digits
    return cnpj_by_matricula.get(matricula, ""), matricula


def extract_date_map_from_excel(path: Path) -> Tuple[Dict[str, str], Dict[str, str], bool, bool, List[CoelhoLog]]:
    logs: List[CoelhoLog] = []
    data_by_matricula: Dict[str, str] = {}
    atendimento_by_matricula: Dict[str, str] = {}
    found_date_layout = False
    has_order_layout = False
    try:
        wb = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        logs.append(CoelhoLog(now_str(), "ERRO", path.name, "Abertura", "Nao foi possivel abrir o Excel.", repr(exc)))
        return data_by_matricula, atendimento_by_matricula, False, False, logs

    try:
        for ws in wb.worksheets:
            if detect_header_row_excel(ws):
                has_order_layout = True
            header = detect_date_layout_header_excel(ws)
            if not header:
                continue
            found_date_layout = True
            header_row, matricula_col, data_col, atendimento_col = header
            linhas = 0
            for row_idx in range(header_row + 1, ws.max_row + 1):
                matricula = only_digits(ws.cell(row_idx, matricula_col + 1).value)
                data = normalize_date_value(ws.cell(row_idx, data_col + 1).value)
                atendimento = cell_text(ws.cell(row_idx, atendimento_col + 1).value) if atendimento_col is not None else ""
                if not matricula and not data:
                    continue
                if not matricula:
                    logs.append(CoelhoLog(now_str(), "AVISO", path.name, ws.title, "Linha de data ignorada sem matricula.", f"Linha {row_idx}"))
                    continue
                if not data:
                    logs.append(CoelhoLog(now_str(), "AVISO", path.name, ws.title, "Matricula sem data no arquivo auxiliar.", f"Matricula {matricula}"))
                    continue
                data_by_matricula[matricula] = data
                if atendimento:
                    atendimento_by_matricula[matricula] = atendimento
                linhas += 1
            logs.append(CoelhoLog(now_str(), "OK", path.name, ws.title, f"Layout de datas identificado com {linhas} matriculas."))
    finally:
        wb.close()
    return data_by_matricula, atendimento_by_matricula, found_date_layout, has_order_layout, logs


def file_has_order_layout(path: Path) -> bool:
    try:
        wb = load_workbook(path, data_only=True, read_only=False)
    except Exception:
        return False
    try:
        return any(detect_header_row_excel(ws) for ws in wb.worksheets)
    finally:
        wb.close()


def parse_order_excel(
    path: Path,
    cnpj_by_matricula: Dict[str, str],
    matricula_by_cnpj: Dict[str, str],
    data_by_matricula: Dict[str, str],
    atendimento_by_matricula: Dict[str, str],
) -> Tuple[List[CoelhoRow], List[CoelhoLog]]:
    rows: List[CoelhoRow] = []
    logs: List[CoelhoLog] = []
    pedido_global = extract_pedido_from_sources(path)
    try:
        wb = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        logs.append(CoelhoLog(now_str(), "ERRO", path.name, "Abertura", "Nao foi possivel abrir o Excel.", repr(exc)))
        return rows, logs

    try:
        if not pedido_global:
            for ws_scan in wb.worksheets[:2]:
                for row in ws_scan.iter_rows(min_row=1, max_row=min(20, ws_scan.max_row), values_only=True):
                    pedido_global = extract_pedido_from_text(" ".join(cell_text(v) for v in row if v is not None))
                    if pedido_global:
                        break
                if pedido_global:
                    break
        pedido_global = normalizar_pedido_coelho_diniz(pedido_global)
        if not pedido_global:
            logs.append(CoelhoLog(now_str(), "AVISO", path.name, "Pedido", "Numero do pedido nao localizado."))
        else:
            logs.append(CoelhoLog(now_str(), "OK", path.name, "Pedido", "Numero do pedido Coelho Diniz normalizado sem truncar digitos.", pedido_global))

        for ws in wb.worksheets:
            header_row = detect_header_row_excel(ws)
            if not header_row:
                continue
            header_values = [ws.cell(header_row, col).value for col in range(1, ws.max_column + 1)]
            store_name_values = [ws.cell(max(header_row - 1, 1), col).value for col in range(1, ws.max_column + 1)]
            ean_col = find_column_by_alias(header_values, ["codigo ean", "ean"])
            desc_col = find_column_by_alias(header_values, ["descricao"])
            sku_col = find_column_by_alias(header_values, ["cod.fab", "cod fab", "codfab", "sku"])
            divisor_col = find_column_by_alias(header_values, ["divisor"])
            if sku_col is None or divisor_col is None:
                logs.append(CoelhoLog(now_str(), "ERRO", path.name, ws.title, "Colunas obrigatorias nao localizadas.", "Necessario localizar Cod.Fab/SKU e Divisor."))
                continue
            store_cols, ignored = detect_store_columns(header_values, store_name_values, divisor_col + 1)
            if not store_cols:
                logs.append(CoelhoLog(now_str(), "ERRO", path.name, ws.title, "Nenhuma coluna de loja identificada apos Divisor."))
                continue
            if ignored:
                logs.append(CoelhoLog(now_str(), "OK", path.name, ws.title, "Colunas de fechamento ignoradas.", ", ".join(ignored[:8])))

            rows_before = len(rows)
            missing_dates = set()
            for row_idx in range(header_row + 1, ws.max_row + 1):
                sku_value = ws.cell(row_idx, sku_col + 1).value
                sku = only_digits(sku_value)
                if not sku:
                    continue
                ean = ws.cell(row_idx, ean_col + 1).value if ean_col is not None else ""
                desc = ws.cell(row_idx, desc_col + 1).value if desc_col is not None else ""
                divisor = ws.cell(row_idx, divisor_col + 1).value if divisor_col is not None else ""
                for store_idx in store_cols:
                    qtd = parse_quantity(ws.cell(row_idx, store_idx + 1).value)
                    if not qtd:
                        continue
                    loja_codigo = header_values[store_idx]
                    loja_nome = store_name_values[store_idx] if store_idx < len(store_name_values) else ""
                    cnpj, matricula = resolve_store(loja_codigo, cnpj_by_matricula, matricula_by_cnpj)
                    data = data_by_matricula.get(matricula, "")
                    if matricula and not data:
                        missing_dates.add(matricula)
                    rows.append(CoelhoRow(
                        cnpj=cnpj,
                        matricula=matricula,
                        sku=sku,
                        qtd=qtd,
                        numero_pedido=pedido_global,
                        data_digitacao=data,
                        atendimento=atendimento_by_matricula.get(matricula, ""),
                        arquivo_origem=path.name,
                        aba=ws.title,
                        linha_origem=str(row_idx),
                        loja=cell_text(loja_nome),
                        codigo_loja_original=cell_text(loja_codigo),
                        ean=only_digits(ean) or cell_text(ean),
                        descricao=cell_text(desc),
                        divisor=cell_text(divisor),
                    ))
            if missing_dates:
                sample = ", ".join(sorted(missing_dates)[:15])
                logs.append(CoelhoLog(now_str(), "AVISO", path.name, ws.title, "Data nao localizada para algumas matriculas.", sample))
            logs.append(CoelhoLog(now_str(), "OK", path.name, ws.title, "Aba processada.", f"Linhas geradas: {len(rows) - rows_before}."))
    finally:
        wb.close()
    return rows, logs


def rows_to_intermediate(rows: List[CoelhoRow], nome_layout: str) -> pd.DataFrame:
    data = []
    for row in rows:
        data.append({
            "matricula_lida": row.matricula,
            "cnpj_lido": row.cnpj,
            "sku_lido": row.sku,
            "codigo_sku_lido": row.sku,
            "ean_lido": row.ean,
            "descricao_lida": row.descricao,
            "quantidade_lida": row.qtd,
            "numero_pedido_lido": normalizar_pedido_coelho_diniz(row.numero_pedido),
            "data_entrega_lida": row.data_digitacao,
            "arquivo_origem": row.arquivo_origem,
            "layout_usado": nome_layout,
            "linha_origem": row.linha_origem,
            "origem_extracao": row.aba,
        })
    df = pd.DataFrame(data)
    return normalize_intermediate_columns(df, layout_usado=nome_layout)


def processar_lote_coelho_diniz(caminhos: Sequence[str | Path], layout_config: Dict[str, str]) -> Dict[str, Any]:
    files = [Path(p) for p in caminhos]
    logs: List[CoelhoLog] = []
    rows: List[CoelhoRow] = []
    alertas: List[str] = []

    cnpj_by_matricula, matricula_by_cnpj, depara_logs = load_de_para()
    logs.extend(depara_logs)

    data_by_matricula: Dict[str, str] = {}
    atendimento_by_matricula: Dict[str, str] = {}
    date_files: List[Path] = []
    order_files: List[Path] = []

    for file_path in files:
        if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            logs.append(CoelhoLog(now_str(), "ERRO", file_path.name, "Entrada", "Formato nao suportado para Coelho Diniz.", "Use .xlsx ou .xlsm."))
            continue
        date_map, atendimento_map, found_date_layout, has_order_layout, date_logs = extract_date_map_from_excel(file_path)
        logs.extend(date_logs)
        if found_date_layout:
            data_by_matricula.update(date_map)
            atendimento_by_matricula.update(atendimento_map)
            date_files.append(file_path)
        if has_order_layout:
            order_files.append(file_path)

    if not order_files or not date_files:
        logs.append(CoelhoLog(now_str(), "ERRO", "PROCESSO", "Entrada", MISSING_PAIR_MESSAGE))
        return {
            "sucesso": False,
            "mensagem": MISSING_PAIR_MESSAGE,
            "df_intermediario": None,
            "qtd_linhas_lidas": 0,
            "alertas": [MISSING_PAIR_MESSAGE],
            "logs": [asdict(log) for log in logs],
            "rows_detail": [],
            "order_files": [str(p) for p in order_files],
            "date_files": [str(p) for p in date_files],
        }

    for order_file in order_files:
        parsed_rows, file_logs = parse_order_excel(
            order_file,
            cnpj_by_matricula,
            matricula_by_cnpj,
            data_by_matricula,
            atendimento_by_matricula,
        )
        rows.extend(parsed_rows)
        logs.extend(file_logs)

    if not rows:
        alertas.append("Coelho Diniz: nenhum item foi extraido do Excel de pedido.")
    if any(not row.data_digitacao for row in rows):
        alertas.append("Coelho Diniz: ha linhas sem data vinculada pelo arquivo de matriculas/datas.")
    if any(not row.numero_pedido for row in rows):
        alertas.append("Coelho Diniz: ha linhas sem numero de pedido identificado.")

    df_intermediario = rows_to_intermediate(rows, layout_config.get("nome_layout", "COELHO DINIZ"))
    return {
        "sucesso": not df_intermediario.empty,
        "mensagem": (
            f"Leitura Coelho Diniz concluida com {len(df_intermediario)} linha(s)"
            if not df_intermediario.empty
            else "Nenhuma linha valida foi extraida da Rede Coelho Diniz"
        ),
        "df_intermediario": df_intermediario if not df_intermediario.empty else None,
        "qtd_linhas_lidas": len(df_intermediario),
        "alertas": sorted(set(alertas)),
        "logs": [asdict(log) for log in logs],
        "rows_detail": [asdict(row) for row in rows],
        "order_files": [str(p) for p in order_files],
        "date_files": [str(p) for p in date_files],
    }


def gerar_excel_validacao_coelho(
    out_path: Path,
    leitura: Dict[str, Any],
    df_final: pd.DataFrame | None = None,
    df_descartadas: pd.DataFrame | None = None,
):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df_lidas = leitura.get("df_intermediario")
    if df_lidas is None:
        df_lidas = pd.DataFrame(columns=INTERMEDIATE_COLUMNS)
    df_detalhe = pd.DataFrame(leitura.get("rows_detail", []))
    df_logs = pd.DataFrame(leitura.get("logs", []))
    df_final = df_final if df_final is not None else pd.DataFrame()
    df_descartadas = df_descartadas if df_descartadas is not None else pd.DataFrame()
    resumo = pd.DataFrame([
        {"indicador": "linhas_lidas", "valor": len(df_lidas)},
        {"indicador": "linhas_validas_fila", "valor": len(df_final)},
        {"indicador": "linhas_pendentes", "valor": len(df_descartadas)},
        {"indicador": "arquivos_pedido", "valor": "; ".join(Path(p).name for p in leitura.get("order_files", []))},
        {"indicador": "arquivos_datas", "valor": "; ".join(Path(p).name for p in leitura.get("date_files", []))},
        {"indicador": "alertas", "valor": " | ".join(leitura.get("alertas", []))},
        {"indicador": "gerado_em", "valor": now_str()},
    ])
    resumo_dict = {str(row["indicador"]): row["valor"] for _, row in resumo.iterrows()}
    resumo_dict["layout"] = "COELHO DINIZ"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        criar_abas_validacao_padrao(
            writer,
            resumo=resumo_dict,
            df_intermediario=df_lidas,
            df_final=df_final,
            df_descartadas=df_descartadas,
            alertas=leitura.get("alertas", []),
        )
        df_lidas.to_excel(writer, index=False, sheet_name="LINHAS_LIDAS")
        df_final.to_excel(writer, index=False, sheet_name="FILA_KOF_PREVIA")
        df_descartadas.to_excel(writer, index=False, sheet_name="PENDENCIAS")
        df_detalhe.to_excel(writer, index=False, sheet_name="DETALHADO")
        df_logs.to_excel(writer, index=False, sheet_name="LOGS")
        resumo.to_excel(writer, index=False, sheet_name="RESUMO")
    aplicar_estilo_validacao(out_path)
    wb = load_workbook(out_path, read_only=True, data_only=True)
    wb.close()


def ler_excel_coelho_diniz(caminho_arquivo: str, layout_config: Dict[str, str], mapeamentos_df=None):
    return {
        "sucesso": False,
        "mensagem": MISSING_PAIR_MESSAGE,
        "df_intermediario": None,
        "qtd_linhas_lidas": 0,
        "alertas": [MISSING_PAIR_MESSAGE],
    }
