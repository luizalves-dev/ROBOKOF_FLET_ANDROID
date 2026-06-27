from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


@dataclass
class MappingConflict:
    cnpj: str
    chosen_matricula: str
    discarded_matriculas: list[str] = field(default_factory=list)


@dataclass
class MappingLoadResult:
    mapping: dict[str, str]
    source_path: Path | None
    row_count: int
    valid_count: int
    conflicts: list[MappingConflict] = field(default_factory=list)
    detected_columns: dict[str, int] = field(default_factory=dict)


class MappingLoaderError(Exception):
    pass


class CNPJMapping:
    def __init__(self, mapping: Dict[str, str], conflicts: Optional[List[MappingConflict]] = None):
        self.mapping = {self.normalize_cnpj(k): str(v).strip() for k, v in mapping.items() if self.normalize_cnpj(k)}
        self.conflicts = conflicts or []

    @staticmethod
    def normalize_cnpj(value) -> str:
        digits = re.sub(r"\D", "", str(value or ""))
        if not digits:
            return ""
        return digits.zfill(14) if len(digits) <= 14 else digits

    @staticmethod
    def normalize_matricula(value) -> str:
        return re.sub(r"\D", "", str(value or ""))

    def get_matricula(self, cnpj) -> str:
        return self.mapping.get(self.normalize_cnpj(cnpj), "")


def load_mapping_file(file_path: str | Path | None) -> MappingLoadResult:
    if not file_path:
        raise MappingLoaderError("Caminho da base BH nao informado.")

    path = Path(file_path)
    if not path.exists():
        raise MappingLoaderError(f"Base BH CNPJ x matricula nao encontrada em: {path}")

    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        rows = _read_excel_rows(path)
    elif path.suffix.lower() in {".csv", ".txt"}:
        rows = _read_delimited_rows(path)
    else:
        raise MappingLoaderError("Formato da base BH nao suportado. Use .xlsx, .xlsm, .csv ou .txt.")

    if not rows:
        raise MappingLoaderError("A base BH CNPJ x matricula esta vazia.")

    cnpj_idx, detected = _detect_columns(rows)
    matricula_idx = detected.get("matricula")
    if cnpj_idx is None or matricula_idx is None:
        raise MappingLoaderError("Nao foi possivel detectar as colunas CNPJ e Matricula na base BH.")

    entries: list[dict[str, str]] = []
    row_count = 0
    for row in rows[1:]:
        row_count += 1
        cnpj = CNPJMapping.normalize_cnpj(_safe_get(row, cnpj_idx))
        matricula = CNPJMapping.normalize_matricula(_safe_get(row, matricula_idx))
        rede = _safe_get(row, detected.get("rede", -1))
        if cnpj and matricula:
            entries.append({"cnpj": cnpj, "matricula": matricula, "rede": rede})

    mapping, conflicts = _resolve_entries(entries)
    return MappingLoadResult(
        mapping=mapping,
        source_path=path,
        row_count=row_count,
        valid_count=len(mapping),
        conflicts=conflicts,
        detected_columns=detected,
    )


def _read_excel_rows(file_path: Path) -> List[List[str]]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return [[_cell_to_str(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _read_delimited_rows(file_path: Path) -> List[List[str]]:
    text = file_path.read_text(encoding="utf-8-sig", errors="replace")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters="\t;,")
        delimiter = dialect.delimiter
    except Exception:
        delimiter = "\t" if "\t" in sample else ";"
    return [[cell.strip() for cell in row] for row in csv.reader(text.splitlines(), delimiter=delimiter)]


def _detect_columns(rows: List[List[str]]) -> Tuple[Optional[int], Dict[str, int]]:
    header_index = 0
    detected: dict[str, int] = {}
    for idx, row in enumerate(rows[:10]):
        normalized = [_normalize_header(cell) for cell in row]
        has_cnpj = any("cnpj" in value for value in normalized)
        has_mat = any("matricula" in value or "matr" in value for value in normalized)
        if has_cnpj and has_mat:
            header_index = idx
            break

    header = [_normalize_header(cell) for cell in rows[header_index]]
    for index, value in enumerate(header):
        if "cnpj" in value and "cnpj" not in detected:
            detected["cnpj"] = index
        elif ("matricula" in value or "matr" in value) and "matricula" not in detected:
            detected["matricula"] = index
        elif "rede" in value and "rede" not in detected:
            detected["rede"] = index

    if header_index > 0:
        del rows[:header_index]

    return detected.get("cnpj"), detected


def _resolve_entries(entries: Iterable[dict[str, str]]) -> tuple[dict[str, str], list[MappingConflict]]:
    grouped: dict[str, list[dict[str, str]]] = {}
    for entry in entries:
        grouped.setdefault(entry["cnpj"], []).append(entry)

    mapping: dict[str, str] = {}
    conflicts: list[MappingConflict] = []
    for cnpj, items in grouped.items():
        unique_mats = sorted({item["matricula"] for item in items if item["matricula"]})
        if len(unique_mats) <= 1:
            mapping[cnpj] = unique_mats[0] if unique_mats else ""
            continue

        chosen = _choose_preferred_entry(items)["matricula"]
        discarded = [mat for mat in unique_mats if mat != chosen]
        mapping[cnpj] = chosen
        conflicts.append(MappingConflict(cnpj=cnpj, chosen_matricula=chosen, discarded_matriculas=discarded))

    return mapping, conflicts


def _choose_preferred_entry(entries: list[dict[str, str]]) -> dict[str, str]:
    """Escolhe a matrícula mais segura em CNPJ duplicado.

    Mesma lógica do bot BH antigo: quando a mesma chave aparece para mais de
    uma razão social, prioriza registros com BH/SUPERMERCADOS e usa a ordem da
    base como desempate estável. Isso evita escolher linhas genéricas apenas
    porque contêm a palavra COMERCIO.
    """
    if not entries:
        return {"cnpj": "", "matricula": "", "rede": ""}

    scored: list[tuple[float, dict[str, str]]] = []
    for index, entry in enumerate(entries):
        rede = str(entry.get("rede", "")).upper()
        score = 0.0
        if "BH" in rede:
            score += 10
        if "SUPERM" in rede or "SUPER BH" in rede:
            score += 3
        if "COMERC" in rede or "ALIM" in rede:
            score += 1
        score += index / 1000
        scored.append((score, entry))

    scored.sort(key=lambda item: (item[0], item[1].get("matricula", "")), reverse=True)
    return scored[0][1]


def _safe_get(row: list[str], index: int | None) -> str:
    if index is None or index < 0 or index >= len(row):
        return ""
    return _cell_to_str(row[index])


def _cell_to_str(value) -> str:
    return "" if value is None else str(value).strip()


def _normalize_header(value) -> str:
    text = unicodedata.normalize("NFKD", _cell_to_str(value).lower())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip()
