from __future__ import annotations

import logging
import json
import hashlib
import re
import shutil
import sys
import unicodedata
import zipfile
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
import gln_service

from .cnpj_mapping import CNPJMapping, MappingLoadResult, load_mapping_file


BH_INVALID_LAYOUT_MESSAGE = "Pedido fora do layout. Por gentileza validar."

ORDER_PATTERNS = [
    # Layout BH legado: P E D I D O   D E   C O M P R A: 123456
    re.compile(r"P\s*E\s*D\s*I\s*D\s*O\s*D\s*E\s*C\s*O\s*M\s*P\s*R\s*A\s*:\s*(\d+)", re.IGNORECASE),
    # Layout RP One / Enterprise usado em alguns pedidos SPAL/BH.
    re.compile(r"N[uú]mero\s+do\s+Pedido\s*:\s*(\d+)", re.IGNORECASE),
    re.compile(r"N[úu]mero\s+Pedido\s*:\s*(\d+)", re.IGNORECASE),
    re.compile(r"N[ºo°]?\s*Pedido\s*:\s*(\d+)", re.IGNORECASE),
    re.compile(r"Numero\s+Pedido\s*:\s*(\d+)", re.IGNORECASE),
    re.compile(r"Pedido\s*[:\-]\s*(\d{4,})", re.IGNORECASE),
]
CNPJ_PATTERN = re.compile(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}|\d{13,14}")
SUPPLIER_CNPJ_PATTERN = re.compile(r"C\.N\.P\.J\.----:\s*([\d./-]+)", re.IGNORECASE)
DELIVERY_CNPJ_LABEL_PATTERNS = [
    re.compile(r"Local\s+de\s+Entrega.*?CNPJ\s*[:\-]*\s*(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})", re.IGNORECASE | re.S),
    re.compile(r"CNPJ\s*[:\-]*\s*(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})", re.IGNORECASE),
]
IE_LABEL_PATTERN = re.compile(r"(?:I\.?E\.?|Inscri[cç][aã]o)\s*[:\-]*\s*(\d{8,14})", re.IGNORECASE)
LEGACY_ITEM_UNIT_FIRST_PATTERN = re.compile(
    r"^\s*(\d{5,6})\s+(.+?)\s+((?:CX|FD)\s+C/\s*\d+|(?:CX|FD)/?\d+)\s+"
    r"(\d+(?:\.\d{3})*,\d{2,4})\s+(\d+(?:\.\d{3})*,\d{3})\s+"
    r"(\d+(?:\.\d{3})*,\d{2})\s*$",
    re.IGNORECASE,
)
ITEM_PATTERN = re.compile(
    r"^\s*(\d{5,6})\s+(.+?)\s+(\d+(?:\.\d{3})*,\d{4})\s+"
    r"(\d+(?:\.\d{3})*,\d{3})\s*((?:CX|FD)\s+C/\s*\d+)\s+(\d+(?:\.\d{3})*,\d{2})\s*$",
    re.IGNORECASE,
)
STOP_MARKERS = (
    "valor da mercadoria",
    "total pedido",
    "local de entrega",
    "observacoes",
    "assinatura fornecedor",
)


@dataclass
class ParsedItem:
    arquivo: str
    numero_pedido: str
    cnpj: str
    matricula: str
    sku: str
    descricao: str
    qtd: int
    unidade: str
    pagina: int
    loja_numero: str = ""
    loja_nome: str = ""


@dataclass
class ParseIssue:
    arquivo: str
    tipo: str
    detalhe: str
    pagina: Optional[int] = None
    linha: Optional[str] = None


@dataclass
class ParseResult:
    pdf_name: str
    layout_valid: bool = True
    numero_pedido: str = ""
    cnpj: str = ""
    matricula: str = ""
    loja_numero: str = ""
    loja_nome: str = ""
    items: List[ParsedItem] = field(default_factory=list)
    issues: List[ParseIssue] = field(default_factory=list)
    output_file: Optional[Path] = None
    pedido_dir: Optional[Path] = None
    source_pdf: Optional[Path] = None
    duration_seconds: float = 0.0


@dataclass
class BatchProcessResult:
    date_dir: Path
    pedidos_dir: Path
    consolidado_file: Path
    log_file: Path
    executive_log_file: Path
    zip_file: Path
    total_pdfs: int
    total_items: int
    total_errors: int
    pedidos_com_sucesso: int
    mapping_summary: str
    lote_data: str
    duplicate_orders: List[str] = field(default_factory=list)
    duplicate_details: List[str] = field(default_factory=list)
    individual_files: List[Path] = field(default_factory=list)
    invalid_layout_files: List[Path] = field(default_factory=list)
    duplicates_dir: Optional[Path] = None
    invalid_layout_dir: Optional[Path] = None
    main_pushes_file: Optional[Path] = None
    push_number: int = 0
    push_label: str = ""


@dataclass
class HistoricalOrderOccurrence:
    pedido: str
    file_path: Path
    push_label: str = ""
    processed_at: str = ""


class BHProcessor:
    def __init__(
        self,
        mapping: CNPJMapping,
        logger: logging.Logger,
        gln_cnpj_map: dict[str, str] | None = None,
        use_gln_fallback: bool = True,
    ):
        self.mapping = mapping
        self.logger = logger
        self.gln_cnpj_map = gln_cnpj_map or {}
        self.use_gln_fallback = use_gln_fallback

    def parse_pdf(self, pdf_path: str | Path) -> ParseResult:
        """Lê um PDF da Rede BH com tolerância operacional.

        A automação antiga da BH não derrubava o lote na primeira divergência
        de cabeçalho: ela tentava extrair pedido, CNPJ/matrícula e itens, e só
        depois classificava como inválido se nada útil fosse encontrado. O Robô
        KOF agora replica esse comportamento para não rejeitar PDFs reais apenas
        por variação de texto/extração do Outlook.
        """
        pdf_path = Path(pdf_path)
        result = ParseResult(pdf_name=pdf_path.name, source_pdf=pdf_path)
        self.logger.info("=" * 110)
        self.logger.info("[BH][PDF] Iniciando leitura: %s", pdf_path.name)

        try:
            page_texts = extract_pdf_text_pages(pdf_path)
        except Exception as exc:
            result.layout_valid = False
            result.issues.append(ParseIssue(pdf_path.name, "ERRO_PDF", f"Falha ao abrir PDF: {exc}"))
            self.logger.exception("[BH][PDF] Falha ao abrir %s", pdf_path.name)
            return result

        full_text = "\n".join(page_texts)
        result.loja_numero, result.loja_nome = extract_store_identity(page_texts, pdf_path)
        if result.loja_numero or result.loja_nome:
            self.logger.info(
                "[BH][PDF] Loja identificada: numero=%s | nome=%s",
                result.loja_numero or "A identificar",
                result.loja_nome or "A identificar",
            )
        confidence = bh_layout_confidence(full_text, self.mapping)
        mapped_layout = self._looks_like_mapped_bh_layout(full_text)
        self.logger.info(
            "[BH][PDF] Confiança layout BH=%s | mapeado=%s | arquivo=%s",
            confidence,
            "SIM" if mapped_layout else "NAO",
            pdf_path.name,
        )

        # IMPORTANTE: não bloquear antes de tentar ler. Muitos PDFs BH vindos do
        # Outlook mudam a forma como o texto é extraído, mas ainda possuem linhas
        # de item válidas. Layout inválido só é decidido ao final.
        result.numero_pedido = extract_order_number(full_text)
        if result.numero_pedido:
            self.logger.info("[BH][PDF] Pedido identificado: %s", result.numero_pedido)
        else:
            result.issues.append(ParseIssue(pdf_path.name, "PEDIDO_NAO_ENCONTRADO", "Numero do pedido nao encontrado."))
            self.logger.warning("[BH][PDF] Numero do pedido nao encontrado em %s", pdf_path.name)

        result.cnpj = extract_delivery_cnpj(page_texts, self.mapping)
        if result.cnpj:
            self.logger.info("[BH][PDF] CNPJ considerado: %s", result.cnpj)
            result.matricula = self._resolve_matricula(result.cnpj, result)
        else:
            result.issues.append(ParseIssue(pdf_path.name, "CNPJ_NAO_ENCONTRADO", "CNPJ do local de entrega nao encontrado."))
            self.logger.warning("[BH][PDF] CNPJ do local de entrega nao encontrado em %s", pdf_path.name)

        for page_number, page_text in enumerate(page_texts, start=1):
            page_items, page_issues = parse_items_from_page(
                page_text=page_text,
                pdf_name=pdf_path.name,
                numero_pedido=result.numero_pedido,
                cnpj=result.cnpj,
                matricula=result.matricula,
                page_number=page_number,
                loja_numero=result.loja_numero,
                loja_nome=result.loja_nome,
            )
            result.items.extend(page_items)
            result.issues.extend(page_issues)
            self.logger.info(
                "[BH][PDF] Pagina %s -> %s item(ns) valido(s) / %s alerta(s)",
                page_number,
                len(page_items),
                len(page_issues),
            )

        if result.items:
            self.logger.info("[BH][PDF] Total de itens extraidos em %s: %s", pdf_path.name, len(result.items))
            return result

        # Só neste ponto classifica como fora do layout. Isso preserva o
        # comportamento do bot BH antigo de tentar ler tudo antes de bloquear,
        # mas impede que PDF sem item/pedido apareça como "Processado para
        # validação" na interface.
        result.layout_valid = False
        result.issues.append(ParseIssue(pdf_path.name, "SEM_ITENS", "Nenhum item foi identificado no PDF."))
        result.issues.append(ParseIssue(pdf_path.name, "PEDIDO_FORA_LAYOUT", BH_INVALID_LAYOUT_MESSAGE))
        self.logger.warning(
            "[BH][PDF] Pedido fora do layout apos tentativa completa de leitura: %s | confiança=%s | mapeado=%s",
            pdf_path.name,
            confidence,
            "SIM" if mapped_layout else "NAO",
        )
        return result

    def _looks_like_mapped_bh_layout(self, text: str) -> bool:
        """Reconhece BH por base/marcadores, sem depender de nome do arquivo.

        A rotina antiga BH aceitava os PDFs reais mesmo quando o assunto/nome do
        anexo era genérico. Por isso, aqui a regra principal é: se existe pedido
        e existe CNPJ que bate na base BH, o layout pode ser processado. O nome
        SPAL/COCA ajuda, mas não é obrigatório.
        """
        normalized = normalize_text(text)
        if not extract_order_number(text):
            return False

        bh_name_tokens = (
            "superm bh",
            "supermercados bh",
            "supemercados bh",
            "bh comercio de alimen",
            "bh com alimentos",
            "bh230",
        )
        if any(token in normalized for token in bh_name_tokens):
            self.logger.info("[BH][PDF] Layout reconhecido por razão social/token BH no texto.")
            return True

        supplier_match = SUPPLIER_CNPJ_PATTERN.search(text or "")
        supplier = normalize_cnpj(supplier_match.group(1)) if supplier_match else ""
        for raw_cnpj in CNPJ_PATTERN.findall(text or ""):
            cnpj = normalize_cnpj(raw_cnpj)
            if not cnpj or cnpj == supplier:
                continue
            if self.mapping.get_matricula(cnpj):
                self.logger.info("[BH][PDF] Layout reconhecido por CNPJ BH no de/para: %s", cnpj)
                return True

        # Último recurso: layout legado completo ou RP One com marcadores fortes.
        if is_bh_layout(text):
            return True
        if ("codigo cod barras" in normalized or "cod barras" in normalized) and ("cod forn" in normalized or "qtde" in normalized):
            return True
        return False

    def _resolve_matricula(self, cnpj: str, result: ParseResult) -> str:
        bh_matricula = self.mapping.get_matricula(cnpj)
        gln_matricula = self.gln_cnpj_map.get(CNPJMapping.normalize_cnpj(cnpj), "")

        if bh_matricula and gln_matricula and bh_matricula != gln_matricula:
            result.issues.append(
                ParseIssue(
                    result.pdf_name,
                    "DIVERGENCIA_BH_GLN",
                    f"Base BH={bh_matricula}; GLN oficial={gln_matricula}. Mantida matricula BH.",
                )
            )
            self.logger.warning("[BH][BASE] Divergencia CNPJ %s | BH=%s | GLN=%s", cnpj, bh_matricula, gln_matricula)

        if bh_matricula:
            return bh_matricula

        if self.use_gln_fallback and gln_matricula:
            result.issues.append(
                ParseIssue(
                    result.pdf_name,
                    "MATRICULA_VIA_GLN",
                    f"CNPJ {cnpj} nao localizado na base BH; usada base GLN oficial como fallback.",
                )
            )
            return gln_matricula

        result.issues.append(ParseIssue(result.pdf_name, "MATRICULA_NAO_ENCONTRADA", f"CNPJ {cnpj} sem matricula na base BH."))
        return ""


def process_batch(
    pdf_paths: List[str | Path],
    output_root: str | Path,
    mapping_path: str | Path | None = None,
    lote_date: str | None = None,
    use_gln_fallback: bool = False,
) -> BatchProcessResult:
    output_root = Path(output_root)
    output_root.mkdir(parents=True, exist_ok=True)

    lote_date_label = normalize_lote_date(lote_date)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    date_dir = output_root / lote_date_label
    pedidos_dir = date_dir / "01_PEDIDOS"
    consolidado_dir = date_dir / "02_CONSOLIDADO"
    log_dir = date_dir / "03_LOG"
    duplicates_dir = date_dir / "04_PEDIDOS_DUPLICADOS"
    invalid_layout_dir = date_dir / "05_LAYOUT_INVALIDO"
    fila_kof_dir = date_dir / "06_FILA_KOF"

    for folder in [pedidos_dir, consolidado_dir, log_dir, duplicates_dir, invalid_layout_dir, fila_kof_dir]:
        folder.mkdir(parents=True, exist_ok=True)

    log_file = log_dir / f"processamento_bh_{lote_date_label.replace('.', '_')}_{timestamp}.log"
    logger = build_logger(log_file)
    logger.info("[BH][LOTE] Data operacional do lote: %s", lote_date_label)
    logger.info("[BH][LOTE] Pasta da data: %s", date_dir)

    mapping_result = load_mapping_file(mapping_path or config.BH_BASE_PATH)
    bh_mapping = CNPJMapping(mapping_result.mapping, conflicts=mapping_result.conflicts)
    gln_cnpj_map = _load_gln_cnpj_map(logger)
    processor = BHProcessor(bh_mapping, logger, gln_cnpj_map=gln_cnpj_map, use_gln_fallback=use_gln_fallback)

    logger.info("[BH][BASE] BH linhas validas: %s", mapping_result.valid_count)
    logger.info("[BH][BASE] GLN apoio carregado: %s CNPJ(s)", len(gln_cnpj_map))

    main_pushes_file = consolidado_dir / f"BH_MAIN_PUSHES_{lote_date_label.replace('.', '_')}.xlsx"
    push_number = get_next_push_number(main_pushes_file)
    push_label = f"PUSH_{push_number:03d}"

    all_items: List[ParsedItem] = []
    all_issues: List[ParseIssue] = []
    parse_results: List[ParseResult] = []
    individual_files: List[Path] = []
    invalid_files: List[Path] = []
    success_count = 0
    duplicate_orders: set[str] = set()
    duplicate_details: list[str] = []
    seen_orders_in_push: dict[str, str] = {}
    existing_order_occurrences = index_existing_order_occurrences(pedidos_dir)

    for idx, pdf_path_raw in enumerate([Path(path) for path in pdf_paths], start=1):
        pdf_path = Path(pdf_path_raw)
        logger.info("[BH][LOTE] Processando arquivo %s/%s: %s", idx, len(pdf_paths), pdf_path.name)
        start_pdf = datetime.now()
        result = processor.parse_pdf(pdf_path)
        result.duration_seconds = (datetime.now() - start_pdf).total_seconds()

        if not result.layout_valid:
            invalid_file = ensure_unique_path(invalid_layout_dir / f"BH_LAYOUT_INVALIDO_{sanitize_filename(pdf_path.stem)}.xlsx")
            build_output_workbook(
                output_file=invalid_file,
                items=[],
                issues=result.issues,
                total_pdfs=1,
                success_count=0,
                mapping_result=mapping_result,
                title=f"LAYOUT INVALIDO: {pdf_path.name}",
                metadata={"PushLabel": push_label, "SourcePDF": pdf_path.name, "Blocked": "SIM"},
            )
            invalid_files.append(invalid_file)
            all_issues.extend(result.issues)
            parse_results.append(result)
            continue

        if result.numero_pedido:
            previous_pdf = seen_orders_in_push.get(result.numero_pedido)
            history_occurrences = existing_order_occurrences.get(result.numero_pedido, [])
            if previous_pdf:
                detail = (
                    f"Pedido {result.numero_pedido} repetido dentro do push atual. "
                    f"Primeira ocorrencia: {previous_pdf}. Ocorrencia atual: {pdf_path.name}."
                )
                result.issues.append(ParseIssue(pdf_path.name, "PEDIDO_REPETIDO_PUSH", detail))
                duplicate_orders.add(result.numero_pedido)
                duplicate_details.append(detail)
            else:
                seen_orders_in_push[result.numero_pedido] = pdf_path.name

            if history_occurrences:
                previous_pushes = sorted({occ.push_label for occ in history_occurrences if occ.push_label})
                previous_pushes_text = ", ".join(previous_pushes) if previous_pushes else "PUSH_ANTERIOR_NAO_MAPEADO"
                detail = (
                    f"Pedido {result.numero_pedido} ja possui historico salvo na pasta da data. "
                    f"Push(es) anterior(es): {previous_pushes_text}."
                )
                result.issues.append(ParseIssue(pdf_path.name, "PEDIDO_REPETIDO_HISTORICO", detail))
                duplicate_orders.add(result.numero_pedido)
                duplicate_details.append(detail)

        parse_results.append(result)
        all_items.extend(result.items)
        all_issues.extend(result.issues)
        if result.items and result.numero_pedido and result.cnpj and result.matricula:
            success_count += 1

        pedido_part = result.numero_pedido or "SEM_PEDIDO"
        pedido_dir = pedidos_dir / sanitize_filename(f"PEDIDO_{pedido_part}")
        pedido_dir.mkdir(parents=True, exist_ok=True)
        individual_file = ensure_unique_path(pedido_dir / f"BH_{pedido_part}_{sanitize_filename(pdf_path.stem)}.xlsx")
        build_output_workbook(
            output_file=individual_file,
            items=result.items,
            issues=result.issues,
            total_pdfs=1,
            success_count=1 if result.items and result.numero_pedido and result.cnpj and result.matricula else 0,
            mapping_result=mapping_result,
            title=f"PDF: {pdf_path.name}",
            duplicate_orders=[result.numero_pedido] if result.numero_pedido in duplicate_orders else [],
            metadata={
                "PushLabel": push_label,
                "PushNumber": str(push_number),
                "LoteDate": lote_date_label,
                "ProcessedAt": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                "SourcePDF": pdf_path.name,
                "SourcePath": str(pdf_path),
                "NumeroPedido": result.numero_pedido or "",
            },
        )
        result.output_file = individual_file
        result.pedido_dir = pedido_dir
        individual_files.append(individual_file)

    historical_items, historical_issues = collect_history_from_individual_files(pedidos_dir)
    duplicate_orders.update(detect_duplicate_orders_in_history(pedidos_dir))
    sync_duplicate_folders(pedidos_dir, duplicates_dir, sorted(duplicate_orders), logger)

    consolidado_file = consolidado_dir / f"BH_CONSOLIDADO_GERAL_{lote_date_label.replace('.', '_')}.xlsx"
    order_status_map = build_order_status_map(pedidos_dir)
    build_output_workbook(
        output_file=consolidado_file,
        items=historical_items,
        issues=historical_issues,
        total_pdfs=len(list(pedidos_dir.rglob("*.xlsx"))),
        success_count=len({item.numero_pedido for item in historical_items if item.numero_pedido}),
        mapping_result=mapping_result,
        title="CONSOLIDADO GERAL DA DATA",
        duplicate_orders=sorted(duplicate_orders),
        order_status_map=order_status_map,
        include_status_sheet=True,
    )

    build_main_pushes_workbook(main_pushes_file, pedidos_dir, lote_date_label, logger)

    executive_log_file = log_dir / f"LOG_EXECUTIVO_{lote_date_label.replace('.', '_')}_{timestamp}.txt"
    build_executive_log(
        executive_log_file,
        lote_date_label,
        date_dir,
        mapping_result,
        parse_results,
        len(pdf_paths),
        len(all_items),
        len(all_issues),
        success_count,
        sorted(duplicate_orders),
        duplicate_details,
        consolidado_file,
        main_pushes_file,
        push_label,
    )

    zip_file = date_dir / f"BH_PACOTE_RESULTADO_{lote_date_label.replace('.', '_')}.zip"
    build_result_zip(zip_file, date_dir)

    return BatchProcessResult(
        date_dir=date_dir,
        pedidos_dir=pedidos_dir,
        consolidado_file=consolidado_file,
        log_file=log_file,
        executive_log_file=executive_log_file,
        zip_file=zip_file,
        total_pdfs=len(pdf_paths),
        total_items=len(all_items),
        total_errors=len(all_issues),
        pedidos_com_sucesso=success_count,
        mapping_summary=f"Base BH: {mapping_result.valid_count} CNPJ(s) validos; GLN apoio: {len(gln_cnpj_map)} CNPJ(s).",
        lote_data=lote_date_label,
        duplicate_orders=sorted(duplicate_orders),
        duplicate_details=duplicate_details,
        individual_files=individual_files,
        invalid_layout_files=invalid_files,
        duplicates_dir=duplicates_dir,
        invalid_layout_dir=invalid_layout_dir,
        main_pushes_file=main_pushes_file,
        push_number=push_number,
        push_label=push_label,
    )


def data_remessa_dmais_um(base: datetime | None = None) -> str:
    """Rede BH: data de remessa sempre D+1 operacional, conforme regra definida."""
    base = base or datetime.now()
    return (base + timedelta(days=1)).strftime("%d.%m.%Y")


def build_kof_queue_from_lote(
    date_dir: str | Path,
    data_remessa: str,
    include_duplicates: bool = False,
    include_orders_with_alerts: bool = False,
) -> dict[str, object]:
    date_dir = Path(date_dir)
    pedidos_dir = date_dir / "01_PEDIDOS"
    fila_dir = date_dir / "06_FILA_KOF"
    fila_dir.mkdir(parents=True, exist_ok=True)
    items, issues = collect_history_from_individual_files(pedidos_dir)
    order_status_map = build_order_status_map(pedidos_dir)
    orders_with_alerts: set[str] = set()
    for workbook_path in pedidos_dir.rglob("*.xlsx"):
        _items, workbook_issues = read_items_and_issues_from_workbook(workbook_path)
        if not workbook_issues:
            continue
        metadata = read_workbook_metadata(workbook_path)
        pedido = metadata.get("NumeroPedido", "") or extract_pedido_from_path(workbook_path)
        if pedido:
            orders_with_alerts.add(pedido)

    rows: list[dict[str, str]] = []
    alerts: list[str] = []
    remessa = data_remessa_dmais_um().replace(".", "/")
    for item in items:
        status = order_status_map.get(item.numero_pedido, "OK")
        if status == "DUPLICADO" and not include_duplicates:
            alerts.append(f"VALIDAR - PEDIDO DUPLICADO | Pedido duplicado nao enviado para fila KOF: {item.numero_pedido}")
            continue
        if item.numero_pedido in orders_with_alerts and not include_orders_with_alerts:
            alerts.append(f"Pedido com alerta nao enviado para fila KOF: {item.numero_pedido}")
            continue
        if not item.matricula or not item.sku or not item.numero_pedido or item.qtd <= 0:
            alerts.append(f"Linha incompleta nao enviada: pedido={item.numero_pedido}, sku={item.sku}")
            continue
        rows.append({
            "Matricula": str(item.matricula),
            "Sku": str(item.sku),
            "Qtd": str(item.qtd),
            "Nº Pedido": str(item.numero_pedido),
            "Data remessa": remessa,
        })

    df = pd.DataFrame(rows, columns=config.FILA_COLUMNS)
    output_file = fila_dir / f"BH_FILA_KOF_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    build_queue_validation_workbook(output_file, df, alerts)
    return {"df": df, "alerts": alerts, "output_file": output_file}


def preview_order_number_from_pdf(pdf_path: str | Path, max_pages: int = 3) -> str:
    try:
        pages = extract_pdf_text_pages(pdf_path, max_pages=max_pages)
    except Exception:
        return ""
    return extract_order_number("\n".join(pages))


def extract_pdf_text_pages(pdf_path: str | Path, max_pages: int | None = None) -> list[str]:
    """Extrai texto com fallback pdfplumber -> pypdf.

    O bot BH antigo usava pypdf. O Robô KOF passou a usar pdfplumber em alguns
    pontos. Como cada biblioteca quebra linhas/colunas de um jeito, testamos as
    duas e escolhemos a extração com maior pontuação de marcadores BH.
    """
    pdf_path = Path(pdf_path)
    candidates: list[tuple[int, str, list[str]]] = []

    try:
        import pdfplumber
        texts: list[str] = []
        with pdfplumber.open(str(pdf_path)) as pdf:
            pages = pdf.pages if max_pages is None else pdf.pages[:max_pages]
            for page in pages:
                texts.append(page.extract_text() or "")
        candidates.append((score_extracted_text_pages(texts), "pdfplumber", texts))
    except Exception:
        pass

    try:
        from pypdf import PdfReader
        reader = PdfReader(str(pdf_path))
        texts = []
        pages = reader.pages if max_pages is None else reader.pages[:max_pages]
        for page in pages:
            try:
                texts.append(page.extract_text() or "")
            except Exception:
                texts.append("")
        candidates.append((score_extracted_text_pages(texts), "pypdf", texts))
    except Exception:
        pass

    if not candidates:
        raise RuntimeError("Nenhuma biblioteca de PDF disponivel para leitura BH. Instale pdfplumber ou pypdf.")

    candidates.sort(key=lambda item: (item[0], sum(len(p or "") for p in item[2])), reverse=True)
    return candidates[0][2]


def score_extracted_text_pages(page_texts: list[str]) -> int:
    text = "\n".join(page_texts or [])
    normalized = normalize_text(text)
    compact = re.sub(r"\s+", "", normalized)
    score = 0
    tokens = [
        "pedido de compra", "local de entrega", "produto descricao", "quantidade", "valor total",
        "numero do pedido", "codigo cod barras", "cod barras", "cod forn", "qtde", "qtde emb",
        "cnpj", "spal", "coca", "supermercados bh", "superm bh", "bh230",
    ]
    for token in tokens:
        if token in normalized:
            score += 3
    if "pedidodecompra" in compact:
        score += 8
    if extract_order_number(text):
        score += 10
    score += min(len(CNPJ_PATTERN.findall(text or "")), 5)
    score += min(len(re.findall(r"^\s*\d{4,8}\s+", text or "", flags=re.M)), 10)
    return score


def bh_layout_confidence(text: str, mapping: CNPJMapping | None = None) -> int:
    normalized = normalize_text(text)
    compact = re.sub(r"\s+", "", normalized)
    score = 0
    if extract_order_number(text):
        score += 3
    if "pedidodecompra" in compact or "pedido de compra" in normalized:
        score += 3
    if "local de entrega" in normalized:
        score += 3
    if "produto" in normalized and "quantidade" in normalized:
        score += 2
    if "codigo cod barras" in normalized or "cod barras" in normalized:
        score += 2
    if "cod forn" in normalized or "qtde emb" in normalized:
        score += 2
    if any(token in normalized for token in ("supermercados bh", "superm bh", "bh230")):
        score += 3
    if mapping is not None:
        supplier_match = SUPPLIER_CNPJ_PATTERN.search(text or "")
        supplier = normalize_cnpj(supplier_match.group(1)) if supplier_match else ""
        for raw in CNPJ_PATTERN.findall(text or ""):
            cnpj = normalize_cnpj(raw)
            if cnpj and cnpj != supplier and mapping.get_matricula(cnpj):
                score += 5
                break
    return score


def is_bh_layout(text: str) -> bool:
    normalized = normalize_text(text)
    if not normalized.strip():
        return False
    if not extract_order_number(text):
        return False
    compact = re.sub(r"\s+", "", normalized)

    legacy_hits = 0
    for token in ("local de entrega", "produto", "quantidade", "valor total"):
        if token in normalized:
            legacy_hits += 1
    if "pedidodecompra" in compact or "pedido de compra" in normalized:
        legacy_hits += 2
    if legacy_hits >= 4:
        return True

    rpone_hits = 0
    for token in ("numero do pedido", "codigo cod barras", "cod barras", "cod forn", "qtde"):
        if token in normalized:
            rpone_hits += 1
    if rpone_hits >= 3:
        return True

    if any(token in normalized for token in ("supermercados bh", "superm bh", "bh230")):
        return True
    return False


def extract_store_identity(page_texts: List[str], pdf_path: Path | None = None) -> tuple[str, str]:
    """Extrai a loja BH do PDF ou dos metadados do Outlook.

    Padrões suportados:
    - Local de Entrega : 139 - MATOZINHOS
    - assunto do e-mail: Pedido loja 139 / Pedido loj 139
    """
    for page_text in page_texts or []:
        block = extract_local_delivery_block(page_text or "") or page_text or ""
        match = re.search(r"Local\s+de\s+Entrega\s*:?\s*(\d{1,5})\s*[-–]\s*([^\r\n]+)", block, re.IGNORECASE)
        if match:
            numero = match.group(1).strip()
            nome = clean_store_name(match.group(2))
            return numero, nome

    metadata = read_outlook_metadata_for_file(pdf_path) if pdf_path else {}
    assunto = str(metadata.get("assunto") or metadata.get("email_assunto") or "")
    match = re.search(r"\bloj(?:a)?\s*(\d{1,5})\b", assunto, re.IGNORECASE)
    if match:
        return match.group(1).strip(), ""
    return "", ""


def clean_store_name(value: str) -> str:
    text = " ".join(str(value or "").split())
    # corta quando a extração junta endereço/CNPJ na mesma linha
    text = re.split(r"\b(?:CNPJ|I\.?E\.?|CEP|Telefone|Fax|Observa[cç][oõ]es)\b", text, maxsplit=1, flags=re.IGNORECASE)[0]
    return text.strip(" -:\t")


def read_outlook_metadata_for_file(file_path: Path | None) -> dict[str, str]:
    if not file_path:
        return {}
    candidates = [
        Path(str(file_path) + ".outlook.json"),
        file_path.with_suffix(file_path.suffix + ".outlook.json"),
        file_path.with_suffix(".outlook.json"),
    ]
    for candidate in candidates:
        if not candidate.exists():
            continue
        try:
            data = json.loads(candidate.read_text(encoding="utf-8"))
            return {str(k): "" if v is None else str(v) for k, v in data.items()}
        except Exception:
            continue
    return {}


def normalize_lote_date(value: str | None) -> str:
    if not value:
        return datetime.now().strftime("%d.%m.%Y")
    raw = str(value).strip()
    if not raw:
        return datetime.now().strftime("%d.%m.%Y")
    for pattern in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d%m%Y"):
        try:
            return datetime.strptime(raw, pattern).strftime("%d.%m.%Y")
        except ValueError:
            continue
    raise ValueError("Data invalida. Use DD.MM.AAAA, DD/MM/AAAA ou AAAA-MM-DD.")


def build_logger(log_file: Path) -> logging.Logger:
    logger = logging.getLogger(f"robokof_bh_{log_file}")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
    logger.addHandler(console_handler)
    logger.propagate = False
    return logger


def extract_order_number(text: str) -> str:
    for pattern in ORDER_PATTERNS:
        match = pattern.search(text or "")
        if match:
            return match.group(1)
    return ""


def extract_delivery_cnpj(page_texts: List[str], mapping: CNPJMapping | None = None) -> str:
    """Identifica o CNPJ correto da loja BH.

    Regra corporativa da BH:
    1) priorizar o CNPJ rotulado no bloco Local de Entrega;
    2) ignorar Inscrição Estadual/I.E. e CNPJ do fornecedor;
    3) só aceitar fallback quando a chave existir na base BH ou for o único CNPJ confiável.
    """
    for page_text in page_texts:
        block = extract_local_delivery_block(page_text)
        if not block:
            continue

        labelled = extract_labelled_delivery_cnpjs(block)
        mapped = _filter_cnpjs_in_bh_mapping(labelled, mapping)
        if mapped:
            return mapped[-1]
        if labelled:
            return labelled[-1]

        # Fallback no bloco, mas removendo IE/Inscrição Estadual.
        ignored = set(extract_ie_like_numbers(block))
        cnpjs = [cnpj for cnpj in _unique_cnpjs(CNPJ_PATTERN.findall(block)) if cnpj not in ignored and is_probable_cnpj_for_pdf(cnpj)]
        mapped = _filter_cnpjs_in_bh_mapping(cnpjs, mapping)
        if mapped:
            return mapped[-1]
        if cnpjs:
            return cnpjs[-1]

    # Alguns PDFs quebram o bloco no OCR/extração. Procura qualquer CNPJ do arquivo que exista na base BH,
    # descartando fornecedor e I.E./Inscrição Estadual.
    for page_text in page_texts:
        supplier_match = SUPPLIER_CNPJ_PATTERN.search(page_text or "")
        supplier = normalize_cnpj(supplier_match.group(1)) if supplier_match else ""
        ignored = set(extract_ie_like_numbers(page_text or ""))
        all_cnpjs = _unique_cnpjs(CNPJ_PATTERN.findall(page_text or ""))
        filtered = [cnpj for cnpj in all_cnpjs if cnpj and cnpj != supplier and cnpj not in ignored and is_probable_cnpj_for_pdf(cnpj)]
        mapped = _filter_cnpjs_in_bh_mapping(filtered, mapping)
        if mapped:
            return mapped[-1]

    # Último fallback: retorna o último CNPJ formatado/confiável diferente do fornecedor.
    for page_text in page_texts:
        supplier_match = SUPPLIER_CNPJ_PATTERN.search(page_text or "")
        supplier = normalize_cnpj(supplier_match.group(1)) if supplier_match else ""
        ignored = set(extract_ie_like_numbers(page_text or ""))
        all_cnpjs = _unique_cnpjs(CNPJ_PATTERN.findall(page_text or ""))
        filtered = [cnpj for cnpj in all_cnpjs if cnpj and cnpj != supplier and cnpj not in ignored and is_probable_cnpj_for_pdf(cnpj)]
        if filtered:
            return filtered[-1]
    return ""


def extract_labelled_delivery_cnpjs(block: str) -> list[str]:
    encontrados: list[str] = []
    for pattern in DELIVERY_CNPJ_LABEL_PATTERNS:
        for match in pattern.finditer(block or ""):
            cnpj = normalize_cnpj(match.group(1))
            if cnpj and cnpj not in encontrados:
                encontrados.append(cnpj)
    return encontrados


def extract_ie_like_numbers(text: str) -> list[str]:
    ignorar: list[str] = []
    for match in IE_LABEL_PATTERN.finditer(text or ""):
        normalized = normalize_cnpj(match.group(1))
        if normalized:
            ignorar.append(normalized)
    return ignorar


def is_probable_cnpj_for_pdf(cnpj: str) -> bool:
    digits = re.sub(r"\D", "", str(cnpj or ""))
    # Em PDFs, CNPJ confiável costuma vir com 14 dígitos. Valores 13 dígitos no texto são,
    # frequentemente, I.E./inscrição; a base cadastral continua aceitando 13 dígitos via zfill.
    return len(digits) == 14

def _unique_cnpjs(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values or []:
        cnpj = normalize_cnpj(value)
        if not cnpj or cnpj in seen:
            continue
        seen.add(cnpj)
        result.append(cnpj)
    return result


def _filter_cnpjs_in_bh_mapping(cnpjs: list[str], mapping: CNPJMapping | None) -> list[str]:
    if mapping is None:
        return []
    return [cnpj for cnpj in cnpjs if mapping.get_matricula(cnpj)]


def extract_local_delivery_block(page_text: str) -> str:
    if not page_text:
        return ""
    start = page_text.find("Local de Entrega")
    if start == -1:
        normalized = normalize_text(page_text)
        if "local de entrega" not in normalized:
            return ""
        return page_text
    end_candidates = [idx for idx in [page_text.find("Assinatura Fornecedor", start), page_text.find("Pagina:", start), page_text.find("Página:", start)] if idx != -1]
    end = min(end_candidates) if end_candidates else len(page_text)
    return page_text[start:end]


def parse_items_from_page(
    page_text: str,
    pdf_name: str,
    numero_pedido: str,
    cnpj: str,
    matricula: str,
    page_number: int,
    loja_numero: str = "",
    loja_nome: str = "",
) -> tuple[List[ParsedItem], List[ParseIssue]]:
    items: List[ParsedItem] = []
    issues: List[ParseIssue] = []
    if not page_text:
        return items, issues

    in_legacy_table = False
    in_rpone_table = False
    rpone_header = ""

    for raw_line in page_text.splitlines():
        line = " ".join((raw_line or "").split())
        if not line:
            continue
        normalized_line = normalize_text(line)

        if is_legacy_table_header(normalized_line):
            in_legacy_table = True
            in_rpone_table = False
            continue

        if is_rpone_table_header(normalized_line):
            in_rpone_table = True
            in_legacy_table = False
            rpone_header = normalized_line
            continue

        if any(marker in normalized_line for marker in STOP_MARKERS):
            in_legacy_table = False
        if in_rpone_table and is_rpone_table_stop(normalized_line):
            in_rpone_table = False

        # RP One precisa ter prioridade: ele começa com código original + EAN,
        # e poderia ser confundido com SKU legado se lido primeiro.
        if in_rpone_table or looks_like_rpone_item_line(line):
            item = parse_rpone_item_line(line, rpone_header, pdf_name, numero_pedido, cnpj, matricula, page_number, issues, loja_numero, loja_nome)
            if item:
                items.append(item)
                continue

        if in_legacy_table or looks_like_legacy_item_line(line):
            item = parse_legacy_item_line(line, pdf_name, numero_pedido, cnpj, matricula, page_number, issues, loja_numero, loja_nome)
            if item:
                items.append(item)
                continue

        if re.match(r"^\s*\d{5,6}\s+", line):
            issues.append(ParseIssue(pdf_name, "LINHA_ITEM_NAO_LIDA", "Linha com SKU encontrada, mas fora dos padrões BH conhecidos.", page_number, line))

    return items, issues


def is_legacy_table_header(normalized_line: str) -> bool:
    if not normalized_line:
        return False
    required = ["produto", "descricao", "quantidade", "valor total"]
    if all(token in normalized_line for token in required):
        return True
    return "produto descricao" in normalized_line and "quantidade" in normalized_line


def is_rpone_table_header(normalized_line: str) -> bool:
    if not normalized_line:
        return False
    return (
        ("codigo cod barras" in normalized_line or "cod barras" in normalized_line)
        and ("cod forn" in normalized_line or "qtde" in normalized_line or "quant" in normalized_line)
    )


def looks_like_legacy_item_line(line: str) -> bool:
    if not re.match(r"^\s*\d{5,6}\s+", line or ""):
        return False
    normalized = normalize_text(line)
    if " c/" in normalized or "cx c/" in normalized or "fd c/" in normalized:
        return True
    return len(re.findall(r"\d+(?:\.\d{3})*,\d{2,4}", line or "")) >= 3


def parse_legacy_item_line(
    line: str,
    pdf_name: str,
    numero_pedido: str,
    cnpj: str,
    matricula: str,
    page_number: int,
    issues: List[ParseIssue],
    loja_numero: str = "",
    loja_nome: str = "",
) -> ParsedItem | None:
    # Layout BH legado real: Produto | Descrição | Unidade | Preço Normal | Quantidade | Valor Total.
    # O padrão anterior podia inverter Unidade/Preço e capturar quantidade errada.
    match_unit_first = LEGACY_ITEM_UNIT_FIRST_PATTERN.match(line)
    if match_unit_first:
        sku, descricao, unidade, _preco, quantidade_bruta, _valor_total = match_unit_first.groups()
    else:
        match = ITEM_PATTERN.match(line)
        if match:
            sku, descricao, _preco, quantidade_bruta, unidade, _valor_total = match.groups()
            issues.append(ParseIssue(pdf_name, "VALIDAR_ORDEM_COLUNAS_BH", "Linha legada lida no padrão alternativo preço/quantidade/unidade.", page_number, line))
        else:
            parsed = parse_legacy_item_line_flex(line)
            if not parsed:
                if re.match(r"^\s*\d{5,6}\s+", line):
                    issues.append(ParseIssue(pdf_name, "LINHA_ITEM_NAO_LIDA", "Linha com SKU fora do padrão esperado.", page_number, line))
                return None
            sku, descricao, quantidade_bruta, unidade = parsed
            issues.append(ParseIssue(pdf_name, "LEITURA_BH_FLEX", "Linha lida pelo fallback flexível da Rede BH.", page_number, line))

    try:
        qtd = normalize_quantity_to_int(quantidade_bruta)
    except ValueError as exc:
        issues.append(ParseIssue(pdf_name, "QTD_INVALIDA", str(exc), page_number, line))
        return None

    return ParsedItem(
        arquivo=pdf_name,
        numero_pedido=numero_pedido,
        cnpj=cnpj,
        matricula=matricula,
        sku=normalize_sku_bh(sku),
        descricao=descricao.strip(),
        qtd=qtd,
        unidade=unidade.strip().upper(),
        pagina=page_number,
        loja_numero=loja_numero,
        loja_nome=loja_nome,
    )

def parse_legacy_item_line_flex(line: str) -> tuple[str, str, str, str] | None:
    """Fallback compatível com o BH antigo quando o PDF muda espaçamento.

    Retorna: sku, descricao, quantidade, unidade.
    A regra prioriza a coluna Quantidade, que no BH legado fica APÓS Unidade e Preço Normal.
    """
    match = re.match(r"^\s*(\d{5,6})\s+(.+?)\s*$", line or "")
    if not match:
        return None
    sku = match.group(1)
    body = match.group(2).strip()

    unit_match = re.search(r"\b((?:CX|FD)\s*C\s*/?\s*\d+|(?:CX|FD)/?\d+)\b", body, flags=re.I)
    if not unit_match:
        return None
    unidade = unit_match.group(1)
    before_unit = body[:unit_match.start()].strip()
    after_unit = body[unit_match.end():].strip()

    # Padrão correto: Unidade, Preço Normal, Quantidade, Valor Total.
    nums_after = list(re.finditer(r"\d+(?:\.\d{3})*,\d{2,4}|\d+", after_unit))
    if len(nums_after) >= 3:
        quantidade_bruta = nums_after[1].group(0)
        descricao = before_unit.strip()
        return sku, descricao, quantidade_bruta, unidade

    # Fallback antigo: se a unidade veio ao final da linha, usa o último número antes da unidade.
    nums_before = list(re.finditer(r"\d+(?:\.\d{3})*,\d{2,4}|\d+", before_unit))
    if len(nums_before) >= 2:
        quantidade_bruta = nums_before[-1].group(0)
        descricao = before_unit[:nums_before[-2].start()].strip() or before_unit[:nums_before[-1].start()].strip()
        return sku, descricao, quantidade_bruta, unidade

    return None

def is_rpone_table_stop(normalized_line: str) -> bool:
    if not normalized_line:
        return False
    stop_tokens = (
        "trocas:", "total das pendencias", "ocorrencias pendentes", "totais",
        "valor total:", "valor ipi:", "peso total:", "comprador", "vendedor",
        "contatos do fornecedor", "atencao:", "favor informar", "informamos aos fornecedores",
        "transacao:", "fornecedor:", "empresa:", "dt. pedido:", "frete:", "pg:",
    )
    return normalized_line.startswith(stop_tokens)


def looks_like_rpone_item_line(line: str) -> bool:
    # Ex.: 1784412 7894900087000 ... 24,000 6,2300 149,52 119143
    # Ex.: 103063 7894900530056 92581 ... 40 80,000 FD/2 ...
    return bool(re.match(r"^\s*\d{3,8}\s+\d{8,14}\s+", line or ""))


def parse_rpone_item_line(
    line: str,
    header: str,
    pdf_name: str,
    numero_pedido: str,
    cnpj: str,
    matricula: str,
    page_number: int,
    issues: List[ParseIssue],
    loja_numero: str = "",
    loja_nome: str = "",
) -> ParsedItem | None:
    if not looks_like_rpone_item_line(line):
        return None

    tokens = line.split()
    if len(tokens) < 7:
        issues.append(ParseIssue(pdf_name, "LINHA_ITEM_NAO_LIDA", "Linha RP One curta demais para leitura segura.", page_number, line))
        return None

    codigo_original = tokens[0]
    ean = tokens[1]
    remaining = tokens[2:]

    # Modelo 1: Código Cod Barras Cod Forn Descrição Qtde Emb Quant Emb...
    # SKU aparece logo após o EAN. A quantidade correta para a fila BH é a
    # quantidade em caixaria/embalagem, normalmente posicionada imediatamente
    # antes da coluna Emb (FD/2, CX/6, etc.).
    if remaining and is_sku_token(remaining[0]):
        sku_raw = remaining[0]
        rest = remaining[1:]

        pack_index = find_packaging_token_index(rest)
        if pack_index is not None:
            qtd_index = find_bh_quantity_index_before_packaging(rest, pack_index)
            if qtd_index is None:
                issues.append(ParseIssue(pdf_name, "LINHA_ITEM_NAO_LIDA", "Linha RP One com embalagem, mas sem Qtde Emb antes da embalagem.", page_number, line))
                return None
            quantidade_bruta = rest[qtd_index]
            unidade = rest[pack_index]
            descricao_tokens = rest[:qtd_index]
        else:
            # Fallback apenas quando a embalagem não foi extraída. Evita capturar
            # números da descrição sempre que possível, priorizando token decimal
            # de quantidade e registrando alerta para conferência.
            qtd_index = find_rpone_quantity_index_without_packaging(rest)
            if qtd_index is None:
                issues.append(ParseIssue(pdf_name, "LINHA_ITEM_NAO_LIDA", "Linha RP One com Cod Forn, mas sem quantidade segura.", page_number, line))
                return None
            quantidade_bruta = rest[qtd_index]
            unidade = "UN"
            descricao_tokens = rest[:qtd_index]
            issues.append(ParseIssue(pdf_name, "VALIDAR_QTD_RPONE", "Linha RP One com Cod Forn sem embalagem clara; validar se a quantidade está em caixa.", page_number, line))

        descricao = " ".join(descricao_tokens).strip() or f"CODIGO {codigo_original}"
    else:
        # Modelo 2: Código Cod Barras Descrição Marca Quant Pr Unit Vl Total Cod Forn
        # SKU fica no fim da linha. Neste padrão não há Qtde Emb; mantemos a
        # quantidade lida e alertamos para validação antes da fila.
        sku_raw = extract_sku_from_tail(tokens)
        if not sku_raw:
            issues.append(ParseIssue(pdf_name, "SKU_NAO_IDENTIFICADO", "Cod Forn/SKU não identificado na linha RP One.", page_number, line))
            return None
        qtd_index = find_rpone_quantity_index_without_packaging(remaining)
        if qtd_index is None:
            issues.append(ParseIssue(pdf_name, "QTD_INVALIDA", "Quantidade RP One não localizada.", page_number, line))
            return None
        quantidade_bruta = remaining[qtd_index]
        descricao_tokens = remaining[:qtd_index]
        descricao = " ".join(descricao_tokens).strip() or f"CODIGO {codigo_original}"
        unidade = "UN"
        if "qtde emb" not in header and "emb" not in header:
            issues.append(ParseIssue(pdf_name, "VALIDAR_QTD_RPONE_SEM_EMBALAGEM", "Layout RP One sem coluna Qtde Emb; quantidade lida deve ser validada antes da fila.", page_number, line))

    try:
        qtd = normalize_quantity_to_int(quantidade_bruta)
    except ValueError as exc:
        issues.append(ParseIssue(pdf_name, "QTD_INVALIDA", str(exc), page_number, line))
        return None

    return ParsedItem(
        arquivo=pdf_name,
        numero_pedido=numero_pedido,
        cnpj=cnpj,
        matricula=matricula,
        sku=normalize_sku_bh(sku_raw),
        descricao=descricao,
        qtd=qtd,
        unidade=unidade.strip().upper(),
        pagina=page_number,
        loja_numero=loja_numero,
        loja_nome=loja_nome,
    )


def is_sku_token(value: str) -> bool:
    cleaned = re.sub(r"\D", "", value or "")
    return 4 <= len(cleaned) <= 6


def normalize_sku_bh(value: str) -> str:
    cleaned = re.sub(r"\D", "", str(value or ""))
    return cleaned.zfill(6) if cleaned else ""


def is_decimal_token(value: str) -> bool:
    return bool(re.fullmatch(r"\d+(?:\.\d{3})*,\d{1,4}|\d+", value or ""))


def find_first_numeric_token_index(tokens: list[str]) -> int | None:
    for idx, tok in enumerate(tokens):
        if is_decimal_token(tok):
            return idx
    return None


def find_packaging_token_index(tokens: list[str]) -> int | None:
    for idx, tok in enumerate(tokens):
        if looks_like_packaging(tok):
            return idx
    return None


def find_numeric_token_before(tokens: list[str], before_index: int) -> int | None:
    for idx in range(before_index - 1, -1, -1):
        if is_decimal_token(tokens[idx]):
            return idx
    return None


def find_bh_quantity_index_before_packaging(tokens: list[str], pack_index: int) -> int | None:
    """Seleciona Qtde Emb na linha RP One da BH.

    Alguns PDFs extraem as colunas como `40 80,000 FD/2` e outros como
    `40 FD/2 80,000`. A quantidade para digitação é a caixaria (`40`), não
    a quantidade unitária/total (`80,000`).
    """
    numeric_indexes = [idx for idx in range(0, pack_index) if is_decimal_token(tokens[idx])]
    if not numeric_indexes:
        return None

    # Quando imediatamente antes da embalagem há número com vírgula e existe
    # outro número antes dele, o primeiro é Qtde Emb e o decimal é Quant Emb.
    last_idx = numeric_indexes[-1]
    if "," in str(tokens[last_idx]) and len(numeric_indexes) >= 2:
        return numeric_indexes[-2]

    return last_idx


def find_rpone_quantity_index_without_packaging(tokens: list[str]) -> int | None:
    # Em layout RP One sem Emb visível, a quantidade costuma vir como 60,000.
    # Preços normalmente têm 2 ou 4 casas decimais. Assim evitamos capturar
    # números soltos da descrição, como 1,5lt, 2L, 350ml etc.
    for idx, tok in enumerate(tokens):
        if re.fullmatch(r"\d+(?:\.\d{3})*,\d{3}", tok or ""):
            return idx
    for idx, tok in enumerate(tokens):
        if is_decimal_token(tok) and "," in tok:
            return idx
    for idx, tok in enumerate(tokens):
        if is_decimal_token(tok):
            return idx
    return None


def looks_like_packaging(value: str) -> bool:
    return bool(re.fullmatch(r"(?:CX|FD|UN|PC|KG|LT)/?\d*|(?:CX|FD)\s*C/?\s*\d+", value or "", flags=re.IGNORECASE))


def extract_sku_from_tail(tokens: list[str]) -> str:
    # Busca do fim para o início para evitar pegar código original/EAN/preço.
    for tok in reversed(tokens):
        clean = re.sub(r"\D", "", tok or "")
        if 4 <= len(clean) <= 6:
            return clean
        if "," in tok:
            candidates = [re.sub(r"\D", "", p) for p in re.split(r"[,;/]", tok)]
            candidates = [p for p in candidates if 4 <= len(p) <= 6]
            if candidates:
                # Quando vier múltiplo Cod Forn, prioriza o último candidato real.
                return candidates[-1]
    return ""


def normalize_cnpj(value: str) -> str:
    digits = re.sub(r"\D", "", value or "")
    if not digits:
        return ""
    return digits.zfill(14) if len(digits) <= 14 else digits


def normalize_quantity_to_int(value: str) -> int:
    cleaned = (value or "").strip().replace(".", "").replace(",", ".")
    if not cleaned:
        raise ValueError("Quantidade vazia.")
    try:
        decimal_value = Decimal(cleaned)
    except InvalidOperation as exc:
        raise ValueError(f"Quantidade invalida: {value}") from exc
    integral = decimal_value.quantize(Decimal("1"))
    if decimal_value != integral:
        raise ValueError(f"Quantidade fracionada detectada ({value}).")
    return int(integral)


def normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).lower().strip()


def _load_gln_cnpj_map(logger: logging.Logger | None = None) -> dict[str, str]:
    try:
        return gln_service.load_cnpj_to_matricula_map(
            config.GLN_BASE_PATH,
            config.GLN_SHEET_NAME,
            config.GLN_COL_CNPJ,
            config.GLN_COL_MATRICULA,
        )
    except Exception as exc:
        if logger:
            logger.warning("[BH][GLN] Base GLN nao carregada para conferencia: %s", exc)
        return {}


def extract_pedido_from_path(file_path: Path) -> str:
    for part in file_path.parts:
        match = re.match(r"PEDIDO_(.+)", part, re.IGNORECASE)
        if match:
            return match.group(1)
    match = re.search(r"BH_(\d+)_", file_path.name)
    return match.group(1) if match else ""


def ensure_unique_path(target: Path) -> Path:
    if not target.exists():
        return target
    return target.with_name(f"{target.stem}__REPETIDO__{datetime.now().strftime('%Y%m%d_%H%M%S')}{target.suffix}")


def validar_workbook_salvo(output_file: Path) -> None:
    wb = load_workbook(output_file, read_only=True, data_only=True)
    wb.close()


def get_next_push_number(main_pushes_file: Path) -> int:
    if not main_pushes_file.exists():
        return 1
    try:
        wb = load_workbook(main_pushes_file, data_only=True, read_only=True)
    except Exception:
        return 1
    try:
        numbers = []
        for sheet_name in wb.sheetnames:
            match = re.fullmatch(r"PUSH_(\d{3})", sheet_name)
            if match:
                numbers.append(int(match.group(1)))
        return max(numbers, default=0) + 1
    finally:
        wb.close()


def collect_history_from_individual_files(pedidos_dir: Path) -> tuple[List[ParsedItem], List[ParseIssue]]:
    all_items: List[ParsedItem] = []
    all_issues: List[ParseIssue] = []
    for workbook_path in sorted(pedidos_dir.rglob("*.xlsx")):
        items, issues = read_items_and_issues_from_workbook(workbook_path)
        all_items.extend(items)
        all_issues.extend(issues)
    return all_items, all_issues


def read_items_and_issues_from_workbook(workbook_path: Path) -> tuple[List[ParsedItem], List[ParseIssue]]:
    items: List[ParsedItem] = []
    issues: List[ParseIssue] = []
    try:
        wb = load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception:
        return items, issues
    try:
        if "PEDIDOS" in wb.sheetnames:
            for row in wb["PEDIDOS"].iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                values = list(row) + [""] * 11
                arquivo, numero_pedido, cnpj, matricula, sku, descricao, qtd, unidade, pagina, loja_numero, loja_nome = values[:11]
                items.append(
                    ParsedItem(
                        arquivo=str(arquivo or workbook_path.name),
                        numero_pedido=str(numero_pedido or ""),
                        cnpj=str(cnpj or ""),
                        matricula=str(matricula or ""),
                        sku=str(sku or ""),
                        descricao=str(descricao or ""),
                        qtd=_safe_int(qtd),
                        unidade=str(unidade or ""),
                        pagina=_safe_int(pagina),
                        loja_numero=str(loja_numero or ""),
                        loja_nome=str(loja_nome or ""),
                    )
                )
        if "ERROS" in wb.sheetnames:
            for row in wb["ERROS"].iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                values = list(row) + [""] * 5
                arquivo, tipo, detalhe, pagina, linha = values[:5]
                issues.append(ParseIssue(str(arquivo or workbook_path.name), str(tipo or ""), str(detalhe or ""), _safe_optional_int(pagina), str(linha or "") or None))
    finally:
        wb.close()
    return items, issues


def read_workbook_metadata(workbook_path: Path) -> dict[str, str]:
    metadata: dict[str, str] = {}
    try:
        wb = load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception:
        return metadata
    try:
        if "METADADOS" not in wb.sheetnames:
            return metadata
        for row in wb["METADADOS"].iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            key, value = row[:2]
            if key not in (None, ""):
                metadata[str(key)] = "" if value is None else str(value)
    finally:
        wb.close()
    return metadata


def resolve_push_label(workbook_path: Path, metadata: dict[str, str] | None = None) -> str:
    metadata = metadata or {}
    push_label = (metadata.get("PushLabel") or "").strip()
    if push_label:
        return push_label
    match = re.search(r"PUSH_(\d{3})", workbook_path.name, re.IGNORECASE)
    if match:
        return f"PUSH_{match.group(1)}"
    return "PUSH_LEGADO"


def index_existing_order_occurrences(pedidos_dir: Path) -> dict[str, List[HistoricalOrderOccurrence]]:
    indexed: dict[str, List[HistoricalOrderOccurrence]] = {}
    if not pedidos_dir.exists():
        return indexed
    for file_path in sorted(pedidos_dir.rglob("*.xlsx")):
        pedido = extract_pedido_from_path(file_path)
        if not pedido:
            continue
        metadata = read_workbook_metadata(file_path)
        indexed.setdefault(pedido, []).append(
            HistoricalOrderOccurrence(
                pedido=pedido,
                file_path=file_path,
                push_label=resolve_push_label(file_path, metadata),
                processed_at=metadata.get("ProcessedAt", ""),
            )
        )
    return indexed


def detect_duplicate_orders_in_history(pedidos_dir: Path) -> List[str]:
    duplicates: List[str] = []
    for pedido_dir in sorted(pedidos_dir.glob("PEDIDO_*")):
        if pedido_dir.is_dir() and len(list(pedido_dir.glob("*.xlsx"))) > 1:
            duplicates.append(pedido_dir.name.replace("PEDIDO_", ""))
    return duplicates


def build_order_status_map(pedidos_dir: Path) -> dict[str, str]:
    occurrences = index_existing_order_occurrences(pedidos_dir)
    return {pedido: ("DUPLICADO" if len(items) > 1 else "OK") for pedido, items in occurrences.items() if pedido}


def sync_duplicate_folders(pedidos_dir: Path, duplicates_dir: Path, duplicate_orders: List[str], logger: logging.Logger | None = None) -> None:
    duplicates_dir.mkdir(parents=True, exist_ok=True)
    expected = {sanitize_filename(f"PEDIDO_{pedido}") for pedido in duplicate_orders if pedido}
    for existing in duplicates_dir.glob("PEDIDO_*"):
        if existing.is_dir() and existing.name not in expected:
            shutil.rmtree(existing, ignore_errors=True)
    for pedido in duplicate_orders:
        source_dir = pedidos_dir / sanitize_filename(f"PEDIDO_{pedido}")
        if not source_dir.exists():
            continue
        target_dir = duplicates_dir / sanitize_filename(f"PEDIDO_{pedido}")
        if target_dir.exists():
            shutil.rmtree(target_dir, ignore_errors=True)
        target_dir.mkdir(parents=True, exist_ok=True)
        copied = 0
        for file_path in sorted(source_dir.glob("*.xlsx")):
            shutil.copy2(file_path, target_dir / file_path.name)
            copied += 1
        if logger:
            logger.info("[BH][DUPLICADOS] Pedido %s espelhado com %s arquivo(s).", pedido, copied)


def build_output_workbook(
    output_file: Path,
    items: List[ParsedItem],
    issues: List[ParseIssue],
    total_pdfs: int,
    success_count: int,
    mapping_result: MappingLoadResult,
    title: str,
    duplicate_orders: List[str] | None = None,
    metadata: dict[str, str] | None = None,
    order_status_map: dict[str, str] | None = None,
    include_status_sheet: bool = False,
) -> None:
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "PEDIDOS"
    ws_errors = wb.create_sheet("ERROS")
    ws_resume = wb.create_sheet("RESUMO")
    ws_duplicates = wb.create_sheet("DUPLICADOS")
    ws_metadata = wb.create_sheet("METADADOS")
    ws_status = wb.create_sheet("STATUS_PEDIDOS") if include_status_sheet and order_status_map else None
    ws_modelo = wb.create_sheet("Modelo Robô KOF para Enviar")
    ws_validacao = wb.create_sheet("Validação do Pedido")
    ws_bloqueados = wb.create_sheet("Itens Bloqueados Fila")
    ws_cadastrar = wb.create_sheet("Cadastrar CNPJ")
    ws_lojas = wb.create_sheet("Lojas")

    data_headers = ["Arquivo", "Numero_Pedido", "CNPJ", "Matricula", "SKU", "Descricao", "Qtd", "Unidade", "Pagina", "Loja Número", "Loja Nome"]
    if order_status_map:
        data_headers.append("Status_Consolidado")
    ws_data.append(data_headers)
    for item in items:
        row = [item.arquivo, item.numero_pedido, item.cnpj, item.matricula, item.sku, item.descricao, item.qtd, item.unidade, item.pagina, getattr(item, "loja_numero", ""), getattr(item, "loja_nome", "")]
        if order_status_map:
            row.append(order_status_map.get(item.numero_pedido, "OK"))
        ws_data.append(row)

    ws_modelo.append(["Matricula", "CNPJ", "Sku", "Qtd", "Nº Pedido", "Data remessa", "Status Conversão"])
    ws_validacao.append(["CNPJ", "Matrícula", "Descrição", "Código SKU", "EAN", "SKU", "QTD", "Nº do Pedido", "Status Conversão", "Observação / Alerta", "Data Remessa", "Loja Número", "Loja Nome"])
    ws_bloqueados.append(["Arquivo", "Nº Pedido", "CNPJ", "Matrícula", "SKU", "QTD", "Motivo Bloqueio", "Observação"])

    duplicate_orders_set = {str(p).strip() for p in (duplicate_orders or []) if str(p).strip()}
    cnpj_pendentes: dict[str, set[str]] = {}
    for item in items:
        observacoes: List[str] = []
        bloqueios: List[str] = []
        pedido_status = str((order_status_map or {}).get(item.numero_pedido, "") or "").strip().upper()
        item_duplicado = bool(item.numero_pedido and (item.numero_pedido in duplicate_orders_set or pedido_status == "DUPLICADO"))
        if item.cnpj and not item.matricula:
            observacoes.append("A CADASTRAR")
            cnpj_pendentes.setdefault(item.cnpj, set()).add(item.numero_pedido or "")
        if item_duplicado:
            observacoes.append("Pedido duplicado no push/histórico da Rede BH")
            observacoes.append("DUPLICADO - NÃO ENVIAR PARA FILA")
            bloqueios.append("PEDIDO_DUPLICADO_BH")

        if item_duplicado:
            status_conversao = "VALIDAR - PEDIDO DUPLICADO"
        else:
            status_conversao = "OK SEM CONVERSÃO"
        obs_text = " | ".join(dict.fromkeys(observacoes)) or "OK"

        if bloqueios:
            ws_bloqueados.append([
                item.arquivo,
                item.numero_pedido,
                item.cnpj,
                item.matricula,
                item.sku,
                item.qtd,
                " | ".join(bloqueios),
                obs_text,
            ])
        else:
            ws_modelo.append([item.matricula, item.cnpj, item.sku, item.qtd, item.numero_pedido, data_remessa_dmais_um().replace(".", "/"), status_conversao])
        ws_validacao.append([
            item.cnpj,
            item.matricula,
            item.descricao,
            item.sku,
            "",
            item.sku,
            item.qtd,
            item.numero_pedido,
            status_conversao,
            obs_text,
            data_remessa_dmais_um().replace(".", "/"),
            getattr(item, "loja_numero", ""),
            getattr(item, "loja_nome", ""),
        ])

    ws_lojas.append(["Loja Número", "Loja Nome", "CNPJ", "Matrícula", "Nº Pedido", "Arquivo Origem", "Observação"])
    lojas_vistas: set[tuple[str, str, str, str, str]] = set()
    for item in items:
        chave = (getattr(item, "loja_numero", "") or "", getattr(item, "loja_nome", "") or "", item.cnpj or "", item.matricula or "", item.numero_pedido or "")
        if chave in lojas_vistas:
            continue
        lojas_vistas.add(chave)
        obs_loja = "OK" if (chave[0] or chave[1]) else "Loja não identificada no PDF/e-mail"
        ws_lojas.append([chave[0], chave[1], chave[2], chave[3], chave[4], item.arquivo, obs_loja])
    if not items:
        ws_lojas.append(["", "", "", "", "", "", "Sem itens para mapear loja."])

    ws_errors.append(["Arquivo", "Tipo", "Detalhe", "Pagina", "Linha"])
    for issue in issues:
        ws_errors.append([issue.arquivo, issue.tipo, issue.detalhe, issue.pagina or "", issue.linha or ""])

    duplicate_orders = sorted(set(duplicate_orders or []))
    ws_resume.append(["Indicador", "Valor"])
    ws_resume.append(["Modelo", title])
    ws_resume.append(["Total de PDFs", total_pdfs])
    ws_resume.append(["Pedidos com sucesso", success_count])
    ws_resume.append(["Itens extraidos", len(items)])
    ws_resume.append(["Erros / alertas", len(issues)])
    ws_resume.append(["Linhas validas da base BH", mapping_result.valid_count])
    ws_resume.append(["Conflitos tratados na base BH", len(mapping_result.conflicts)])
    ws_resume.append(["Pedidos repetidos identificados", len(duplicate_orders)])

    ws_duplicates.append(["Numero_Pedido", "Status"])
    if duplicate_orders:
        for pedido in duplicate_orders:
            ws_duplicates.append([pedido, "REPETIDO"])
    else:
        ws_duplicates.append(["-", "SEM_PEDIDOS_REPETIDOS"])

    ws_metadata.append(["Chave", "Valor"])
    for key, value in (metadata or {}).items():
        ws_metadata.append([key, value])

    if ws_status is not None:
        ws_status.append(["Numero_Pedido", "Status_Consolidado"])
        for pedido in sorted(order_status_map or {}):
            ws_status.append([pedido, (order_status_map or {}).get(pedido, "OK")])

    ws_cadastrar.append(["CNPJ", "Matrícula Encontrada", "Status", "Rede/Layout", "Nº do Pedido", "Observação"])
    if cnpj_pendentes:
        for cnpj, pedidos in sorted(cnpj_pendentes.items()):
            pedidos_texto = " | ".join(sorted(p for p in pedidos if p))
            ws_cadastrar.append([cnpj, "", "A CADASTRAR", "Rede BH", pedidos_texto, "Cadastrar de/para CNPJ x matrícula antes da etapa final."])
    else:
        ws_cadastrar.append(["", "", "OK", "Rede BH", "", "Nenhum CNPJ pendente de cadastro."])

    for sheet in wb.worksheets:
        apply_sheet_style(sheet)
    for row in ws_data.iter_rows(min_row=2, max_col=11):
        for cell in [row[1], row[2], row[3], row[4], row[9] if len(row) > 9 else row[4]]:
            textify_cell(cell)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    validar_workbook_salvo(output_file)


def build_main_pushes_workbook(output_file: Path, pedidos_dir: Path, lote_date: str, logger: logging.Logger | None = None) -> None:
    occurrences = index_existing_order_occurrences(pedidos_dir)
    occurrence_summary = {
        pedido: {"count": len(items), "pushes": sorted({item.push_label for item in items if item.push_label})}
        for pedido, items in occurrences.items()
    }
    push_rows: dict[str, list[list[str | int]]] = {}
    push_duplicates: dict[str, list[list[str]]] = {}
    push_summary: dict[str, dict[str, object]] = {}
    main_rows: list[list[str | int]] = []
    main_duplicate_rows: list[list[str]] = []

    for workbook_path in sorted(pedidos_dir.rglob("*.xlsx")):
        metadata = read_workbook_metadata(workbook_path)
        push_label = resolve_push_label(workbook_path, metadata)
        items, issues = read_items_and_issues_from_workbook(workbook_path)
        info = push_summary.setdefault(push_label, {"arquivos": set(), "pedidos": set(), "itens": 0, "processed_at": metadata.get("ProcessedAt", "")})
        info["arquivos"].add(workbook_path.name)
        for item in items:
            pedido_info = occurrence_summary.get(item.numero_pedido, {"count": 0, "pushes": []})
            duplicate_status = "DUPLICADO" if pedido_info.get("count", 0) > 1 else "OK"
            duplicate_origin = ", ".join(pedido_info.get("pushes", [])) if duplicate_status == "DUPLICADO" else "-"
            row = [item.cnpj, item.matricula, item.sku, item.qtd, item.numero_pedido, push_label, item.arquivo, duplicate_status, duplicate_origin]
            push_rows.setdefault(push_label, []).append(row)
            main_rows.append(row)
            info["pedidos"].add(item.numero_pedido)
            info["itens"] = int(info["itens"]) + 1
        for issue in issues:
            if issue.tipo not in {"PEDIDO_REPETIDO_PUSH", "PEDIDO_REPETIDO_HISTORICO"}:
                continue
            pedido = extract_order_number_from_detail(issue.detalhe) or metadata.get("NumeroPedido", "")
            row = [push_label, pedido, issue.tipo, issue.arquivo, issue.detalhe]
            push_duplicates.setdefault(push_label, []).append(row)
            main_duplicate_rows.append(row)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "RESUMO_PUSHES"
    ws_summary.append(["Push", "Arquivos", "Pedidos_Unicos", "Itens", "Pedidos_Duplicados", "Processado_Em", "Data_Lote"])
    for push_label in sorted(push_summary.keys(), key=push_sort_key):
        info = push_summary[push_label]
        duplicate_count = len({row[1] for row in push_duplicates.get(push_label, []) if len(row) > 1 and row[1] not in ("", "-")})
        ws_summary.append([push_label, len(info["arquivos"]), len(info["pedidos"]), int(info["itens"]), duplicate_count, str(info.get("processed_at", "")), lote_date])

    headers = ["CNPJ", "Matricula", "SKU", "Qtd", "Numero_Pedido", "Push", "Arquivo", "Status_Duplicidade", "Origem_Duplicidade"]
    ws_main = wb.create_sheet("MAIN_PEDIDOS")
    ws_main.append(headers)
    for row in main_rows:
        ws_main.append(row)

    ws_main_dup = wb.create_sheet("MAIN_DUPLICADAS")
    ws_main_dup.append(["Push", "Numero_Pedido", "Tipo_Alerta", "Arquivo", "Detalhe"])
    if main_duplicate_rows:
        for row in main_duplicate_rows:
            ws_main_dup.append(row)
    else:
        ws_main_dup.append(["-", "-", "SEM_DUPLICADAS", "-", "Nenhum alerta de duplicidade."])

    for push_label in sorted(push_summary.keys(), key=push_sort_key):
        ws_push = wb.create_sheet(safe_sheet_title(push_label))
        ws_push.append(headers)
        for row in push_rows.get(push_label, []):
            ws_push.append(row)

    for sheet in wb.worksheets:
        apply_sheet_style(sheet)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    validar_workbook_salvo(output_file)
    if logger:
        logger.info("[BH][PUSH] Main reconstruido com %s push(es).", len(push_summary))


def build_executive_log(
    output_file: Path,
    lote_date: str,
    lote_dir: Path,
    mapping_result: MappingLoadResult,
    parse_results: List[ParseResult],
    total_pdfs: int,
    total_items: int,
    total_errors: int,
    success_count: int,
    duplicate_orders: List[str],
    duplicate_details: List[str],
    consolidado_file: Path,
    main_pushes_file: Path,
    push_label: str,
) -> None:
    issue_counter = Counter(issue.tipo for result in parse_results for issue in result.issues)
    lines = [
        "ROBO KOF - REDE BH - LOG EXECUTIVO",
        "=" * 100,
        f"Data do lote: {lote_date}",
        f"Push atual: {push_label}",
        f"Pasta da data: {lote_dir}",
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        f"Consolidado geral: {consolidado_file}",
        f"Main de pushes: {main_pushes_file}",
        "",
        "RESUMO GERAL",
        "-" * 100,
        f"PDFs recebidos: {total_pdfs}",
        f"Pedidos com sucesso: {success_count}",
        f"Itens extraidos: {total_items}",
        f"Erros / alertas: {total_errors}",
        f"Linhas validas da base BH: {mapping_result.valid_count}",
        f"Pedidos repetidos identificados: {len(duplicate_orders)}",
    ]
    if duplicate_orders:
        lines.extend(["", "PEDIDOS REPETIDOS", "-" * 100])
        lines.extend([f"Pedido repetido: {pedido}" for pedido in duplicate_orders])
    if duplicate_details:
        lines.extend(["", "DETALHE DOS DUPLICADOS", "-" * 100])
        lines.extend([f"- {detail}" for detail in duplicate_details])
    if issue_counter:
        lines.extend(["", "TIPOS DE ALERTA", "-" * 100])
        lines.extend([f"{issue_type}: {qty}" for issue_type, qty in issue_counter.most_common()])
    lines.extend(["", "DETALHE POR PDF", "-" * 100])
    for index, result in enumerate(parse_results, start=1):
        status = "OK" if result.items and result.numero_pedido and result.cnpj and result.matricula else "ATENCAO"
        if not result.layout_valid:
            status = "BLOQUEADO_LAYOUT_INVALIDO"
        lines.extend([
            f"[{index}/{len(parse_results)}] {result.pdf_name}",
            f"Status: {status}",
            f"Pedido: {result.numero_pedido or 'NAO IDENTIFICADO'}",
            f"CNPJ: {result.cnpj or 'NAO IDENTIFICADO'}",
            f"Matricula: {result.matricula or 'NAO IDENTIFICADA'}",
            f"Itens extraidos: {len(result.items)}",
            f"Alertas: {len(result.issues)}",
            f"Excel individual: {result.output_file or 'NAO GERADO'}",
        ])
        for issue in result.issues[:10]:
            lines.append(f"  - {issue.tipo}: {issue.detalhe}")
        lines.append("-" * 100)
    output_file.write_text("\n".join(lines), encoding="utf-8")


def build_queue_validation_workbook(output_file: Path, df: pd.DataFrame, alerts: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "FILA_KOF"
    ws.append(list(config.FILA_COLUMNS))
    for _, row in df.iterrows():
        ws.append([row.get(col, "") for col in config.FILA_COLUMNS])
    ws_alerts = wb.create_sheet("ALERTAS")
    ws_alerts.append(["Tipo", "Mensagem"])
    if alerts:
        for alert in alerts:
            ws_alerts.append(["ALERTA", alert])
    else:
        ws_alerts.append(["OK", "Nenhum alerta para a fila KOF gerada."])
    for sheet in wb.worksheets:
        apply_sheet_style(sheet)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    validar_workbook_salvo(output_file)


def build_result_zip(zip_file: Path, lote_dir: Path) -> None:
    with zipfile.ZipFile(zip_file, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in sorted(lote_dir.rglob("*")):
            if path == zip_file or path.is_dir():
                continue
            zf.write(path, arcname=path.relative_to(lote_dir))


def textify_cell(cell) -> None:
    cell.value = "" if cell.value is None else str(cell.value)
    cell.number_format = "@"


def apply_sheet_style(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="C00000")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="2B2B2B")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def sanitize_filename(value: str, limit: int = 48) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", value or "arquivo").strip("_") or "arquivo"
    if len(text) <= limit:
        return text
    digest = hashlib.sha1(text.encode("utf-8", errors="ignore")).hexdigest()[:8]
    return f"{text[: max(8, limit - 9)].rstrip('_')}_{digest}"


def safe_sheet_title(value: str) -> str:
    return re.sub(r"[\\/*?:\[\]]+", "_", value or "Sheet")[:31] or "Sheet"


def push_sort_key(push_label: str) -> tuple[int, int | str]:
    match = re.fullmatch(r"PUSH_(\d{3})", push_label)
    if match:
        return (1, int(match.group(1)))
    if push_label == "PUSH_LEGADO":
        return (0, 0)
    return (2, push_label)


def extract_order_number_from_detail(detail: str) -> str:
    match = re.search(r"Pedido\s+(\d+)", detail or "", re.IGNORECASE)
    return match.group(1) if match else ""


def _safe_int(value) -> int:
    try:
        return int(value) if value not in (None, "") else 0
    except Exception:
        return 0


def _safe_optional_int(value) -> int | None:
    try:
        return int(value) if value not in (None, "") else None
    except Exception:
        return None
