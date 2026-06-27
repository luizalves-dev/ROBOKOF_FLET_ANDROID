# -*- coding: utf-8 -*-
"""
Layout GRANCOFFEE / SPAL - Robô KOF
Autor: Kauê Melo

Objetivo:
- Processar e-mails .msg da rede Grancoffee com anexos Excel .xlsm/.xlsx.
- Usar a DATA DA REMESSA do CORPO DO E-MAIL como data oficial, cruzando por matrícula + pedido.
- Extrair os itens dos anexos da aba "Pedido" com rastreabilidade completa.
- Gerar primeiro o Excel de validação, sem gerar TXT/fila KOF automaticamente.

Regra de negócio homologada:
- Rede: Grancoffee.
- Entrada principal: .msg do Outlook com assunto semelhante a "Pedidos Grancoffee | SPAL | dd.mm.aaaa".
- Anexos válidos: PEDIDO SPAL - <regional> <data>.xlsm/.xlsx.
- Corpo do e-mail possui a tabela: REGIONAL, MATRÍCULA, TIPO DE FATURAMENTO, PEDIDO, DATA DA REMESSA.
- A data oficial é a DATA DA REMESSA do corpo do e-mail, não apenas a DATA DE ENTREGA do anexo.
- Cruzamento principal: matrícula + pedido. Fallback controlado: matrícula/regional com alerta.
- SKU oficial: coluna SKU da aba Pedido.
- Quantidade final para Modelo Robô KOF: QTDE FARDO.
- EAN, Unidades por Fardo e QTDE Unitário são rastreabilidade/auditoria.
- Não há conversão unidade -> caixaria neste layout; a quantidade do pedido já vem em fardos/caixas.
- CNPJ x matrícula é opcional; se não houver CNPJ, o pedido continua no modelo e fica com alerta.
"""
from __future__ import annotations

import csv
import datetime as _dt
import html as _html
import json
import os
import re
import shutil
import struct
import tempfile
import traceback
import unicodedata
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except Exception as exc:  # pragma: no cover
    raise RuntimeError("O layout Grancoffee precisa da biblioteca openpyxl instalada.") from exc


LAYOUT_NOME = "GRANCOFFEE"
LAYOUT_ALIASES = ("GRANCOFFEE", "GRAN COFFEE", "PEDIDOS GRANCOFFEE", "PEDIDO SPAL")
EXTENSOES_EXCEL = {".xlsx", ".xlsm", ".xls"}
EXTENSOES_MSG = {".msg"}


@dataclass
class LogEntry:
    nivel: str
    etapa: str
    arquivo: str
    mensagem: str
    detalhe: str = ""
    timestamp: _dt.datetime = field(default_factory=_dt.datetime.now)


@dataclass
class RemessaEmail:
    regional: str
    regional_original: str
    matricula: str
    tipo_faturamento: str
    pedido: str
    data_remessa: Optional[_dt.date]
    origem: str

    @property
    def chave(self) -> Tuple[str, str]:
        return (somente_digitos(self.matricula), somente_digitos(self.pedido))


@dataclass
class ItemGrancoffee:
    arquivo_origem: str
    aba: str
    linha: int
    regional: str
    matricula: str
    cnpj: str
    pedido: str
    data_entrega_anexo: Optional[_dt.date]
    data_remessa_email: Optional[_dt.date]
    tipo_faturamento_email: str
    codigo_grancoffee: str
    sku: str
    ean: str
    unidades_por_fardo: Optional[float]
    descricao: str
    qtd_fardo: Optional[float]
    qtd_unitario: Optional[float]
    status: str
    alerta: str


class GrancoffeeErro(Exception):
    """Erro controlado do layout Grancoffee."""


# ---------------------------------------------------------------------------
# Normalização / utilitários
# ---------------------------------------------------------------------------


def sem_acentos(texto: object) -> str:
    s = "" if texto is None else str(texto)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s


def norm(texto: object) -> str:
    s = sem_acentos(texto).upper()
    s = re.sub(r"[\r\n\t]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def somente_digitos(valor: object) -> str:
    return re.sub(r"\D+", "", "" if valor is None else str(valor))


def texto_limpo(valor: object) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    if isinstance(valor, int):
        return str(valor)
    return str(valor).strip()


def limpar_corpo_email(corpo: object) -> str:
    """Normaliza corpo de e-mail do Outlook/MSG/HTML para leitura da Grancoffee.

    O problema mais comum neste layout é a data oficial estar no corpo do e-mail,
    enquanto o anexo também possui uma DATA DE ENTREGA. Esta função prepara texto
    puro e HTML para que a tabela REGIONAL/MATRÍCULA/PEDIDO/DATA DA REMESSA seja
    lida de forma consistente.
    """
    texto = "" if corpo is None else str(corpo)
    if not texto:
        return ""
    texto = texto.replace("\x00", " ").replace("\xa0", " ")
    # Quando vier HTMLBody, preserva separação de células/linhas antes de remover tags.
    if "<" in texto and ">" in texto:
        texto = re.sub(r"(?is)<\s*(br|/p|/div|/tr|/td|/th|/li)\b[^>]*>", "\n", texto)
        texto = re.sub(r"(?is)<\s*(p|div|tr|td|th|li)\b[^>]*>", "\n", texto)
        texto = re.sub(r"(?is)<style.*?</style>|<script.*?</script>", " ", texto)
        texto = re.sub(r"(?is)<[^>]+>", " ", texto)
    texto = _html.unescape(texto)
    texto = re.sub(r"[ \t;]+", " ", texto)
    texto = re.sub(r" *\n+ *", "\n", texto)
    texto = re.sub(r"\n{3,}", "\n\n", texto)
    return texto.strip()


def numero_limpo(valor: object) -> Optional[float]:
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip()
    if not s:
        return None
    s = s.replace(".", "").replace(",", ".") if re.search(r"\d,\d", s) else s
    s = re.sub(r"[^0-9.\-]", "", s)
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def quantidade_para_saida(valor: Optional[float]) -> object:
    if valor is None:
        return ""
    if float(valor).is_integer():
        return int(valor)
    return round(float(valor), 4)


def excel_serial_para_data(valor: object) -> Optional[_dt.date]:
    if valor is None or valor == "":
        return None
    if isinstance(valor, _dt.datetime):
        return valor.date()
    if isinstance(valor, _dt.date):
        return valor
    if isinstance(valor, (int, float)):
        # Sistema padrão 1900 do Excel.
        return (_dt.datetime(1899, 12, 30) + _dt.timedelta(days=float(valor))).date()
    s = str(valor).strip()
    m = re.search(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})", s)
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100:
        y += 2000
    try:
        return _dt.date(y, mo, d)
    except ValueError:
        return None


def fmt_data(data: Optional[_dt.date]) -> str:
    """Formata datas da Grancoffee no padrão aceito pelo Robô KOF: dd.mm.aaaa."""
    return data.strftime("%d.%m.%Y") if data else ""


def nome_seguro(nome: str) -> str:
    s = sem_acentos(nome)
    s = re.sub(r"[^A-Za-z0-9_. -]+", "_", s).strip(" ._")
    return s.replace(" ", "_") or f"arquivo_{uuid.uuid4().hex[:8]}"


def regional_canonica(valor: object) -> str:
    s = norm(valor)
    if not s:
        return ""
    if "HORT" in s or s in {"CTL", "SP HORT", "SP-HORT", "SP HORTOLANDIA", "SP - HORTOLANDIA"}:
        return "SP-HORT"
    if s in {"POA", "PORTO ALEGRE"}:
        return "RS"
    if s in {"CTBA", "CURITIBA"}:
        return "PR"
    if s in {"SC", "PR", "SP", "MG", "RS"}:
        return s
    return s.replace(" ", "-")


def parece_regional(linha: str) -> bool:
    s = norm(linha)
    if not s:
        return False
    if s in {"PR", "SC", "SP", "MG", "RS", "POA", "CTBA", "CTL", "SP-HORT"}:
        return True
    return "HORT" in s or "SUMARE" in s or "CURITIBA" in s or "PORTO ALEGRE" in s


def parece_matricula(linha: str) -> bool:
    d = somente_digitos(linha)
    return len(d) >= 7 and len(d) <= 12


def parece_pedido(linha: str) -> bool:
    d = somente_digitos(linha)
    return len(d) >= 6 and len(d) <= 12


# ---------------------------------------------------------------------------
# Identificação do layout
# ---------------------------------------------------------------------------


def identificar_grancoffee(caminho: str, texto_previo: str = "") -> bool:
    """Retorna True se o arquivo/texto aparenta ser da rede Grancoffee.

    Pode ser usado no roteador geral do Robô KOF antes do processamento.
    """
    nome = norm(Path(caminho).name)
    texto = norm(texto_previo)
    ext = Path(caminho).suffix.lower()
    if any(alias in nome for alias in LAYOUT_ALIASES) or any(alias in texto for alias in LAYOUT_ALIASES):
        return True
    if ext in EXTENSOES_MSG and "GRANCOFFEE" in nome:
        return True
    if ext in EXTENSOES_EXCEL and "PEDIDO SPAL" in nome:
        return True
    return False




def _candidatos_metadata_outlook(caminho_excel: str) -> List[Path]:
    """Localiza sidecars `.outlook.json` salvos pelo importador do Outlook.

    A Grancoffee depende do corpo do e-mail para a data oficial de remessa.
    Quando o usuário importa somente o anexo pelo Outlook, o corpo fica no
    JSON de metadados ao lado do arquivo ou na pasta `_metadata_outlook`.
    """
    caminho = Path(caminho_excel)
    candidatos = [
        Path(str(caminho) + ".outlook.json"),
        caminho.with_suffix(caminho.suffix + ".outlook.json"),
    ]
    meta_dir = caminho.parent / "_metadata_outlook"
    if meta_dir.exists():
        candidatos.extend(sorted(meta_dir.glob("*.json")))
    # Remove duplicados preservando ordem.
    vistos = set()
    unicos: List[Path] = []
    for cand in candidatos:
        chave = str(cand.resolve()) if cand.exists() else str(cand)
        if chave not in vistos:
            vistos.add(chave)
            unicos.append(cand)
    return unicos


def extrair_remessas_metadata_outlook(caminho_excel: str, logs: List[LogEntry]) -> List[RemessaEmail]:
    """Extrai remessas do corpo do e-mail gravado no sidecar do Outlook.

    Isso corrige o fluxo onde o Outlook importa apenas o anexo `.xlsm/.xlsx`,
    mas a data oficial precisa vir do corpo do e-mail, não da data interna do anexo.
    """
    arquivo = Path(caminho_excel).name
    remessas: List[RemessaEmail] = []
    for meta in _candidatos_metadata_outlook(caminho_excel):
        if not meta.exists() or not meta.is_file():
            continue
        try:
            payload = json.loads(meta.read_text(encoding="utf-8", errors="ignore"))
        except Exception as exc:
            logs.append(LogEntry("WARN", "OUTLOOK", arquivo, "Metadata Outlook inválida para Grancoffee", f"{meta.name}: {exc}"))
            continue
        arquivo_salvo = str(payload.get("arquivo_salvo", "") or "")
        nome_original = str(payload.get("nome_original", "") or "")
        if arquivo_salvo and Path(arquivo_salvo).name != arquivo and Path(arquivo_salvo).resolve() != Path(caminho_excel).resolve():
            continue
        if nome_original and nome_seguro(nome_original) != nome_seguro(arquivo) and not arquivo_salvo:
            continue
        corpo = "\n".join(
            str(payload.get(chave, "") or "")
            for chave in ["corpo_email", "body", "email_body", "corpo", "corpo_html_email", "assunto", "email_assunto"]
        ).strip()
        if not corpo:
            logs.append(LogEntry("WARN", "OUTLOOK", arquivo, "Metadata Outlook sem corpo do e-mail", meta.name))
            continue
        extraidas = extrair_remessas_do_email(corpo, f"{arquivo}::{meta.name}", logs)
        if extraidas:
            remessas.extend(extraidas)
            logs.append(LogEntry("INFO", "OUTLOOK", arquivo, "Remessas Grancoffee extraídas do metadata Outlook", f"{len(extraidas)} remessa(s) | {meta.name}"))
    return remessas


# ---------------------------------------------------------------------------
# Leitura .msg: Outlook COM quando possível + fallback binário puro
# ---------------------------------------------------------------------------


def extrair_msg(caminho_msg: str, pasta_temp: str, logs: List[LogEntry]) -> Tuple[str, List[str]]:
    """Extrai corpo e anexos válidos de um .msg.

    Em Windows com Outlook instalado, tenta COM primeiro. Se falhar, usa fallback
    binário puro, importante para rodar em servidor/ambiente sem Outlook.
    """
    try:
        corpo, anexos = _extrair_msg_via_outlook(caminho_msg, pasta_temp, logs)
        if corpo or anexos:
            logs.append(LogEntry("INFO", "EMAIL", Path(caminho_msg).name, "Arquivo .msg lido via Outlook COM", f"Anexos extraídos: {len(anexos)}"))
            return corpo, anexos
    except Exception as exc:
        logs.append(LogEntry("WARN", "EMAIL", Path(caminho_msg).name, "Não foi possível ler .msg via Outlook COM; usando fallback binário", str(exc)))
    corpo, anexos = _extrair_msg_fallback_binario(caminho_msg, pasta_temp, logs)
    logs.append(LogEntry("INFO", "EMAIL", Path(caminho_msg).name, "Arquivo .msg lido por fallback binário", f"Anexos extraídos: {len(anexos)}"))
    return corpo, anexos


def _extrair_msg_via_outlook(caminho_msg: str, pasta_temp: str, logs: List[LogEntry]) -> Tuple[str, List[str]]:
    """Integração COM removida na edição Android; o fallback binário é usado."""
    raise RuntimeError("Leitura via Outlook COM não está disponível na edição Android.")


class _OleCF:
    """Leitor mínimo de Compound File Binary para .msg sem dependência externa."""

    ENDOFCHAIN = 0xFFFFFFFE
    FREESECT = 0xFFFFFFFF
    NOSTREAM = 0xFFFFFFFF

    def __init__(self, caminho: str):
        self.caminho = caminho
        with open(caminho, "rb") as fh:
            self.data = fh.read()
        self._parse_header()
        self._load_fat()
        self._load_directory()

    def _parse_header(self) -> None:
        h = self.data[:512]
        if h[:8] != bytes.fromhex("D0CF11E0A1B11AE1"):
            raise GrancoffeeErro("Arquivo .msg inválido ou não é Compound File Binary.")
        self.sector_size = 1 << struct.unpack_from("<H", h, 30)[0]
        self.mini_sector_size = 1 << struct.unpack_from("<H", h, 32)[0]
        self.num_fat_sectors = struct.unpack_from("<I", h, 44)[0]
        self.first_dir_sector = struct.unpack_from("<I", h, 48)[0]
        self.mini_cutoff = struct.unpack_from("<I", h, 56)[0]
        self.first_minifat_sector = struct.unpack_from("<I", h, 60)[0]
        self.first_difat_sector = struct.unpack_from("<I", h, 68)[0]
        self.num_difat_sectors = struct.unpack_from("<I", h, 72)[0]
        self.header_difat = list(struct.unpack_from("<109I", h, 76))

    def _sector(self, sid: int) -> bytes:
        ini = (sid + 1) * self.sector_size
        return self.data[ini : ini + self.sector_size]

    def _chain(self, start: int, fat: Optional[List[int]] = None) -> List[int]:
        fat = fat or self.fat
        if start in {self.ENDOFCHAIN, self.FREESECT, self.NOSTREAM}:
            return []
        out: List[int] = []
        seen = set()
        sid = start
        while sid not in {self.ENDOFCHAIN, self.FREESECT, self.NOSTREAM} and sid < len(fat) and sid not in seen:
            out.append(sid)
            seen.add(sid)
            sid = fat[sid]
        return out

    def _read_chain(self, start: int, fat: Optional[List[int]] = None, mini: bool = False, size: Optional[int] = None) -> bytes:
        if mini:
            blob = b"".join(self.mini_stream[s * self.mini_sector_size : (s + 1) * self.mini_sector_size] for s in self._chain(start, fat))
        else:
            blob = b"".join(self._sector(s) for s in self._chain(start, fat))
        return blob[:size] if size is not None else blob

    def _load_fat(self) -> None:
        difat = [x for x in self.header_difat if x not in {self.FREESECT, self.ENDOFCHAIN}]
        sid = self.first_difat_sector
        for _ in range(self.num_difat_sectors):
            entries = list(struct.unpack("<%dI" % (self.sector_size // 4), self._sector(sid)))
            difat.extend(x for x in entries[:-1] if x not in {self.FREESECT, self.ENDOFCHAIN})
            sid = entries[-1]
            if sid == self.ENDOFCHAIN:
                break
        self.fat: List[int] = []
        for sector_id in difat[: self.num_fat_sectors]:
            self.fat.extend(struct.unpack("<%dI" % (self.sector_size // 4), self._sector(sector_id)))

    def _load_directory(self) -> None:
        dir_bytes = self._read_chain(self.first_dir_sector)
        self.entries: List[Dict[str, object]] = []
        for offset in range(0, len(dir_bytes), 128):
            entry = dir_bytes[offset : offset + 128]
            if len(entry) < 128:
                continue
            name_len = struct.unpack_from("<H", entry, 64)[0]
            name = entry[: max(0, name_len - 2)].decode("utf-16le", "ignore") if name_len >= 2 else ""
            typ = entry[66]
            left, right, child = struct.unpack_from("<III", entry, 68)
            start = struct.unpack_from("<I", entry, 116)[0]
            size = struct.unpack_from("<Q", entry, 120)[0]
            self.entries.append({"id": offset // 128, "name": name, "type": typ, "left": left, "right": right, "child": child, "start": start, "size": size})
        root = self.entries[0]
        self.mini_stream = self._read_chain(int(root["start"]), size=int(root["size"])) if int(root["start"]) not in {self.ENDOFCHAIN, self.FREESECT, self.NOSTREAM} else b""
        self.minifat: List[int] = []
        for sid in self._chain(self.first_minifat_sector):
            self.minifat.extend(struct.unpack("<%dI" % (self.sector_size // 4), self._sector(sid)))

    def stream(self, idx: int) -> bytes:
        e = self.entries[idx]
        if int(e["type"]) == 2 and int(e["size"]) < self.mini_cutoff:
            return self._read_chain(int(e["start"]), self.minifat, mini=True, size=int(e["size"]))
        return self._read_chain(int(e["start"]), size=int(e["size"]))

    def paths(self) -> List[Tuple[int, str, Dict[str, object]]]:
        out: List[Tuple[int, str, Dict[str, object]]] = []

        def walk(idx: int, prefix: List[str]) -> None:
            if idx in {self.FREESECT, self.NOSTREAM} or idx >= len(self.entries):
                return
            e = self.entries[idx]
            walk(int(e["left"]), prefix)
            path = prefix + [str(e["name"])]
            out.append((idx, "/".join(path), e))
            if int(e["child"]) not in {self.FREESECT, self.NOSTREAM}:
                walk(int(e["child"]), path)
            walk(int(e["right"]), prefix)

        walk(int(self.entries[0]["child"]), [])
        return out


def _decode_utf16(blob: bytes) -> str:
    return blob.decode("utf-16le", "ignore").rstrip("\x00")


def _extrair_msg_fallback_binario(caminho_msg: str, pasta_temp: str, logs: List[LogEntry]) -> Tuple[str, List[str]]:
    ole = _OleCF(caminho_msg)
    streams = {path: (idx, e) for idx, path, e in ole.paths() if int(e["type"]) == 2}
    corpo = ""
    # PR_BODY_W Unicode: __substg1.0_1000001F
    body_key = "__substg1.0_1000001F"
    for path, (idx, _e) in streams.items():
        if path.endswith(body_key):
            corpo = _decode_utf16(ole.stream(idx))
            break
    anexos: List[str] = []
    anexos_dirs = sorted(set(path.split("/")[0] for path in streams if path.startswith("__attach")))
    for ad in anexos_dirs:
        nome = ""
        for prop in ("3707001F", "3001001F", "3704001F"):
            key = f"{ad}/__substg1.0_{prop}"
            if key in streams:
                nome = _decode_utf16(ole.stream(streams[key][0]))
                if nome:
                    break
        data_key = f"{ad}/__substg1.0_37010102"
        if not nome or data_key not in streams:
            continue
        ext = Path(nome).suffix.lower()
        if ext not in EXTENSOES_EXCEL:
            logs.append(LogEntry("INFO", "OUTLOOK", nome_seguro(nome), "Anexo ignorado por extensão não permitida", f"MSG: {Path(caminho_msg).name}"))
            continue
        destino = os.path.join(pasta_temp, nome_seguro(nome))
        with open(destino, "wb") as fh:
            fh.write(ole.stream(streams[data_key][0]))
        anexos.append(destino)
        logs.append(LogEntry("INFO", "OUTLOOK", Path(destino).name, "Anexo extraído do e-mail/.msg", f"MSG: {Path(caminho_msg).name}"))
    return corpo, anexos


# ---------------------------------------------------------------------------
# Corpo do e-mail: datas oficiais de remessa
# ---------------------------------------------------------------------------


def _extrair_remessas_linhas_compactas(linhas: List[str], origem: str, logs: List[LogEntry], vistos: set) -> List[RemessaEmail]:
    """Lê linhas compactas/tabulares do corpo do e-mail.

    Exemplo aceito:
    PR  7120041366  ENTREGA 24H  152489216  01/06/2026

    Esse fallback é importante quando o Outlook cola a tabela em uma única linha
    por regional, ou quando o corpo vem de HTML/MSG com separadores diferentes.
    """
    remessas: List[RemessaEmail] = []
    for linha in linhas:
        linha_original = linha.strip()
        if not linha_original:
            continue
        linha_norm = norm(linha_original)
        if "ENTREGA" not in linha_norm:
            continue
        data_match = re.search(r"(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})", linha_original)
        if not data_match:
            continue
        numeros = re.findall(r"\b\d{6,12}\b", linha_original)
        if len(numeros) < 2:
            continue
        matricula = somente_digitos(numeros[0])
        pedido = somente_digitos(numeros[1])
        if not parece_matricula(matricula) or not parece_pedido(pedido):
            continue
        tipo_match = re.search(r"ENTREGA\s*\d{1,3}\s*H", linha_original, flags=re.I)
        tipo = tipo_match.group(0).upper().replace("  ", " ") if tipo_match else ""
        antes_matricula = linha_original.split(numeros[0], 1)[0].strip(" -:;|\t")
        regional_original = antes_matricula or ""
        regional = regional_canonica(regional_original)
        data = excel_serial_para_data(data_match.group(1))
        chave = (matricula, pedido)
        if chave in vistos:
            continue
        vistos.add(chave)
        rem = RemessaEmail(
            regional=regional,
            regional_original=regional_original,
            matricula=matricula,
            tipo_faturamento=tipo,
            pedido=pedido,
            data_remessa=data,
            origem=origem,
        )
        remessas.append(rem)
        logs.append(LogEntry("INFO", "EMAIL", origem, "Remessa capturada do corpo do e-mail", f"Regional {rem.regional} | Matrícula {rem.matricula} | Pedido {rem.pedido} | Remessa {fmt_data(rem.data_remessa)} | Tipo {rem.tipo_faturamento}"))
    return remessas


def extrair_remessas_do_email(corpo: str, origem: str, logs: List[LogEntry]) -> List[RemessaEmail]:
    corpo = limpar_corpo_email(corpo)
    if not corpo:
        logs.append(LogEntry("WARN", "EMAIL", origem, "Corpo do e-mail vazio; não foi possível extrair datas oficiais", ""))
        return []
    linhas = [ln.strip() for ln in re.split(r"[\r\n]+", corpo) if ln and ln.strip()]
    # Começa depois de DATA DA REMESSA, quando existir.
    inicio = 0
    for i, linha in enumerate(linhas):
        if "DATA" in norm(linha) and "REMESSA" in norm(linha):
            inicio = i + 1
            break
    fim = len(linhas)
    for i in range(inicio, len(linhas)):
        if "OBSERV" in norm(linhas[i]) or "HORARIO" in norm(linhas[i]) or "FAVOR CONFIRMAR" in norm(linhas[i]):
            fim = i
            break
    bloco = linhas[inicio:fim]
    remessas: List[RemessaEmail] = []
    vistos = set()

    # Primeiro tenta linhas compactas/tabulares; se a tabela veio inteira em uma linha,
    # este caminho resolve sem depender de cada célula estar em uma linha separada.
    remessas.extend(_extrair_remessas_linhas_compactas(bloco, origem, logs, vistos))

    regional_atual = ""
    regional_original_atual = ""
    i = 0
    while i < len(bloco):
        linha = bloco[i]
        if parece_regional(linha):
            regional_atual = regional_canonica(linha)
            regional_original_atual = linha
            i += 1
            continue
        if parece_matricula(linha):
            matricula = somente_digitos(linha)
            tipo = ""
            pedido = ""
            data = None
            j = i + 1
            # Tipo de faturamento vem como ENTREGA 24H/48H.
            while j < len(bloco) and not tipo:
                if "ENTREGA" in norm(bloco[j]):
                    tipo = bloco[j].strip()
                    j += 1
                    break
                if parece_regional(bloco[j]):
                    break
                j += 1
            while j < len(bloco) and not pedido:
                if parece_pedido(bloco[j]) and not excel_serial_para_data(bloco[j]):
                    pedido = somente_digitos(bloco[j])
                    j += 1
                    break
                j += 1
            while j < len(bloco) and data is None:
                data = excel_serial_para_data(bloco[j])
                j += 1
            if matricula and pedido:
                chave = (matricula, pedido)
                if chave not in vistos:
                    vistos.add(chave)
                    rem = RemessaEmail(regional=regional_atual, regional_original=regional_original_atual if regional_atual else "", matricula=matricula, tipo_faturamento=tipo, pedido=pedido, data_remessa=data, origem=origem)
                    remessas.append(rem)
                    logs.append(LogEntry("INFO", "EMAIL", origem, "Remessa capturada do corpo do e-mail", f"Regional {rem.regional} | Matrícula {rem.matricula} | Pedido {rem.pedido} | Remessa {fmt_data(rem.data_remessa)} | Tipo {rem.tipo_faturamento}"))
            i = max(j, i + 1)
            continue
        i += 1
    logs.append(LogEntry("INFO", "EMAIL", origem, f"Remessas extraídas do corpo do e-mail: {len(remessas)}", ""))
    return remessas

# ---------------------------------------------------------------------------
# Anexo Excel Grancoffee
# ---------------------------------------------------------------------------


def _valor_ao_lado(ws, label_norm: str) -> object:
    for row in ws.iter_rows():
        for cell in row:
            if label_norm in norm(cell.value):
                return ws.cell(cell.row, cell.column + 1).value
    return None


def _encontrar_linha_cabecalho(ws) -> Tuple[int, Dict[str, int]]:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50)):
        valores = [norm(c.value) for c in row]
        if any(v == "SKU" for v in valores) and any("EAN" == v for v in valores) and any("QTDE" in v and "FARDO" in v for v in valores):
            cols: Dict[str, int] = {}
            for c in row:
                h = norm(c.value)
                if not h:
                    continue
                if "CODIGO" in h and ("GRAN" in h or "COFFE" in h or "COFFEE" in h):
                    cols["codigo_grancoffee"] = c.column
                elif h == "SKU":
                    cols["sku"] = c.column
                elif h == "EAN":
                    cols["ean"] = c.column
                elif "UNIDADES" in h and "FARDO" in h:
                    cols["unidades_por_fardo"] = c.column
                elif "DESCRICAO" in h and "ITEM" in h:
                    cols["descricao"] = c.column
                elif "QTDE" in h and "FARDO" in h:
                    cols["qtd_fardo"] = c.column
                elif "QTDE" in h and "UNIT" in h:
                    cols["qtd_unitario"] = c.column
            obrig = {"sku", "qtd_fardo"}
            if not obrig.issubset(cols):
                raise GrancoffeeErro(f"Cabeçalho da aba Pedido encontrado, mas faltam colunas obrigatórias: {sorted(obrig - set(cols))}")
            return row[0].row, cols
    raise GrancoffeeErro("Não encontrei o cabeçalho de itens da Grancoffee na aba Pedido.")


def ler_anexo_grancoffee(caminho_excel: str, remessas_por_chave: Dict[Tuple[str, str], RemessaEmail], remessas_por_matricula: Dict[str, List[RemessaEmail]], cnpj_por_matricula: Dict[str, str], logs: List[LogEntry]) -> List[ItemGrancoffee]:
    arquivo = Path(caminho_excel).name
    try:
        wb = load_workbook(caminho_excel, data_only=True, read_only=False)
    except Exception as exc:
        logs.append(LogEntry("ERROR", "EXCEL", arquivo, "Falha ao abrir anexo Excel", str(exc)))
        return []
    if "Pedido" not in wb.sheetnames:
        logs.append(LogEntry("ERROR", "EXCEL", arquivo, "Aba 'Pedido' não encontrada no anexo", ", ".join(wb.sheetnames)))
        return []
    ws = wb["Pedido"]
    matricula = somente_digitos(_valor_ao_lado(ws, "MATRICULA"))
    pedido = somente_digitos(_valor_ao_lado(ws, "PEDIDO"))
    data_entrega_anexo = excel_serial_para_data(_valor_ao_lado(ws, "DATA DE ENTREGA"))
    regional_anexo = regional_canonica(_valor_ao_lado(ws, "REGIONAL"))
    if not matricula or not pedido:
        logs.append(LogEntry("ERROR", "EXCEL", arquivo, "Cabeçalho do pedido sem matrícula ou número do pedido", f"Matrícula={matricula} | Pedido={pedido}"))
        return []
    remessa = remessas_por_chave.get((matricula, pedido))
    match_alerta = ""
    if not remessa:
        possiveis = remessas_por_matricula.get(matricula, [])
        if len(possiveis) == 1:
            remessa = possiveis[0]
            match_alerta = "Data do e-mail vinculada por fallback de matrícula; validar pedido."
        elif possiveis:
            # tenta regional também
            for cand in possiveis:
                if cand.regional == regional_anexo:
                    remessa = cand
                    match_alerta = "Data do e-mail vinculada por fallback de matrícula + regional; validar pedido."
                    break
    cnpj = cnpj_por_matricula.get(matricula, "")
    if not cnpj:
        logs.append(LogEntry("WARN", "CADASTRO", arquivo, "CNPJ não encontrado na base opcional matricula_cnpj.csv", f"Matrícula: {matricula}"))
    try:
        header_row, cols = _encontrar_linha_cabecalho(ws)
    except Exception as exc:
        logs.append(LogEntry("ERROR", "EXCEL", arquivo, "Falha ao localizar cabeçalho de itens", str(exc)))
        return []

    itens: List[ItemGrancoffee] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        sku = texto_limpo(ws.cell(row_idx, cols.get("sku", 0)).value if cols.get("sku") else "")
        qtd_fardo = numero_limpo(ws.cell(row_idx, cols.get("qtd_fardo", 0)).value if cols.get("qtd_fardo") else None)
        codigo = texto_limpo(ws.cell(row_idx, cols.get("codigo_grancoffee", 0)).value if cols.get("codigo_grancoffee") else "")
        descricao = texto_limpo(ws.cell(row_idx, cols.get("descricao", 0)).value if cols.get("descricao") else "")
        # Linha realmente vazia encerra/ignora.
        if not sku and qtd_fardo is None and not codigo and not descricao:
            continue
        ean = texto_limpo(ws.cell(row_idx, cols.get("ean", 0)).value if cols.get("ean") else "")
        unidades = numero_limpo(ws.cell(row_idx, cols.get("unidades_por_fardo", 0)).value if cols.get("unidades_por_fardo") else None)
        qtd_unitario = numero_limpo(ws.cell(row_idx, cols.get("qtd_unitario", 0)).value if cols.get("qtd_unitario") else None)
        alertas: List[str] = []
        status = "OK SEM CONVERSÃO"
        if not sku:
            status = "ALERTA"
            alertas.append("Linha com quantidade/descrição sem SKU; validar antes de gerar fila KOF.")
        if qtd_fardo is None:
            if qtd_unitario is not None and unidades:
                qtd_fardo = qtd_unitario / unidades
                alertas.append("QTDE FARDO calculada por fallback usando QTDE Unitário / Unidades por Fardo; validar.")
                status = "ALERTA"
            else:
                status = "ALERTA"
                alertas.append("Quantidade em fardo não encontrada.")
        if not remessa or not remessa.data_remessa:
            status = "ALERTA"
            alertas.append("Data oficial de remessa não encontrada no corpo do e-mail.")
        if match_alerta:
            status = "ALERTA"
            alertas.append(match_alerta)
        if not cnpj:
            status = "ALERTA"
            alertas.append("CNPJ não informado na base opcional.")
        item = ItemGrancoffee(
            arquivo_origem=arquivo,
            aba="Pedido",
            linha=row_idx,
            regional=regional_anexo or (remessa.regional if remessa else ""),
            matricula=matricula,
            cnpj=cnpj,
            pedido=pedido,
            data_entrega_anexo=data_entrega_anexo,
            data_remessa_email=remessa.data_remessa if remessa else None,
            tipo_faturamento_email=remessa.tipo_faturamento if remessa else "",
            codigo_grancoffee=codigo,
            sku=sku,
            ean=ean,
            unidades_por_fardo=unidades,
            descricao=descricao,
            qtd_fardo=qtd_fardo,
            qtd_unitario=qtd_unitario,
            status=status,
            alerta=" | ".join(dict.fromkeys(alertas)),
        )
        itens.append(item)
    logs.append(LogEntry("INFO", "EXCEL", arquivo, "Anexo Grancoffee processado", f"Itens extraídos: {len(itens)} | Matrícula {matricula} | Pedido {pedido} | Regional {regional_anexo}"))
    return itens


# ---------------------------------------------------------------------------
# Base opcional CNPJ x matrícula
# ---------------------------------------------------------------------------


def carregar_cnpj_por_matricula(caminho_base: Optional[str], logs: List[LogEntry]) -> Dict[str, str]:
    if not caminho_base or not os.path.exists(caminho_base):
        logs.append(LogEntry("INFO", "CONFIG", "matricula_cnpj.csv", "Base matrícula x CNPJ não informada; CNPJ ficará em branco com alerta", ""))
        return {}
    base: Dict[str, str] = {}
    try:
        with open(caminho_base, "r", encoding="utf-8-sig", newline="") as fh:
            amostra = fh.read(4096)
            fh.seek(0)
            dialect = csv.Sniffer().sniff(amostra, delimiters=";,\t|") if amostra else csv.excel
            reader = csv.DictReader(fh, dialect=dialect)
            campos = {norm(c): c for c in (reader.fieldnames or [])}
            col_mat = campos.get("MATRICULA") or campos.get("MATRÍCULA") or campos.get("CLIENTE")
            col_cnpj = campos.get("CNPJ") or campos.get("CNPJ LOJA")
            if not col_mat or not col_cnpj:
                raise GrancoffeeErro("CSV de CNPJ precisa ter colunas Matrícula e CNPJ.")
            for row in reader:
                mat = somente_digitos(row.get(col_mat, ""))
                cnpj = somente_digitos(row.get(col_cnpj, ""))
                if mat and cnpj:
                    base[mat] = cnpj.zfill(14)
        logs.append(LogEntry("INFO", "CONFIG", Path(caminho_base).name, f"Base matrícula x CNPJ carregada: {len(base)} registros", ""))
    except Exception as exc:
        logs.append(LogEntry("WARN", "CONFIG", Path(caminho_base).name, "Falha ao carregar base matrícula x CNPJ; seguindo com CNPJ em branco", str(exc)))
    return base


# ---------------------------------------------------------------------------
# Geração do Excel de validação
# ---------------------------------------------------------------------------


def _setup_sheet(ws, titulo: str, subtitulo: str, max_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(1, 1, titulo)
    ws.cell(2, 1, subtitulo)
    ws.cell(1, 1).font = Font(bold=True, color="FFFFFF", size=14)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor="1F4E78")
    ws.cell(1, 1).alignment = Alignment(horizontal="center")
    ws.cell(2, 1).font = Font(italic=True, color="666666")
    ws.cell(2, 1).alignment = Alignment(horizontal="center")


def _write_header(ws, row: int, headers: Sequence[str]) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    side = Side(style="thin", color="B7B7B7")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        c.font = Font(bold=True)
        c.fill = fill
        c.border = Border(top=side, bottom=side, left=side, right=side)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 55))
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, 42))
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 18
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A5"


def gerar_excel_validacao(itens: List[ItemGrancoffee], remessas: List[RemessaEmail], logs: List[LogEntry], pasta_saida: str) -> str:
    os.makedirs(pasta_saida, exist_ok=True)
    wb = Workbook()
    ws_modelo = wb.active
    ws_modelo.title = "Modelo Robô KOF para Enviar"
    ws_val = wb.create_sheet("Validação do Pedido")
    ws_resumo = wb.create_sheet("Resumo por Loja")
    ws_email = wb.create_sheet("Email x Remessa")
    ws_alertas = wb.create_sheet("Alertas e Erros")
    ws_logs = wb.create_sheet("Logs")

    _setup_sheet(ws_modelo, "GRANCOFFEE - MODELO ROBÔ KOF PARA ENVIAR", "Primeiro output do lote: validar antes de gerar TXT/fila KOF", 10)
    headers_modelo = ["CNPJ", "MATRÍCULA", "SKU", "QTD", "NÚMERO DO PEDIDO", "DATA REMESSA", "REGIONAL", "ARQUIVO ORIGEM", "STATUS", "ALERTA"]
    _write_header(ws_modelo, 4, headers_modelo)
    for item in itens:
        ws_modelo.append([
            item.cnpj,
            item.matricula,
            item.sku,
            quantidade_para_saida(item.qtd_fardo),
            item.pedido,
            fmt_data(item.data_remessa_email),
            item.regional,
            item.arquivo_origem,
            item.status,
            item.alerta,
        ])

    _setup_sheet(ws_val, "GRANCOFFEE - VALIDAÇÃO DO PEDIDO", "Rastreabilidade completa: e-mail, anexo, linha, SKU, quantidade e data", 19)
    headers_val = ["Arquivo Origem", "Aba", "Linha", "Regional", "Matrícula", "CNPJ", "Pedido", "Data Entrega Anexo", "Data Remessa E-mail", "Tipo Faturamento E-mail", "Código GranCoffee", "SKU", "EAN", "Unid. por Fardo", "Descrição", "Qtd Fardo", "Qtd Unitário", "Status", "Alerta"]
    _write_header(ws_val, 4, headers_val)
    for item in itens:
        ws_val.append([
            item.arquivo_origem, item.aba, item.linha, item.regional, item.matricula, item.cnpj, item.pedido,
            fmt_data(item.data_entrega_anexo), fmt_data(item.data_remessa_email), item.tipo_faturamento_email, item.codigo_grancoffee,
            item.sku, item.ean, quantidade_para_saida(item.unidades_por_fardo), item.descricao,
            quantidade_para_saida(item.qtd_fardo), quantidade_para_saida(item.qtd_unitario), item.status, item.alerta,
        ])

    _setup_sheet(ws_resumo, "GRANCOFFEE - RESUMO POR LOJA", "Conferência executiva do lote processado", 10)
    headers_resumo = ["Regional", "Matrícula", "CNPJ", "Pedido", "Data Remessa", "Itens", "Qtd Fardo Total", "Qtd Unitário Total", "Status Geral", "Alertas"]
    _write_header(ws_resumo, 4, headers_resumo)
    grupos: Dict[Tuple[str, str, str], List[ItemGrancoffee]] = {}
    for it in itens:
        grupos.setdefault((it.matricula, it.pedido, it.regional), []).append(it)
    for (_mat, _pedido, _reg), grupo in sorted(grupos.items(), key=lambda kv: (kv[0][2], kv[0][0], kv[0][1])):
        alertas = " | ".join(dict.fromkeys(a for item in grupo for a in item.alerta.split(" | ") if a))
        status = "VALIDAR" if any(item.status != "OK SEM CONVERSÃO" for item in grupo) else "OK"
        ws_resumo.append([
            grupo[0].regional, grupo[0].matricula, grupo[0].cnpj, grupo[0].pedido, fmt_data(grupo[0].data_remessa_email),
            len(grupo), sum((g.qtd_fardo or 0) for g in grupo), sum((g.qtd_unitario or 0) for g in grupo), status, alertas,
        ])

    _setup_sheet(ws_email, "GRANCOFFEE - EMAIL X REMESSA", "Datas oficiais extraídas do corpo do e-mail", 6)
    headers_email = ["Regional", "Matrícula", "Tipo Faturamento", "Pedido", "Data Remessa", "Origem"]
    _write_header(ws_email, 4, headers_email)
    for rem in remessas:
        ws_email.append([rem.regional, rem.matricula, rem.tipo_faturamento, rem.pedido, fmt_data(rem.data_remessa), rem.origem])

    _setup_sheet(ws_alertas, "GRANCOFFEE - ALERTAS E ERROS", "Nada é descartado silenciosamente; todo ponto de atenção aparece aqui", 5)
    headers_alertas = ["Nível", "Etapa", "Arquivo", "Mensagem", "Detalhe"]
    _write_header(ws_alertas, 4, headers_alertas)
    for log in logs:
        if log.nivel in {"WARN", "ERROR"}:
            ws_alertas.append([log.nivel, log.etapa, log.arquivo, log.mensagem, log.detalhe])

    _setup_sheet(ws_logs, "GRANCOFFEE - LOGS COMPLETOS", "Registro técnico para manutenção e auditoria do processamento", 6)
    headers_logs = ["Timestamp", "Nível", "Etapa", "Arquivo", "Mensagem", "Detalhe"]
    _write_header(ws_logs, 4, headers_logs)
    for log in logs:
        ws_logs.append([log.timestamp.strftime("%Y-%m-%d %H:%M:%S"), log.nivel, log.etapa, log.arquivo, log.mensagem, log.detalhe])

    # Formatação de datas e texto sensível.
    for ws in (ws_modelo, ws_val, ws_resumo, ws_email):
        for row in ws.iter_rows(min_row=5):
            for cell in row:
                if isinstance(cell.value, _dt.date):
                    cell.number_format = "dd.mm.yyyy"
    # Força as colunas de data da Grancoffee como texto para evitar conversão automática para barra.
    for ws, cols in ((ws_modelo, [6]), (ws_val, [8, 9]), (ws_resumo, [5]), (ws_email, [5])):
        for row_idx in range(5, ws.max_row + 1):
            for col_idx in cols:
                ws.cell(row_idx, col_idx).number_format = "@"
    for ws in wb.worksheets:
        _autosize(ws)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    out = os.path.join(pasta_saida, f"VALIDACAO_GRANCOFFEE_{stamp}_{uuid.uuid4().hex[:8]}.xlsx")
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# Fallback Outlook selecionado
# ---------------------------------------------------------------------------


def extrair_remessas_outlook_selecionado(logs: List[LogEntry]) -> List[RemessaEmail]:
    """Outlook removido: a edição Android usa apenas arquivos e metadados fornecidos."""
    logs.append(LogEntry("INFO", "ENTRADA", "Android", "Fallback Outlook desativado", "Use PDF/Excel pelo seletor de arquivos."))
    return []


# ---------------------------------------------------------------------------
# Processamento principal
# ---------------------------------------------------------------------------


def processar_grancoffee(arquivos: Sequence[str], pasta_saida: str, caminho_base_cnpj: Optional[str] = None, corpo_email_interface: Optional[str] = None, assunto_email_interface: str = "") -> str:
    """Processa arquivos Grancoffee e retorna o caminho do Excel de validação.

    Esta função é a entrada recomendada para integrar ao Robô KOF.
    Ela NÃO gera TXT/fila KOF automaticamente.
    """
    logs: List[LogEntry] = [LogEntry("INFO", "INICIO", "", "Processamento Grancoffee iniciado", f"Arquivos recebidos: {len(arquivos)}")]
    cnpj_por_matricula = carregar_cnpj_por_matricula(caminho_base_cnpj, logs)
    remessas: List[RemessaEmail] = []
    if corpo_email_interface:
        origem_interface = f"Texto colado na interface - {assunto_email_interface}".strip(" -")
        logs.append(LogEntry("INFO", "EMAIL", "interface", "Texto de e-mail recebido pela interface", origem_interface))
        remessas.extend(extrair_remessas_do_email(corpo_email_interface, origem_interface, logs))
    anexos_excel: List[str] = []
    temp_root = tempfile.mkdtemp(prefix="grancoffee_msg_")
    try:
        for caminho in arquivos:
            if not caminho:
                continue
            ext = Path(caminho).suffix.lower()
            nome = Path(caminho).name
            if ext in EXTENSOES_MSG:
                corpo, anexos = extrair_msg(caminho, temp_root, logs)
                remessas.extend(extrair_remessas_do_email(corpo, nome_seguro(nome), logs))
                anexos_excel.extend(anexos)
            elif ext in EXTENSOES_EXCEL:
                anexos_excel.append(caminho)
                remessas_sidecar = extrair_remessas_metadata_outlook(caminho, logs)
                if remessas_sidecar:
                    remessas.extend(remessas_sidecar)
                    logs.append(LogEntry("INFO", "ENTRADA", nome, "Excel recebido com corpo do e-mail via metadata Outlook", f"Remessas extraídas: {len(remessas_sidecar)}"))
                else:
                    logs.append(LogEntry("INFO", "ENTRADA", nome, "Excel recebido diretamente para processamento", "Sem metadata/corpo de e-mail associado."))
            else:
                logs.append(LogEntry("WARN", "ENTRADA", nome, "Arquivo ignorado por extensão não suportada no layout Grancoffee", ext))
        if not remessas and anexos_excel:
            remessas_auto = extrair_remessas_outlook_selecionado(logs)
            if remessas_auto:
                remessas.extend(remessas_auto)
                logs.append(LogEntry("INFO", "OUTLOOK", "seleção atual", "Aplicando segunda tentativa com e-mail selecionado no Outlook", f"Remessas: {len(remessas_auto)}"))

        # Remove remessas duplicadas mantendo ordem.
        rem_unicas: List[RemessaEmail] = []
        vistos = set()
        for rem in remessas:
            # O mesmo corpo de e-mail pode acompanhar vários anexos .xlsm via sidecar.
            # A deduplicação deve considerar a regra de negócio (matrícula + pedido + data),
            # não o nome do arquivo de origem do metadata.
            k = (somente_digitos(rem.matricula), somente_digitos(rem.pedido), rem.data_remessa)
            if k not in vistos:
                vistos.add(k)
                rem_unicas.append(rem)
        remessas = rem_unicas
        remessas_por_chave = {rem.chave: rem for rem in remessas if rem.matricula and rem.pedido}
        remessas_por_matricula: Dict[str, List[RemessaEmail]] = {}
        for rem in remessas:
            remessas_por_matricula.setdefault(somente_digitos(rem.matricula), []).append(rem)
        itens: List[ItemGrancoffee] = []
        for anexo in anexos_excel:
            itens.extend(ler_anexo_grancoffee(anexo, remessas_por_chave, remessas_por_matricula, cnpj_por_matricula, logs))
        if not itens:
            logs.append(LogEntry("ERROR", "PROCESSAMENTO", "GRANCOFFEE", "Nenhum item Grancoffee foi extraído", "Verificar corpo do e-mail/anexos e layout da aba Pedido."))
        saida = gerar_excel_validacao(itens, remessas, logs, pasta_saida)
        logs.append(LogEntry("INFO", "FIM", Path(saida).name, "Excel de validação Grancoffee gerado", saida))
        return saida
    except Exception as exc:
        logs.append(LogEntry("ERROR", "FATAL", "GRANCOFFEE", "Falha fatal no processamento Grancoffee", f"{exc}\n{traceback.format_exc()}"))
        return gerar_excel_validacao([], remessas, logs, pasta_saida)
    finally:
        try:
            shutil.rmtree(temp_root, ignore_errors=True)
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Compatibilidade com roteadores diferentes do projeto
# ---------------------------------------------------------------------------


def processar_layout(arquivos: Sequence[str], pasta_saida: str, **kwargs) -> str:
    """Alias padrão para roteadores que chamam processar_layout(...)."""
    return processar_grancoffee(arquivos, pasta_saida, kwargs.get("caminho_base_cnpj") or kwargs.get("base_cnpj"), kwargs.get("corpo_email") or kwargs.get("body"), kwargs.get("assunto_email") or kwargs.get("assunto") or "")


def detectar_layout(caminho: str, texto_previo: str = "") -> bool:
    """Alias padrão para roteadores que chamam detectar_layout(...)."""
    return identificar_grancoffee(caminho, texto_previo)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Processador Grancoffee / SPAL para Robô KOF")
    parser.add_argument("arquivos", nargs="+", help="Arquivos .msg/.xlsm/.xlsx da Grancoffee")
    parser.add_argument("--saida", default="saida_grancoffee", help="Pasta de saída do Excel de validação")
    parser.add_argument("--base-cnpj", default=None, help="CSV opcional com colunas Matrícula e CNPJ")
    args = parser.parse_args()
    print(processar_grancoffee(args.arquivos, args.saida, args.base_cnpj))
