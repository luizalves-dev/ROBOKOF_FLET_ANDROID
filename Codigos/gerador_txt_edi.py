from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Tuple
import re
import openpyxl

import config


def clean_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def only_digits(v) -> str:
    return re.sub(r"\D+", "", clean_str(v))


def alfa_fixo(valor, tamanho: int) -> str:
    return str(valor or "")[:tamanho].ljust(tamanho)


def num_fixo(valor, tamanho: int) -> str:
    dig = only_digits(valor)
    return dig.zfill(tamanho)[:tamanho]


def decimal_fixo(valor, tamanho: int, decimais: int) -> str:
    """
    Ex.: tamanho=17, decimais=4
    1   -> 00000000000010000
    1.5 -> 00000000000015000
    """
    if valor is None or clean_str(valor) == "":
        return "0" * tamanho

    texto = str(valor).strip().replace(",", ".")
    numero = float(texto)
    inteiro = int(round(numero * (10 ** decimais)))
    return str(inteiro).zfill(tamanho)[:tamanho]


def data_yyyymmddhhmm(valor=None, usar_agora=False, somar_dias=0) -> str:
    if usar_agora:
        return datetime.now().strftime("%Y%m%d%H%M")

    if valor is None:
        return "0" * 12

    if isinstance(valor, datetime):
        dt = valor
    else:
        texto = clean_str(valor)
        dt = None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                dt = datetime.strptime(texto, fmt)
                break
            except ValueError:
                pass

        if dt is None:
            return "0" * 12

    if somar_dias:
        dt = dt + timedelta(days=somar_dias)

    # quando não houver hora/minuto reais, completa com 0000
    return dt.strftime("%Y%m%d") + "0000"


def _find_header_row_and_cols(
    ws,
    header_candidates: Dict[str, List[str]],
    search_rows: int = 80
) -> Tuple[int, Dict[str, int]]:
    best_row = None
    best_map = {}

    keys = list(header_candidates.keys())

    for r in range(1, min(search_rows, ws.max_row) + 1):
        row_vals = [clean_str(c.value) for c in ws[r]]
        col_map = {}

        for key in keys:
            for cand in header_candidates[key]:
                if cand in row_vals:
                    col_map[key] = row_vals.index(cand) + 1
                    break

        if len(col_map) > len(best_map):
            best_map = col_map
            best_row = r

    if not best_row:
        raise ValueError("Não encontrei a linha de cabeçalho na sheet 'Ordem' para gerar o TXT.")

    required = ["Matricula", "Sku", "Qtd", "Pedido", "Data"]
    missing = [k for k in required if k not in best_map]
    if missing:
        raise ValueError(f"Cabeçalhos obrigatórios não encontrados para TXT: {missing}")

    return best_row, best_map


def _read_gln_from_sheet(ws) -> str:
    gln_col = 10  # J

    for r in range(1, ws.max_row + 1):
        if clean_str(ws.cell(r, gln_col).value).upper() == "GLN":
            return only_digits(ws.cell(r + 1, gln_col).value)

    return ""


def _read_rows_from_robokof_excel(excel_path: Path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if config.TEMPLATE_ORDEM_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{config.TEMPLATE_ORDEM_SHEET}' não encontrada no arquivo {excel_path.name}")

    ws = wb[config.TEMPLATE_ORDEM_SHEET]
    header_row, col_map = _find_header_row_and_cols(ws, config.HEADER_MAP)
    data_start = header_row + 1

    gln = _read_gln_from_sheet(ws)
    rows = []

    for r in range(data_start, ws.max_row + 1):
        matricula = clean_str(ws.cell(r, col_map["Matricula"]).value)
        sku = clean_str(ws.cell(r, col_map["Sku"]).value)
        qtd = ws.cell(r, col_map["Qtd"]).value
        pedido = clean_str(ws.cell(r, col_map["Pedido"]).value)
        data = ws.cell(r, col_map["Data"]).value

        if not any([matricula, sku, qtd, pedido, data]):
            break

        rows.append({
            "matricula": only_digits(matricula),
            "sku": only_digits(sku),
            "qtd": qtd,
            "pedido": clean_str(pedido),
            "data": data,
        })

    if not rows:
        raise ValueError(f"Nenhuma linha encontrada no Excel RoboKOF: {excel_path.name}")

    return rows, gln


def montar_registro_01(pedido: str, matricula: str, gln: str, data_entrega_valor) -> str:
    """
    Header 1:
    - pedido em 20 posições
    - tipo pedido fixo = 001
    - lista de preço = 20 zeros
    - data final entrega = data de emissão + 7 dias
    """
    partes = []

    data_emissao_agora = datetime.now()
    data_emissao = data_emissao_agora.strftime("%Y%m%d%H%M")

    # data inicial continua vindo da data do pedido/entrega do Excel
    data_inicial_entrega = data_yyyymmddhhmm(data_entrega_valor)

    # data final = data da emissão + 7 dias
    data_final_entrega = (data_emissao_agora + timedelta(days=7)).strftime("%Y%m%d%H%M")

    # 001-002
    partes.append("01")

    # 003-008
    partes.append("ORDERS")

    # 009-028 - 20 posições
    partes.append(alfa_fixo(pedido, 20))

    # 029-036
    partes.append(" " * 8)

    # 037-039
    partes.append("9  ")

    # 040-051 - Data da Emissão do Pedido
    partes.append(data_emissao)

    # 052-063 - Data Inicial para Entrega
    partes.append(data_inicial_entrega)

    # 064-075 - Data Final para Entrega
    partes.append(data_final_entrega)

    # 076-083 Horário Descarga Dia Útil
    partes.append("00000000")

    # 084-091 Horário Descarga Sábado
    partes.append("00000000")

    # 092-094 Tipo do Pedido
    partes.append(config.TXT_REG01_TIPO_PEDIDO)

    # 095-114 Lista de Preço
    partes.append(config.TXT_REG01_LISTA_PRECO)

    # 115-127 EAN Comprador
    partes.append(num_fixo(gln, 13))

    # 128-142 CGC Comprador
    partes.append(num_fixo(matricula, 15))

    # 143-155 EAN Fornecedor
    partes.append(config.TXT_REG01_BLOCO_FIXO_MEIO[:13])

    # 156-170 CGC Fornecedor
    partes.append(num_fixo(config.TXT_REG01_BLOCO_FIXO_MEIO[13:], 15))

    # 171-183 EAN Entrega
    partes.append(num_fixo(gln, 13))

    # 184-198 CGC Entrega
    partes.append(num_fixo(matricula, 15))

    # 199-211 EAN Cobrança
    partes.append(num_fixo(gln, 13))

    # 212-226 CGC Cobrança
    partes.append(num_fixo(matricula, 15))

    # 227-260
    partes.append(" " * 34)

    # 261-295
    partes.append(alfa_fixo("", 35))

    # 296-330
    partes.append(alfa_fixo("", 35))

    # 331-365
    partes.append(alfa_fixo("", 35))

    # 366-390
    partes.append(alfa_fixo("", 25))

    # 391-400
    partes.append(" " * 10)

    linha = "".join(partes)
    return linha[:400].ljust(400)


def montar_registro_02() -> str:
    return config.TXT_REG02_FIXO.ljust(config.TXT_REG02_TAMANHO)


def montar_registro_03() -> str:
    linha = "03" + (" " * 17) + config.TXT_REG03_CONDICAO
    return linha.ljust(config.TXT_REG03_TAMANHO)


def montar_registro_04(seq: int, sku: str, qtd, gln: str, matricula: str) -> str:
    partes = []

    # 001-002
    partes.append("04")

    # 003-008
    partes.append(str(seq).zfill(6))

    # 009-022
    partes.append(alfa_fixo(sku.zfill(14), 14))

    # 023-060
    partes.append(" " * 38)

    # 061-095
    partes.append(alfa_fixo(".", config.TXT_REG04_DESC_TAMANHO))

    # 096-112 - 17 com 4 decimais implícitos
    partes.append(decimal_fixo(qtd, 17, 4))

    # 113-129 - bonificada zerada
    partes.append("0" * config.TXT_REG04_QTD_BONIF_TAMANHO)

    # 130-142 - EAN Entrega
    partes.append(num_fixo(gln, 13))

    # 143-157 - CGC Entrega
    partes.append(num_fixo(matricula, 15))

    # 158-163
    partes.append(" " * config.TXT_REG04_ESPACOS_APOS_CLIENTE)

    # 164-180 - preço líquido
    partes.append("0" * 17)

    # 181-197 - preço bruto
    partes.append("0" * 17)

    # 198-213 - IPI
    partes.append("0" * 16)

    # 214-229 - ICMR
    partes.append("0" * 16)

    # 230-237 - % desconto comercial (8,4)
    partes.append("0" * 8)

    # 238-250 - valor desconto comercial (13,2)
    partes.append("0" * 13)

    # 251-265 - % desconto em nota (15,2)
    partes.append("0" * 15)

    # 266-278 - valor outros encargos (13,2)
    partes.append("0" * 13)

    # 279-293 - valor frete (15,2)
    partes.append("0" * 15)

    # 294-400
    partes.append(" " * config.TXT_REG04_FILLER_FINAL)

    linha = "".join(partes)
    return linha[:400].ljust(400)


def montar_registro_09() -> str:
    partes = []

    # 001-002
    partes.append("09")

    # 003-024 Valor Total Pedido 22,6
    partes.append("0" * 22)

    # 025-046 Valor Total Mercadorias 22,6
    partes.append("0" * 22)

    # 047-064 Valor Total Desconto Comercial 18,2
    partes.append("0" * 18)

    # 065-082 Valor Total Outros Encargos 18,2
    partes.append("0" * 18)

    # 083-400 filler
    partes.append(" " * 318)

    linha = "".join(partes)
    return linha[:400].ljust(400)


def gerar_txt_edi_do_excel(excel_robokof_path: Path, txt_out_path: Path):
    rows, gln = _read_rows_from_robokof_excel(excel_robokof_path)

    pedido = rows[0]["pedido"]
    matricula = rows[0]["matricula"]
    data_entrega = rows[0]["data"]

    linhas = [
        montar_registro_01(pedido, matricula, gln, data_entrega),
        montar_registro_02(),
        montar_registro_03(),
    ]

    for i, row in enumerate(rows, start=1):
        linhas.append(
            montar_registro_04(
                seq=i,
                sku=row["sku"],
                qtd=row["qtd"],
                gln=gln,
                matricula=row["matricula"],
            )
        )

    linhas.append(montar_registro_09())

    txt_out_path.parent.mkdir(parents=True, exist_ok=True)

    with open(txt_out_path, "w", encoding=config.TXT_ENCODING, newline="") as f:
        for linha in linhas:
            f.write(linha + "\n")