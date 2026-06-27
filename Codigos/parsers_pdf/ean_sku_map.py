from __future__ import annotations

import re
import unicodedata
from functools import lru_cache
from pathlib import Path
from typing import Dict, Iterable

from openpyxl import load_workbook

import config


def only_digits(value) -> str:
    return re.sub(r"\D+", "", str(value or "")).strip()


def clean_cell(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text.strip()


def normalize_header(value) -> str:
    text = unicodedata.normalize("NFKD", clean_cell(value).lower())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip()


def resolve_ean_to_sku(ean: str, gabarito_names: Iterable[str], overrides: Dict[str, str] | None = None) -> str:
    ean_key = only_digits(ean)
    if not ean_key:
        return ""
    overrides = overrides or {}
    if ean_key in overrides:
        return clean_cell(overrides[ean_key])

    for gabarito_name in gabarito_names:
        mapping = load_ean_sku_map(str(config.ARQUIVOS_BASE_DIR / gabarito_name))
        sku = mapping.get(ean_key)
        if sku:
            return sku
    return ""


@lru_cache(maxsize=16)
def load_ean_sku_map(gabarito_path: str) -> Dict[str, str]:
    path = Path(gabarito_path)
    if not path.exists():
        return {}

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["LISTA DE PRODUTOS"] if "LISTA DE PRODUTOS" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return {}

    header_index = 0
    for idx, row in enumerate(rows[:20]):
        values = [normalize_header(cell) for cell in row]
        has_ean = any(_is_ean_header(value) for value in values)
        has_sku = any(_is_sku_header(value) for value in values)
        if has_ean and has_sku:
            header_index = idx
            break

    headers = [normalize_header(cell) for cell in rows[header_index]]
    ean_idx = None
    sku_idx = None
    for idx, header in enumerate(headers):
        if ean_idx is None and _is_ean_header(header):
            ean_idx = idx
        if sku_idx is None and _is_sku_header(header):
            sku_idx = idx

    if ean_idx is None or sku_idx is None:
        if len(headers) >= 2:
            ean_idx = ean_idx if ean_idx is not None else 0
            sku_idx = sku_idx if sku_idx is not None else 1
        else:
            return {}

    mapping: Dict[str, str] = {}
    for row in rows[header_index + 1:]:
        ean = only_digits(row[ean_idx] if ean_idx < len(row) else "")
        sku = clean_cell(row[sku_idx] if sku_idx < len(row) else "")
        if ean and sku:
            mapping.setdefault(ean, sku)
    return mapping


def _is_ean_header(value: str) -> bool:
    return any(token in value for token in ("ean", "dun", "barra", "codigo de barras", "cod barras"))


def _is_sku_header(value: str) -> bool:
    return value == "sku" or "sku ecc" in value or ("sku" in value and "ecc" in value)

