from __future__ import annotations

import re
import unicodedata
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from parsers_pdf.pdf_utils import build_intermediate_df, clean_text, extract_pages_text, only_digits
from terminal_logger import get_terminal_logger


BASE_ITALO_PATH = Path(__file__).resolve().parents[2] / "Arquivos Base" / "BASE PRODUTOS ITALO -atualizada.xlsx"
terminal_log = get_terminal_logger("pdf_rede_italo")

BRANDS = [
    "Estrella Galicia",
    "Coca Cola",
    "Del Valle",
    "Schweppes",
    "Crystal",
    "Monster",
    "Reign",
    "Cerpa",
    "Eisenbahn",
    "Kaiser",
    "Sol",
    "Ades",
]
BRAND_RE = "|".join(re.escape(b) for b in sorted(BRANDS, key=len, reverse=True))

RE_PEDIDO = re.compile(r"N[uú]mero\s+do\s+Pedido:\s*(\d+)", re.I)
RE_CNPJ = re.compile(r"CNPJ:\s*([\d./-]+)", re.I)
RE_ITEM_START = re.compile(r"^\s*(?P<codigo>\d{3,9})\s+(?P<ean>\d{8,14})\s+(?P<rest>.+)$")
RE_QTD_AFTER_BRAND = re.compile(
    rf"\s(?P<marca>{BRAND_RE})\s+(?P<qtd>(?:\d{{1,3}}(?:\.\d{{3}})+|\d+),\d{{3}})\s+"
    r"(?P<preco>\d+(?:[,.]\d+)?)",
    re.I,
)
RE_QTD_PRICE_FALLBACK = re.compile(
    r"\s(?P<qtd>(?:\d{1,3}(?:\.\d{3})+|\d+),\d{3})\s+(?P<preco>\d+(?:[,.]\d+)?)",
    re.I,
)


def _parse_br_decimal(value) -> Optional[Decimal]:
    text = str(value or "").strip().replace("R$", "").replace(" ", "")
    if not text:
        return None
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _format_qtd(value: Decimal) -> str:
    if value == value.to_integral_value():
        return str(int(value))
    normalized = value.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP).normalize()
    return format(normalized, "f").replace(".", ",")


def _normalize_text(value) -> str:
    text = str(value or "").lower().strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _ean_keys(value) -> List[str]:
    digits = only_digits(value)
    if not digits:
        return []
    stripped = digits.lstrip("0") or "0"
    keys = [digits, stripped]
    if len(stripped) <= 13:
        keys.append(stripped.zfill(13))
    if len(stripped) > 8:
        keys.append(stripped[-8:])
    out = []
    seen = set()
    for key in keys:
        if key and key not in seen:
            seen.add(key)
            out.append(key)
    return out


def _header_key(value) -> str:
    text = _normalize_text(value)
    replacements = {
        "cod barras": "cod barras",
        "codigo barras": "cod barras",
        "codigo de barras": "cod barras",
        "ean": "cod barras",
        "cod ean": "cod barras",
        "codigo ean": "cod barras",
        "codigo femsa": "codigo femsa",
        "sku femsa": "codigo femsa",
        "sku": "codigo femsa",
        "conversao": "conversao",
        "conversao un cx": "conversao",
        "unidade conversao": "conversao",
        "codigo": "codigo",
    }
    return replacements.get(text, text)


def _load_base_italo() -> Dict[str, Dict[str, Decimal | str]]:
    if not BASE_ITALO_PATH.exists():
        raise FileNotFoundError(f"Base de produtos Italo nao encontrada: {BASE_ITALO_PATH}")

    wb = load_workbook(BASE_ITALO_PATH, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
                headers = [_header_key(v) for v in row]
                joined = "|".join(headers)
                if "cod barras" not in joined or "codigo femsa" not in joined or "conversao" not in joined:
                    continue

                header = {value: idx for idx, value in enumerate(headers) if value}
                col_ean = header["cod barras"]
                col_sku = header["codigo femsa"]
                col_conv = header["conversao"]
                index: Dict[str, Dict[str, Decimal | str]] = {}

                for data_row in ws.iter_rows(min_row=row_idx + 1, values_only=True):
                    ean = only_digits(data_row[col_ean] if col_ean < len(data_row) else "")
                    sku = only_digits(data_row[col_sku] if col_sku < len(data_row) else "")
                    conversao = _parse_br_decimal(data_row[col_conv] if col_conv < len(data_row) else "")
                    if not ean or not sku or not conversao or conversao == 0:
                        continue
                    item = {"sku": sku, "conversao": conversao}
                    for key in _ean_keys(ean):
                        index.setdefault(key, item)

                if index:
                    return index
    finally:
        wb.close()

    raise ValueError("Base de produtos Italo sem colunas validas de EAN, SKU Femsa e Conversao")


def _extract_header(page_text: str) -> Tuple[str, str]:
    pedido = ""
    cnpj = ""

    match_pedido = RE_PEDIDO.search(page_text)
    if match_pedido:
        pedido = clean_text(match_pedido.group(1))

    if re.search(r"^\s*Empresa:", page_text, flags=re.I | re.M):
        cnpjs = RE_CNPJ.findall(page_text)
        if cnpjs:
            cnpj = only_digits(cnpjs[-1])

    return pedido, cnpj


def _parse_item_line(line: str) -> Optional[Tuple[str, str, Decimal]]:
    start = RE_ITEM_START.match(line)
    if not start:
        return None

    codigo_pdf = only_digits(start.group("codigo"))
    ean = only_digits(start.group("ean"))
    rest = start.group("rest")

    matches = list(RE_QTD_AFTER_BRAND.finditer(rest))
    if matches:
        qtd = _parse_br_decimal(matches[-1].group("qtd"))
        return codigo_pdf, ean, qtd if qtd is not None else Decimal("0")

    fallback = RE_QTD_PRICE_FALLBACK.search(rest)
    if fallback:
        qtd = _parse_br_decimal(fallback.group("qtd"))
        return codigo_pdf, ean, qtd if qtd is not None else Decimal("0")

    return None


def ler_pdf_rede_italo(caminho_arquivo: str, layout_config: dict, mapeamentos_df=None):
    paginas = extract_pages_text(caminho_arquivo)
    base_italo = _load_base_italo()

    linhas_saida: List[Dict[str, str]] = []
    alertas: List[str] = []
    pedidos_encontrados = {}
    cnpjs_encontrados = {}
    linhas_lidas = 0
    linhas_validas = 0
    linhas_descartadas = 0
    sem_conversao = 0
    current_pedido = ""
    current_cnpj = ""

    for page_idx, texto in enumerate(paginas, start=1):
        pedido, cnpj = _extract_header(texto)
        if pedido:
            current_pedido = pedido
            pedidos_encontrados[pedido] = pedidos_encontrados.get(pedido, 0) + 1
        if cnpj:
            current_cnpj = cnpj
            cnpjs_encontrados[cnpj] = cnpjs_encontrados.get(cnpj, 0) + 1

        for linha in [clean_text(l) for l in texto.splitlines() if clean_text(l)]:
            if not RE_ITEM_START.match(linha):
                continue

            linhas_lidas += 1
            parsed = _parse_item_line(linha)
            if not parsed:
                linhas_descartadas += 1
                alertas.append(f"Pagina {page_idx}: linha de item nao reconhecida | {linha[:120]}")
                continue

            codigo_pdf, ean, qtd_unidades = parsed
            produto = None
            for key in _ean_keys(ean):
                produto = base_italo.get(key)
                if produto:
                    break

            if not produto:
                sem_conversao += 1
                alerta = f"Pagina {page_idx}: EAN sem conversao na base Italo | ean={ean} | {linha[:120]}"
                alertas.append(alerta)
                # Regra corporativa: item não pode sumir por falta de conversão.
                # Ele segue para o Excel com status ALERTA - NÃO CONVERTIDO.
                linhas_saida.append(
                    {
                        "matricula_lida": "",
                        "cnpj_lido": current_cnpj,
                        "sku_lido": codigo_pdf,
                        "codigo_sku_lido": codigo_pdf,
                        "ean_lido": ean,
                        "codigo_origem_lido": codigo_pdf,
                        "pagina_pdf": str(page_idx),
                        "linha_bruta": linha,
                        "origem_extracao": "PDF_REDE_ITALO",
                        "quantidade_lida": _format_qtd(qtd_unidades),
                        "qtd_original": _format_qtd(qtd_unidades),
                        "tipo_qtd_original": "UNIDADE",
                        "fator_conversao": "",
                        "qtd_convertida": "",
                        "qtd_final": _format_qtd(qtd_unidades),
                        "status_conversao": "ALERTA - NÃO CONVERTIDO",
                        "tipo_regra_conversao": "BASE_ITALO_PRIORITARIA",
                        "origem_regra_conversao": str(BASE_ITALO_PATH.name),
                        "observacao_conversao": alerta,
                        "numero_pedido_lido": current_pedido,
                        "data_entrega_lida": "",
                    }
                )
                continue

            conversao = produto["conversao"]
            qtd_convertida = qtd_unidades / conversao
            sku = str(produto.get("sku", "") or codigo_pdf)

            if not current_pedido or not current_cnpj or not sku or qtd_convertida <= 0:
                linhas_descartadas += 1
                alertas.append(
                    f"Pagina {page_idx}: item descartado por campo obrigatorio ausente "
                    f"(pedido={current_pedido or '-'}, cnpj={current_cnpj or '-'}, codigo={codigo_pdf or '-'}, "
                    f"ean={ean or '-'}, sku={sku or '-'}, qtd={_format_qtd(qtd_convertida) if qtd_convertida else '-'})"
                )
                continue

            linhas_validas += 1
            linhas_saida.append(
                {
                    "matricula_lida": "",
                    "cnpj_lido": current_cnpj,
                    "sku_lido": sku,
                    "codigo_sku_lido": sku,
                    "ean_lido": ean,
                    "codigo_origem_lido": codigo_pdf,
                    "pagina_pdf": str(page_idx),
                    "linha_bruta": linha,
                    "origem_extracao": "PDF_REDE_ITALO",
                    "quantidade_lida": _format_qtd(qtd_convertida),
                    "qtd_original": _format_qtd(qtd_unidades),
                    "tipo_qtd_original": "UNIDADE",
                    "fator_conversao": _format_qtd(conversao),
                    "qtd_convertida": _format_qtd(qtd_convertida),
                    "qtd_final": _format_qtd(qtd_convertida),
                    "status_conversao": "OK CONVERTIDO",
                    "tipo_regra_conversao": "BASE_ITALO_PRIORITARIA",
                    "origem_regra_conversao": str(BASE_ITALO_PATH.name),
                    "observacao_conversao": "Conversão Ítalo aplicada pela base prioritária: Cod Barras -> Código Femsa / Conversão.",
                    "numero_pedido_lido": current_pedido,
                    "data_entrega_lida": "",
                }
            )

    df_intermediario = build_intermediate_df(
        linhas_saida,
        caminho_arquivo,
        layout_config.get("nome_layout", ""),
    )

    terminal_log.info(
        "[REDE ITALO] pedidos=%s | cnpjs=%s | linhas_brutas=%s | validas=%s | descartes=%s | sem_conversao=%s | intermediarias=%s",
        sorted(pedidos_encontrados.keys()),
        sorted(cnpjs_encontrados.keys()),
        linhas_lidas,
        linhas_validas,
        linhas_descartadas,
        sem_conversao,
        len(df_intermediario),
    )

    if df_intermediario.empty:
        return {
            "sucesso": False,
            "mensagem": "Nenhuma linha valida foi extraida do PDF Rede Italo",
            "df_intermediario": df_intermediario,
            "qtd_linhas_lidas": linhas_lidas,
            "qtd_itens_extraidos": len(df_intermediario),
            "qtd_itens_ignorados": linhas_descartadas,
            "qtd_itens_nao_convertidos": sem_conversao,
            "alertas": sorted(set(alertas)),
        }

    return {
        "sucesso": True,
        "mensagem": f"Leitura PDF Rede Italo concluida com {len(df_intermediario)} linha(s)",
        "df_intermediario": df_intermediario,
        "qtd_linhas_lidas": linhas_lidas,
        "qtd_itens_extraidos": len(df_intermediario),
        "qtd_itens_ignorados": linhas_descartadas,
        "qtd_itens_nao_convertidos": sem_conversao,
        "alertas": sorted(set(alertas)),
    }
