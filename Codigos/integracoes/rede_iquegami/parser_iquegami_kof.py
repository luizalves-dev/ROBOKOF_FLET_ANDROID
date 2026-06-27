# Autor: Kauê Melo
# Integração segura Rede Iquegami -> Robô KOF-Projeto
from __future__ import annotations

import csv
import json
import os
import re
import traceback
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

try:
    from .xls_biff_fallback import read_xls_cells, cells_to_matrix
except Exception:  # execução direta
    from xls_biff_fallback import read_xls_cells, cells_to_matrix



DE_PARA_PADRAO_MATRICULA_CNPJ: dict[str, str] = {
    "7120136174": "01915625000146",
    "7120148030": "01915625000227",
    "7120021397": "01915625000308",
    "7120041167": "01915625000499",
    "7120041165": "01915625000570",
    "7120043621": "01915625000650",
    "7110317767": "01915625000731",
    "7120067250": "01915625000812",
    "7120183043": "01915625000901",
    "7120224787": "01915625001037",
    "7120219208": "01915625001118",
    "7120254231": "01915625001207",
    "7120275165": "01915625001380",
    "7120291224": "01915625001460",
    "7120297546": "01915625001541",
    "7120498825": "01915625001622",
    "7120321089": "01915625001703",
    "7120360096": "01915625001894",
    "7120362796": "01915625001975",
    "7120440837": "01915625002009",
    "7120523572": "01915625002190",
    "7120502631": "01915625002270",
    "7120510870": "01915625002351",
}

@dataclass
class ItemPedido:
    cnpj: str
    matricula_loja: str
    sku: str
    qtd: Any
    numero_pedido: str
    arquivo_origem: str
    aba_origem: str
    linha_origem: int
    status: str = "OK"
    observacao: str = ""


@dataclass
class AlertaProcessamento:
    tipo: str
    arquivo: str
    aba: str
    linha: int | str
    pedido: str
    loja: str
    sku: str
    qtd: Any
    mensagem: str


class IquegamiParser:
    """Parser corporativo para pedidos Iquegami em XLS, XLSX e PDF textual."""

    def __init__(self):
        self.itens: list[ItemPedido] = []
        self.alertas: list[AlertaProcessamento] = []
        self.logs: list[dict[str, Any]] = []
        self.de_para_matricula_cnpj = self.carregar_de_para_matricula_cnpj()
        self.de_para_cnpj_matricula = {cnpj: mat for mat, cnpj in self.de_para_matricula_cnpj.items()}

    def log(self, nivel: str, evento: str, detalhe: str = "", **extra: Any) -> None:
        self.logs.append({
            "data_hora": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "nivel": nivel,
            "evento": evento,
            "detalhe": detalhe,
            **extra,
        })

    @classmethod
    def carregar_de_para_matricula_cnpj(cls) -> dict[str, str]:
        """Carrega o de/para matrícula -> CNPJ.

        O CSV fica em ../data/de_para_matricula_cnpj.csv para manutenção simples.
        Caso o arquivo não exista ou tenha problema, usa o de/para padrão embutido.
        """
        mapa = dict(DE_PARA_PADRAO_MATRICULA_CNPJ)
        csv_path = Path(__file__).resolve().parents[1] / "data" / "de_para_matricula_cnpj.csv"
        if not csv_path.exists():
            return mapa
        try:
            with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f, delimiter=";")
                for row in reader:
                    matricula = cls.somente_digitos(row.get("matricula") or row.get("MATRICULA") or "")
                    cnpj = cls.somente_digitos(row.get("cnpj") or row.get("CNPJ") or "")
                    if matricula and cnpj:
                        mapa[matricula] = cnpj.zfill(14)
        except Exception:
            # Mantém o robô funcionando mesmo se o CSV for editado incorretamente.
            return mapa
        return mapa

    def resolver_matricula_cnpj(self, loja: str) -> tuple[str, str, str]:
        """Resolve a coluna Loja para matrícula e CNPJ.

        Retorna: (matricula, cnpj, observação).
        - Se Loja tiver 14 dígitos, entende como CNPJ e tenta achar a matrícula inversa.
        - Se Loja tiver outro tamanho, entende como matrícula e procura no de/para.
        """
        loja = self.somente_digitos(loja)
        if not loja:
            return "", "", "Loja/Matrícula vazia"
        if len(loja) == 14:
            matricula = self.de_para_cnpj_matricula.get(loja, "")
            return matricula or loja, loja, "" if matricula else "CNPJ informado, matrícula não localizada no de/para"
        cnpj = self.de_para_matricula_cnpj.get(loja, "")
        if cnpj:
            return loja, cnpj, ""
        return loja, "", f"Matrícula {loja} sem CNPJ cadastrado no de/para"

    @staticmethod
    def somente_digitos(valor: Any) -> str:
        if valor is None:
            return ""
        if isinstance(valor, float) and valor.is_integer():
            return str(int(valor))
        if isinstance(valor, int):
            return str(valor)
        txt = str(valor).strip()
        if not txt:
            return ""
        if re.fullmatch(r"\d+\.0", txt):
            txt = txt[:-2]
        return re.sub(r"\D", "", txt)

    @staticmethod
    def normalizar_texto(valor: Any) -> str:
        if valor is None:
            return ""
        txt = str(valor).strip().lower()
        mapa = str.maketrans("áàãâäéèêëíìîïóòõôöúùûüç", "aaaaaeeeeiiiiooooouuuuc")
        return txt.translate(mapa)

    @staticmethod
    def parse_qtd(valor: Any) -> int | float | str | None:
        if valor is None or str(valor).strip() == "":
            return None
        if isinstance(valor, int):
            return valor
        if isinstance(valor, float):
            return int(valor) if valor.is_integer() else valor
        txt = str(valor).strip().replace(" ", "")
        if not txt:
            return None
        txt = txt.replace("−", "-")
        if "," in txt and "." in txt:
            txt = txt.replace(".", "").replace(",", ".")
        elif "," in txt:
            txt = txt.replace(",", ".")
        elif re.fullmatch(r"\d{1,3}(\.\d{3})+", txt):
            txt = txt.replace(".", "")
        try:
            num = float(txt)
            return int(num) if num.is_integer() else num
        except Exception:
            return valor

    @staticmethod
    def detectar_pedido_em_texto(texto: str, nome_arquivo: str = "") -> str:
        """Detecta o número do pedido em cabeçalhos e nomes de arquivo.

        Cobre os layouts Iquegami já observados:
        - Pedido 236369
        - Pedido de Compra Número 313487(...)
        - Pedido de Compra Numero 313487(...)
        - Cabeçalho compacto com o número imediatamente antes do parêntese.

        Retorna somente os dígitos do pedido para manter a saída padronizada.
        """
        texto = str(texto or "")
        padroes = [
            r"Pedido\s+de\s+Compra\s+N[úu]mero\s*[:\-]?\s*([0-9]{4,})",
            r"Pedido\s+de\s+Compra\s+Numero\s*[:\-]?\s*([0-9]{4,})",
            r"Pedido\s*[:\-]?\s*([0-9]{4,})",
            r"N[úu]mero\s*[:\-]?\s*([0-9]{4,})",
            r"\b[A-Z]?([0-9]{5,})\s*\(",
        ]
        for padrao in padroes:
            candidatos = re.findall(padrao, texto, flags=re.I)
            if candidatos:
                return re.sub(r"\D", "", str(candidatos[0]))

        # Fallback por nome do arquivo. Ignora datas simples e usa somente números com 5+ dígitos.
        candidatos = re.findall(r"\b([0-9]{5,})\b", Path(nome_arquivo).stem)
        return candidatos[0] if candidatos else ""

    @staticmethod
    def _eh_coluna_loja_ou_matricula(cell: str) -> bool:
        cell = cell.strip()
        return (
            cell == "loja"
            or "loja" in cell
            or cell in {"matricula", "matricula loja", "matricula_loja", "cnpj"}
            or ("matricula" in cell and "sku" not in cell and "codigo" not in cell)
        )

    @staticmethod
    def _eh_coluna_sku(cell: str) -> bool:
        cell = cell.strip()
        return (
            ("ref" in cell and ("fornec" in cell or "fornecedor" in cell))
            or cell in {"sku", "cod", "codigo", "codigo produto", "cod produto", "codigo externo"}
            or (cell.startswith("codigo") and "barra" not in cell and "ean" not in cell)
        )

    @staticmethod
    def _eh_coluna_qtd(cell: str) -> bool:
        cell = cell.strip()
        return "qtde" in cell or "qtd" in cell or "quant" in cell

    def linha_parece_item_compacto(self, row: list[Any]) -> bool:
        """Identifica linhas do layout compacto sem cabeçalho.

        Exemplo real observado: primeira coluna matrícula, segunda SKU/código e
        terceira quantidade, com linhas em branco separando blocos por loja.
        """
        if len(row) < 3:
            return False
        loja = self.somente_digitos(row[0])
        sku = self.somente_digitos(row[1])
        qtd = self.parse_qtd(row[2])
        return len(loja) >= 8 and len(sku) >= 4 and isinstance(qtd, (int, float))

    def detectar_colunas(self, matrix: list[list[Any]]) -> tuple[int | None, dict[str, int]]:
        """Retorna índice da linha cabeçalho e colunas de Loja/Matrícula, SKU e Qtde.

        Layouts Iquegami mapeados:
        1) Loja | Ref. Fornecedor | Qtde. | Pedido
        2) Loja | Descrição do Produto | Ref. Fornec/Fornecedor | Qtde.
        3) MATRÍCULA | SKU | QUANTIDADE, com cabeçalhos repetidos por bloco
        4) MATRÍCULA | CÓDIGO | QUANTIDADE
        5) Compacto sem cabeçalho: matrícula | sku/código | quantidade
        """
        for i, row in enumerate(matrix[:150]):
            norm = [self.normalizar_texto(v) for v in row]
            tem_loja = any(self._eh_coluna_loja_ou_matricula(c) for c in norm)
            tem_qtd = any(self._eh_coluna_qtd(c) for c in norm)
            tem_ref = any(self._eh_coluna_sku(c) for c in norm)
            if tem_loja and tem_qtd and tem_ref:
                cols: dict[str, int] = {}
                for idx, cell in enumerate(norm):
                    if "loja" not in cols and self._eh_coluna_loja_ou_matricula(cell):
                        cols["loja"] = idx
                    if "sku" not in cols and self._eh_coluna_sku(cell):
                        cols["sku"] = idx
                    if "qtd" not in cols and self._eh_coluna_qtd(cell):
                        cols["qtd"] = idx
                if {"loja", "sku", "qtd"}.issubset(cols):
                    return i, cols

        # Fallback para arquivos Iquegami compactos sem cabeçalho.
        # Ex.: algumas planilhas alcoólicos vêm apenas com matrícula, sku e quantidade,
        # iniciando com uma linha de título "iquegami" e linhas em branco entre lojas.
        amostra = matrix[:200]
        hits = sum(1 for row in amostra if self.linha_parece_item_compacto(row))
        if hits >= 2:
            return -1, {"loja": 0, "sku": 1, "qtd": 2}
        return None, {}

    def linha_eh_cabecalho(self, row: list[Any]) -> bool:
        norm = [self.normalizar_texto(v) for v in row]
        return (
            any(self._eh_coluna_loja_ou_matricula(c) for c in norm)
            and any(self._eh_coluna_qtd(c) for c in norm)
            and any(self._eh_coluna_sku(c) for c in norm)
        )

    def linha_eh_total(self, row: list[Any]) -> bool:
        norm = [self.normalizar_texto(v) for v in row]
        return any("total" in c or "totais" in c for c in norm)

    @staticmethod
    def linha_vazia(row: list[Any]) -> bool:
        return not any(str(v).strip() for v in row if v is not None)

    def primeira_qtd_numerica(self, row: list[Any]) -> int | float | None:
        for valor in row:
            qtd = self.parse_qtd(valor)
            if isinstance(qtd, (int, float)):
                return qtd
        return None

    def pedido_col_apos_qtd(self, matrix: list[list[Any]], header_idx: int, qtd_col: int) -> int | None:
        col = qtd_col + 1
        hits = 0
        for row in matrix[header_idx + 1: header_idx + 20]:
            if col < len(row) and re.fullmatch(r"\d{4,}", self.somente_digitos(row[col])):
                hits += 1
        return col if hits >= 2 else None

    def processar_matrix(self, matrix: list[list[Any]], arquivo: str, aba: str = "Planilha") -> None:
        texto_geral = "\n".join(" ".join(str(x or "") for x in row) for row in matrix[:40])
        pedido_header = self.detectar_pedido_em_texto(texto_geral, arquivo)
        pedido_atual = pedido_header
        header_idx, cols = self.detectar_colunas(matrix)
        if header_idx is None:
            self.alertas.append(AlertaProcessamento(
                tipo="LAYOUT_NAO_IDENTIFICADO", arquivo=arquivo, aba=aba, linha="-", pedido=pedido_header,
                loja="", sku="", qtd="", mensagem="Não encontrei layout Iquegami válido: Loja/Ref./Qtde., Matrícula/SKU/Quantidade, Matrícula/Código/Quantidade ou compacto sem cabeçalho."
            ))
            self.log("ERRO", "layout_nao_identificado", arquivo=arquivo, aba=aba)
            return

        pedido_col = self.pedido_col_apos_qtd(matrix, header_idx, cols["qtd"])
        self.log("INFO", "layout_identificado", arquivo=arquivo, aba=aba, linha_cabecalho=header_idx + 1,
                 col_loja=cols["loja"] + 1, col_sku=cols["sku"] + 1, col_qtd=cols["qtd"] + 1,
                 col_pedido=(pedido_col + 1 if pedido_col is not None else "cabeçalho/linha"), pedido=pedido_header)

        linhas_lidas = 0
        linhas_totais_ignoradas = 0
        cabecalhos_repetidos_ignorados = 0
        ultima_loja = ""

        for r_idx in range(header_idx + 1, len(matrix)):
            row = matrix[r_idx]

            if self.linha_vazia(row):
                continue

            row_text = " ".join(str(x or "") for x in row)
            pedido_linha = self.detectar_pedido_em_texto(row_text, "")
            if pedido_linha and pedido_linha != pedido_atual:
                pedido_atual = pedido_linha
                self.log("INFO", "pedido_identificado_na_linha", arquivo=arquivo, aba=aba, linha=r_idx + 1, pedido=pedido_atual)

            if self.linha_eh_cabecalho(row):
                cabecalhos_repetidos_ignorados += 1
                self.log("INFO", "cabecalho_repetido_ignorado", arquivo=arquivo, aba=aba, linha=r_idx + 1)
                continue

            if self.linha_eh_total(row):
                linhas_totais_ignoradas += 1
                self.log("INFO", "linha_total_ignorada", arquivo=arquivo, aba=aba, linha=r_idx + 1, qtd_total=self.primeira_qtd_numerica(row))
                continue

            def get(col: int):
                return row[col] if col < len(row) else None

            loja_original = self.somente_digitos(get(cols["loja"]))
            sku = self.somente_digitos(get(cols["sku"]))
            qtd = self.parse_qtd(get(cols["qtd"]))
            pedido = self.somente_digitos(get(pedido_col)) if pedido_col is not None else ""
            pedido = pedido or pedido_linha or pedido_atual or pedido_header

            # Regra operacional Rede Iquegami:
            # alguns arquivos chegam sem número de pedido na origem.
            # Para essa rede, o marcador técnico oficial é ponto (.),
            # permitindo que a linha siga para o Modelo/Fila quando matrícula, SKU,
            # quantidade e data estiverem válidos.
            if not str(pedido or "").strip():
                pedido = "."

            # Alguns arquivos vêm com linhas quebradas: a matrícula aparece na primeira
            # linha do bloco e as linhas seguintes podem carregar somente SKU/Qtd.
            loja = loja_original
            herdou_loja = False
            if loja:
                ultima_loja = loja
            elif ultima_loja and (sku or isinstance(qtd, (int, float))):
                loja = ultima_loja
                herdou_loja = True

            # Linhas sem conteúdo de item real são ignoradas. Isso evita somar títulos,
            # espaçadores e resíduos do Excel como se fossem SKUs.
            if not loja and not sku and (qtd is None or qtd == ""):
                continue
            if not loja and not sku and isinstance(qtd, (int, float)):
                # Número isolado sem Loja/SKU não é item. Total já foi tratado acima;
                # este fallback evita falso positivo quando o layout quebra visualmente.
                self.log("WARN", "linha_numerica_sem_loja_sku_ignorada", arquivo=arquivo, aba=aba, linha=r_idx + 1, qtd=qtd)
                continue

            matricula_loja, cnpj, obs_de_para = self.resolver_matricula_cnpj(loja)
            linhas_lidas += 1

            problemas = []
            if not loja:
                problemas.append("Loja/Matrícula vazia")
            if herdou_loja:
                problemas.append(f"Loja/Matrícula herdada da linha anterior: {loja}")
            if loja and not cnpj:
                problemas.append(obs_de_para or "CNPJ não localizado para a matrícula informada")
            if not sku:
                problemas.append("SKU/Ref. Fornecedor vazio ou não identificado")
            if qtd is None or qtd == "":
                problemas.append("Quantidade vazia")
            elif not isinstance(qtd, (int, float)):
                problemas.append(f"Quantidade inválida/não numérica: {qtd}")
            # Pedido ausente já foi normalizado para "." na Rede Iquegami.
            # Portanto, não gera pendência/bloqueio somente por falta do número.

            status = "OK" if not problemas else "ALERTA"
            obs = "; ".join(problemas)
            item = ItemPedido(
                cnpj=cnpj,
                matricula_loja=matricula_loja,
                sku=sku,
                qtd=qtd,
                numero_pedido=pedido,
                arquivo_origem=arquivo,
                aba_origem=aba,
                linha_origem=r_idx + 1,
                status=status,
                observacao=obs,
            )
            self.itens.append(item)

            if problemas:
                self.alertas.append(AlertaProcessamento(
                    tipo="DADO_PENDENTE", arquivo=arquivo, aba=aba, linha=r_idx + 1, pedido=pedido,
                    loja=matricula_loja, sku=sku, qtd=qtd, mensagem=obs
                ))
        self.log("INFO", "arquivo_processado", arquivo=arquivo, aba=aba, linhas_lidas=linhas_lidas,
                 linhas_totais_ignoradas=linhas_totais_ignoradas, cabecalhos_repetidos_ignorados=cabecalhos_repetidos_ignorados)

    def ler_xlsx(self, path: str) -> Iterable[tuple[str, list[list[Any]]]]:
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            matrix = [list(row) for row in ws.iter_rows(values_only=True)]
            yield ws.title, matrix

    def ler_xls(self, path: str) -> Iterable[tuple[str, list[list[Any]]]]:
        # Preferência: xlrd. Fallback: leitor BIFF simples embutido.
        try:
            import xlrd  # type: ignore
            book = xlrd.open_workbook(path)
            for sheet in book.sheets():
                matrix = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
                yield sheet.name, matrix
            return
        except Exception as e:
            self.log("WARN", "xlrd_indisponivel_ou_falhou", detalhe=str(e), arquivo=os.path.basename(path))
        cells = read_xls_cells(path)
        yield "Planilha", cells_to_matrix(cells)

    def ler_pdf_textual(self, path: str) -> Iterable[tuple[str, list[list[Any]]]]:
        try:
            import pdfplumber  # type: ignore
        except Exception:
            self.alertas.append(AlertaProcessamento(
                tipo="PDF_NAO_PROCESSADO", arquivo=os.path.basename(path), aba="PDF", linha="-", pedido="",
                loja="", sku="", qtd="", mensagem="pdfplumber não instalado. Instale requirements.txt para habilitar leitura de PDF textual."
            ))
            return
        texto_total = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                texto_total.append(page.extract_text() or "")
        texto = "\n".join(texto_total)
        pedido = self.detectar_pedido_em_texto(texto, path)
        matrix = [[f"Pedido {pedido}"], ["Loja", "Ref. Fornecedor", "Qtde."]]
        for line in texto.splitlines():
            # Ex.: 7120291224 92522 25 236369
            m = re.search(r"\b(\d{8,14})\b\s+\b(\d{4,8})\b\s+(\d+(?:[\.,]\d+)?)\s*(\d{4,})?", line)
            if m:
                loja, sku, qtd, ped_linha = m.groups()
                matrix.append([loja, sku, qtd, ped_linha or pedido])
        yield "PDF", matrix

    def processar_arquivo(self, path: str) -> None:
        nome = os.path.basename(path)
        ext = Path(path).suffix.lower()
        self.log("INFO", "iniciando_arquivo", arquivo=nome)
        try:
            if ext == ".xlsx":
                for aba, matrix in self.ler_xlsx(path):
                    self.processar_matrix(matrix, nome, aba)
            elif ext == ".xls":
                for aba, matrix in self.ler_xls(path):
                    self.processar_matrix(matrix, nome, aba)
            elif ext == ".pdf":
                for aba, matrix in self.ler_pdf_textual(path):
                    self.processar_matrix(matrix, nome, aba)
            else:
                self.alertas.append(AlertaProcessamento(
                    tipo="EXTENSAO_NAO_SUPORTADA", arquivo=nome, aba="-", linha="-", pedido="",
                    loja="", sku="", qtd="", mensagem=f"Extensão não suportada: {ext}. Use .xls, .xlsx ou PDF textual."
                ))
                self.log("ERRO", "extensao_nao_suportada", arquivo=nome, extensao=ext)
        except Exception as e:
            self.alertas.append(AlertaProcessamento(
                tipo="ERRO_PROCESSAMENTO", arquivo=nome, aba="-", linha="-", pedido="",
                loja="", sku="", qtd="", mensagem=str(e)
            ))
            self.log("ERRO", "falha_arquivo", arquivo=nome, detalhe=str(e), traceback=traceback.format_exc())


    def gerar_excel(self, output_path: str) -> str:
        """Gera Excel no padrão Robô KOF: primeiro o modelo para enviar/digitar,
        depois validação, cadastros pendentes, alertas e logs.

        Regra preservada: nenhum item estruturalmente lido é descartado. Mesmo com
        alerta de pedido, matrícula/CNPJ ou SKU, o item aparece na validação e, se
        possuir os campos mínimos, também fica visível no modelo para conferência.
        """
        wb = Workbook()

        ws = wb.active
        ws.title = "Modelo Robô KOF para Enviar"
        headers_modelo = ["CNPJ", "MATRICULA", "SKU", "QTD", "NUMERO_PEDIDO"]
        ws.append(headers_modelo)
        for item in self.itens:
            # Mantém o item visível para validação/digitação quando houver campos mínimos.
            # CNPJ ausente fica em branco; a pendência vai para Cadastrar CNPJ.
            ws.append([item.cnpj, item.matricula_loja, item.sku, item.qtd, item.numero_pedido])
        self._formatar_planilha(ws, cor="#7A0000")
        for col in ["A", "B", "C", "E"]:
            for cell in ws[col][1:]:
                cell.number_format = "@"

        wv = wb.create_sheet("Validação do Pedido")
        headers_val = [
            "REDE", "CNPJ", "MATRICULA", "SKU", "QTD", "NUMERO_PEDIDO",
            "ARQUIVO_ORIGEM", "ABA_ORIGEM", "LINHA_ORIGEM", "STATUS", "OBSERVACAO"
        ]
        wv.append(headers_val)
        for item in self.itens:
            wv.append(["IQUEGAMI", item.cnpj, item.matricula_loja, item.sku, item.qtd, item.numero_pedido,
                       item.arquivo_origem, item.aba_origem, item.linha_origem, item.status, item.observacao])
        self._formatar_planilha(wv, cor="#111111")
        for col in ["B", "C", "D", "F"]:
            for cell in wv[col][1:]:
                cell.number_format = "@"

        wc = wb.create_sheet("Cadastrar CNPJ")
        wc.append(["REDE", "MATRICULA", "CNPJ", "SKU", "QTD", "NUMERO_PEDIDO", "ARQUIVO_ORIGEM", "LINHA_ORIGEM", "MOTIVO"])
        for item in self.itens:
            if item.matricula_loja and not item.cnpj:
                wc.append(["IQUEGAMI", item.matricula_loja, "A CADASTRAR", item.sku, item.qtd, item.numero_pedido,
                           item.arquivo_origem, item.linha_origem, "Matrícula sem CNPJ no de/para Iquegami"])
        self._formatar_planilha(wc, cor="#7A0000")
        for col in ["B", "C", "D", "F"]:
            for cell in wc[col][1:]:
                cell.number_format = "@"

        wa = wb.create_sheet("Alertas/Erros")
        wa.append(["TIPO", "ARQUIVO", "ABA", "LINHA", "PEDIDO", "MATRICULA/LOJA", "SKU", "QTD", "MENSAGEM"])
        for a in self.alertas:
            wa.append([a.tipo, a.arquivo, a.aba, a.linha, a.pedido, a.loja, a.sku, a.qtd, a.mensagem])
        self._formatar_planilha(wa, cor="#111111")

        wl = wb.create_sheet("LOG_PROCESSAMENTO")
        wl.append(["DATA_HORA", "NIVEL", "EVENTO", "DETALHE", "EXTRAS_JSON"])
        for l in self.logs:
            base = [l.get("data_hora", ""), l.get("nivel", ""), l.get("evento", ""), l.get("detalhe", "")]
            extras = {k: v for k, v in l.items() if k not in {"data_hora", "nivel", "evento", "detalhe"}}
            wl.append(base + [json.dumps(extras, ensure_ascii=False)])
        self._formatar_planilha(wl, cor="#7A0000")

        wr = wb.create_sheet("RESUMO")
        total_ok = sum(1 for i in self.itens if i.status == "OK")
        pedidos = sorted({i.numero_pedido for i in self.itens if i.numero_pedido})
        arquivos = sorted({i.arquivo_origem for i in self.itens})
        wr.append(["INDICADOR", "VALOR"])
        wr.append(["Rede", "IQUEGAMI"])
        wr.append(["Arquivos processados", len(arquivos)])
        wr.append(["Pedidos identificados", len(pedidos)])
        wr.append(["Itens totais", len(self.itens)])
        wr.append(["Itens OK", total_ok])
        wr.append(["Itens com alerta", len(self.itens) - total_ok])
        wr.append(["Alertas", len(self.alertas)])
        wr.append(["Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
        wr.append([])
        wr.append(["PEDIDOS", ", ".join(pedidos)])
        self._formatar_planilha(wr, cor="#111111")

        wd = wb.create_sheet("DE_PARA_IQUEGAMI")
        wd.append(["MATRICULA", "CNPJ"])
        for matricula, cnpj in sorted(self.de_para_matricula_cnpj.items()):
            wd.append([matricula, cnpj])
        self._formatar_planilha(wd, cor="#7A0000")
        for col in ["A", "B"]:
            for cell in wd[col][1:]:
                cell.number_format = "@"

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    def modelo_robo_kof_rows(self) -> list[dict[str, Any]]:
        """Retorna linhas prontas para o writer central do Robô KOF.
        Colunas mantidas no padrão operacional: CNPJ, matrícula, SKU, quantidade e pedido.
        """
        return [{
            "REDE": "IQUEGAMI",
            "CNPJ": item.cnpj,
            "MATRICULA": item.matricula_loja,
            "SKU": item.sku,
            "QTD": item.qtd,
            "NUMERO_PEDIDO": item.numero_pedido,
            "STATUS": item.status,
            "OBSERVACAO": item.observacao,
            "ARQUIVO_ORIGEM": item.arquivo_origem,
            "ABA_ORIGEM": item.aba_origem,
            "LINHA_ORIGEM": item.linha_origem,
        } for item in self.itens]

    @staticmethod
    def _formatar_planilha(ws, cor: str) -> None:
        header_fill = PatternFill("solid", fgColor=cor.replace("#", ""))
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.freeze_panes = "A2"
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            max_len = max(len(str(c.value or "")) for c in column_cells[:200])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)
        ws.auto_filter.ref = ws.dimensions



REDE = "IQUEGAMI"
NOMES_REDE = {"iquegami", "rede iquegami", "pedido iquegami"}


def identificar_rede_iquegami(nome_arquivo: str = "", texto: str = "", matrix: list[list[Any]] | None = None) -> bool:
    """Identificador leve para o roteador central do Robô KOF.
    Não interfere em outros layouts: só retorna True quando há evidência da rede.
    """
    nome = IquegamiParser.normalizar_texto(Path(str(nome_arquivo or "")).name)
    if "iquegami" in nome:
        return True
    texto_norm = IquegamiParser.normalizar_texto(texto or "")
    if "iquegami" in texto_norm:
        return True
    if matrix:
        amostra = " ".join(" ".join(str(c or "") for c in row[:8]) for row in matrix[:20])
        amostra_norm = IquegamiParser.normalizar_texto(amostra)
        if "iquegami" in amostra_norm:
            return True
        parser = IquegamiParser()
        header_idx, _ = parser.detectar_colunas(matrix)
        if header_idx is not None:
            # Só confirma sem nome quando o layout tem matrícula/SKU/quantidade típico.
            return True
    return False

# Aliases comuns usados por roteadores diferentes.
identificar_layout = identificar_rede_iquegami
can_handle = identificar_rede_iquegami
is_layout = identificar_rede_iquegami


def processar_lote(arquivos: list[str], output_dir: str | None = None) -> dict[str, Any]:
    """Processa lote Iquegami sem gerar TXT/fila KOF automaticamente.
    Primeiro output é Excel de validação no padrão Robô KOF.
    """
    parser = IquegamiParser()
    for arquivo in arquivos:
        parser.processar_arquivo(arquivo)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = output_dir or os.getcwd()
    excel_path = os.path.join(output_dir, f"VALIDACAO_IQUEGAMI_{stamp}.xlsx")
    parser.gerar_excel(excel_path)
    log_json = os.path.join(output_dir, f"LOG_IQUEGAMI_{stamp}.json")
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    with open(log_json, "w", encoding="utf-8") as f:
        json.dump({"itens": [asdict(i) for i in parser.itens], "alertas": [asdict(a) for a in parser.alertas], "logs": parser.logs}, f, ensure_ascii=False, indent=2)
    log_txt = os.path.join(output_dir, f"LOG_IQUEGAMI_{stamp}.txt")
    with open(log_txt, "w", encoding="utf-8") as f:
        for l in parser.logs:
            f.write(f"[{l.get('data_hora')}] {l.get('nivel')} - {l.get('evento')} - {l.get('detalhe')} - {l}\n")
    return {
        "rede": "IQUEGAMI",
        "layout": "IQUEGAMI_EXCEL_MULTI_LAYOUT",
        "excel": excel_path,
        "log_json": log_json,
        "log_txt": log_txt,
        "total_itens": len(parser.itens),
        "total_alertas": len(parser.alertas),
        "pedidos": sorted({i.numero_pedido for i in parser.itens if i.numero_pedido}),
        "arquivos": [os.path.basename(a) for a in arquivos],
        "modelo_robo_kof": parser.modelo_robo_kof_rows(),
        "alertas": [asdict(a) for a in parser.alertas],
        "logs": parser.logs,
    }


def processar_arquivo(path: str, output_dir: str | None = None) -> dict[str, Any]:
    return processar_lote([path], output_dir)


def extrair_itens(path: str) -> list[dict[str, Any]]:
    return processar_lote([path], None)["modelo_robo_kof"]

# Mais aliases para facilitar encaixe em estruturas antigas do Robô KOF.
processar = processar_lote
processar_lote_iquegami = processar_lote
processar_iquegami = processar_lote
extrair_pedido = processar_arquivo
extrair_pedidos = processar_lote
