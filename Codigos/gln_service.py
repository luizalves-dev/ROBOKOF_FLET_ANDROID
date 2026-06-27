from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, Optional
import math
import re

import openpyxl

from terminal_logger import get_terminal_logger


terminal_log = get_terminal_logger("gln")
AMBIGUOUS_VALUE = "__AMBIGUO__"
_LOGGED_MISSING_CNPJ: set[str] = set()
_LOGGED_AMBIGUOUS_CNPJ: set[str] = set()
_LOGGED_AMBIGUOUS_KEYS_ON_LOAD: set[str] = set()
_GLN_MAP_CACHE: dict[tuple[str, str, str, str, str, str], Dict[str, str]] = {}
_CNPJ_MAP_CACHE: dict[tuple[str, str, str, str, str, str], Dict[str, str]] = {}


def _workbook_cache_key(path: Path, sheet_name: str, col_a: str, col_b: str) -> tuple[str, str, str, str, str, str]:
    path = Path(path)
    try:
        resolved = str(path.resolve())
    except Exception:
        resolved = str(path)
    try:
        stat = path.stat()
        return resolved, str(stat.st_mtime_ns), str(stat.st_size), str(sheet_name), str(col_a), str(col_b)
    except Exception:
        return resolved, "", "", str(sheet_name), str(col_a), str(col_b)



def clean_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return ""
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return ""
        return str(int(v)) if v.is_integer() else str(v)
    return str(v).strip()


def _number_text_to_digits(text: str) -> str:
    raw = str(text or "").strip()
    if not raw:
        return ""
    normalizado = raw.replace(",", ".")
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?", normalizado):
        try:
            decimal = Decimal(normalizado)
            if decimal == decimal.to_integral_value():
                return str(int(decimal))
        except (InvalidOperation, ValueError):
            pass
    if re.fullmatch(r"\d+\.0+", raw):
        return raw.split(".", 1)[0]
    return ""


def only_digits(v) -> str:
    texto = clean_str(v)
    convertido = _number_text_to_digits(texto)
    if convertido:
        return convertido
    return re.sub(r"\D+", "", texto)


def normalize_cnpj_key(v) -> str:
    digits = only_digits(v)
    if not digits:
        return ""
    return digits.lstrip("0") or "0"


def normalize_cnpj_keys(v) -> list[str]:
    digits = only_digits(v)
    if not digits:
        return []
    candidatos = [
        digits,
        digits.zfill(14) if len(digits) <= 14 else digits,
        digits.lstrip("0") or "0",
    ]
    unicos: list[str] = []
    for candidato in candidatos:
        if candidato and candidato not in unicos:
            unicos.append(candidato)
    return unicos


def _registrar_chave_mapa(cnpj_map: Dict[str, str], chave: str, matricula: str):
    if not chave or not matricula:
        return
    existente = cnpj_map.get(chave)
    if existente and existente not in {matricula, AMBIGUOUS_VALUE}:
        cnpj_map[chave] = AMBIGUOUS_VALUE
        if chave not in _LOGGED_AMBIGUOUS_KEYS_ON_LOAD:
            _LOGGED_AMBIGUOUS_KEYS_ON_LOAD.add(chave)
            terminal_log.warning("[GLN] Chave CNPJ ambigua no de/para: %s", chave)
        return
    if existente != AMBIGUOUS_VALUE:
        cnpj_map[chave] = matricula


def load_gln_map(
    gln_base_path: Path,
    sheet_name: str,
    col_gln: str,
    col_matricula: str,
) -> Dict[str, str]:
    if not gln_base_path.exists():
        raise FileNotFoundError(f"Base de GLNs nao encontrada: {gln_base_path}")

    cache_key = _workbook_cache_key(gln_base_path, sheet_name, col_gln, col_matricula)
    cached = _GLN_MAP_CACHE.get(cache_key)
    if cached is not None:
        return cached

    wb = openpyxl.load_workbook(gln_base_path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' nao existe em {gln_base_path}. Sheets: {wb.sheetnames}")

        ws = wb[sheet_name]
        idx_gln = openpyxl.utils.column_index_from_string(col_gln) - 1
        idx_mat = openpyxl.utils.column_index_from_string(col_matricula) - 1

        gln_map: Dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            matricula = only_digits(row[idx_mat]) if idx_mat < len(row) else ""
            gln = only_digits(row[idx_gln]) if idx_gln < len(row) else ""
            if matricula and gln:
                gln_map[matricula] = gln
        _GLN_MAP_CACHE.clear()
        _GLN_MAP_CACHE[cache_key] = gln_map
        return gln_map
    finally:
        wb.close()


def load_cnpj_to_matricula_map(
    gln_base_path: Path,
    sheet_name: str,
    col_cnpj: str,
    col_matricula: str,
) -> Dict[str, str]:
    if not gln_base_path.exists():
        raise FileNotFoundError(f"Base de GLNs nao encontrada: {gln_base_path}")

    cache_key = _workbook_cache_key(gln_base_path, sheet_name, col_cnpj, col_matricula)
    cached = _CNPJ_MAP_CACHE.get(cache_key)
    if cached is not None:
        return cached

    wb = openpyxl.load_workbook(gln_base_path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' nao existe em {gln_base_path}. Sheets: {wb.sheetnames}")

        ws = wb[sheet_name]
        idx_cnpj = openpyxl.utils.column_index_from_string(col_cnpj) - 1
        idx_mat = openpyxl.utils.column_index_from_string(col_matricula) - 1

        cnpj_map: Dict[str, str] = {}
        linhas_validas = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue

            matricula = only_digits(row[idx_mat]) if idx_mat < len(row) else ""
            cnpj_raw = row[idx_cnpj] if idx_cnpj < len(row) else ""
            chaves = normalize_cnpj_keys(cnpj_raw)

            if matricula and chaves:
                linhas_validas += 1
                for chave in chaves:
                    _registrar_chave_mapa(cnpj_map, chave, matricula)

        terminal_log.info(
            "[GLN] De/para CNPJ x matricula carregado/cacheado | linhas=%s | chaves=%s",
            linhas_validas,
            len(cnpj_map),
        )
        _CNPJ_MAP_CACHE.clear()
        _CNPJ_MAP_CACHE[cache_key] = cnpj_map
        return cnpj_map
    finally:
        wb.close()


def buscar_matricula_por_cnpj(
    cnpj: str,
    gln_base_path: Path,
    sheet_name: str,
    col_cnpj: str,
    col_matricula: str,
    mapa_cache: Optional[Dict[str, str]] = None,
) -> str:
    cnpj_original = only_digits(cnpj)
    chaves = normalize_cnpj_keys(cnpj)

    if not chaves:
        terminal_log.warning("[GLN] CNPJ vazio/ilegivel para busca de matricula: %s", cnpj)
        return ""

    mapa = mapa_cache or load_cnpj_to_matricula_map(
        gln_base_path=gln_base_path,
        sheet_name=sheet_name,
        col_cnpj=col_cnpj,
        col_matricula=col_matricula,
    )

    for chave in chaves:
        matricula = mapa.get(chave, "")
        if matricula == AMBIGUOUS_VALUE:
            if chave not in _LOGGED_AMBIGUOUS_CNPJ:
                _LOGGED_AMBIGUOUS_CNPJ.add(chave)
                terminal_log.warning("[GLN] CNPJ ambiguo no de/para | original=%s | chave=%s", cnpj_original, chave)
            return ""
        if matricula:
            terminal_log.debug("[GLN] CNPJ localizado | original=%s | chave=%s | matricula=%s", cnpj_original, chave, matricula)
            return matricula

    chave_log = chaves[0]
    if chave_log not in _LOGGED_MISSING_CNPJ:
        _LOGGED_MISSING_CNPJ.add(chave_log)
        terminal_log.warning("[GLN] CNPJ sem matricula no de/para | original=%s | chaves=%s", cnpj_original, chaves)
    return ""
