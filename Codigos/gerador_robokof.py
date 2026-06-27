from pathlib import Path
from typing import Dict, List, Tuple, Optional
import openpyxl


def clean_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def set_cell_general(ws, row: int, col: int, value, force_text: bool = False):
    cell = ws.cell(row, col)

    if force_text:
        cell.value = str(value) if value is not None else ""
        cell.number_format = "@"
        return

    try:
        if value is not None and str(value).isdigit():
            # evita converter números muito grandes para int no Excel
            if len(str(value)) >= 15:
                cell.value = str(value)
                cell.number_format = "@"
            else:
                cell.value = int(value)
                cell.number_format = "General"
        else:
            cell.value = str(value)
            cell.number_format = "General"
    except Exception:
        cell.value = str(value)
        cell.number_format = "General"


def _find_header_row_and_cols(ws, header_candidates: Dict[str, List[str]], search_rows: int = 80) -> Tuple[int, Dict[str, int]]:
    best_row = None
    best_map: Dict[str, int] = {}

    keys = list(header_candidates.keys())

    for r in range(1, min(search_rows, ws.max_row) + 1):
        row_vals = [clean_str(c.value) for c in ws[r]]
        col_map: Dict[str, int] = {}

        for key in keys:
            for cand in header_candidates[key]:
                if cand in row_vals:
                    col_map[key] = row_vals.index(cand) + 1
                    break

        if len(col_map) > len(best_map):
            best_map = col_map
            best_row = r

    if not best_row:
        raise ValueError("Não encontrei a linha de cabeçalho na sheet 'Ordem' do template.")

    required = ["Matricula", "Sku", "Qtd", "Pedido", "Data"]
    missing = [k for k in required if k not in best_map]
    if missing:
        raise ValueError(f"Template: cabeçalhos obrigatórios não encontrados: {missing}. Achados: {best_map}")

    return best_row, best_map


def _clear_data_area(ws, start_row: int, cols_to_clear: List[int], max_rows: int = 5000):
    end = min(ws.max_row, start_row + max_rows)

    for r in range(start_row, end + 1):
        if all(ws.cell(r, c).value in (None, "") for c in cols_to_clear):
            break

        for c in cols_to_clear:
            ws.cell(r, c).value = None


def _write_gln(ws, gln_value: str):
    if not gln_value:
        return

    gln_col = 10  # J

    for r in range(1, ws.max_row + 1):
        if clean_str(ws.cell(r, gln_col).value).upper() == "GLN":
            cell = ws.cell(r + 1, gln_col)

            try:
                cell.value = int(gln_value)
            except Exception:
                cell.value = gln_value

            cell.number_format = "0"
            return


def gerar_arquivo_robokof(
    template_path: Path,
    out_path: Path,
    sheet_name: str,
    header_candidates: Dict[str, List[str]],
    rows: List[dict],
    gln_value: Optional[str],
    tipo_solicitacao_value: str,
    forma_pagamento_value: str
):
    if not template_path.exists():
        raise FileNotFoundError(f"Template não encontrado: {template_path}")

    wb = openpyxl.load_workbook(template_path)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' não existe no template. Sheets: {wb.sheetnames}")

    ws = wb[sheet_name]

    header_row, col_map = _find_header_row_and_cols(ws, header_candidates)
    data_start = header_row + 1

    cols_to_clear = [
        col_map["Matricula"],
        col_map["Sku"],
        col_map["Qtd"],
        col_map["Pedido"],
        col_map["Data"],
    ]

    if "DescricaoSku" in col_map:
        cols_to_clear.append(col_map["DescricaoSku"])

    if "TipoSolicitacao" in col_map:
        cols_to_clear.append(col_map["TipoSolicitacao"])

    if "FormaPagamento" in col_map:
        cols_to_clear.append(col_map["FormaPagamento"])

    _clear_data_area(ws, data_start, cols_to_clear)

    for i, r in enumerate(rows):
        rr = data_start + i

        set_cell_general(ws, rr, col_map["Matricula"], r["Matricula"])
        set_cell_general(ws, rr, col_map["Sku"], r["Sku"], force_text=True)

        cell_qtd = ws.cell(rr, col_map["Qtd"])
        cell_qtd.value = int(r["Qtd"])
        cell_qtd.number_format = "0"

        if "DescricaoSku" in col_map:
            cell_desc = ws.cell(rr, col_map["DescricaoSku"])
            cell_desc.value = "."
            cell_desc.number_format = "General"

        # Pedido deve ser texto para não virar notação científica
        set_cell_general(ws, rr, col_map["Pedido"], r["Pedido"], force_text=True)

        ws.cell(rr, col_map["Data"]).value = r["Data"]

        if "TipoSolicitacao" in col_map:
            ws.cell(rr, col_map["TipoSolicitacao"]).value = tipo_solicitacao_value

        if "FormaPagamento" in col_map:
            ws.cell(rr, col_map["FormaPagamento"]).value = forma_pagamento_value

    if gln_value:
        _write_gln(ws, gln_value)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)