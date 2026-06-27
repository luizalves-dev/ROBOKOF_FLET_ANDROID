from __future__ import annotations

import csv
import os
import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

import pandas as pd
from openpyxl import load_workbook

import config
from terminal_logger import get_terminal_logger


terminal_log = get_terminal_logger("conversao")

STATUS_OK_CONVERTIDO = "OK CONVERTIDO"
STATUS_OK_SEM_CONVERSAO = "OK SEM CONVERSÃO"
STATUS_ALERTA_NAO_CONVERTIDO = "ALERTA - NÃO CONVERTIDO"
STATUS_VALIDAR_CONVERSAO = "VALIDAR CONVERSÃO"

CSV_COLUMNS = [
    "regra_id",
    "rede_layout",
    "layout_id",
    "centro",
    "centro_referencia",
    "centro_alternativo",
    "sku",
    "ean",
    "descricao",
    "fator_conversao",
    "tipo_regra",
    "usa_mapa_produtos",
    "origem_regra",
    "prioridade",
    "ativo",
    "sku_destino",
    "arredondamento",
    "observacao",
    "data_atualizacao",
]


@dataclass
class RegraConversao:
    regra_id: str = ""
    rede_layout: str = ""
    layout_id: str = ""
    centro: str = ""
    centro_referencia: str = ""
    centro_alternativo: str = ""
    sku: str = ""
    ean: str = ""
    descricao: str = ""
    fator_conversao: str = ""
    tipo_regra: str = ""
    usa_mapa_produtos: str = ""
    origem_regra: str = ""
    prioridade: int = 999
    ativo: bool = True
    sku_destino: str = ""
    arredondamento: str = ""
    observacao: str = ""
    data_atualizacao: str = ""

    @property
    def tipo(self) -> str:
        return normalize_key(self.tipo_regra).upper()

    @property
    def fator_decimal(self) -> Optional[Decimal]:
        return parse_decimal(self.fator_conversao)

    @property
    def usa_mapa(self) -> bool:
        valor = str(self.usa_mapa_produtos or "").strip().upper()
        return valor in {"1", "S", "SIM", "Y", "YES", "TRUE", "VERDADEIRO"} or self.tipo in {
            "MAPA_PRODUTOS", "MAPA_PRODUTOS_FALLBACK", "FALLBACK_MAPA", "MAPA"
        }


@dataclass
class MapaProduto:
    centro: str
    ean: str
    sku: str
    descricao: str
    apresentacao: Decimal
    origem: str
    linha_mapa: int
    dun: str = ""
    sku_base: str = ""


def normalize_text(value: Any) -> str:
    text = str(value or "").strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_key(value: Any) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def only_digits(value: Any) -> str:
    return re.sub(r"\D", "", str(value or ""))


def parse_decimal(value: Any) -> Optional[Decimal]:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null", "*"}:
        return None
    text = text.replace("R$", "").replace(" ", "")
    # 1.234,56 -> 1234.56 | 12,000 -> 12.000 | 12.0 ok
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def format_decimal(value: Any) -> str:
    dec = value if isinstance(value, Decimal) else parse_decimal(value)
    if dec is None:
        return ""
    if dec == dec.to_integral_value():
        return str(int(dec))
    out = dec.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP).normalize()
    return format(out, "f").replace(".", ",")


def _layout_id(layout_config: Dict[str, Any]) -> str:
    return str(layout_config.get("layout_id", "") or "").strip()


def _layout_name(layout_config: Dict[str, Any]) -> str:
    return str(layout_config.get("nome_layout", "") or "").strip()


def _match_text(value: str, pattern: str) -> bool:
    pattern = str(pattern or "").strip()
    if not pattern or pattern == "*":
        return True
    v_key = normalize_key(value)
    p_key = normalize_key(pattern)
    if not p_key:
        return True
    return p_key == v_key or p_key in v_key or v_key in p_key


def _match_digits(value: Any, pattern: Any) -> bool:
    pattern_text = str(pattern or "").strip()
    if not pattern_text or pattern_text == "*":
        return True
    value_digits = only_digits(value)
    pattern_digits = only_digits(pattern_text)
    if not pattern_digits:
        return True
    if not value_digits:
        return False
    return value_digits == pattern_digits or value_digits.lstrip("0") == pattern_digits.lstrip("0")


def _is_active(value: Any) -> bool:
    text = str(value or "").strip().upper()
    return text not in {"0", "FALSE", "FALSO", "N", "NAO", "NÃO", "INATIVO"}


def _to_int(value: Any, default: int = 999) -> int:
    try:
        return int(float(str(value or "").replace(",", ".")))
    except Exception:
        return default


def _row_to_rule(row: Dict[str, Any]) -> RegraConversao:
    normalized = {normalize_key(k): v for k, v in row.items()}
    return RegraConversao(
        regra_id=str(normalized.get("regra_id", "") or ""),
        rede_layout=str(normalized.get("rede_layout", normalized.get("nome_layout", "")) or ""),
        layout_id=str(normalized.get("layout_id", "") or ""),
        centro=str(normalized.get("centro", "") or "").strip().upper(),
        centro_referencia=str(
            normalized.get("centro_referencia", normalized.get("centro_de_referencia", normalized.get("centro_ref", ""))) or ""
        ).strip().upper(),
        centro_alternativo=str(
            normalized.get("centro_alternativo", normalized.get("centro_de_apoio", normalized.get("centro_alterno", ""))) or ""
        ).strip(),
        sku=only_digits(normalized.get("sku", normalized.get("sku_origem", ""))),
        ean=only_digits(normalized.get("ean", "")),
        descricao=str(normalized.get("descricao", "") or ""),
        fator_conversao=str(normalized.get("fator_conversao", "") or ""),
        tipo_regra=str(normalized.get("tipo_regra", "") or ""),
        usa_mapa_produtos=str(normalized.get("usa_mapa_produtos", normalized.get("usa_mapa", "")) or ""),
        origem_regra=str(normalized.get("origem_regra", "") or ""),
        prioridade=_to_int(normalized.get("prioridade", "999")),
        ativo=_is_active(normalized.get("ativo", "1")),
        sku_destino=only_digits(normalized.get("sku_destino", "")),
        arredondamento=str(normalized.get("arredondamento", "") or ""),
        observacao=str(normalized.get("observacao", normalized.get("observacoes", "")) or ""),
        data_atualizacao=str(normalized.get("data_atualizacao", "") or ""),
    )


def carregar_regras(path: Path | None = None) -> list[RegraConversao]:
    path = path or getattr(config, "REGRAS_CONVERSAO_PATH", config.CADASTROS_DIR / "regras_conversao.csv")
    path = Path(path)
    if not path.exists():
        terminal_log.warning("[CONVERSAO] CSV de regras nao encontrado: %s", path)
        return []

    cache_key = _path_cache_key(path)
    cached = _REGRAS_CACHE.get(cache_key)
    if cached is not None:
        return cached

    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as fh:
                rows = list(csv.DictReader(fh))
            regras = [regra for regra in (_row_to_rule(row) for row in rows) if regra.ativo]
            _REGRAS_CACHE.clear()
            _REGRAS_CACHE[cache_key] = regras
            terminal_log.info("[CONVERSAO] Regras de conversao carregadas/cacheadas | arquivo=%s | regras=%s", path, len(regras))
            return regras
        except UnicodeDecodeError:
            continue
    terminal_log.exception("[CONVERSAO] Falha ao carregar CSV de regras: %s", path)
    return []


def regras_do_layout(layout_config: Dict[str, Any], regras: Optional[list[RegraConversao]] = None) -> list[RegraConversao]:
    regras = regras if regras is not None else carregar_regras()
    lid = _layout_id(layout_config)
    nome = _layout_name(layout_config)
    selecionadas: list[RegraConversao] = []
    for regra in regras:
        match_id = bool(regra.layout_id) and regra.layout_id == lid
        match_nome = bool(regra.rede_layout) and _match_text(nome, regra.rede_layout)
        if match_id or match_nome:
            selecionadas.append(regra)
    return sorted(selecionadas, key=lambda r: (r.prioridade, -_specificity_score(r)))


def layout_tem_conversao(layout_config: Dict[str, Any]) -> bool:
    return any(regra.tipo not in {"SEM_CONVERSAO", "SEM CONVERSAO", "SEM_CONVERSÃO"} for regra in regras_do_layout(layout_config))


def _specificity_score(regra: RegraConversao) -> int:
    score = 0
    for attr in ("centro", "sku", "ean", "descricao"):
        if str(getattr(regra, attr, "") or "").strip() not in {"", "*"}:
            score += 1
    return score


def _linha_centro(row: pd.Series, regra: RegraConversao | None = None) -> str:
    for col in ["centro_lido", "centro", "Centro", "centro_sap_lido", "codigo_centro_lido"]:
        value = str(row.get(col, "") or "").strip().upper()
        if value:
            return value
    if regra and regra.centro:
        if regra.centro == "*":
            return ""
        return regra.centro
    return str(getattr(config, "CENTRO_CONVERSAO_PADRAO", "") or "").strip().upper()





def _centro_operacional_lido(row: pd.Series) -> str:
    for col in ["centro_lido", "centro", "Centro", "centro_sap_lido", "codigo_centro_lido"]:
        value = str(row.get(col, "") or "").strip().upper()
        if value:
            return value
    return ""

def _centro_referencia_mapa(row: pd.Series, regra: RegraConversao | None = None) -> tuple[str, str]:
    """Retorna (codigo_centro_referencia, label) usado para consulta no mapa.

    A regra de negócio é: o centro operacional lido do arquivo pode ser diferente do centro
    de referência para conversão. Ex.: Rede Celeiro deve usar Chapecó/BFDZ mesmo que o
    PDF não informe esse centro. Esta função mantém a regra rastreável sem misturar layouts.
    """
    if regra and regra.centro_referencia:
        codigo = regra.centro_referencia.strip().upper()
        label = regra.centro_alternativo.strip() if regra.centro_alternativo else ""
        return codigo, f"{codigo} - {label}" if label else codigo
    if regra and regra.centro and regra.centro != "*":
        return regra.centro.strip().upper(), regra.centro.strip().upper()
    centro_linha = _linha_centro(row, None)
    if centro_linha:
        return centro_linha.strip().upper(), centro_linha.strip().upper()
    padrao = str(getattr(config, "CENTRO_CONVERSAO_PADRAO", "") or "").strip().upper()
    return padrao, padrao


def _is_mapa_rule(regra: RegraConversao | None) -> bool:
    if not regra:
        return False
    return regra.tipo in {"MAPA_PRODUTOS", "MAPA_PRODUTOS_FALLBACK", "FALLBACK_MAPA", "MAPA"} or regra.usa_mapa


def _rule_matches_row(regra: RegraConversao, row: pd.Series, layout_config: Dict[str, Any]) -> bool:
    centro_row = _linha_centro(row, regra)
    sku_row = row.get("sku_lido", "") or row.get("codigo_sku_lido", "")
    ean_row = row.get("ean_lido", "")
    descricao_row = row.get("descricao_lida", "")

    if regra.centro and regra.centro != "*" and centro_row and regra.centro.upper() != centro_row.upper():
        return False
    if regra.sku and regra.sku != "*" and not _match_digits(sku_row, regra.sku):
        return False
    if regra.ean and regra.ean != "*" and not _match_digits(ean_row, regra.ean):
        return False
    if regra.descricao and regra.descricao != "*" and not _match_text(str(descricao_row), regra.descricao):
        return False
    return True


def _best_rule_for_row(row: pd.Series, layout_config: Dict[str, Any], regras_layout: list[RegraConversao]) -> Optional[RegraConversao]:
    matches = [regra for regra in regras_layout if _rule_matches_row(regra, row, layout_config)]
    if not matches:
        return None
    return sorted(matches, key=lambda r: (r.prioridade, -_specificity_score(r)))[0]


def _header_key(value: Any) -> str:
    text = normalize_key(value).replace("_", " ")
    replacements = {
        "codigo ean": "codigo ean",
        "cod ean": "codigo ean",
        "ean": "codigo ean",
        "cod barras": "codigo ean",
        "codigo barras": "codigo ean",
        "codigo de barras": "codigo ean",
        "codigo dun": "codigo dun",
        "cod dun": "codigo dun",
        "dun": "codigo dun",
        "sku ecc": "sku ecc",
        "sku": "sku ecc",
        "codigo femsa": "sku ecc",
        "codigo": "sku ecc",
        "sku base": "sku base",
        "descricao sku ecc": "descricao sku ecc",
        "descricao": "descricao sku ecc",
        "descrição sku ecc": "descricao sku ecc",
        "apresentacao": "apresentacao",
        "apresentação": "apresentacao",
        "conversao": "apresentacao",
        "centro": "centro",
        "regional": "regional",
        "nome uo": "nome uo",
    }
    return replacements.get(text, text)


def _map_paths() -> list[Path]:
    paths: list[Path] = []
    env = os.getenv("ROBOKOF_MAPA_PRODUTOS", "")
    if env:
        paths.append(Path(env))
    for attr in [
        "MAPA_PRODUTOS_PATH",
        "MAPA_PRODUTOS_PADRAO",
        "MAPA_PRODUTOS_OFICIAL_PATH",
        "MAPA_PRODUTOS_PROJETO_PATH",
        "MAPA_PRODUTOS_LEGADO_PATH",
    ]:
        value = getattr(config, attr, None)
        if value:
            paths.append(Path(str(value)))
    # Fallback defensivo: se o config.py antigo não tiver as constantes novas,
    # procura uma cópia local em Arquivos Base sem quebrar o fluxo.
    root_dir = Path(getattr(config, "ROOT_DIR", Path(__file__).resolve().parents[1]))
    paths.append(root_dir / "Arquivos Base" / "Mapa de Produtos" / "mapa de produtos mais atualizado 08.04.xlsx")
    paths.append(root_dir / "Arquivos Base" / "mapa de produtos mais atualizado 08.04.xlsx")
    return list(dict.fromkeys(paths))


_REGRAS_CACHE: dict[tuple[str, str, str], list[RegraConversao]] = {}
_MAP_CACHE: dict[tuple[str, str], list[MapaProduto]] = {}
_MAP_INDEX_CACHE: dict[int, dict[str, dict[str, list[MapaProduto]]]] = {}

# Log linha a linha deixa o processamento perceptivelmente mais lento em lotes
# grandes. Mantemos auditoria completa no Excel e no resumo, mas reduzimos o
# volume de INFO no terminal para evitar gargalo de I/O.
_CONVERSAO_ITEM_LOG_LIMIT = 25
_CONVERSAO_ITEM_LOG_EVERY = 100


def _path_cache_key(path: Path) -> tuple[str, str, str]:
    try:
        resolved = str(path.resolve())
    except Exception:
        resolved = str(path)
    try:
        stat = path.stat()
        return resolved, str(stat.st_mtime_ns), str(stat.st_size)
    except Exception:
        return resolved, "", ""


def _deve_logar_item_conversao(idx: int) -> bool:
    linha = int(idx) + 1
    return idx < _CONVERSAO_ITEM_LOG_LIMIT or (linha % _CONVERSAO_ITEM_LOG_EVERY == 0)


def _log_conversao_item(idx: int, mensagem: str, *args):
    if _deve_logar_item_conversao(idx):
        terminal_log.info(mensagem, *args)
    else:
        terminal_log.debug(mensagem, *args)



def _load_mapa_produtos(path: Path | None = None) -> list[MapaProduto]:
    chosen = path
    if chosen is None:
        for candidate in _map_paths():
            if candidate.exists():
                chosen = candidate
                break
    if chosen is None:
        return []
    chosen = Path(chosen)
    key = _path_cache_key(chosen)
    if key in _MAP_CACHE:
        return _MAP_CACHE[key]
    if not chosen.exists():
        return []
    wb = load_workbook(chosen, read_only=True, data_only=True)
    produtos: list[MapaProduto] = []
    try:
        header_info = None
        for ws in wb.worksheets:
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
                headers = [_header_key(v) for v in row]
                joined = "|".join(headers)
                if "sku ecc" in joined and "apresentacao" in joined and ("codigo ean" in joined or "centro" in joined):
                    header = {value: idx for idx, value in enumerate(headers) if value}
                    header_info = (ws, row_idx, header)
                    break
            if header_info:
                break
        if not header_info:
            return []
        ws, header_row, header = header_info
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            centro = str(_get(row, header, "centro") or "").strip().upper()
            ean = only_digits(_get(row, header, "codigo ean"))
            dun = only_digits(_get(row, header, "codigo dun"))
            sku = only_digits(_get(row, header, "sku ecc"))
            sku_base = only_digits(_get(row, header, "sku base"))
            if not sku and sku_base:
                sku = sku_base
            apresentacao = parse_decimal(_get(row, header, "apresentacao"))
            if not apresentacao or apresentacao == 0:
                continue
            if not sku and not ean and not dun:
                continue
            produtos.append(
                MapaProduto(
                    centro=centro,
                    ean=ean,
                    sku=sku,
                    descricao=str(_get(row, header, "descricao sku ecc") or "").strip(),
                    apresentacao=apresentacao,
                    origem=f"{chosen.name}::{ws.title}",
                    linha_mapa=row_idx,
                    dun=dun,
                    sku_base=sku_base,
                )
            )
    finally:
        wb.close()
    _MAP_CACHE[key] = produtos
    terminal_log.info("[CONVERSAO] Mapa de produtos carregado | arquivo=%s | linhas=%s", chosen, len(produtos))
    return produtos


def _get(row: tuple, header: dict[str, int], name: str) -> Any:
    idx = header.get(name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _codigo_matches(valor: Any, *candidatos: str) -> bool:
    val = only_digits(valor)
    if not val:
        return False
    for cand in candidatos:
        cand_digits = only_digits(cand)
        if cand_digits and _match_digits(val, cand_digits):
            return True
    return False


def _indexar_mapa_produtos(produtos: list[MapaProduto]) -> dict[str, dict[str, list[MapaProduto]]]:
    """Cria índices em memória para evitar varrer o mapa inteiro a cada linha.

    Antes, cada item fazia list comprehension sobre todos os produtos do mapa
    para EAN, DUN e SKU. Em lotes grandes isso vira um gargalo forte. Os índices
    abaixo mantêm a mesma regra de comparação por dígitos/zero à esquerda, mas
    reduzem a busca para poucas linhas candidatas.
    """
    cache_key = id(produtos)
    cached = _MAP_INDEX_CACHE.get(cache_key)
    if cached is not None:
        return cached

    indices: dict[str, dict[str, list[MapaProduto]]] = {
        "ean": {},
        "dun": {},
        "sku": {},
        "sku_base": {},
    }

    def add(campo: str, valor: Any, produto: MapaProduto):
        digits = only_digits(valor)
        if not digits:
            return
        variantes = [digits, digits.lstrip("0") or "0"]
        for chave in variantes:
            bucket = indices[campo].setdefault(chave, [])
            bucket.append(produto)

    for produto in produtos:
        add("ean", produto.ean, produto)
        add("dun", produto.dun, produto)
        add("sku", produto.sku, produto)
        add("sku_base", produto.sku_base, produto)

    _MAP_INDEX_CACHE.clear()
    _MAP_INDEX_CACHE[cache_key] = indices
    terminal_log.info(
        "[CONVERSAO] Indices do mapa de produtos gerados | produtos=%s | eans=%s | duns=%s | skus=%s",
        len(produtos), len(indices["ean"]), len(indices["dun"]), len(indices["sku"]),
    )
    return indices


def _buscar_no_indice_mapa(indices: dict[str, dict[str, list[MapaProduto]]], campos: tuple[str, ...], valor: Any) -> list[MapaProduto]:
    digits = only_digits(valor)
    if not digits:
        return []
    variantes = [digits, digits.lstrip("0") or "0"]
    encontrados: list[MapaProduto] = []
    vistos: set[int] = set()
    for campo in campos:
        mapa = indices.get(campo, {})
        for chave in variantes:
            for produto in mapa.get(chave, []):
                ident = id(produto)
                if ident not in vistos:
                    encontrados.append(produto)
                    vistos.add(ident)
    return encontrados


def _find_mapa(row: pd.Series, centro: str) -> tuple[Optional[MapaProduto], list[str]]:
    produtos = _load_mapa_produtos()
    alertas: list[str] = []
    if not produtos:
        paths = ", ".join(str(p) for p in _map_paths()) or "MAPA_PRODUTOS_PATH nao configurado"
        return None, [f"Mapa de produtos não localizado para fallback. Caminhos avaliados: {paths}"]

    ean = only_digits(row.get("ean_lido", ""))
    sku = only_digits(row.get("sku_lido", "") or row.get("codigo_sku_lido", ""))
    centro = str(centro or "").strip().upper()

    indices = _indexar_mapa_produtos(produtos)
    candidatos: list[MapaProduto] = []
    if ean:
        candidatos.extend(_buscar_no_indice_mapa(indices, ("ean", "dun"), ean))
    # Quando EAN e SKU vierem juntos, o SKU deve desempatar o mesmo EAN em SKUs/fatores diferentes.
    # Isso evita marcar VALIDAR CONVERSÃO sem necessidade quando o layout informa os dois campos.
    if sku and candidatos:
        filtrados_por_sku = [p for p in candidatos if _codigo_matches(sku, p.sku, p.sku_base)]
        if filtrados_por_sku:
            candidatos = filtrados_por_sku
    if sku and not candidatos:
        candidatos.extend(_buscar_no_indice_mapa(indices, ("sku", "sku_base"), sku))

    if centro and centro != "*":
        centro_candidatos = [p for p in candidatos if p.centro == centro]
    else:
        centro_candidatos = candidatos
        if candidatos:
            alertas.append("Centro não definido na regra/linha; mapa consultado sem filtro de centro. Validar se o centro operacional está correto.")

    if not centro_candidatos:
        if candidatos:
            centros = sorted({p.centro for p in candidatos if p.centro})[:18]
            alertas.append(f"Item existe no mapa, mas não para o centro {centro or '-'}; centros encontrados: {', '.join(centros)}")
        else:
            alertas.append("EAN/SKU não encontrado no mapa de produtos.")
        return None, alertas

    # Maior similaridade de descrição quando houver múltiplos candidatos.
    desc = normalize_key(row.get("descricao_lida", ""))

    def rank(p: MapaProduto):
        desc_norm = normalize_key(p.descricao)
        desc_score = 1 if desc and desc_norm and (desc in desc_norm or desc_norm in desc) else 0
        ean_score = 1 if ean and _codigo_matches(ean, p.ean) else 0
        dun_score = 1 if ean and _codigo_matches(ean, p.dun) else 0
        sku_score = 1 if sku and _codigo_matches(sku, p.sku, p.sku_base) else 0
        centro_score = 1 if centro and p.centro == centro else 0
        return (centro_score, ean_score, dun_score, sku_score, desc_score, len(p.descricao or ""), -p.linha_mapa)

    chosen = sorted(centro_candidatos, key=rank, reverse=True)[0]
    distinct = {(p.sku, str(p.apresentacao), p.descricao, p.centro) for p in centro_candidatos}
    distinct_fatores = {(p.sku, str(p.apresentacao)) for p in centro_candidatos}
    if len(distinct_fatores) > 1:
        alertas.append(
            "Mais de uma conversão possível no mapa; escolha automática por centro/descrição/linha. "
            + "; ".join(f"Centro {p.centro or '-'} SKU {p.sku}/Fator {p.apresentacao}/Linha {p.linha_mapa}" for p in centro_candidatos[:8])
        )
    elif len(distinct) > 1:
        alertas.append(
            "Mapa possui múltiplas linhas equivalentes para o item; fator único mantido, mas validar centro/descrição. "
            + "; ".join(f"Centro {p.centro or '-'} SKU {p.sku}/Fator {p.apresentacao}/Linha {p.linha_mapa}" for p in centro_candidatos[:8])
        )
    return chosen, alertas

def _calcular_qtd(qtd_original: Decimal, fator: Decimal) -> Decimal:
    if not fator or fator == 0:
        return qtd_original
    return qtd_original / fator




def _aviso_exige_validacao(alerta: str) -> bool:
    """Define quais avisos de mapa devem bloquear fila/TXT.

    Nem todo aviso significa risco de conversão. Ex.: o mapa pode ter duas linhas
    equivalentes com o mesmo SKU/fator; nesse caso registramos a rastreabilidade,
    mas não forçamos VALIDAR CONVERSÃO. Bloqueamos somente quando há dúvida real
    de centro, fator, arredondamento ou escolha entre fatores diferentes.
    """
    texto = normalize_text(alerta).upper()
    texto = texto.replace("NAO", "NÃO").replace("CONVERSAO", "CONVERSÃO")
    termos_bloqueantes = [
        "MAIS DE UMA CONVERSÃO POSSÍVEL",
        "MAIS DE UMA CONVERSAO POSSIVEL",
        "CENTRO NÃO DEFINIDO",
        "CENTRO NAO DEFINIDO",
        "SEM FILTRO DE CENTRO",
        "QUANTIDADE FRACIONADA",
        "ARREDONDAMENTO",
        "AMBÍGU",
        "AMBIGU",
    ]
    return any(termo in texto for termo in termos_bloqueantes)


def _avisos_exigem_validacao(alertas: Iterable[str] | None) -> bool:
    return any(_aviso_exige_validacao(str(alerta)) for alerta in (alertas or []))

def _aplicar_resultado(row: pd.Series, fator: Decimal, regra: RegraConversao | None, origem: str, sku_destino: str = "", alertas: Optional[list[str]] = None) -> dict[str, str]:
    qtd_original = parse_decimal(row.get("qtd_original", "")) or parse_decimal(row.get("quantidade_lida", "")) or Decimal("0")
    qtd_convertida = _calcular_qtd(qtd_original, fator)
    status = STATUS_OK_CONVERTIDO
    observacoes = list(alertas or [])
    if qtd_convertida != qtd_convertida.to_integral_value():
        status = STATUS_VALIDAR_CONVERSAO
        observacoes.append("Conversão resultou em quantidade fracionada; validar arredondamento antes de seguir para TXT/fila.")
    if observacoes and _avisos_exigem_validacao(observacoes):
        # Só bloqueia quando o aviso indicar risco real de fator/centro/arredondamento.
        # Duplicidades equivalentes do mapa ficam rastreadas na observação, mas não bloqueiam.
        status = STATUS_VALIDAR_CONVERSAO if status == STATUS_OK_CONVERTIDO else status
    regra_label = "MAPA_PRODUTOS"
    if regra:
        regra_label = f"{regra.regra_id or '-'} | {regra.tipo_regra or 'REGRA_CSV'} | prioridade {regra.prioridade}"
    return {
        "sku_lido": sku_destino or only_digits(row.get("sku_lido", "")) or only_digits(row.get("codigo_sku_lido", "")),
        "codigo_sku_lido": sku_destino or only_digits(row.get("codigo_sku_lido", "")) or only_digits(row.get("sku_lido", "")),
        "qtd_original": format_decimal(qtd_original),
        "tipo_qtd_original": str(row.get("tipo_qtd_original", "") or "UNIDADE").strip() or "UNIDADE",
        "fator_conversao": format_decimal(fator),
        "qtd_convertida": format_decimal(qtd_convertida),
        "qtd_final": format_decimal(qtd_convertida),
        "quantidade_lida": format_decimal(qtd_convertida),
        "status_conversao": status,
        "tipo_regra_conversao": regra.tipo_regra if regra else "MAPA_PRODUTOS",
        "regra_aplicada_conversao": regra_label,
        "origem_regra_conversao": origem,
        "prioridade_regra_conversao": str(regra.prioridade if regra else "MAPA"),
        "observacao_conversao": " | ".join(dict.fromkeys([v for v in observacoes if v])) or "Conversão aplicada com sucesso.",
    }


def aplicar_regras_conversao(df: pd.DataFrame, layout_config: Dict[str, Any]) -> tuple[pd.DataFrame, dict[str, Any]]:
    if df is None or df.empty:
        return df, {
            "layout_possui_conversao": False,
            "itens_convertidos": 0,
            "itens_nao_convertidos": 0,
            "itens_validar_conversao": 0,
            "itens_sem_conversao": 0,
            "alertas_conversao": [],
        }

    work = df.copy().fillna("")
    for col in [
        "centro_lido", "centro_referencia_conversao", "qtd_original", "tipo_qtd_original", "fator_conversao", "qtd_convertida", "qtd_final",
        "status_conversao", "tipo_regra_conversao", "regra_aplicada_conversao", "origem_regra_conversao", "observacao_conversao", "prioridade_regra_conversao",
    ]:
        if col not in work.columns:
            work[col] = ""

    regras_layout = regras_do_layout(layout_config)
    possui_conversao = any(regra.tipo not in {"SEM_CONVERSAO", "SEM_CONVERSÃO", "SEM CONVERSAO", "SEM CONVERSÃO"} for regra in regras_layout)
    alertas_conversao: list[str] = []

    for idx, row in work.iterrows():
        status_previo = str(row.get("status_conversao", "") or "").strip().upper()
        if status_previo in {STATUS_OK_CONVERTIDO, STATUS_OK_SEM_CONVERSAO}:
            # Parser dedicado já resolveu com segurança. Só completa campos de auditoria.
            if not str(work.at[idx, "qtd_original"]).strip():
                work.at[idx, "qtd_original"] = work.at[idx, "quantidade_lida"]
            if not str(work.at[idx, "qtd_final"]).strip():
                work.at[idx, "qtd_final"] = work.at[idx, "quantidade_lida"]
            if not str(work.at[idx, "tipo_qtd_original"]).strip():
                work.at[idx, "tipo_qtd_original"] = "CAIXARIA" if status_previo == STATUS_OK_SEM_CONVERSAO else "UNIDADE"
            regra_previa = _best_rule_for_row(row, layout_config, regras_layout)
            if not str(work.at[idx, "centro_referencia_conversao"]).strip():
                centro_ref, centro_ref_label = _centro_referencia_mapa(row, regra_previa)
                if centro_ref_label:
                    work.at[idx, "centro_referencia_conversao"] = centro_ref_label
            if not str(work.at[idx, "regra_aplicada_conversao"]).strip():
                tipo_previo = str(work.at[idx, "tipo_regra_conversao"] or "PARSER_DEDICADO")
                origem_previa = str(work.at[idx, "origem_regra_conversao"] or "PARSER_DEDICADO")
                prioridade = str(regra_previa.prioridade if regra_previa else "PARSER")
                regra_id = str(regra_previa.regra_id if regra_previa else "-")
                work.at[idx, "regra_aplicada_conversao"] = f"{regra_id} | {tipo_previo} | prioridade {prioridade}"
                if not str(work.at[idx, "origem_regra_conversao"]).strip():
                    work.at[idx, "origem_regra_conversao"] = origem_previa
            continue

        regra = _best_rule_for_row(row, layout_config, regras_layout)
        # Se a primeira regra for apenas uma base prioritária sem fator direto
        # e o parser ainda marcou alerta/validação, permite cair para a próxima
        # regra de mapa/fallback do próprio layout. Isso preserva o fluxo da
        # Rede Ítalo: base prioritária primeiro, mapa por centro depois.
        if regra and not regra.fator_decimal and regra.tipo not in {
            "SEM_CONVERSAO", "SEM_CONVERSÃO", "SEM CONVERSAO", "SEM CONVERSÃO",
            "MAPA_PRODUTOS", "MAPA_PRODUTOS_FALLBACK", "FALLBACK_MAPA", "MAPA",
        } and not regra.usa_mapa:
            regra_fallback = _best_rule_for_row(
                row,
                layout_config,
                [r for r in regras_layout if r.regra_id != regra.regra_id and _is_mapa_rule(r)],
            )
            if regra_fallback:
                regra = regra_fallback
        qtd_original = parse_decimal(row.get("quantidade_lida", "")) or Decimal("0")
        if not str(work.at[idx, "qtd_original"]).strip():
            work.at[idx, "qtd_original"] = format_decimal(qtd_original)

        if regra and regra.tipo in {"SEM_CONVERSAO", "SEM_CONVERSÃO", "SEM CONVERSAO", "SEM CONVERSÃO"}:
            work.at[idx, "tipo_qtd_original"] = str(row.get("tipo_qtd_original", "") or "CAIXARIA")
            work.at[idx, "qtd_final"] = format_decimal(qtd_original)
            work.at[idx, "status_conversao"] = STATUS_OK_SEM_CONVERSAO
            work.at[idx, "tipo_regra_conversao"] = regra.tipo_regra or "SEM_CONVERSAO"
            work.at[idx, "regra_aplicada_conversao"] = f"{regra.regra_id or '-'} | {regra.tipo_regra or 'SEM_CONVERSAO'} | prioridade {regra.prioridade}"
            work.at[idx, "origem_regra_conversao"] = regra.origem_regra or "CSV_REGRAS_CONVERSAO"
            work.at[idx, "observacao_conversao"] = regra.observacao or "Layout já trafega em caixaria; sem divisão por fator."
            work.at[idx, "prioridade_regra_conversao"] = str(regra.prioridade)
            continue

        if not possui_conversao and not regra:
            work.at[idx, "tipo_qtd_original"] = str(row.get("tipo_qtd_original", "") or "CAIXARIA")
            work.at[idx, "qtd_final"] = format_decimal(qtd_original)
            work.at[idx, "status_conversao"] = STATUS_OK_SEM_CONVERSAO
            work.at[idx, "tipo_regra_conversao"] = "SEM_CONVERSAO"
            work.at[idx, "regra_aplicada_conversao"] = "PADRAO_LAYOUT_SEM_REGRA_CONVERSAO"
            work.at[idx, "origem_regra_conversao"] = "PADRAO_LAYOUT_SEM_REGRA_CONVERSAO"
            work.at[idx, "observacao_conversao"] = "Nenhuma regra de conversão ativa para o layout; quantidade mantida como caixaria."
            continue

        if regra:
            fator = regra.fator_decimal
            if fator and fator > 0:
                updates = _aplicar_resultado(row, fator, regra, regra.origem_regra or "CSV_REGRAS_CONVERSAO", regra.sku_destino)
                centro_ref, centro_ref_label = _centro_referencia_mapa(row, regra)
                updates["centro_referencia_conversao"] = centro_ref_label
                if not str(work.at[idx, "centro_lido"]).strip():
                    updates["centro_lido"] = _linha_centro(row, regra) or centro_ref
                for col, value in updates.items():
                    work.at[idx, col] = value
                _log_conversao_item(
                    idx,
                    "[CONVERSAO] Item convertido | layout=%s | linha=%s | SKU=%s | EAN=%s | fator=%s | origem=%s | status=%s",
                    _layout_name(layout_config), idx + 2, updates.get("sku_lido", ""), row.get("ean_lido", ""), updates.get("fator_conversao", ""), updates.get("origem_regra_conversao", ""), updates.get("status_conversao", ""),
                )
                continue

            if _is_mapa_rule(regra):
                centro_operacional = _centro_operacional_lido(row)
                centro_ref, centro_ref_label = _centro_referencia_mapa(row, regra)
                produto, avisos = _find_mapa(row, centro_ref)
                if produto:
                    origem_mapa = (
                        f"{regra.origem_regra or 'MAPA_PRODUTOS'}::centro_ref={centro_ref_label}::"
                        f"{produto.origem}::linha_{produto.linha_mapa}"
                    )
                    updates = _aplicar_resultado(
                        row,
                        produto.apresentacao,
                        regra,
                        origem_mapa,
                        produto.sku,
                        avisos,
                    )
                    updates["centro_lido"] = centro_operacional or centro_ref
                    updates["centro_referencia_conversao"] = centro_ref_label
                    for col, value in updates.items():
                        work.at[idx, col] = value
                    _log_conversao_item(
                        idx,
                        "[CONVERSAO] Item convertido via mapa | layout=%s | linha=%s | centro_operacional=%s | centro_ref=%s | SKU=%s | EAN=%s | fator=%s | origem=%s | status=%s",
                        _layout_name(layout_config), idx + 2, centro_operacional, centro_ref_label, updates.get("sku_lido", ""), row.get("ean_lido", ""), updates.get("fator_conversao", ""), updates.get("origem_regra_conversao", ""), updates.get("status_conversao", ""),
                    )
                    for aviso in avisos:
                        alertas_conversao.append(f"Linha {idx + 2}: {aviso}")
                    continue
                else:
                    work.at[idx, "centro_lido"] = centro_operacional or centro_ref
                    work.at[idx, "centro_referencia_conversao"] = centro_ref_label
                    if avisos:
                        alertas_conversao.extend([f"Linha {idx + 2}: {aviso}" for aviso in avisos])

        # Se chegou aqui, deveria converter, mas não há fator confiável.
        obs = "Deveria converter unidade para caixaria, mas não encontrou regra/fator específico no CSV nem fallback confiável no mapa."
        work.at[idx, "tipo_qtd_original"] = str(row.get("tipo_qtd_original", "") or "UNIDADE")
        work.at[idx, "qtd_convertida"] = ""
        work.at[idx, "qtd_final"] = format_decimal(qtd_original)
        work.at[idx, "status_conversao"] = STATUS_ALERTA_NAO_CONVERTIDO
        work.at[idx, "tipo_regra_conversao"] = regra.tipo_regra if regra else "SEM_REGRA_CONVERSAO"
        work.at[idx, "regra_aplicada_conversao"] = (f"{regra.regra_id or '-'} | {regra.tipo_regra or 'REGRA_CSV'} | prioridade {regra.prioridade}" if regra else "SEM_REGRA_CONVERSAO")
        work.at[idx, "origem_regra_conversao"] = regra.origem_regra if regra else "CSV_REGRAS_CONVERSAO"
        work.at[idx, "observacao_conversao"] = obs
        if regra:
            centro_ref, centro_ref_label = _centro_referencia_mapa(row, regra)
            if centro_ref_label:
                work.at[idx, "centro_referencia_conversao"] = centro_ref_label
            if not str(work.at[idx, "centro_lido"]).strip():
                work.at[idx, "centro_lido"] = _linha_centro(row, regra) or centro_ref
        if regra:
            work.at[idx, "prioridade_regra_conversao"] = str(regra.prioridade)
        alerta = f"Linha {idx + 2}: {obs} | SKU={row.get('sku_lido', '')} | EAN={row.get('ean_lido', '')}"
        alertas_conversao.append(alerta)
        terminal_log.warning("[CONVERSAO] %s", alerta)

    counts = work["status_conversao"].astype(str).str.upper().value_counts().to_dict()
    resumo = {
        "layout_possui_conversao": possui_conversao,
        "itens_convertidos": int(counts.get(STATUS_OK_CONVERTIDO, 0)),
        "itens_nao_convertidos": int(counts.get(STATUS_ALERTA_NAO_CONVERTIDO, 0)),
        "itens_validar_conversao": int(counts.get(STATUS_VALIDAR_CONVERSAO, 0)),
        "itens_sem_conversao": int(counts.get(STATUS_OK_SEM_CONVERSAO, 0)),
        "alertas_conversao": sorted(set(alertas_conversao)),
    }
    terminal_log.info(
        "[CONVERSAO] Resumo | layout=%s | possui_conversao=%s | convertidos=%s | nao_convertidos=%s | validar=%s | sem_conversao=%s",
        _layout_name(layout_config),
        possui_conversao,
        resumo["itens_convertidos"],
        resumo["itens_nao_convertidos"],
        resumo["itens_validar_conversao"],
        resumo["itens_sem_conversao"],
    )
    return work, resumo
